"""
excel_export_enhanced.py — Phase 3 enhanced Excel export engine.

WHY THIS EXISTS
---------------
The validated exporter (``excel_export.py`` + ``excel_generator_core.py``)
produces the exact Daniel-format workbooks the coordinator signed off on. Those
files MUST stay byte-for-byte stable, so this module never touches them.

Instead, this is an *additive* layer that produces richer, analysis-oriented
workbooks on demand:

    * Color-coded group export (one stable color per group, legend sheet,
      conditional formatting for over-/under-subscribed groups, alternating
      rows, Loyola-branded headers, frozen panes, auto-filters, cell comments).
    * Five detailed analysis sheets (Room Utilization, Professor Workload,
      Student Placement, Time Slot Analysis heatmap, Quality Metrics) — several
      with native Excel charts.
    * Advanced formatting helpers (auto-filters, dropdown data-validation,
      named ranges, sheet protection with unlocked filter cells).

DESIGN NOTES
------------
* Every colour uses a colour-blind-safe palette (Okabe-Ito derived). Status is
  additionally encoded with a text label and a fill *pattern*, never colour
  alone, so the workbook stays accessible.
* All file access is routed through :mod:`app_paths` so the module behaves the
  same from source and inside the packaged .exe.
* Pure builder functions take DataFrames and return / mutate an openpyxl
  ``Workbook``; thin ``load_*`` helpers read the real optimisation outputs.

PUBLIC API
----------
    load_schedule(semester=None) -> pandas.DataFrame
    load_groups(semester=None) -> pandas.DataFrame
    load_kpi() -> dict
    load_solver_stats() -> list[dict]
    build_enhanced_workbook(schedule_df, groups_df, *, kpi=None,
                            solver_stats=None, options=None,
                            color_scheme='loyola') -> openpyxl.Workbook
    export_enhanced(semester=None, *, out_path=None, options=None,
                    color_scheme='loyola') -> dict
"""

from __future__ import annotations

import colorsys
import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

import app_paths

# ────────────────────────────────────────────────────────────
# Business thresholds (mirror pipeline.py — single source of truth here for the
# export layer; kept in sync with PREFERRED_GROUP_SIZE / MAX / MIN).
# ────────────────────────────────────────────────────────────
PREFERRED_GROUP_SIZE = 12
MAX_GROUP_SIZE = 15
MIN_GROUP_SIZE = 7

# ────────────────────────────────────────────────────────────
# Loyola brand colours (ARGB hex, no leading '#').
# ────────────────────────────────────────────────────────────
LOYOLA_NAVY = "FF003366"
LOYOLA_GOLD = "FFFFCC00"
WHITE = "FFFFFFFF"
BLACK = "FF000000"
LIGHT_ROW = "FFF2F5FA"   # very light navy tint for alternating rows

# ────────────────────────────────────────────────────────────
# Colour-blind-safe status palette (Okabe-Ito derived). We keep traffic-light
# SEMANTICS but pick accessible hues, and always pair colour with a label +
# fill pattern so meaning survives greyscale / colour-blindness.
# ────────────────────────────────────────────────────────────
STATUS_OK = "FF009E73"       # bluish green  -> optimal / good
STATUS_WARN = "FFE69F00"     # orange        -> under-utilised / watch
STATUS_CRIT = "FFD55E00"     # vermillion    -> over-subscribed / critical
STATUS_OK_LIGHT = "FFB7E4D3"
STATUS_WARN_LIGHT = "FFF7DFA8"
STATUS_CRIT_LIGHT = "FFF3C6AC"

# Qualitative colour-blind-safe base palette for per-group colours (Okabe-Ito
# 8-colour set, minus pure black). Extended deterministically via HSL when a
# workbook contains more groups than base colours.
_BASE_GROUP_PALETTE = [
    "E69F00",  # orange
    "56B4E9",  # sky blue
    "009E73",  # bluish green
    "F0E442",  # yellow
    "0072B2",  # blue
    "D55E00",  # vermillion
    "CC79A7",  # reddish purple
    "999999",  # grey
]

# ────────────────────────────────────────────────────────────
# Colour scheme presets for headers / accents.
# ────────────────────────────────────────────────────────────
COLOR_SCHEMES: Dict[str, Dict[str, str]] = {
    "loyola": {
        "header_bg": LOYOLA_NAVY,
        "header_fg": WHITE,
        "accent_bg": LOYOLA_GOLD,
        "accent_fg": BLACK,
        "alt_row": LIGHT_ROW,
    },
    "default": {
        "header_bg": "FF1F4E78",
        "header_fg": WHITE,
        "accent_bg": "FFDCE6F1",
        "accent_fg": BLACK,
        "alt_row": "FFF2F2F2",
    },
    "monochrome": {
        "header_bg": "FF333333",
        "header_fg": WHITE,
        "accent_bg": "FFCCCCCC",
        "accent_fg": BLACK,
        "alt_row": "FFF5F5F5",
    },
}

_THIN = Side(style="thin", color="FFBFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ════════════════════════════════════════════════════════════
# Options container
# ════════════════════════════════════════════════════════════
@dataclass
class ExportOptions:
    """User-tunable switches controlling which sheets get generated.

    Every flag defaults to ``True`` so a bare ``ExportOptions()`` yields the
    full enhanced workbook. The Streamlit UI maps its checkboxes onto these.
    """

    color_coded_groups: bool = True
    legend: bool = True
    room_utilization: bool = True
    professor_workload: bool = True
    student_placement: bool = True
    time_slot_analysis: bool = True
    quality_metrics: bool = True
    # formatting toggles
    conditional_formatting: bool = True
    auto_filter: bool = True
    freeze_panes: bool = True
    data_validation: bool = True
    named_ranges: bool = True
    cell_comments: bool = True
    protect_sheets: bool = False

    @classmethod
    def from_dict(cls, data: Optional[Dict[str, Any]]) -> "ExportOptions":
        """Build options from a plain dict, ignoring unknown keys."""
        if not data:
            return cls()
        known = {f for f in cls.__dataclass_fields__}  # type: ignore[attr-defined]
        return cls(**{k: bool(v) for k, v in data.items() if k in known})


# ════════════════════════════════════════════════════════════
# Data loading helpers (route through app_paths; tolerant of missing files)
# ════════════════════════════════════════════════════════════
def load_schedule(semester: Optional[int] = None):
    """Load the optimised schedule CSV as a DataFrame.

    Parameters
    ----------
    semester:
        If given (1 or 2), the frame is filtered to that semester.

    Returns
    -------
    pandas.DataFrame
        Empty frame if the file is absent (never raises for a missing file).
    """
    import pandas as pd

    path = app_paths.resolve_existing(
        "outputs/optimization/optimized_schedule_v5.csv"
    )
    if not path:
        return pd.DataFrame()
    df = pd.read_csv(path)
    df.columns = [c.lstrip("\ufeff") for c in df.columns]
    if semester is not None and "semester" in df.columns:
        df = df[df["semester"] == semester].copy()
    return df


def load_groups(semester: Optional[int] = None):
    """Load the per-student group composition CSV as a DataFrame."""
    import pandas as pd

    path = app_paths.resolve_existing(
        "outputs/optimization/group_composition.csv"
    )
    if not path:
        return pd.DataFrame()
    df = pd.read_csv(path)
    df.columns = [c.lstrip("\ufeff") for c in df.columns]
    if semester is not None and "semester" in df.columns:
        target = f"S{semester}"
        df = df[df["semester"].astype(str) == target].copy()
    return df


def load_kpi() -> Dict[str, Any]:
    """Load ``reports/kpi_report.json`` (empty dict if absent/invalid)."""
    path = app_paths.resolve_existing("reports/kpi_report.json")
    if not path:
        return {}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, ValueError):
        return {}


def load_solver_stats() -> List[Dict[str, Any]]:
    """Load ``reports/solver_stats.json`` (empty list if absent/invalid)."""
    path = app_paths.resolve_existing("reports/solver_stats.json")
    if not path:
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, list) else []
    except (OSError, ValueError):
        return []


# ════════════════════════════════════════════════════════════
# Colour helpers
# ════════════════════════════════════════════════════════════
def group_color(index: int) -> str:
    """Return a stable, colour-blind-safe ARGB fill for a group by index.

    The first :pydata:`_BASE_GROUP_PALETTE` colours are used directly; beyond
    that, extra distinct hues are generated deterministically in HSL space so a
    given index always maps to the same colour across every sheet.
    """
    n = len(_BASE_GROUP_PALETTE)
    if index < n:
        return "FF" + _BASE_GROUP_PALETTE[index]
    # Deterministic extension: rotate hue by the golden-ratio conjugate for
    # maximal separation, with mid lightness / moderate saturation.
    step = (index - n + 1) * 0.61803398875
    hue = step % 1.0
    r, g, b = colorsys.hls_to_rgb(hue, 0.62, 0.55)
    return "FF{:02X}{:02X}{:02X}".format(int(r * 255), int(g * 255), int(b * 255))


def build_group_color_map(groups: Sequence[Any]) -> Dict[Any, str]:
    """Map each distinct group key to a stable colour (consistent across sheets).

    Parameters
    ----------
    groups:
        Iterable of group keys (any hashable). Order is preserved; sorting is
        the caller's responsibility if a specific colour order is desired.
    """
    seen: List[Any] = []
    for g in groups:
        if g not in seen:
            seen.append(g)
    return {g: group_color(i) for i, g in enumerate(seen)}


def _fill(color: str, pattern: str = "solid") -> PatternFill:
    """Convenience PatternFill builder."""
    return PatternFill(start_color=color, end_color=color, fill_type=pattern)


def _scheme(name: str) -> Dict[str, str]:
    """Return a colour-scheme preset, falling back to Loyola."""
    return COLOR_SCHEMES.get((name or "loyola").lower(), COLOR_SCHEMES["loyola"])


def group_status(nb_students: int) -> str:
    """Classify a group size as 'over', 'under' or 'optimal'."""
    if nb_students > MAX_GROUP_SIZE:
        return "over"
    if nb_students < MIN_GROUP_SIZE:
        return "under"
    return "optimal"


_STATUS_STYLE = {
    "over": (STATUS_CRIT_LIGHT, "Over-subscribed"),
    "under": (STATUS_WARN_LIGHT, "Under-utilized"),
    "optimal": (STATUS_OK_LIGHT, "Optimal"),
}


# ════════════════════════════════════════════════════════════
# Low-level styling helpers
# ════════════════════════════════════════════════════════════
def style_header_row(
    ws: Worksheet,
    headers: Sequence[str],
    *,
    row: int = 1,
    scheme: str = "loyola",
    accent: bool = False,
) -> None:
    """Write and style a bold, branded header row.

    Parameters
    ----------
    accent:
        If True use the accent (gold) colours instead of the primary header
        colours — handy for sub-headers / section bands.
    """
    s = _scheme(scheme)
    bg = s["accent_bg"] if accent else s["header_bg"]
    fg = s["accent_fg"] if accent else s["header_fg"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, color=fg, size=11)
        cell.fill = _fill(bg)
        cell.alignment = _CENTER
        cell.border = _BORDER


def apply_alternating_rows(
    ws: Worksheet,
    *,
    first_row: int,
    last_row: int,
    n_cols: int,
    scheme: str = "loyola",
) -> None:
    """Shade every other data row for readability (skips already-filled cells)."""
    tint = _scheme(scheme)["alt_row"]
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill is None or cell.fill.fill_type != "solid":
                    cell.fill = _fill(tint)


def autosize_columns(ws: Worksheet, *, min_width: int = 8, max_width: int = 48) -> None:
    """Approximate auto-fit of column widths from cell content length."""
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max(len(part) for part in str(cell.value).split("\n"))
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(
            min_width, min(max_width, w + 2)
        )


def add_comment(cell, text: str, author: str = "LabScheduling") -> None:
    """Attach a cell comment (no-op on falsy text)."""
    if text:
        cell.comment = Comment(text, author)


def freeze_header(ws: Worksheet, cell: str = "A2") -> None:
    """Freeze panes above/left of ``cell`` (default: freeze the header row)."""
    ws.freeze_panes = cell


def enable_auto_filter(ws: Worksheet, last_col: int, last_row: int, first_row: int = 1) -> None:
    """Turn on an auto-filter spanning the given rectangle."""
    if last_col < 1 or last_row < first_row:
        return
    ref = f"A{first_row}:{get_column_letter(last_col)}{last_row}"
    ws.auto_filter.ref = ref


def protect_sheet_keep_filters(ws: Worksheet, password: str = "loyola") -> None:
    """Protect a sheet but leave auto-filter and sort usable."""
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.autoFilter = False
    ws.protection.sort = False


# ════════════════════════════════════════════════════════════
# Room capacity proxy (no explicit capacity data in the pipeline outputs, so we
# derive a defensible proxy: the max students ever scheduled in a room).
# ════════════════════════════════════════════════════════════
def derive_room_capacity(schedule_df) -> Dict[str, int]:
    """Return {room -> capacity proxy} = max nb_students observed in that room."""
    caps: Dict[str, int] = {}
    if schedule_df is None or len(schedule_df) == 0:
        return caps
    if "lab_rooms" not in schedule_df.columns:
        return caps
    grouped = schedule_df.groupby("lab_rooms")["nb_students"].max()
    for room, cap in grouped.items():
        caps[str(room)] = int(cap)
    return caps


# ════════════════════════════════════════════════════════════
# Sheet builder: colour-coded groups
# ════════════════════════════════════════════════════════════
_GROUP_HEADERS = [
    "Semester", "Subject", "Program", "Group", "Sessions",
    "Students", "Status", "Day", "Time Block", "Room", "Professor",
]


def _group_summary(schedule_df):
    """Aggregate the raw schedule into one row per (semester, subject, group).

    Returns a list of dicts with the columns used by the group sheet and the
    legend. Uses the first session's day/block/room/professor as the group's
    representative slot (all sessions of a group share the same slot).
    """
    import pandas as pd

    if schedule_df is None or len(schedule_df) == 0:
        return []
    rows: List[Dict[str, Any]] = []
    keys = ["semester", "subject", "grupo"]
    for (sem, subj, grp), block in schedule_df.groupby(keys):
        first = block.iloc[0]
        rows.append(
            {
                "semester": int(sem),
                "subject": str(subj),
                "program": str(first.get("program", "")),
                "grupo": int(grp),
                "sessions": int(len(block)),
                "students": int(first.get("nb_students", 0)),
                "day": str(first.get("day", "")),
                "time_block": str(first.get("time_block", "")),
                "room": str(first.get("lab_rooms", "")),
                "professor": str(first.get("professor", "")),
            }
        )
    rows.sort(key=lambda r: (r["semester"], r["subject"], r["grupo"]))
    return rows


def build_color_coded_groups_sheet(
    wb: Workbook,
    schedule_df,
    *,
    options: ExportOptions,
    scheme: str = "loyola",
    color_map: Optional[Dict[Any, str]] = None,
) -> Dict[Any, str]:
    """Build the "Groups" sheet with one stable colour per group.

    Returns the colour map actually used so the legend sheet can reuse it.
    """
    summary = _group_summary(schedule_df)
    ws = wb.create_sheet("Groups")
    style_header_row(ws, _GROUP_HEADERS, scheme=scheme)

    keys = [(r["semester"], r["subject"], r["grupo"]) for r in summary]
    if color_map is None:
        color_map = build_group_color_map(keys)

    caps = derive_room_capacity(schedule_df)

    r = 2
    for row in summary:
        key = (row["semester"], row["subject"], row["grupo"])
        color = color_map.get(key, "FFFFFFFF")
        status = group_status(row["students"])
        status_fill, status_label = _STATUS_STYLE[status]

        values = [
            f"S{row['semester']}", row["subject"], row["program"],
            row["grupo"], row["sessions"], row["students"], status_label,
            row["day"], row["time_block"], row["room"], row["professor"],
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            cell.alignment = _LEFT if c in (2, 10, 11) else _CENTER

        # Colour the Group cell with the group's stable colour.
        gcell = ws.cell(row=r, column=4)
        gcell.fill = _fill(color)
        gcell.font = Font(bold=True, color=_readable_fg(color))

        # Status cell: colour + pattern (accessible; not colour-alone).
        scell = ws.cell(row=r, column=7)
        if options.conditional_formatting:
            pattern = "solid" if status == "optimal" else "lightUp"
            scell.fill = PatternFill(
                start_color=status_fill, end_color=status_fill, fill_type=pattern
            )

        # Cell comments: professor + room-capacity proxy.
        if options.cell_comments:
            add_comment(
                ws.cell(row=r, column=11),
                f"Professor: {row['professor']}",
            )
            cap = caps.get(row["room"])
            if cap:
                add_comment(
                    ws.cell(row=r, column=10),
                    f"Room: {row['room']}\nObserved max occupancy: {cap} students",
                )
        r += 1

    last_row = r - 1
    if options.conditional_formatting and last_row >= 2:
        _apply_group_size_rules(ws, first_row=2, last_row=last_row, size_col=6)
    if last_row >= 2:
        apply_alternating_rows(
            ws, first_row=2, last_row=last_row, n_cols=len(_GROUP_HEADERS),
            scheme=scheme,
        )
    if options.freeze_panes:
        freeze_header(ws, "A2")
    if options.auto_filter:
        enable_auto_filter(ws, len(_GROUP_HEADERS), last_row)
    autosize_columns(ws)
    return color_map


def _apply_group_size_rules(ws: Worksheet, *, first_row: int, last_row: int, size_col: int) -> None:
    """Conditional formatting on the Students column: over/under/optimal."""
    col = get_column_letter(size_col)
    rng = f"{col}{first_row}:{col}{last_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="greaterThan", formula=[str(MAX_GROUP_SIZE)],
                   fill=_fill(STATUS_CRIT_LIGHT)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="lessThan", formula=[str(MIN_GROUP_SIZE)],
                   fill=_fill(STATUS_WARN_LIGHT)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="between",
                   formula=[str(MIN_GROUP_SIZE), str(MAX_GROUP_SIZE)],
                   fill=_fill(STATUS_OK_LIGHT)),
    )


def _readable_fg(argb: str) -> str:
    """Pick black or white text for legibility over a given background colour."""
    try:
        r = int(argb[2:4], 16)
        g = int(argb[4:6], 16)
        b = int(argb[6:8], 16)
    except (ValueError, IndexError):
        return BLACK
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return BLACK if luminance > 0.6 else WHITE


# ════════════════════════════════════════════════════════════
# Sheet builder: legend
# ════════════════════════════════════════════════════════════
def build_legend_sheet(
    wb: Workbook,
    schedule_df,
    color_map: Dict[Any, str],
    *,
    scheme: str = "loyola",
) -> None:
    """Build a legend sheet documenting group colours and status semantics."""
    ws = wb.create_sheet("Legend")

    # Section 1 — group colours.
    style_header_row(ws, ["Group (Sem / Subject / #)", "Colour", "Students", "Status"],
                     scheme=scheme)
    summary = {(_r["semester"], _r["subject"], _r["grupo"]): _r
               for _r in _group_summary(schedule_df)}
    r = 2
    for key, color in color_map.items():
        info = summary.get(key, {})
        label = f"S{key[0]} / {key[1]} / G{key[2]}"
        ws.cell(row=r, column=1, value=label).border = _BORDER
        swatch = ws.cell(row=r, column=2, value="")
        swatch.fill = _fill(color)
        swatch.border = _BORDER
        n = info.get("students", 0)
        ws.cell(row=r, column=3, value=n).border = _BORDER
        _, status_label = _STATUS_STYLE[group_status(n)]
        ws.cell(row=r, column=4, value=status_label).border = _BORDER
        r += 1

    # Section 2 — status semantics key.
    r += 1
    ws.cell(row=r, column=1, value="Status key").font = Font(bold=True, size=12)
    r += 1
    style_header_row(ws, ["Status", "Colour", "Meaning", "Rule"], row=r, scheme=scheme)
    r += 1
    legend_rows = [
        ("Optimal", STATUS_OK_LIGHT,
         "Group size within target band",
         f"{MIN_GROUP_SIZE} <= students <= {MAX_GROUP_SIZE}"),
        ("Under-utilized", STATUS_WARN_LIGHT,
         "Group smaller than minimum", f"students < {MIN_GROUP_SIZE}"),
        ("Over-subscribed", STATUS_CRIT_LIGHT,
         "Group exceeds room / policy max", f"students > {MAX_GROUP_SIZE}"),
    ]
    for name, fill, meaning, rule in legend_rows:
        ws.cell(row=r, column=1, value=name).border = _BORDER
        sc = ws.cell(row=r, column=2, value="")
        sc.fill = _fill(fill)
        sc.border = _BORDER
        ws.cell(row=r, column=3, value=meaning).border = _BORDER
        ws.cell(row=r, column=4, value=rule).border = _BORDER
        r += 1

    ws.cell(row=r + 1, column=1,
            value="Colours use a colour-blind-safe palette; status is also "
                  "encoded by label and fill pattern.")
    autosize_columns(ws)
    if scheme:
        freeze_header(ws, "A2")


# ════════════════════════════════════════════════════════════
# Sheet builder: room utilization (with bar chart)
# ════════════════════════════════════════════════════════════
def build_room_utilization_sheet(
    wb: Workbook,
    schedule_df,
    *,
    options: ExportOptions,
    scheme: str = "loyola",
) -> None:
    """Room utilisation: sessions & students per room, with a bar chart.

    Utilisation % = sessions in room / total scheduled sessions. Conditional
    formatting: >80% green, 50-80% amber, <50% red (colour-blind-safe hues).
    """
    ws = wb.create_sheet("Room Utilization")
    headers = ["Room", "Sessions", "Total Students", "Avg Students/Session",
               "Utilization %"]
    style_header_row(ws, headers, scheme=scheme)

    rows: List[Dict[str, Any]] = []
    if schedule_df is not None and len(schedule_df) and "lab_rooms" in schedule_df.columns:
        total_sessions = len(schedule_df)
        grp = schedule_df.groupby("lab_rooms")
        for room, block in grp:
            sess = int(len(block))
            students = int(block["nb_students"].sum())
            rows.append({
                "room": str(room),
                "sessions": sess,
                "students": students,
                "avg": round(students / sess, 1) if sess else 0.0,
                "util": round(100.0 * sess / total_sessions, 1) if total_sessions else 0.0,
            })
        rows.sort(key=lambda x: x["sessions"], reverse=True)

    r = 2
    for row in rows:
        for c, val in enumerate(
            [row["room"], row["sessions"], row["students"], row["avg"], row["util"]],
            start=1,
        ):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            cell.alignment = _LEFT if c == 1 else _CENTER
        r += 1
    last_row = r - 1

    if options.conditional_formatting and last_row >= 2:
        rng = f"E2:E{last_row}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThanOrEqual", formula=["80"],
                            fill=_fill(STATUS_OK_LIGHT)))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="between", formula=["50", "79.999"],
                            fill=_fill(STATUS_WARN_LIGHT)))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["50"],
                            fill=_fill(STATUS_CRIT_LIGHT)))

    if last_row >= 2:
        apply_alternating_rows(ws, first_row=2, last_row=last_row,
                               n_cols=len(headers), scheme=scheme)
        _add_bar_chart(ws, title="Sessions per room", cat_col=1, val_col=2,
                       first_row=2, last_row=last_row, anchor="H2")
    if options.freeze_panes:
        freeze_header(ws, "A2")
    if options.auto_filter:
        enable_auto_filter(ws, len(headers), last_row)
    autosize_columns(ws)


def _add_bar_chart(
    ws: Worksheet,
    *,
    title: str,
    cat_col: int,
    val_col: int,
    first_row: int,
    last_row: int,
    anchor: str,
) -> None:
    """Insert a simple bar chart referencing an existing data block."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 8
    chart.width = 16
    data = Reference(ws, min_col=val_col, min_row=first_row - 1, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)


# ════════════════════════════════════════════════════════════
# Sheet builder: professor workload (with chart)
# ════════════════════════════════════════════════════════════
def build_professor_workload_sheet(
    wb: Workbook,
    schedule_df,
    *,
    options: ExportOptions,
    scheme: str = "loyola",
) -> None:
    """Professor workload: sessions, subjects, students per professor + chart."""
    ws = wb.create_sheet("Professor Workload")
    headers = ["Professor", "Sessions", "Subjects", "Groups", "Total Students"]
    style_header_row(ws, headers, scheme=scheme)

    rows: List[Dict[str, Any]] = []
    if schedule_df is not None and len(schedule_df) and "professor" in schedule_df.columns:
        for prof, block in schedule_df.groupby("professor"):
            rows.append({
                "professor": str(prof),
                "sessions": int(len(block)),
                "subjects": int(block["subject"].nunique()),
                "groups": int(block.groupby(["subject", "grupo"]).ngroups),
                "students": int(block["nb_students"].sum()),
            })
        rows.sort(key=lambda x: x["sessions"], reverse=True)

    r = 2
    for row in rows:
        for c, val in enumerate(
            [row["professor"], row["sessions"], row["subjects"],
             row["groups"], row["students"]],
            start=1,
        ):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            cell.alignment = _LEFT if c == 1 else _CENTER
        r += 1
    last_row = r - 1

    if last_row >= 2:
        apply_alternating_rows(ws, first_row=2, last_row=last_row,
                               n_cols=len(headers), scheme=scheme)
        _add_bar_chart(ws, title="Sessions per professor", cat_col=1, val_col=2,
                       first_row=2, last_row=last_row, anchor="G2")
    if options.freeze_panes:
        freeze_header(ws, "A2")
    if options.auto_filter:
        enable_auto_filter(ws, len(headers), last_row)
    autosize_columns(ws)


# ════════════════════════════════════════════════════════════
# Sheet builder: student placement
# ════════════════════════════════════════════════════════════
def build_student_placement_sheet(
    wb: Workbook,
    schedule_df,
    kpi: Dict[str, Any],
    *,
    options: ExportOptions,
    scheme: str = "loyola",
) -> None:
    """Per-subject placement view with green/amber/red status.

    Status per subject uses group sizing: red if any group over-subscribed,
    amber if any under-utilised, green otherwise. A summary band at the top
    reflects the global placement KPI.
    """
    ws = wb.create_sheet("Student Placement")

    placement = (kpi or {}).get("placement", {})
    ws.cell(row=1, column=1, value="Global placement").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value="Enrolled")
    ws.cell(row=2, column=2, value=placement.get("enrolled", "n/a"))
    ws.cell(row=3, column=1, value="Placed")
    ws.cell(row=3, column=2, value=placement.get("placed", "n/a"))
    ws.cell(row=4, column=1, value="Unplaced")
    ws.cell(row=4, column=2, value=placement.get("unplaced", "n/a"))
    ws.cell(row=5, column=1, value="Placement %")
    pct_cell = ws.cell(row=5, column=2, value=placement.get("placement_pct", "n/a"))
    if isinstance(placement.get("placement_pct"), (int, float)):
        pct = placement["placement_pct"]
        fill = (STATUS_OK_LIGHT if pct >= 99 else
                STATUS_WARN_LIGHT if pct >= 90 else STATUS_CRIT_LIGHT)
        pct_cell.fill = _fill(fill)

    header_row = 7
    headers = ["Semester", "Subject", "Groups", "Students",
               "Min Group", "Max Group", "Status"]
    style_header_row(ws, headers, row=header_row, scheme=scheme)

    rows: List[Dict[str, Any]] = []
    if schedule_df is not None and len(schedule_df):
        for (sem, subj), block in schedule_df.groupby(["semester", "subject"]):
            sizes = block.groupby("grupo")["nb_students"].first()
            n_over = int((sizes > MAX_GROUP_SIZE).sum())
            n_under = int((sizes < MIN_GROUP_SIZE).sum())
            status = ("over" if n_over else "under" if n_under else "optimal")
            rows.append({
                "semester": int(sem),
                "subject": str(subj),
                "groups": int(len(sizes)),
                "students": int(sizes.sum()),
                "min": int(sizes.min()),
                "max": int(sizes.max()),
                "status": status,
            })
        rows.sort(key=lambda x: (x["semester"], x["subject"]))

    r = header_row + 1
    for row in rows:
        status_fill, status_label = _STATUS_STYLE[row["status"]]
        values = [f"S{row['semester']}", row["subject"], row["groups"],
                  row["students"], row["min"], row["max"], status_label]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            cell.alignment = _LEFT if c == 2 else _CENTER
        if options.conditional_formatting:
            pattern = "solid" if row["status"] == "optimal" else "lightUp"
            ws.cell(row=r, column=7).fill = PatternFill(
                start_color=status_fill, end_color=status_fill, fill_type=pattern)
        r += 1
    last_row = r - 1

    if last_row > header_row:
        apply_alternating_rows(ws, first_row=header_row + 1, last_row=last_row,
                               n_cols=len(headers), scheme=scheme)
    if options.freeze_panes:
        freeze_header(ws, f"A{header_row + 1}")
    if options.auto_filter:
        enable_auto_filter(ws, len(headers), last_row, first_row=header_row)
    autosize_columns(ws)


# ════════════════════════════════════════════════════════════
# Sheet builder: time slot analysis (heatmap)
# ════════════════════════════════════════════════════════════
DAYS_ORDER = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]


def build_time_slot_analysis_sheet(
    wb: Workbook,
    schedule_df,
    *,
    options: ExportOptions,
    scheme: str = "loyola",
) -> None:
    """Days x Time-blocks heatmap of session counts (aggregated over rooms).

    A per-room breakdown table follows the heatmap. Heatmap cells are shaded on
    a colour-blind-safe intensity ramp (light -> Loyola navy).
    """
    ws = wb.create_sheet("Time Slot Analysis")
    ws.cell(row=1, column=1, value="Sessions by day and time block "
                                   "(all rooms)").font = Font(bold=True, size=12)

    import pandas as pd

    blocks: List[str] = []
    counts: Dict[Any, int] = {}
    if schedule_df is not None and len(schedule_df):
        blocks = sorted(schedule_df["time_block"].dropna().astype(str).unique())
        for (day, blk), block in schedule_df.groupby(["day", "time_block"]):
            counts[(str(day), str(blk))] = int(len(block))

    header_row = 2
    style_header_row(ws, ["Time Block"] + DAYS_ORDER, row=header_row, scheme=scheme)
    max_count = max(counts.values()) if counts else 0
    r = header_row + 1
    for blk in blocks:
        ws.cell(row=r, column=1, value=blk).border = _BORDER
        for ci, day in enumerate(DAYS_ORDER, start=2):
            val = counts.get((day, blk), 0)
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _BORDER
            cell.alignment = _CENTER
            if val and max_count:
                cell.fill = _heat_fill(val / max_count)
                cell.font = Font(color=_readable_fg(_heat_argb(val / max_count)))
        r += 1
    last_heat_row = r - 1

    # Per-room breakdown table.
    r += 1
    ws.cell(row=r, column=1, value="Per-room session count").font = Font(bold=True, size=12)
    r += 1
    room_header = r
    style_header_row(ws, ["Room", "Day", "Time Block", "Sessions", "Students"],
                     row=room_header, scheme=scheme)
    r += 1
    if schedule_df is not None and len(schedule_df) and "lab_rooms" in schedule_df.columns:
        agg = (schedule_df.groupby(["lab_rooms", "day", "time_block"])
               .agg(sessions=("session", "count"),
                    students=("nb_students", "sum"))
               .reset_index())
        agg = agg.sort_values(["lab_rooms", "day", "time_block"])
        for _, row in agg.iterrows():
            for c, val in enumerate(
                [str(row["lab_rooms"]), str(row["day"]), str(row["time_block"]),
                 int(row["sessions"]), int(row["students"])],
                start=1,
            ):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = _BORDER
                cell.alignment = _LEFT if c in (1, 2, 3) else _CENTER
            r += 1
    last_room_row = r - 1

    if options.freeze_panes:
        freeze_header(ws, f"A{header_row + 1}")
    if options.auto_filter and last_room_row > room_header:
        enable_auto_filter(ws, 5, last_room_row, first_row=room_header)
    autosize_columns(ws)


def _heat_argb(fraction: float) -> str:
    """Interpolate a light tint -> Loyola navy for a 0..1 intensity."""
    fraction = max(0.0, min(1.0, fraction))
    # light blue (230,238,248) -> navy (0,51,102)
    r = int(230 + (0 - 230) * fraction)
    g = int(238 + (51 - 238) * fraction)
    b = int(248 + (102 - 248) * fraction)
    return "FF{:02X}{:02X}{:02X}".format(r, g, b)


def _heat_fill(fraction: float) -> PatternFill:
    argb = _heat_argb(fraction)
    return _fill(argb)


# ════════════════════════════════════════════════════════════
# Sheet builder: quality metrics
# ════════════════════════════════════════════════════════════
def build_quality_metrics_sheet(
    wb: Workbook,
    kpi: Dict[str, Any],
    solver_stats: List[Dict[str, Any]],
    *,
    options: ExportOptions,
    scheme: str = "loyola",
) -> None:
    """KPI + solver quality dashboard with traffic-light colours."""
    ws = wb.create_sheet("Quality Metrics")
    ws.cell(row=1, column=1, value="Quality metrics").font = Font(bold=True, size=14)

    r = 3
    style_header_row(ws, ["Metric", "Value", "Status"], row=r, scheme=scheme)
    r += 1

    kpi = kpi or {}
    groups = kpi.get("groups", {})
    placement = kpi.get("placement", {})

    def _emit(metric: str, value: Any, status: Optional[str]) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=metric).border = _BORDER
        ws.cell(row=r, column=2, value=value).border = _BORDER
        scell = ws.cell(row=r, column=3, value="")
        scell.border = _BORDER
        if status and options.conditional_formatting:
            fill, label = _STATUS_STYLE[status]
            pattern = "solid" if status == "optimal" else "lightUp"
            scell.value = label
            scell.fill = PatternFill(start_color=fill, end_color=fill, fill_type=pattern)
        r += 1

    total_groups = groups.get("total", 0)
    overflow = groups.get("overflow", 0)
    overflow_status = ("optimal" if overflow == 0 else
                       "under" if total_groups and overflow / total_groups < 0.15
                       else "over")
    _emit("Total groups", total_groups, None)
    _emit("Over-subscribed groups", overflow, overflow_status)
    _emit("Group size min", groups.get("size_min", "n/a"), None)
    _emit("Group size max", groups.get("size_max", "n/a"), None)
    _emit("Group size mean", groups.get("size_mean", "n/a"), None)

    pct = placement.get("placement_pct")
    placement_status = None
    if isinstance(pct, (int, float)):
        placement_status = ("optimal" if pct >= 99 else
                            "under" if pct >= 90 else "over")
    _emit("Placement %", pct if pct is not None else "n/a", placement_status)
    _emit("Enrolled", placement.get("enrolled", "n/a"), None)
    _emit("Unplaced", placement.get("unplaced", "n/a"),
          "optimal" if placement.get("unplaced") == 0 else "over")
    _emit("Total sessions", kpi.get("total_sessions", "n/a"), None)

    # Solver runs table.
    r += 1
    ws.cell(row=r, column=1, value="Solver runs").font = Font(bold=True, size=12)
    r += 1
    solver_header = r
    style_header_row(
        ws,
        ["Semester", "Status", "Sessions", "Wall time (s)", "Objective",
         "Best bound", "Gap"],
        row=solver_header, scheme=scheme,
    )
    r += 1
    for run in (solver_stats or []):
        status = str(run.get("status", ""))
        status_key = "optimal" if status == "OPTIMAL" else (
            "under" if status == "FEASIBLE" else "over")
        values = [
            run.get("label", f"S{run.get('semester', '')}"),
            status,
            run.get("n_sessions", "n/a"),
            run.get("wall_time_s", "n/a"),
            run.get("objective", "n/a"),
            run.get("best_bound", "n/a"),
            run.get("gap", "n/a"),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            cell.alignment = _CENTER
        if options.conditional_formatting:
            fill, _ = _STATUS_STYLE[status_key]
            ws.cell(row=r, column=2).fill = _fill(fill)
        r += 1

    ws.cell(row=r + 1, column=1,
            value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    if options.freeze_panes:
        freeze_header(ws, "A4")
    autosize_columns(ws)


# ════════════════════════════════════════════════════════════
# Advanced formatting: data validation, named ranges
# ════════════════════════════════════════════════════════════
def add_dropdown_validations(ws: Worksheet, schedule_df) -> None:
    """Attach dropdown data-validation lists to a helper 'Filters' block.

    The validations live in a dedicated top-right area (columns N+) so they do
    not interfere with the data grid; each is a convenience picker the user can
    reference. Values are drawn from the schedule's distinct entries.
    """
    if schedule_df is None or len(schedule_df) == 0:
        return
    specs = [
        ("Semester", "semester"),
        ("Subject", "subject"),
        ("Professor", "professor"),
        ("Room", "lab_rooms"),
        ("Day", "day"),
        ("Block", "time_block"),
    ]
    col = 14  # column N
    for label, field_name in specs:
        if field_name not in schedule_df.columns:
            continue
        values = sorted({str(v) for v in schedule_df[field_name].dropna().unique()})
        # Excel data-validation list literal cap ~255 chars; truncate safely.
        joined = ",".join(values)
        if len(joined) > 250:
            joined = joined[:250].rsplit(",", 1)[0]
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=label).font = Font(bold=True)
        dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
        dv.prompt = f"Pick a {label}"
        dv.promptTitle = label
        ws.add_data_validation(dv)
        dv.add(f"{letter}2")
        col += 1


def add_named_ranges(wb: Workbook, ws_title: str, headers: Sequence[str],
                     last_row: int) -> None:
    """Define a workbook-level named range per column of a data table."""
    from openpyxl.workbook.defined_name import DefinedName

    if last_row < 2:
        return
    for idx, header in enumerate(headers, start=1):
        safe = _safe_name(header)
        if not safe:
            continue
        letter = get_column_letter(idx)
        ref = f"'{ws_title}'!${letter}$2:${letter}${last_row}"
        name = f"{_safe_name(ws_title)}_{safe}"
        if name in wb.defined_names:
            continue
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def _safe_name(text: str) -> str:
    """Sanitise a string into a valid Excel defined-name token."""
    out = "".join(ch if ch.isalnum() else "_" for ch in str(text))
    out = out.strip("_")
    if out and out[0].isdigit():
        out = "_" + out
    return out


# ════════════════════════════════════════════════════════════
# Orchestration
# ════════════════════════════════════════════════════════════
def _add_cover_sheet(wb: Workbook, *, scheme: str, semester: Optional[int]) -> None:
    """Add a branded cover / summary sheet as the first tab."""
    ws = wb.create_sheet("Overview", 0)
    s = _scheme(scheme)
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="Universidad Loyola — Lab Scheduling")
    title.font = Font(bold=True, size=16, color=s["header_fg"])
    title.fill = _fill(s["header_bg"])
    title.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    scope = "All semesters" if semester is None else f"Semester {semester}"
    info = [
        ("Report", "Enhanced schedule export"),
        ("Scope", scope),
        ("Colour scheme", scheme),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40


def build_enhanced_workbook(
    schedule_df,
    groups_df=None,
    *,
    kpi: Optional[Dict[str, Any]] = None,
    solver_stats: Optional[List[Dict[str, Any]]] = None,
    options: Optional[ExportOptions] = None,
    color_scheme: str = "loyola",
    semester: Optional[int] = None,
) -> Workbook:
    """Assemble the full enhanced workbook and return it.

    Parameters
    ----------
    schedule_df:
        Optimised schedule DataFrame (may be empty — sheets degrade gracefully).
    groups_df:
        Per-student composition (currently informational; reserved for future
        per-student sheets). Accepted for API completeness.
    kpi, solver_stats:
        Loaded report dicts; fetched via ``load_*`` when omitted.
    options:
        :class:`ExportOptions` controlling which sheets/formatting to include.
    color_scheme:
        One of ``COLOR_SCHEMES`` ('loyola', 'default', 'monochrome').
    """
    options = options or ExportOptions()
    kpi = kpi if kpi is not None else load_kpi()
    solver_stats = solver_stats if solver_stats is not None else load_solver_stats()

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    _add_cover_sheet(wb, scheme=color_scheme, semester=semester)

    color_map: Dict[Any, str] = {}
    if options.color_coded_groups:
        color_map = build_color_coded_groups_sheet(
            wb, schedule_df, options=options, scheme=color_scheme)
        if options.data_validation:
            add_dropdown_validations(wb["Groups"], schedule_df)
        if options.named_ranges:
            add_named_ranges(wb, "Groups", _GROUP_HEADERS,
                             wb["Groups"].max_row)
        if options.protect_sheets:
            protect_sheet_keep_filters(wb["Groups"])

    if options.legend and color_map:
        build_legend_sheet(wb, schedule_df, color_map, scheme=color_scheme)

    if options.room_utilization:
        build_room_utilization_sheet(wb, schedule_df, options=options,
                                     scheme=color_scheme)
    if options.professor_workload:
        build_professor_workload_sheet(wb, schedule_df, options=options,
                                       scheme=color_scheme)
    if options.student_placement:
        build_student_placement_sheet(wb, schedule_df, kpi, options=options,
                                      scheme=color_scheme)
    if options.time_slot_analysis:
        build_time_slot_analysis_sheet(wb, schedule_df, options=options,
                                       scheme=color_scheme)
    if options.quality_metrics:
        build_quality_metrics_sheet(wb, kpi, solver_stats, options=options,
                                    scheme=color_scheme)

    # Guarantee at least one visible sheet.
    if not wb.sheetnames:
        wb.create_sheet("Overview")
    return wb


def export_enhanced(
    semester: Optional[int] = None,
    *,
    out_path: Optional[str] = None,
    options: Optional[Dict[str, Any] | ExportOptions] = None,
    color_scheme: str = "loyola",
) -> Dict[str, Any]:
    """Load real data, build the enhanced workbook and save it.

    Returns a result dict ``{ok, file, error}``. Never raises for expected
    conditions (missing inputs) — reports them in ``error`` instead.
    """
    try:
        opts = options if isinstance(options, ExportOptions) else ExportOptions.from_dict(options)
        schedule_df = load_schedule(semester)
        groups_df = load_groups(semester)
        kpi = load_kpi()
        solver_stats = load_solver_stats()

        wb = build_enhanced_workbook(
            schedule_df, groups_df, kpi=kpi, solver_stats=solver_stats,
            options=opts, color_scheme=color_scheme, semester=semester,
        )

        if out_path is None:
            tag = "all" if semester is None else f"S{semester}"
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = app_paths.workspace_path(
                "outputs", "optimization",
                f"enhanced_export_{tag}_{stamp}.xlsx",
            )
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb.save(out_path)
        return {"ok": True, "file": out_path, "error": None}
    except Exception as exc:  # pragma: no cover - defensive
        import traceback
        return {"ok": False, "file": None, "error": str(exc),
                "trace": traceback.format_exc()}
