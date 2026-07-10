"""
traceability_sheet.py
======================

Build a "Parameters" worksheet that is appended to every generated workbook.

Purpose (Point 10 — "everything must be visible in the Excel output")
---------------------------------------------------------------------
For every optimization run, this sheet surfaces — inside the deliverable Excel
itself — the *actual* configuration and constraints the solver used:

* Run metadata (generation timestamp, solver status / time / objective / gap).
* Solver profile and the soft-constraint weights that were active.
* Global generation parameters (group sizes, lab maxima, calendar).
* The per-subject configuration table (sessions, capacity, week window, rooms).
* Teacher availability / rule counts.

This lets anyone opening the workbook confirm which settings produced the plan
— proving the system honoured the configuration, without opening the app.

Design goals
------------
* Purely additive: it only *adds* one sheet to a workbook handed to it.
* Read-only: it reads ``config/applied_config.json`` (written by the pipeline
  at the end of every run) plus ``reports/solver_stats.json`` and the live
  ``solver_config``. Every source is best-effort; missing sources degrade
  gracefully and never break Excel generation.
* Content in English (matches the recently anglicised Validation sheet).
* Loyola visual identity (navy + gold), no bright green/red.

Usage
-----
    import traceability_sheet
    traceability_sheet.build_traceability_sheet(wb)          # auto-loads sources
    traceability_sheet.build_traceability_sheet(wb, applied=cfg, stats=stats)
"""

from __future__ import annotations

import json
import os
from datetime import datetime
from typing import Any, Optional

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Loyola palette (shared with validation_sheet) ───────────────────────────
NAVY = "1B3A6F"
NAVY_DEEP = "0F2344"
GOLD = "FFCC00"
CYAN = "6FAED9"
GREY_BG = "F2F5FA"
BAND_BG = "E4F0F6"
WHITE = "FFFFFF"
INK = "1B2A44"

_THIN = Side(style="thin", color="C7D2E4")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FONT = Font(name="Calibri", size=15, bold=True, color=WHITE)
_SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=INK)
_VALUE_FONT = Font(name="Calibri", size=10, color=INK)
_HELP_FONT = Font(name="Calibri", size=9, italic=True, color="5B6B84")
_HDR_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)

_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Best-effort source loaders ──────────────────────────────────────────────
def _load_json(path: str) -> Optional[Any]:
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return None


def _load_applied_config() -> dict:
    """Read config/applied_config.json (written by the pipeline each run)."""
    data = _load_json(os.path.join("config", "applied_config.json"))
    return data if isinstance(data, dict) else {}


def _load_solver_stats() -> list:
    data = _load_json(os.path.join("reports", "solver_stats.json"))
    return data if isinstance(data, list) else []


def _load_solver_config_summary() -> Optional[dict]:
    """Pull the active solver profile + soft-constraint weights, if available."""
    try:
        import solver_config  # local module, optional
        cfg = solver_config.load_config()
        return {
            "profile": solver_config.detect_profile(cfg),
            "soft_constraints": {
                k: {
                    "enabled": solver_config.is_enabled(cfg, k),
                    "weight": solver_config.get_weight(cfg, k),
                }
                for k in getattr(solver_config, "SOFT_CONSTRAINT_KEYS", [])
            },
        }
    except Exception:
        return None


# ── Low-level cell helpers ──────────────────────────────────────────────────
def _title_band(ws, row: int, text: str, span: int = 4) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _TITLE_FONT
    c.alignment = _LEFT
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY_DEEP)
    ws.row_dimensions[row].height = 26
    return row + 1


def _section(ws, row: int, text: str, span: int = 4) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _SECTION_FONT
    c.alignment = _LEFT
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[row].height = 20
    return row + 1


def _kv(ws, row: int, label: str, value: Any, help_text: str = "") -> int:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _LABEL_FONT
    lc.alignment = _LEFT
    lc.fill = PatternFill("solid", fgColor=GREY_BG)
    lc.border = BORDER
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = _VALUE_FONT
    vc.alignment = _LEFT
    vc.border = BORDER
    if help_text:
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        hc = ws.cell(row=row, column=3, value=help_text)
        hc.font = _HELP_FONT
        hc.alignment = _LEFT
    return row + 1


def _table_header(ws, row: int, headers: list) -> int:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
    ws.row_dimensions[row].height = 18
    return row + 1


def _table_row(ws, row: int, values: list, band: bool = False) -> int:
    fill = PatternFill("solid", fgColor=BAND_BG) if band else None
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _VALUE_FONT
        c.alignment = _CENTER if i > 1 else _LEFT
        c.border = BORDER
        if fill:
            c.fill = fill
    return row + 1


# ── Public API ──────────────────────────────────────────────────────────────
def build_traceability_sheet(wb, applied: Optional[dict] = None,
                             stats: Optional[list] = None,
                             solver_summary: Optional[dict] = None,
                             sheet_name: str = "Parameters"):
    """Append a "Parameters" sheet documenting the run configuration.

    Args:
        wb: an openpyxl Workbook.
        applied: applied_config.json dict (auto-loaded if None).
        stats: solver_stats.json list (auto-loaded if None).
        solver_summary: solver profile/weights dict (auto-loaded if None).
        sheet_name: worksheet title.

    Returns:
        The created worksheet.
    """
    applied = applied if applied is not None else _load_applied_config()
    stats = stats if stats is not None else _load_solver_stats()
    solver_summary = (solver_summary if solver_summary is not None
                      else _load_solver_config_summary())

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 30

    r = 1
    r = _title_band(ws, r, "Universidad Loyola - Run parameters & traceability")
    intro = ws.cell(row=r, column=1,
                    value="These are the configuration values and constraints "
                          "actually applied by the solver for this plan. They "
                          "let you verify the output matches the intended setup.")
    intro.font = _HELP_FONT
    intro.alignment = _LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 30
    r += 2

    # ── Run metadata ────────────────────────────────────────────────────────
    r = _section(ws, r, "Run metadata")
    meta = applied.get("meta", {}) if isinstance(applied, dict) else {}
    applied_at = meta.get("applied_at") or datetime.now().isoformat(timespec="seconds")
    r = _kv(ws, r, "Generated at", applied_at,
            "Timestamp when the configuration was applied to this run.")
    if not applied:
        r = _kv(ws, r, "Configuration source", "live pipeline defaults",
                "config/applied_config.json was not found; showing built-in "
                "values. Run an optimization to capture the exact applied set.")
    r += 1

    # ── Solver profile + soft constraints ────────────────────────────────────
    r = _section(ws, r, "Solver configuration (soft constraints)")
    if solver_summary:
        r = _kv(ws, r, "Active profile", solver_summary.get("profile", "Balanced"),
                "Preset that groups the soft-constraint weights below.")
        soft = solver_summary.get("soft_constraints", {})
        if soft:
            r = _table_header(ws, r, ["Soft constraint", "Enabled", "Weight", ""])
            band = False
            for key, val in soft.items():
                r = _table_row(
                    ws, r,
                    [key, "Yes" if val.get("enabled") else "No",
                     val.get("weight", 0), ""],
                    band=band)
                band = not band
    else:
        r = _kv(ws, r, "Solver configuration", "default (Balanced)",
                "solver_config module unavailable; historical defaults used.")
    r += 1

    # Hard constraints are always enforced — state them explicitly.
    r = _kv(ws, r, "Hard constraints (always on)",
            "C1 room, C4 student, C5 professor",
            "Non-negotiable: no room/student/professor double-booking. "
            "These cannot be disabled.")
    r += 1

    # ── Global generation parameters ─────────────────────────────────────────
    r = _section(ws, r, "Global generation parameters")
    g = applied.get("global", {}) if isinstance(applied, dict) else {}
    _fallback_globals = {
        "preferred_size": _live_const("PREFERRED_GROUP_SIZE", 12),
        "default_max": _live_const("MAX_GROUP_SIZE", 15),
        "min_size": _live_const("MIN_GROUP_SIZE", 7),
        "computer_lab_max": _live_const("COMPUTER_LAB_MAX", 24),
        "reduced_max_size": _live_const("REDUCED_MAX_SIZE", 12),
        "s1_total_weeks": _live_const("SEMESTER_1_WEEKS", 14),
        "s2_total_weeks": _live_const("SEMESTER_2_WEEKS", 20),
    }
    for key, gg in _fallback_globals.items():
        if key not in g:
            g[key] = gg
    labels = [
        ("preferred_size", "Preferred group size", "Target students per lab group."),
        ("default_max", "Default max group size", "Upper bound for a standard group."),
        ("min_size", "Min group size", "Groups smaller than this are consolidated."),
        ("computer_lab_max", "Computer lab max", "Capacity for computer-room subjects."),
        ("reduced_max_size", "Reduced lab max", "Capacity for reduced-capacity rooms."),
        ("s1_total_weeks", "S1 total weeks", "Length of the first-semester window."),
        ("s2_total_weeks", "S2 total weeks", "Length of the second-semester window."),
    ]
    for key, lab, hlp in labels:
        r = _kv(ws, r, lab, g.get(key, "-"), hlp)
    r += 1

    # ── Solver run results (per semester) ─────────────────────────────────────
    if stats:
        r = _section(ws, r, "Solver run results")
        r = _table_header(ws, r, ["Semester", "Status", "Sessions",
                                  "Time (s) / gap"])
        band = False
        for s in stats:
            if not isinstance(s, dict):
                continue
            time_gap = f"{s.get('wall_time_s', '-')}s / {s.get('gap', '-')}"
            r = _table_row(ws, r, [
                s.get("label", s.get("semester", "-")),
                s.get("status", "-"),
                s.get("n_sessions", "-"),
                time_gap,
            ], band=band)
            band = not band
        r += 1

    # ── Per-subject configuration ─────────────────────────────────────────────
    subjects = applied.get("subjects", {}) if isinstance(applied, dict) else {}
    if not subjects:
        subjects = _live_subjects()
    if subjects:
        r = _section(ws, r, "Per-subject configuration")
        r = _table_header(ws, r, ["Subject", "Sessions", "Max students",
                                  "Week window"])
        band = False
        for name, cfg in subjects.items():
            if not isinstance(cfg, dict):
                continue
            wk = f"{cfg.get('min_week', '-')}-{cfg.get('max_week', '-')}"
            r = _table_row(ws, r, [
                name,
                cfg.get("num_sessions", "-"),
                cfg.get("max_students", "-"),
                wk,
            ], band=band)
            band = not band
        r += 1

    # ── Teacher availability / rules ──────────────────────────────────────────
    blocked = applied.get("teachers_blocked_slots", {}) if isinstance(applied, dict) else {}
    rules = applied.get("teacher_rules", {}) if isinstance(applied, dict) else {}
    r = _section(ws, r, "Teacher constraints")
    r = _kv(ws, r, "Teachers with blocked slots", len(blocked),
            "Number of teachers with explicit unavailability windows.")
    r = _kv(ws, r, "Teachers with preference rules", len(rules),
            "Preferred blocks / max days-per-week signals applied.")

    ws.sheet_view.zoomScale = 100
    return ws


# ── Live fallbacks (used only when applied_config.json is absent) ────────────
def _live_const(name: str, default):
    try:
        import pipeline
        return getattr(pipeline, name, default)
    except Exception:
        return default


def _live_subjects() -> dict:
    try:
        import pipeline
        return {
            k: {
                "num_sessions": v.get("num_sessions"),
                "max_students": v.get("max_students"),
                "min_week": v.get("min_week"),
                "max_week": v.get("max_week"),
            }
            for k, v in getattr(pipeline, "LAB_CONFIG", {}).items()
        }
    except Exception:
        return {}
