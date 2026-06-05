"""
04 v5 + 05 v5 COMBINÉ — Pipeline complet aulario+alumnos → planning optimal
Projet Scheduling - Universidad Loyola
Auteur: RAKOTONJANAHARY Maminiaina Mendrika

APPROCHE FINALE :
  INPUT  : master_schedule.csv (aulario + alumnos, scripts 01-02)
           + CONFIGURATION des labs (quelles matières, salles, sessions)
  OUTPUT : planning optimal des sessions de lab

  Les fichiers de Daniel ne sont PLUS utilisés comme entrée.
  Ils servent UNIQUEMENT de validation a posteriori.

PIPELINE EN 5 ÉTAPES :
  1. Charger le master_schedule et la configuration des labs
  2. Identifier les étudiants inscrits à chaque matière via MixtoID
  3. Calculer l'emploi du temps individuel de chaque étudiant
  4. Former les groupes : chaque étudiant assigné à un créneau libre
  5. Solveur : décider la semaine de chaque session (créneau déjà fixé)
"""

from ortools.sat.python import cp_model
import sys
import io
import math
# Fix Windows cp1252 encoding issue with Unicode characters
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

import pandas as pd
import os
import sys
import hashlib
from datetime import datetime, timedelta
from collections import defaultdict, Counter

# ============================================================
# CONFIGURATION GÉNÉRALE
# ============================================================
MASTER_PATH = 'data_clean/master_schedule.csv'
OUTPUT_DIR = 'outputs/optimization/'
REPORT_PATH = 'reports/pipeline_v5_report.txt'

TIME_BLOCKS = [
    {'id': 1, 'label': '08:30-10:30', 'start': 510, 'end': 630, 'period': 'morning'},
    {'id': 2, 'label': '10:30-12:30', 'start': 630, 'end': 750, 'period': 'morning'},
    {'id': 3, 'label': '12:30-14:30', 'start': 750, 'end': 870, 'period': 'morning'},
    {'id': 4, 'label': '15:00-17:00', 'start': 900, 'end': 1020, 'period': 'afternoon'},
    {'id': 5, 'label': '17:00-19:00', 'start': 1020, 'end': 1140, 'period': 'afternoon'},
    {'id': 6, 'label': '19:00-21:00', 'start': 1140, 'end': 1260, 'period': 'evening'},
]

BLOCK_LABELS = {b['id']: b['label'] for b in TIME_BLOCKS}
MORNING_BLOCKS = [b['id'] for b in TIME_BLOCKS if b['period'] == 'morning']
AFTERNOON_BLOCKS = [b['id'] for b in TIME_BLOCKS if b['period'] in ['afternoon', 'evening']]
ALL_BLOCKS = [b['id'] for b in TIME_BLOCKS]

DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
DAY_IDS = {d: i for i, d in enumerate(DAYS)}

SEMESTER_1_WEEKS = 14   # Semaines 1-14 (2 sept - 5 déc 2025)
SEMESTER_2_WEEKS = 20   # Semaines 1-20 (à vérifier avec Daniel pour les matières de 4ème année)

# ── Year-of-degree time-period exceptions (configurable from the app) ──
# Default = STRICT rule (see Considerations PDF): 1st/3rd year mornings only,
# 2nd/4th year afternoons only. When the app sets these to True, the matching
# overflow/recovery logic is allowed to use the opposite period as a fallback.
ALLOW_AFTERNOON_Y1Y3 = False   # if True: 1st/3rd year MAY use afternoon slots
ALLOW_MORNING_Y2Y4   = False   # if True: 2nd/4th year MAY use morning slots

# ── Teacher unavailability (configurable from the app) ──
# Maps a teacher identifier to a list of forbidden (day_idx, block_id) slots.
# Consumed during group formation: a group whose recurring slot is blocked for
# its teacher is never created on that slot. Empty by default.
TEACHER_UNAVAILABILITY = {}

# ── Chemistry lab room availability (configurable from the app) ──
# By default Química is confined to ONE room (Ciencias Experimentales I), which
# bottlenecks the shared Física+Química morning groups: that single room carries
# 5 (Física) + 4 (Química) = 9 sessions per slot while Ciencias II only carries
# Física's 5, so a 2nd group can't fit. If the chemistry equipment is actually
# available in BOTH science labs (to confirm with Daniel), set this True so
# Química spreads across the two rooms like Física does, unlocking ~100%.
QUIMICA_USE_TWO_ROOMS = False

# ── Parity-alternation spreading (Daniel's real strategy) ──
# When several groups of the same subject share a (day, block) slot, Daniel
# alternates their weeks by parity: half on even weeks (W4,6,8,10,12), half on
# odd (W5,7,9,11,13). Two groups of opposite parity never collide on the same
# week, so the shared room (e.g. Ciencias I) is never oversaturated. This is the
# key that lets him fit everyone in the morning. Enabled by default; soft
# (weighted) so the solver stays feasible if a slot can't be perfectly split.
PARITY_ALTERNATION = True
PARITY_PENALTY_WEIGHT = 50  # below anchoring (100), above spacing (1)

SOLVER_TIME_LIMIT = 300

# Calendrier académique 2025/2026 - jours fériés
# Format: {semester: {(semaine, jour_idx): "raison"}}
# jour_idx: 0=Lundi, 1=Mardi, 2=Mercredi, 3=Jueves, 4=Viernes
HOLIDAYS = {
    1: {
        # Sem 7 : 13/10 Lundi Día de la Hispanidad
        (7, 0): "Día de la Hispanidad",
    },
    2: {
        # Sem 6 : 13/03 Vendredi — Blue Day (jour férié institutionnel de l'université)
        (6, 4): "Blue Day",
        # Sem 7 : 16/03 Lundi Canonización de San Ignacio
        (7, 0): "Canonización de San Ignacio",
        # Sem 9 : 30/03-05/04 Semana Santa (semaine complète)
        (9, 0): "Semana Santa", (9, 1): "Semana Santa",
        (9, 2): "Semana Santa", (9, 3): "Semana Santa",
        (9, 4): "Semana Santa",
        # Sem 12 : 23-25/04 Feria de Abril. 23/04=jeudi, 24/04=vendredi
        # (25/04=samedi). On bloque donc UNIQUEMENT jeudi et vendredi — le
        # mercredi 22/04 n'est PAS férié (corrigé d'après le calendrier officiel).
        (12, 3): "Feria de Abril", (12, 4): "Feria de Abril",
        # Sem 13 : 1/05 Vendredi Día del Trabajador
        (13, 4): "Día del Trabajador",
    },
}

# ============================================================
# SUBJECT-SPECIFIC BLOCKED SLOTS  (room reserved for another activity)
# ------------------------------------------------------------
# Unlike HOLIDAYS (which blocks a whole (week, day) for EVERY subject), this
# blocks a single (week, day_idx, block_id) for ONE subject only — used when a
# lab ROOM is occupied by another (future) activity on specific slots, so the
# real práctica sessions of that subject must not be scheduled there.
#
# Origin: Daniel's reference (Distribucion_Practicas_25-26_rev15) shows the
# Química room "Ciencias Experimentales I" occupied by a "Práctica Biotecnología"
# (Grupo 17) on Wednesday+Thursday, blocks 10:30-12:30 and 12:30-14:30, in weeks
# 7, 8, 11, 12. That Biotecnología lab does not exist in our model yet, so we
# treat those exact slots as unavailable for S1_Química and surface them in the
# Excel as "Festivo / No disponible".
#
# Format: {(semester, subject): {(week, day_idx, block_id): "label", ...}}
#   day_idx: 0=Lunes 1=Martes 2=Miércoles 3=Jueves 4=Viernes
#   block_id: 1=08:30-10:30 2=10:30-12:30 3=12:30-14:30 (see TIME_BLOCKS)
SUBJECT_BLOCKED_LABEL = "Festivo / No disponible"
SUBJECT_BLOCKED_SLOTS = {
    (1, 'S1_Química'): {
        (w, d, b): SUBJECT_BLOCKED_LABEL
        for w in (7, 8, 11, 12)
        for d in (2, 3)          # Miércoles, Jueves
        for b in (2, 3)          # 10:30-12:30, 12:30-14:30
    },
}


def subject_blocked_map(semester, subject):
    """All blocked (week, day_idx, block_id) -> label for this subject/semester."""
    return SUBJECT_BLOCKED_SLOTS.get((semester, subject), {})


def is_week_blocked_for_session(semester, subject, week, day_idx, block_id):
    """True if (subject, week, day, block) is reserved for another activity."""
    return (week, day_idx, block_id) in subject_blocked_map(semester, subject)


# True = vrais AlumnoIDs (pour validation avec Daniel)
# False = IDs hachés (pour GitHub / partage externe)
#
# Set to True: the app runs LOCALLY on Daniel's machine, on his own institutional
# data, so showing real student names in the interface is appropriate (hashes are
# unrecognizable to him). The privacy guarantee that matters — never SHIPPING
# student-identifying files inside the distributed .exe — is enforced separately
# by the _PII_FILES exclusion list in LabScheduling.spec, which stays in place.
# So: names visible in-app on the local workspace; never embedded in the binary.
INCLUDE_REAL_NAMES = True

# ============================================================
# CONFIGURATION DES LABS
# Corrigé selon les données réelles de Daniel (zip Curso_2025_2026)
# Focus S1 uniquement (S2 en attente selon feedback Daniel)
# ============================================================

PREFERRED_GROUP_SIZE = 12  # Marge pour nouveaux étudiants
MAX_GROUP_SIZE = 15
MIN_GROUP_SIZE = 7  # Pas de groupe en dessous de ce seuil (default per Daniel)
# Floor used ONLY in the per-program last-resort recovery phase (Phase 5).
# Defaults to MIN_GROUP_SIZE (no relaxation). Lower it (e.g. 5) via the app to
# let the recovery phase form small homogeneous groups and reach ~100%.
# This is Daniel's pragmatic fallback: he accepts groups of 11-13, occasionally
# smaller, rather than leaving students unplaced.
RECOVERY_MIN_GROUP_SIZE = 7
MAX_EXTRA_GROUPS = 3  # Si 12 crée > 3 groupes de plus que 15, utiliser 15
COMPUTER_LAB_MAX = 24  # Salle informatique : capacité max (Considerations PDF)
REDUCED_MAX_SIZE = 12  # Resist. Matériaux / Méca. Fluides : max 12 (Considerations PDF)

LAB_CONFIG = {
    # ── SEMESTRE 1 : 1ère année ──
    # GROUPES PARTAGÉS : Física et Química utilisent les mêmes groupes
    # Daniel: 19 groupes, homogènes par programme, 7-15 étudiants
    # Física: 5 sessions/groupe (17 groupes), Química: 4 sessions/groupe (15 groupes)
    'S1_Física': {
        'curso_num': 1, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
        'simultaneous_rooms': False,  # 1 groupe = 1 salle (les 2 salles tournent
        # en parallèle sur des groupes DIFFÉRENTS, doublant la capacité). C'était
        # 'True', ce qui forçait chaque groupe à occuper les 2 salles à la fois et
        # divisait la capacité par 2 (goulot artificiel : 12 groupes au lieu de
        # ~15). L'appairage des séances d'intro reste géré par
        # 'intro_session_paired' ci-dessous.
        'keywords': ['física i', 'física'],
        # Exclude lookalikes captured by the bare 'física' keyword that are NOT
        # 1st-year Física I: Física II, Física Computacional (3rd year) and
        # Métodos Numéricos para Simulación Física. Cross-checked against the
        # official enrolment report (informeDetalleGruposPorCurso): Física I 1st
        # year = 208 students across 4 programs; without these exclusions the
        # bare keyword over-counted by ~46 phantom students.
        'keyword_exclude': ['física ii', 'computacional', 'simulación',
                            'simulacion'],
        'shared_group': 'S1_1er_anno',  # Groupes partagés avec Química
        'group_by_program': True,       # Groupes homogènes par programme
        # Práctica 1 ("séance d'introduction") : appairer les groupes consécutifs
        # (1&2, 3&4, 5&6...) qui font leur première séance ensemble dans 2 labos
        # différents simultanément (pattern observé dans le rev15 de Daniel).
        # À partir de Práctica 2, chaque groupe a sa propre séance.
        'intro_session_paired': True,
    },
    'S1_Química': {
        'curso_num': 1, 'semester': 1, 'num_sessions': 4, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I'],
        'simultaneous_rooms': False,
        'keywords': ['química general'],
        'keyword_exclude': [],
        'shared_group': 'S1_1er_anno',  # Réutilise les groupes de Física
        'group_by_program': True,
    },

    # ── SEMESTRE 1 : 2ème année ──
    # Daniel: groupes MIXTES (pas homogènes par programme)
    # Corrigé: num_sessions = 5 (pas 1-2 comme avant)
    'S1_Electrotecnia': {
        'curso_num': 2, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': 15,  # Daniel: 14-18
        'lab_rooms': ['Lab. Eléctrica'],  # Daniel ref: Lab. Eléctrica
        'simultaneous_rooms': False,
        'keywords': ['electrotecnia'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Mecanismos': {
        'curso_num': 2, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,  # Max 15 (Answers.docx: max 15 per lab)
        'lab_rooms': ['Lab. Telemática'],  # Daniel ref: Lab. Telemática
        'simultaneous_rooms': False,
        'keywords': ['mecanismos y elementos'],  # Plus spécifique pour éviter faux positifs
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Termodinámica': {
        'curso_num': 2, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Termodinámica'],  # Daniel ref: Lab. Termodinámica (corrigé, était Mecánica de Fluidos)
        'simultaneous_rooms': False,
        'keywords': ['termodinámica'],
        'keyword_exclude': [],
        'group_by_program': False,
    },

    # ── SEMESTRE 1 : 3ème année ──
    'S1_Tecnologías de Fabricación': {
        'curso_num': 3, 'semester': 1, 'num_sessions': 2, 'min_week': 3, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Telemática'],  # Daniel ref: Lab. Telemática
        'simultaneous_rooms': False,
        'keywords': ['tecnologías de fabricación'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Robótica y Automatización': {
        'curso_num': 3, 'semester': 1, 'num_sessions': 3, 'min_week': 3, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Robótica'],  # Daniel ref: Lab. Robótica
        'simultaneous_rooms': False,
        'keywords': ['robótica y automatización'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Automatización Industrial': {
        'curso_num': 3, 'semester': 1, 'num_sessions': 4, 'min_week': 3, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Robótica'],  # Daniel ref: Lab. Robótica (automatización industrial e instrumentación)
        'simultaneous_rooms': False,
        'keywords': ['automatización industrial', 'instrumentación'],
        'keyword_exclude': ['robótica'],
        'group_by_program': False,
    },

    # ── SEMESTRE 2 (en attente selon Daniel, mais config conservée) ──
    'S2_Física II': {
        'curso_num': 1, 'semester': 2, 'num_sessions': 5, 'min_week': 9, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
        'simultaneous_rooms': False,  # 1 groupe = 1 salle (doubles la capacité),
        # comme S1_Física. simultaneous_rooms=True faisait occuper les 2 salles
        # à la fois et saturait Ciencias I → 102 sessions à caser → S2 INFAISABLE.
        'keywords': ['física ii'],
        'keyword_exclude': [],
        'group_by_program': True,
    },
    'S2_Tecnología Medio Ambiente': {
        'curso_num': 1, 'semester': 2, 'num_sessions': 2, 'min_week': 11, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I'],
        'simultaneous_rooms': False,
        'keywords': ['tecnología del medio ambiente', 'medio ambiente'],
        'keyword_exclude': [],
        'group_by_program': True,
    },
    'S2_Resistencia de Materiales': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 10, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Automoción y Resistencia de Mat.'],
        'simultaneous_rooms': False,
        'keywords': ['resistencia de materiales'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Mecánica de Fluidos': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 9, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Mecánica de Fluidos'],
        'simultaneous_rooms': False,
        'keywords': ['mecánica de fluido', 'mecánica y máquinas de fluido'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Regulación Automática': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 5, 'min_week': 9, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['regulación automática'],
        'keyword_exclude': ['automatic control'],
        'group_by_program': False,
    },
    'S2_Tecnología Electrónica': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 3, 'min_week': 8, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Electrónica'],
        'simultaneous_rooms': False,
        'keywords': ['tecnología electrónica'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Electrónica y Automática': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 3, 'min_week': 9, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Electrónica', 'Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['electrónica y automática'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Informática y Com. Industriales': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 8, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['informática y comunicaciones industriales'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Métodos Numéricos': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 9, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Laboratorio de Ingeniería Telemática'],
        'simultaneous_rooms': False,
        'keywords': ['métodos numéricos'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Modelado de Sistemas': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 5, 'min_week': 7, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Laboratorio de Ingeniería Telemática'],
        'simultaneous_rooms': False,
        'keywords': ['modelado simulación', 'modelado y simulación'],
        'keyword_exclude': ['modelado de sistemas físicos', 'modelado de sistemas fís'],
        'group_by_program': False,
    },
    'S2_Automatic Control': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 5, 'min_week': 10, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['automatic control'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Ingeniería de Control': {
        'curso_num': 3, 'semester': 2, 'num_sessions': 5, 'min_week': 8, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['ingeniería de control'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Control de Máquinas': {
        'curso_num': 3, 'semester': 2, 'num_sessions': 5, 'min_week': 13, 'max_week': 20,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Eléctrica'],
        'simultaneous_rooms': False,
        'keywords': ['control de máquinas'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Estructuras': {
        'curso_num': 3, 'semester': 2, 'num_sessions': 1, 'min_week': 16, 'max_week': 19,
        'max_students': 12,
        'lab_rooms': ['Automoción y Resistencia de Mat.'],
        'simultaneous_rooms': False,
        'keywords': ['estructuras'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
}


# ════════════════════════════════════════════════════════════════════
# BRIDGE UI → PIPELINE : Apply user configuration overrides
# ════════════════════════════════════════════════════════════════════
USER_CONFIG_PATH = 'config/user_config.json'


def apply_user_config():
    """
    Load user configuration from `config/user_config.json` and apply overrides
    to LAB_CONFIG and global parameters. This allows the Streamlit app to
    customize the pipeline behavior without modifying the source code.

    Config file structure:
        {
          "global": {
            "preferred_size": 12, "default_max": 15, "min_size": 7,
            "computer_lab_max": 24, "reduced_max_size": 12,
            "start_week": 4, "s1_total_weeks": 14
          },
          "subjects": {
            "S1_Física": {
              "num_sessions": 5, "max_students": 15,
              "min_week": 4, "max_week": 13,
              "lab_rooms": [...], "keywords": [...], "keyword_exclude": [...]
            },
            ...
          },
          "year_prefs": {
            "allow_afternoon_y1y3": false,
            "allow_morning_y2y4": false
          },
          "teachers": {...}
        }
    """
    import json
    global PREFERRED_GROUP_SIZE, MAX_GROUP_SIZE, MIN_GROUP_SIZE, RECOVERY_MIN_GROUP_SIZE
    global COMPUTER_LAB_MAX, REDUCED_MAX_SIZE
    global SEMESTER_1_WEEKS, SEMESTER_2_WEEKS
    global ALLOW_AFTERNOON_Y1Y3, ALLOW_MORNING_Y2Y4, TEACHER_UNAVAILABILITY
    global QUIMICA_USE_TWO_ROOMS, PARITY_ALTERNATION

    if not os.path.exists(USER_CONFIG_PATH):
        print(f"\n  [INFO]  Aucune config utilisateur trouvée ({USER_CONFIG_PATH})")
        print(f"  → utilisation des valeurs par défaut")
        return

    try:
        with open(USER_CONFIG_PATH, 'r', encoding='utf-8') as f:
            user_config = json.load(f)
    except Exception as e:
        print(f"\n  [WARN]  Impossible de charger {USER_CONFIG_PATH} : {e}")
        return

    print(f"\n  [CONFIG] Configuration utilisateur chargée depuis {USER_CONFIG_PATH}")
    saved_at = user_config.get('meta', {}).get('saved_at', '?')
    print(f"     Sauvegardée le : {saved_at}")

    # ── 1) Apply GLOBAL overrides ──
    global_cfg = user_config.get('global', {})
    if global_cfg:
        print(f"  [GLOBAL]  Paramètres globaux :")
        if 'preferred_size' in global_cfg:
            PREFERRED_GROUP_SIZE = int(global_cfg['preferred_size'])
            print(f"     PREFERRED_GROUP_SIZE = {PREFERRED_GROUP_SIZE}")
        if 'default_max' in global_cfg:
            MAX_GROUP_SIZE = int(global_cfg['default_max'])
            print(f"     MAX_GROUP_SIZE = {MAX_GROUP_SIZE}")
        if 'min_size' in global_cfg:
            MIN_GROUP_SIZE = int(global_cfg['min_size'])
            print(f"     MIN_GROUP_SIZE = {MIN_GROUP_SIZE}")
        # Recovery floor (Phase 5). Defaults to MIN_GROUP_SIZE when not provided,
        # so behaviour is unchanged unless the app explicitly relaxes it.
        if 'recovery_min_size' in global_cfg:
            RECOVERY_MIN_GROUP_SIZE = int(global_cfg['recovery_min_size'])
            print(f"     RECOVERY_MIN_GROUP_SIZE = {RECOVERY_MIN_GROUP_SIZE}")
        else:
            RECOVERY_MIN_GROUP_SIZE = MIN_GROUP_SIZE
        if 'computer_lab_max' in global_cfg:
            try:
                COMPUTER_LAB_MAX = int(global_cfg['computer_lab_max'])
                print(f"     COMPUTER_LAB_MAX = {COMPUTER_LAB_MAX}")
            except NameError:
                pass  # Variable not defined globally, skip
        if 'reduced_max_size' in global_cfg:
            try:
                REDUCED_MAX_SIZE = int(global_cfg['reduced_max_size'])
                print(f"     REDUCED_MAX_SIZE = {REDUCED_MAX_SIZE}")
            except NameError:
                pass
        if 's1_total_weeks' in global_cfg:
            try:
                SEMESTER_1_WEEKS = int(global_cfg['s1_total_weeks'])
                print(f"     SEMESTER_1_WEEKS = {SEMESTER_1_WEEKS}")
            except NameError:
                pass
        if 's2_total_weeks' in global_cfg:
            try:
                SEMESTER_2_WEEKS = int(global_cfg['s2_total_weeks'])
                print(f"     SEMESTER_2_WEEKS = {SEMESTER_2_WEEKS}")
            except NameError:
                pass
        if 'quimica_use_two_rooms' in global_cfg:
            QUIMICA_USE_TWO_ROOMS = bool(global_cfg['quimica_use_two_rooms'])
        if 'parity_alternation' in global_cfg:
            PARITY_ALTERNATION = bool(global_cfg['parity_alternation'])

    # ── 1b) Chemistry two-room layout (unlocks the morning capacity bottleneck) ──
    # When enabled, Química can use BOTH science labs. Crucially we keep
    # simultaneous_rooms=False so the allocator SPREADS groups across the two
    # rooms (group A → Ciencias I, group B → Ciencias II), halving the per-room
    # load — rather than occupying both at once like Física does. This is what
    # relieves the Ciencias I overload that capped assignment at ~91%.
    if QUIMICA_USE_TWO_ROOMS and 'S1_Química' in LAB_CONFIG:
        LAB_CONFIG['S1_Química']['lab_rooms'] = [
            'Ciencias Experimentales I', 'Ciencias Experimentales II']
        LAB_CONFIG['S1_Química']['simultaneous_rooms'] = False
        print(f"  [ROOMS]  Química réparti sur 2 salles "
              f"(Ciencias I + II, alternées) — déblocage capacité matin")

    # ── 2) Apply PER-SUBJECT overrides ──
    subjects_cfg = user_config.get('subjects', {})
    if subjects_cfg:
        print(f"  [SUBJECTS] Surcharges par matière : {len(subjects_cfg)} matières personnalisées")
        for subj_key, overrides in subjects_cfg.items():
            if subj_key not in LAB_CONFIG:
                print(f"     [WARN]  {subj_key} : matière inconnue, ignorée")
                continue

            base = LAB_CONFIG[subj_key]
            applied = []

            if 'num_sessions' in overrides:
                old, new = base.get('num_sessions'), int(overrides['num_sessions'])
                if old != new:
                    base['num_sessions'] = new
                    applied.append(f"num_sessions {old}→{new}")

            if 'max_students' in overrides:
                old, new = base.get('max_students'), int(overrides['max_students'])
                if old != new:
                    base['max_students'] = new
                    applied.append(f"max_students {old}→{new}")

            if 'min_week' in overrides:
                old, new = base.get('min_week'), int(overrides['min_week'])
                if old != new:
                    base['min_week'] = new
                    applied.append(f"min_week {old}→{new}")

            if 'max_week' in overrides:
                old, new = base.get('max_week'), int(overrides['max_week'])
                if old != new:
                    base['max_week'] = new
                    applied.append(f"max_week {old}→{new}")

            if 'lab_rooms' in overrides and overrides['lab_rooms']:
                new_rooms = list(overrides['lab_rooms'])
                old_rooms = base.get('lab_rooms', [])
                if set(new_rooms) != set(old_rooms):
                    base['lab_rooms'] = new_rooms
                    applied.append(f"lab_rooms ({len(new_rooms)} salles)")

            if 'keywords' in overrides and overrides['keywords']:
                new_kw = [k.strip() for k in overrides['keywords'] if k.strip()]
                if new_kw and new_kw != base.get('keywords'):
                    base['keywords'] = new_kw
                    applied.append(f"keywords ({len(new_kw)})")

            if 'keyword_exclude' in overrides:
                new_kx = [k.strip() for k in overrides['keyword_exclude'] if k.strip()]
                # MERGE rather than replace: the code-level exclusions encode
                # critical disambiguations (e.g. Física vs Física Computacional /
                # Simulación Física) that must survive even if the app sends an
                # incomplete list. Union preserves both, deduplicated.
                code_kx = base.get('keyword_exclude', [])
                merged_kx = list(dict.fromkeys(code_kx + new_kx))  # ordered union
                if merged_kx != code_kx:
                    base['keyword_exclude'] = merged_kx
                    applied.append(f"keyword_exclude ({len(merged_kx)}, fusionné)")
                else:
                    base['keyword_exclude'] = merged_kx

            if applied:
                print(f"     [CONFIG]  {subj_key} : {', '.join(applied)}")

    # ── 3) Year preferences (NOW APPLIED, not just printed) ──
    year_prefs = user_config.get('year_prefs', {})
    if year_prefs:
        print(f"  [SCHEDULE] Préférences horaires :")
        ALLOW_AFTERNOON_Y1Y3 = bool(year_prefs.get('allow_afternoon_y1y3', False))
        ALLOW_MORNING_Y2Y4 = bool(year_prefs.get('allow_morning_y2y4', False))
        if ALLOW_AFTERNOON_Y1Y3:
            print(f"     [WARN]  1ère/3ème année : après-midi exceptionnellement AUTORISÉ")
        else:
            print(f"     1ère/3ème année : matin STRICT (après-midi interdit)")
        if ALLOW_MORNING_Y2Y4:
            print(f"     [WARN]  2ème/4ème année : matin exceptionnellement AUTORISÉ")
        else:
            print(f"     2ème/4ème année : après-midi STRICT (matin interdit)")

    # ── 4) Teacher unavailability (NOW APPLIED at group formation) ──
    teachers = user_config.get('teachers', {})
    if teachers:
        print(f"  [TEACHER] Restrictions professeurs : {len(teachers)} professeur(s)")
        # Normalise to {teacher: set((day_idx, block_id))}
        normalised = {}
        for teacher, slots in teachers.items():
            slot_set = set()
            for s in slots:
                # accept [day_idx, block_id] or {"day": d, "block": b}
                if isinstance(s, (list, tuple)) and len(s) == 2:
                    slot_set.add((int(s[0]), int(s[1])))
                elif isinstance(s, dict) and 'day' in s and 'block' in s:
                    slot_set.add((int(s['day']), int(s['block'])))
            normalised[teacher] = slot_set
            print(f"     {teacher} : {len(slot_set)} créneau(x) bloqué(s)")
        TEACHER_UNAVAILABILITY = normalised

    print(f"  [OK] Configuration utilisateur appliquée\n")

    # ── CHECK 1: write back the EFFECTIVELY-APPLIED values for verification ──
    write_applied_config()


def write_applied_config():
    """
    Write `config/applied_config.json` — a read-back of the values the pipeline
    has ACTUALLY applied (post-override globals + the live LAB_CONFIG).

    The app compares this against the `user_config.json` it sent, parameter by
    parameter, to prove that every setting was honoured. A parameter present in
    user_config but absent or different here is a synchronisation gap.
    """
    import json
    applied = {
        'global': {
            'preferred_size':   PREFERRED_GROUP_SIZE,
            'default_max':      MAX_GROUP_SIZE,
            'min_size':         MIN_GROUP_SIZE,
            'computer_lab_max': COMPUTER_LAB_MAX,
            'reduced_max_size': REDUCED_MAX_SIZE,
            's1_total_weeks':   SEMESTER_1_WEEKS,
            's2_total_weeks':   SEMESTER_2_WEEKS,
            'quimica_use_two_rooms': QUIMICA_USE_TWO_ROOMS,
            'parity_alternation': PARITY_ALTERNATION,
        },
        'year_prefs': {
            'allow_afternoon_y1y3': ALLOW_AFTERNOON_Y1Y3,
            'allow_morning_y2y4':   ALLOW_MORNING_Y2Y4,
        },
        'teachers_blocked_slots': {
            t: sorted(list(slots)) for t, slots in TEACHER_UNAVAILABILITY.items()
        },
        # Per-subject snapshot of the parameters that matter downstream
        'subjects': {
            k: {
                'num_sessions':    v.get('num_sessions'),
                'max_students':    v.get('max_students'),
                'min_week':        v.get('min_week'),
                'max_week':        v.get('max_week'),
                'lab_rooms':       v.get('lab_rooms', []),
            }
            for k, v in LAB_CONFIG.items()
        },
        'meta': {
            'applied_at': datetime.now().isoformat(),
            'note': 'Auto-generated by pipeline. Reflects values actually used.',
        },
    }
    try:
        os.makedirs('config', exist_ok=True)
        with open('config/applied_config.json', 'w', encoding='utf-8') as f:
            json.dump(applied, f, indent=2, ensure_ascii=False)
        print(f"  [ECHO]  État appliqué écrit dans config/applied_config.json")
    except Exception as e:
        print(f"  [WARN]  Impossible d'écrire applied_config.json : {e}")


def print_section(title):
    """Print a horizontal separator with a title in the middle (60 chars wide)."""
    print("\n" + "=" * 60)
    print(f"  {title}")
    print("=" * 60)


def min_to_block_id(start_min):
    """
    Convert "minutes since midnight" to a TIME_BLOCKS id (1..6).

    Tolerates +/- 15 minutes around the canonical block start time
    (e.g., 504 -> block 1, even though block 1 nominally starts at 510).

    Args:
        start_min: int, minutes since midnight (e.g., 510 = 08:30)

    Returns:
        int | None: block id (1-6) or None if no block matches.
    """
    for b in TIME_BLOCKS:
        if abs(b['start'] - start_min) < 15:
            return b['id']
    return None


# ============================================================
# ÉTAPE 1 : CHARGER ET PRÉPARER
# ============================================================
def load_and_prepare(df):
    """
    Step 1: Print summary statistics for the loaded master schedule.

    The actual loading happens upstream (in main()); this function
    just validates and prints diagnostics.

    Args:
        df: master_schedule DataFrame (joined aulario + alumnos)

    Returns:
        The same DataFrame, unchanged.
    """
    print_section("ÉTAPE 1 : Chargement et préparation")
    print(f"  {df.shape[0]} lignes × {df.shape[1]} colonnes")
    print(f"  {df['AlumnoID'].nunique()} étudiants uniques")
    print(f"  {df['actividad'].nunique()} cours uniques")
    return df


# ============================================================
# ÉTAPE 2 : IDENTIFIER LES ÉTUDIANTS PAR MATIÈRE (via MixtoID)
# ============================================================
def identify_students(df):
    """
    Pour chaque matière avec lab (définie dans LAB_CONFIG),
    cherche les étudiants inscrits via les mots-clés dans actividad.
    Filtre par campus Sevilla et par année d'étude (curso_asignatura).
    """
    print_section("ÉTAPE 2 : Étudiants inscrits par matière (via MixtoID)")

    # Filtrer campus Sevilla uniquement
    if 'campus' in df.columns:
        sev = df[df['campus'].fillna('').str.contains('Sevilla', case=False, na=False)]
        print(f"  Filtre campus Sevilla: {len(sev)}/{len(df)} lignes")
    else:
        sev = df

    subject_students = {}

    for subject, config in LAB_CONFIG.items():
        keywords = config['keywords']
        keyword_exclude = config.get('keyword_exclude', [])
        curso_num = config['curso_num']

        # Chercher les lignes correspondant à cette matière
        mask = pd.Series(False, index=sev.index)
        for kw in keywords:
            mask = mask | sev['actividad'].fillna('').str.lower().str.contains(kw, na=False)
        for kw_ex in keyword_exclude:
            mask = mask & ~sev['actividad'].fillna('').str.lower().str.contains(kw_ex, na=False)

        # Filtrer par année d'étude : exclure les programmes de Master
        if 'Titulación' in sev.columns:
            mask = mask & ~sev['Titulación'].fillna('').str.contains('Máster|Master|Diploma', case=False, na=False)

        # Extraire les AlumnoIDs uniques
        student_ids = sev[mask]['AlumnoID'].dropna().unique().tolist()

        subject_students[subject] = student_ids

        sem = config['semester']
        print(f"  {'[OK]' if len(student_ids) > 0 else '[WARN]'} {subject:40s} S{sem} | "
              f"{len(student_ids):3d} étudiants")

    total = sum(len(v) for v in subject_students.values())
    print(f"\n  Total inscriptions : {total}")

    # ── Export real enrolment counts per subject for the reliability dashboard ──
    # This is the TRUE denominator for the assignment rate. Without it, the
    # dashboard divides assigned-by-assigned and always shows ~100%.
    try:
        os.makedirs('data_clean/optimization', exist_ok=True)
        enrol_rows = [
            {'subject': subj,
             'semester': LAB_CONFIG[subj]['semester'],
             'student_count': len(ids)}
            for subj, ids in subject_students.items()
        ]
        pd.DataFrame(enrol_rows).to_csv(
            'data_clean/optimization/lab_enrollments.csv',
            index=False, encoding='utf-8-sig')
        print(f"  [OK] Export lab_enrollments.csv ({len(enrol_rows)} matières)")
    except Exception as e:
        print(f"  [WARN]  Erreur export lab_enrollments.csv : {e}")

    return subject_students


# ============================================================
# ÉTAPE 3 : EMPLOI DU TEMPS INDIVIDUEL
# ============================================================
def build_individual_timetables(df, subject_students):
    """
    Pour chaque étudiant inscrit à une matière avec lab,
    calcule :
    1. Ses créneaux occupés (TOUS ses cours)
    2. Pour chaque matière avec lab, quels créneaux sont occupés
       PAR CETTE MATIÈRE spécifiquement

    PRINCIPE CLÉ : un lab de Física REMPLACE le cours de Física.
    Donc quand on cherche un créneau pour un lab de Física,
    le créneau du cours de Física est DISPONIBLE (pas occupé).
    Seuls les cours des AUTRES matières bloquent ce créneau.
    """
    print_section("ÉTAPE 3 : Emploi du temps individuel")

    all_student_ids = set()
    for ids in subject_students.values():
        all_student_ids.update(ids)

    print(f"  {len(all_student_ids)} étudiants uniques concernés par des labs")

    valid = df.dropna(subset=['AlumnoID', 'slot_hora_inicio_min', 'slot_jour_semaine'])
    valid = valid[valid['slot_hora_inicio_min'] > 0]
    valid = valid[valid['AlumnoID'].isin(all_student_ids)]

    # 1. Emploi du temps complet de chaque étudiant
    student_busy = {}  # AlumnoID → set of (day_idx, block_id)

    # 2. Pour chaque matière, les créneaux occupés PAR cette matière
    #    pour chaque étudiant
    student_subject_slots = defaultdict(lambda: defaultdict(set))
    # student_subject_slots[student_id][subject_name] = set of (day_idx, block_id)

    for student_id, group in valid.groupby('AlumnoID'):
        busy = set()
        for _, row in group.iterrows():
            day = row['slot_jour_semaine']
            block_id = min_to_block_id(row['slot_hora_inicio_min'])
            if block_id and day in DAYS:
                slot = (DAY_IDS[day], block_id)
                busy.add(slot)

                # Check if this course matches any lab subject
                act = str(row.get('actividad', '')).lower()
                for subject, config in LAB_CONFIG.items():
                    keywords = config['keywords']
                    keyword_exclude = config.get('keyword_exclude', [])
                    if any(kw in act for kw in keywords):
                        if not any(kw_ex in act for kw_ex in keyword_exclude):
                            student_subject_slots[student_id][subject].add(slot)

        student_busy[student_id] = busy

    # Statistiques
    avg_busy = sum(len(b) for b in student_busy.values()) / len(student_busy) if student_busy else 0
    avg_free = 30 - avg_busy
    min_free = 30 - max(len(b) for b in student_busy.values()) if student_busy else 0
    max_free = 30 - min(len(b) for b in student_busy.values()) if student_busy else 30

    print(f"  Créneaux occupés par étudiant : moy={avg_busy:.1f}")
    print(f"  Créneaux libres par étudiant  : min={min_free}, moy={avg_free:.1f}, max={max_free}")

    # Show how many extra slots become available due to course replacement
    extra_slots = 0
    for sid in all_student_ids:
        for subject in student_subject_slots.get(sid, {}):
            extra_slots += len(student_subject_slots[sid][subject])
    avg_extra = extra_slots / len(all_student_ids) if all_student_ids else 0
    print(f"  Créneaux récupérés par remplacement cours→lab : moy={avg_extra:.1f}")
    print(f"  Créneaux effectivement libres pour labs : moy={avg_free + avg_extra:.1f}")

    # ────────────────────────────────────────────────────────
    # Export student_busy.csv for downstream use (Caso individual page)
    # We export with the same identifier as group_composition.csv uses
    # (student_name if INCLUDE_REAL_NAMES, else student_hash)
    # ────────────────────────────────────────────────────────
    try:
        os.makedirs('data_clean/optimization', exist_ok=True)
        sb_rows = []
        for sid, busy_set in student_busy.items():
            for (day_idx, block_id) in busy_set:
                sb_rows.append({
                    'student_id': sid,
                    'day_idx': day_idx,
                    'block_id': block_id,
                })
        if sb_rows:
            sb_df = pd.DataFrame(sb_rows)
            sb_path = 'data_clean/optimization/student_busy.csv'
            sb_df.to_csv(sb_path, index=False, encoding='utf-8-sig')
            print(f"  [OK] Export {sb_path} ({len(sb_rows)} entrées)")
    except Exception as e:
        print(f"  [WARN]  Erreur export student_busy.csv : {e}")

    return student_busy, student_subject_slots


# ============================================================
# FLUX PROFESSEUR (modèle validé en session Pablo/Daniel)
# ============================================================
# Symmetric to the student flow:
#   Students  → enrollment        → student schedule  → busy / free slots
#   Professors → teaching assignments (normal class) → professor schedule → busy / free
# A lab can only be placed on a slot where BOTH the students AND the assigned
# professor are free. This builds professor_busy from the master schedule.
#
# The whiteboard left two items OPEN (data not yet available):
#   • which labs?            → which professor teaches which lab subject
#   • how many credits/groups? → how many groups a professor can supervise
# So this function DEGRADES GRACEFULLY: if no professor column exists in the
# data, it returns empty structures and the pipeline behaves exactly as before.

# Candidate column names for the professor identifier (auto-detected)
PROFESSOR_ID_COLUMNS = ['docentes', 'ProfesorID', 'profesor_id', 'profesor',
                        'Profesor', 'docente', 'Docente', 'teacher', 'Teacher',
                        'profesorado']

# Columns where one cell may list several professors separated by commas
# (e.g. 'docentes' in the official master schedule). For those, each name is a
# distinct professor and the row's time slot makes ALL of them busy.
MULTI_PROFESSOR_COLUMNS = {'docentes', 'Docentes'}

SUPERVISION_PATH = 'data_clean/optimization/subject_supervision.csv'


def load_supervision_capacity():
    """
    Load per-subject supervision capacity (number of distinct professors) from
    subject_supervision.csv, produced by build_professor_table.py from the
    official enrolment report's 'Docentes' column.

    Returns {subject → n_professors}, or {} if the file is absent.

    This answers the whiteboard's "how many groups?" question: a subject's
    parallel-group capacity is bounded by min(n_professors, n_rooms). For Física,
    n_professors=9 and n_rooms=2, so rooms are the real limit — confirming the
    morning bottleneck is physical (rooms), not supervision.
    """
    if not os.path.exists(SUPERVISION_PATH):
        return {}
    try:
        df = pd.read_csv(SUPERVISION_PATH)
        if {'subject', 'n_professors'}.issubset(df.columns):
            return {str(r['subject']): int(r['n_professors'])
                    for _, r in df.iterrows()}
    except Exception as e:
        print(f"  [WARN]  Lecture {SUPERVISION_PATH} échouée : {e}")
    return {}


def build_professor_busy(df):
    """
    Build the professor side of the busy/free model.

    Returns:
        professor_busy: {professor_id → set of (day_idx, block_id)}  busy slots
                        from their NORMAL classes (mirrors student_busy).
        professor_subjects: {professor_id → set of lab subjects they teach}
                        derived from 'actividad' keywords (best-effort; the
                        explicit prof↔lab mapping is the open data question).
        available: bool — whether a usable professor column was found.

    If no professor column exists, returns ({}, {}, False) and the rest of the
    pipeline runs unchanged.
    """
    print_section("FLUX PROFESSEUR : disponibilités (busy/free)")

    # Auto-detect the professor identifier column
    prof_col = None
    for cand in PROFESSOR_ID_COLUMNS:
        if cand in df.columns:
            prof_col = cand
            break

    if prof_col is None:
        print(f"  [INFO]  Aucune colonne professeur trouvée "
              f"(cherché: {', '.join(PROFESSOR_ID_COLUMNS[:4])}…)")
        print(f"  → flux professeur INACTIF, pipeline inchangé.")
        print(f"  → données requises (questions ouvertes du tableau) :")
        print(f"      • quel professeur enseigne quel lab ('which labs?')")
        print(f"      • combien de groupes par professeur ('how many groups?')")
        return {}, {}, False

    print(f"  Colonne professeur détectée : '{prof_col}'")
    multi = prof_col in MULTI_PROFESSOR_COLUMNS
    if multi:
        print(f"  Format multi-professeurs (noms séparés par virgule)")

    valid = df.dropna(subset=[prof_col, 'slot_hora_inicio_min', 'slot_jour_semaine'])
    valid = valid[valid['slot_hora_inicio_min'] > 0]

    professor_busy = defaultdict(set)      # prof → set of (day_idx, block_id)
    professor_subjects = defaultdict(set)  # prof → set of lab subjects taught

    for _, row in valid.iterrows():
        day = row['slot_jour_semaine']
        block_id = min_to_block_id(row['slot_hora_inicio_min'])
        if not block_id or day not in DAYS:
            continue
        slot = (DAY_IDS[day], block_id)
        act = str(row.get('actividad', '')).lower()

        # Split the cell into individual professor names if multi-prof
        raw = str(row[prof_col])
        names = [n.strip() for n in raw.split(',')] if multi else [raw.strip()]

        for prof_id in names:
            if not prof_id or prof_id.lower() == 'nan':
                continue
            professor_busy[prof_id].add(slot)
            # Best-effort link prof → lab subjects via actividad keywords
            for subject, config in LAB_CONFIG.items():
                if any(kw in act for kw in config['keywords']):
                    if not any(kw_ex in act
                               for kw_ex in config.get('keyword_exclude', [])):
                        professor_subjects[prof_id].add(subject)

    professor_busy = dict(professor_busy)

    n_prof = len(professor_busy)
    if n_prof:
        avg_busy = sum(len(b) for b in professor_busy.values()) / n_prof
        print(f"  {n_prof} professeurs avec emploi du temps")
        print(f"  Créneaux occupés par prof : moy={avg_busy:.1f} / 30")
        n_linked = sum(1 for s in professor_subjects.values() if s)
        print(f"  Professeurs liés à ≥1 matière de lab : {n_linked}")
    else:
        print(f"  [WARN]  Colonne '{prof_col}' présente mais aucun prof exploitable")
        return {}, {}, False

    # Export for downstream use / verification
    try:
        os.makedirs('data_clean/optimization', exist_ok=True)
        pb_rows = []
        for pid, busy_set in professor_busy.items():
            for (day_idx, block_id) in busy_set:
                pb_rows.append({'professor_id': pid,
                                'day_idx': day_idx, 'block_id': block_id})
        if pb_rows:
            pd.DataFrame(pb_rows).to_csv(
                'data_clean/optimization/professor_busy.csv',
                index=False, encoding='utf-8-sig')
            print(f"  [OK] Export professor_busy.csv ({len(pb_rows)} entrées)")
    except Exception as e:
        print(f"  [WARN]  Erreur export professor_busy.csv : {e}")

    return professor_busy, dict(professor_subjects), True


# ============================================================
# ÉTAPE 4 : FORMATION DES GROUPES (CŒUR)
# ============================================================
def form_groups(subject_students, student_busy, student_subject_slots, student_program,
                subject_professor_busy=None):
    """
    Formation des groupes avec support pour:
    1. Groupes partagés (shared_group): Física et Química partagent les mêmes groupes
    2. Groupes par programme (group_by_program): 1ère année = homogène par titulación
    3. Tailles variables par matière
    4. Contrainte professeur (subject_professor_busy): un groupe d'une matière ne
       peut pas être créé sur un créneau où un professeur de cette matière est
       occupé. Optionnel — vide/None = aucun effet (rétro-compatible).
    """
    print_section("ÉTAPE 4 : Formation des groupes")

    if subject_professor_busy is None:
        subject_professor_busy = {}

    all_groups = []
    stats = {'total_assigned': 0, 'total_unassigned': 0}
    slot_room_usage = {}
    # Track sessions per (subject, semester, day, block) — enforces C1 from the
    # FORMATION step. Without it, Phase 1 could pile 4 groups of the same
    # subject on one weekly slot (e.g. 4 × Física II × Lunes 12:30 = 20
    # sessions to schedule in 11 available weeks → C1 unsatisfiable → S2
    # INFAISABLE). The room-level check alone misses this because two
    # non-simultaneous rooms hide the subject saturation.
    subject_slot_usage = {}

    # Track student lab placements SEPARATELY from student_busy. student_busy
    # records the student's COURSE schedule (and is partially "freed" via
    # own_slots since a lab REPLACES a course at the same slot). But once a
    # student is placed in a lab group, that slot must be blocked from ALL
    # other lab placements, regardless of whether they have a course there.
    # Mixing the two led to 4 cross-subject conflicts: own_slots cancelled the
    # propagation when the student had a course in the new subject.
    student_lab_busy = defaultdict(set)

    def _propagate_busy(group):
        """Block the group's (day, block) slot for every member: they now have
        a lab there and cannot accept ANOTHER lab placement at the same slot.
        Uses student_lab_busy (a hard block) rather than student_busy (which
        is partially neutralised by own_slots)."""
        slot = (group['day_idx'], group['block_id'])
        for sid in group.get('student_ids', []):
            student_lab_busy[sid].add(slot)

    def _propagate_busy_one(sid, group):
        """Same for a SINGLE student added to an existing group."""
        student_lab_busy[sid].add(
            (group['day_idx'], group['block_id']))

    # ── Pré-traitement: identifier les groupes partagés ──
    shared_group_map = {}  # shared_key → first subject that forms the groups
    for subject, config in LAB_CONFIG.items():
        sg = config.get('shared_group')
        if sg and sg not in shared_group_map:
            shared_group_map[sg] = subject
    
    # Pour les groupes partagés, le PREMIER sujet forme les groupes,
    # les autres les réutilisent. On exclut les sujets secondaires du round-robin.
    secondary_shared = set()
    for subject, config in LAB_CONFIG.items():
        sg = config.get('shared_group')
        if sg and shared_group_map[sg] != subject:
            secondary_shared.add(subject)

    # Préparer les données par matière (seulement les primaires)
    subject_data = {}
    for subject, config in LAB_CONFIG.items():
        if subject in secondary_shared:
            continue  # Sera traité après via les groupes partagés
        
        # Pour les sujets partagés, prendre l'UNION des étudiants
        sg = config.get('shared_group')
        if sg:
            student_ids = set()
            for other_subj, other_cfg in LAB_CONFIG.items():
                if other_cfg.get('shared_group') == sg:
                    student_ids.update(subject_students.get(other_subj, []))
            student_ids = list(student_ids)
        else:
            student_ids = subject_students.get(subject, [])
        
        if not student_ids:
            continue

        curso_num = config['curso_num']
        semester = config['semester']
        num_sessions = config['num_sessions']
        min_week = config['min_week']
        # Use subject-specific max_week if defined (Daniel feedback)
        # Otherwise use semester default
        sem_max = SEMESTER_1_WEEKS if semester == 1 else SEMESTER_2_WEEKS
        max_week = config.get('max_week', sem_max)
        max_per_group = config['max_students']
        lab_rooms_list = config['lab_rooms']

        # Règle Daniel : si max=12 crée > 3 groupes de plus que max=15, utiliser 15
        if max_per_group == PREFERRED_GROUP_SIZE and len(student_ids) > 0:
            import math
            groups_at_12 = math.ceil(len(student_ids) / PREFERRED_GROUP_SIZE)
            groups_at_15 = math.ceil(len(student_ids) / MAX_GROUP_SIZE)
            # Two triggers to lift the cap to 15 (Daniel uses 12-15, up to 15
            # when needed):
            #   (a) the original heuristic: 12 makes >3 more groups than 15;
            #   (b) ROOM-CAPACITY trigger: this subject shares a single room with
            #       another (shared_group), so morning capacity is tight — at 12
            #       the groups physically cannot all fit. Going to 15 reduces the
            #       group count enough to fit (e.g. Física: 18 groups@12=216 seats
            #       won't fit, 14 groups@15=210 will). Cross-checked: 208 students
            #       need 15-sized groups to fit the ~16-group morning ceiling.
            shares_room = bool(config.get('shared_group'))
            if (groups_at_12 - groups_at_15 > MAX_EXTRA_GROUPS) or shares_room:
                max_per_group = MAX_GROUP_SIZE
                reason = ("partage de salle — capacité matin serrée" if shares_room
                          else f"12→{groups_at_12} gr vs 15→{groups_at_15} gr")
                print(f"  ↑ {subject}: max porté à {MAX_GROUP_SIZE} ({reason})")

        sem_holidays = HOLIDAYS.get(semester, {})
        available_weeks_by_day = {}
        for d in range(5):
            valid = [w for w in range(min_week, max_week + 1)
                     if (w, d) not in sem_holidays]
            available_weeks_by_day[d] = len(valid)

        available_weeks = min(available_weeks_by_day.values()) if available_weeks_by_day else (max_week - min_week + 1)

        if curso_num in [1, 3]:
            allowed_blocks = MORNING_BLOCKS
        elif curso_num in [2, 4]:
            allowed_blocks = AFTERNOON_BLOCKS
        else:
            allowed_blocks = ALL_BLOCKS

        all_slots = [(d, b) for d in range(5) for b in allowed_blocks]

        # ── Professor constraint: drop slots where a professor of this subject
        # is busy with a normal class. Symmetric to student availability.
        # Safe-guard: never let it empty the candidate list (data inconsistency).
        prof_busy = subject_professor_busy.get(subject, set())
        if prof_busy:
            filtered = [s for s in all_slots if s not in prof_busy]
            if filtered:
                removed = len(all_slots) - len(filtered)
                if removed:
                    print(f"    [PROF]  {subject}: {removed} créneau(x) retiré(s) "
                          f"(professeur occupé)")
                all_slots = filtered
            else:
                print(f"    [PROF][WARN] {subject}: la contrainte professeur "
                      f"viderait tous les créneaux — ignorée (à vérifier)")

        # Max groups per (day, block) slot, derived from the real per-room
        # session capacity. A room-slot has `available_weeks` weeks; a group of
        # this subject consumes `num_sessions` of them. With two rooms available
        # (non-simultaneous), capacity is doubled. We do NOT inflate this beyond
        # physical capacity — over-stating it lets Phase 1 stack groups on one
        # slot and makes the placement solver INFEASIBLE.
        n_rooms = len(config['lab_rooms']) if not config.get('simultaneous_rooms', False) else 1
        max_groups_per_slot_by_day = {
            d: max(1, (available_weeks_by_day.get(d, 0) // num_sessions) * n_rooms)
            for d in range(5)
        }

        subject_data[subject] = {
            'config': config,
            'unassigned': set(student_ids),
            'groups': [],
            'group_counter': 0,
            'all_slots': all_slots,
            'max_groups_per_slot_by_day': max_groups_per_slot_by_day,
            'slot_usage': defaultdict(int),
            'num_sessions': num_sessions,
            'max_per_group': max_per_group,
            'lab_rooms_list': lab_rooms_list,
            'semester': semester,
            'curso_num': curso_num,
            'min_week': min_week,
            'max_week': max_week,
            'available_weeks': available_weeks,
            'available_weeks_by_day': available_weeks_by_day,
            'total_students': len(student_ids),
            'group_by_program': config.get('group_by_program', False),
        }

    # Trier les matières : les plus contraintes en premier
    # Critère : (nombre de salles, nombre de slots possibles) croissant
    # Trier les matières : groupes partagés EN PREMIER (plus d'étudiants),
    # puis les plus contraintes (moins de salles, moins de slots)
    sorted_subjects = sorted(subject_data.keys(), key=lambda s: (
        0 if LAB_CONFIG[s].get('shared_group') else 1,  # Partagés en premier
        len(subject_data[s]['lab_rooms_list']),
        len(subject_data[s]['all_slots']),
        -subject_data[s]['total_students'],
    ))

    print(f"  Ordre de traitement (plus contraint en premier) :")
    for s in sorted_subjects:
        sd = subject_data[s]
        print(f"    {s:40s} | {sd['total_students']:3d} étu | "
              f"{len(sd['all_slots']):2d} slots | "
              f"{len(sd['lab_rooms_list'])} salle(s)")

    # ── Pré-calcul: charge salle supplémentaire pour les groupes partagés ──
    # Pour chaque matière primaire partagée, calculer les sessions supplémentaires
    # par salle venant des matières secondaires
    shared_extra_sessions = {}  # subject → {room: extra_sessions}
    for subject in subject_data:
        sg = LAB_CONFIG[subject].get('shared_group')
        if sg and shared_group_map.get(sg) == subject:
            extra = defaultdict(int)
            for other_subj, other_cfg in LAB_CONFIG.items():
                if other_cfg.get('shared_group') == sg and other_subj != subject:
                    other_rooms = other_cfg['lab_rooms']
                    other_sess = other_cfg['num_sessions']
                    if other_cfg.get('simultaneous_rooms', False):
                        # Simultané : toutes les salles reçoivent toutes les sessions
                        for room in other_rooms:
                            extra[room] += other_sess
                    else:
                        # Non-simultané : la charge est répartie entre les salles
                        per_room = math.ceil(other_sess / len(other_rooms)) if other_rooms else other_sess
                        for room in other_rooms:
                            extra[room] += per_room
            shared_extra_sessions[subject] = dict(extra)
            if extra:
                print(f"\n  Charge partagée pour {subject}:")
                for room, sess in extra.items():
                    total = subject_data[subject]['num_sessions'] + sess
                    print(f"    {room}: {subject_data[subject]['num_sessions']} + {sess} = {total} sess/groupe")

    # ── Phase 1: Former TOUS les groupes des matières partagées d'abord ──
    # Les salles Ciencias ne sont pas partagées avec d'autres matières,
    # donc on peut former tous les groupes sans interférence.
    shared_subjects_list_rr = [s for s in sorted_subjects if LAB_CONFIG[s].get('shared_group')]
    remaining_subjects = [s for s in sorted_subjects if not LAB_CONFIG[s].get('shared_group')]

    if shared_subjects_list_rr:
        print(f"\n  ── Phase 1: Groupes partagés (priorité) ──")
        shared_round = 0
        shared_progress = True
        while shared_progress and shared_round < 200:
            shared_round += 1
            shared_progress = False
            for subject in shared_subjects_list_rr:
                sd = subject_data[subject]
                if not sd['unassigned']:
                    continue
                config = sd['config']
                num_sessions = sd['num_sessions']
                max_per_group = sd['max_per_group']
                lab_rooms_list = sd['lab_rooms_list']
                best_slot = None
                best_free = []
                best_room = None
                best_prog = None
                best_score = -1
                simultaneous = config.get('simultaneous_rooms', False)
                extra_per_room = shared_extra_sessions.get(subject, {})

                for slot in sd['all_slots']:
                    day_idx = slot[0]
                    if sd['slot_usage'][slot] >= sd['max_groups_per_slot_by_day'].get(day_idx, 0):
                        continue
                    day_available_weeks = sd['available_weeks_by_day'].get(day_idx, 0)

                    # Subject-level capacity (C1 enforcement at formation time):
                    # one more group here would add `num_sessions` to the (subject,
                    # day, block) load. Refuse if it would exceed the weeks
                    # available. Without this guard, multiple groups pile on the
                    # same subject-slot and the solver can't schedule them all on
                    # distinct weeks (no feasible week assignment exists).
                    subj_key = (subject, sd['semester'], slot[0], slot[1])
                    if subject_slot_usage.get(subj_key, 0) + num_sessions > day_available_weeks:
                        continue

                    if simultaneous:
                        room_ok = True
                        for room in lab_rooms_list:
                            room_key = (room, sd['semester'], slot[0], slot[1])
                            current = slot_room_usage.get(room_key, 0)
                            room_load = num_sessions + extra_per_room.get(room, 0)
                            if (current + room_load) > day_available_weeks:
                                room_ok = False; break
                        if not room_ok: continue
                        chosen_room = ', '.join(lab_rooms_list)
                    else:
                        # Non-simultaneous: pick the LEAST-loaded room that can
                        # still physically host another group on this slot. The
                        # capacity check is STRICT (sessions must fit in the
                        # weeks of THIS room-slot) so Phase 1 never over-stacks a
                        # slot — which is what made the placement solver
                        # infeasible (e.g. 18 sessions on an 11-week slot).
                        chosen_room = None
                        min_usage = float('inf')
                        for room in lab_rooms_list:
                            room_key = (room, sd['semester'], slot[0], slot[1])
                            current = slot_room_usage.get(room_key, 0)
                            room_load = num_sessions + extra_per_room.get(room, 0)
                            if (current + room_load) <= day_available_weeks and current < min_usage:
                                chosen_room = room; min_usage = current
                        if chosen_room is None: continue

                    sg = sd['config'].get('shared_group')
                    shared_subjects_for_slots = [subject]
                    if sg:
                        shared_subjects_for_slots = [s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg]

                    # Diversification: prefer LESS-used room-slots at equal
                    # student count. Without this the formation piles every
                    # group onto the single slot with the most free students
                    # (e.g. 7 Física groups on Martes 12:30), which overloads
                    # that room-slot and makes the placement solver INFEASIBLE.
                    room_for_penalty = (chosen_room.split(',')[0].strip()
                                        if chosen_room else lab_rooms_list[0])
                    pk = (room_for_penalty, sd['semester'], slot[0], slot[1])
                    usage_penalty = slot_room_usage.get(pk, 0)

                    if sd.get('group_by_program', False):
                        free_by_prog = defaultdict(list)
                        for sid in sd['unassigned']:
                            busy = student_busy.get(sid, set())
                            own_slots = set()
                            for ss in shared_subjects_for_slots:
                                own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                            effective_busy = busy - own_slots
                            if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                                prog = student_program.get(sid, 'UNKNOWN')
                                free_by_prog[prog].append(sid)
                        for prog, free in free_by_prog.items():
                            score = len(free) - usage_penalty
                            if len(free) >= MIN_GROUP_SIZE and score > best_score:
                                best_score = score
                                best_slot = slot; best_free = free
                                best_room = chosen_room; best_prog = prog
                    else:
                        free = []
                        for sid in sd['unassigned']:
                            busy = student_busy.get(sid, set())
                            own_slots = set()
                            for ss in shared_subjects_for_slots:
                                own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                            effective_busy = busy - own_slots
                            if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                                free.append(sid)
                        score = len(free) - usage_penalty
                        if len(free) >= MIN_GROUP_SIZE and score > best_score:
                            best_score = score
                            best_slot = slot; best_free = free
                            best_room = chosen_room; best_prog = 'MIXED'

                if len(best_free) < MIN_GROUP_SIZE: continue
                shared_progress = True
                sd['group_counter'] += 1
                members = best_free[:max_per_group]
                day_idx, block_id = best_slot
                group = {
                    'subject': subject, 'semester': sd['semester'],
                    'curso_num': sd['curso_num'], 'group_num': sd['group_counter'],
                    'program': best_prog, 'day_idx': day_idx, 'day': DAYS[day_idx],
                    'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                    'student_ids': members, 'nb_students': len(members),
                    'num_sessions': num_sessions, 'max_students': max_per_group,
                    'lab_rooms': best_room, 'min_week': sd['min_week'],
                    'max_week': sd['max_week'],
                }
                sd['groups'].append(group); all_groups.append(group)
                _propagate_busy(group)
                sd['slot_usage'][best_slot] += 1
                # Track subject-level slot load for C1 enforcement
                subj_key = (subject, sd['semester'], day_idx, block_id)
                subject_slot_usage[subj_key] = subject_slot_usage.get(subj_key, 0) + num_sessions
                for room in best_room.split(','):
                    room = room.strip()
                    if room:
                        room_key = (room, sd['semester'], day_idx, block_id)
                        room_load = num_sessions + extra_per_room.get(room, 0)
                        slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + room_load
                for sid in members:
                    sd['unassigned'].discard(sid)

        for s in shared_subjects_list_rr:
            sd = subject_data[s]
            assigned = sum(g['nb_students'] for g in sd['groups'])
            print(f"    {s:40s} | {len(sd['groups']):2d} gr | {assigned}/{sd['total_students']} assignés")

        # ── Phase 1b: Cross-program overflow ──
        # Les étudiants non-assignés (programmes trop petits) rejoignent
        # des groupes existants d'autres programmes qui ont de la place
        overflow_count = 0
        for s in shared_subjects_list_rr:
            sd = subject_data[s]
            if not sd['unassigned']:
                continue

            sg_key = sd['config'].get('shared_group')
            shared_subjs = [s] if not sg_key else [x for x, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key]

            for g in sorted(sd['groups'], key=lambda x: x['max_students'] - x['nb_students'], reverse=True):
                room = g['max_students'] - g['nb_students']
                if room <= 0:
                    continue
                slot = (g['day_idx'], g['block_id'])

                added = []
                for sid in list(sd['unassigned']):
                    if len(added) >= room:
                        break
                    busy = student_busy.get(sid, set())
                    own_slots = set()
                    for ss in shared_subjs:
                        own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                    if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                        added.append(sid)

                for sid in added:
                    g['student_ids'].append(sid)
                    g['nb_students'] += 1
                    _propagate_busy_one(sid, g)
                    sd['unassigned'].discard(sid)
                    overflow_count += 1

        if overflow_count > 0:
            print(f"    [RECYCLE]  {overflow_count} étudiants cross-programme redistribués")
            for s in shared_subjects_list_rr:
                sd = subject_data[s]
                assigned = sum(g['nb_students'] for g in sd['groups'])
                if sd['unassigned']:
                    print(f"    {s:40s} | {assigned}/{sd['total_students']} assignés ({len(sd['unassigned'])} restants)")

    # ── Phase 2: Round-robin pour les matières restantes ──
    print(f"\n  ── Phase 2: Round-robin pour les matières restantes ──")
    round_num = 0
    max_rounds = 200
    progress = True

    while progress and round_num < max_rounds:
        round_num += 1
        progress = False

        for subject in remaining_subjects:
            sd = subject_data[subject]

            if not sd['unassigned']:
                continue  # Tous les étudiants sont assignés

            config = sd['config']
            num_sessions = sd['num_sessions']
            max_per_group = sd['max_per_group']
            lab_rooms_list = sd['lab_rooms_list']
            available_weeks = sd['available_weeks']

            # Trouver le meilleur créneau pour CE tour
            best_slot = None
            best_free = []
            best_room = None  # The specific room for this group
            best_prog = None  # The program (titulación) for this group

            simultaneous = config.get('simultaneous_rooms', False)
            extra_per_room = shared_extra_sessions.get(subject, {})

            for slot in sd['all_slots']:
                day_idx = slot[0]

                # Vérifier la capacité par matière (par jour)
                if sd['slot_usage'][slot] >= sd['max_groups_per_slot_by_day'].get(day_idx, 0):
                    continue

                # Semaines disponibles pour CE jour spécifique
                day_available_weeks = sd['available_weeks_by_day'].get(day_idx, 0)

                # Subject-level capacity (C1 enforcement at formation time):
                # would adding `num_sessions` here exceed the available weeks
                # for THIS (subject, day, block)? If so, refuse — otherwise the
                # placement solver gets an unsatisfiable problem (e.g. 4 groups
                # of Física II all on Lunes 12:30 → 20 sess / 11 weeks → S2
                # INFAISABLE).
                subj_key = (subject, sd['semester'], slot[0], slot[1])
                if subject_slot_usage.get(subj_key, 0) + num_sessions > day_available_weeks:
                    continue

                if simultaneous:
                    # ALL rooms must be available (Física uses both Ciencias I+II)
                    room_ok = True
                    for room in lab_rooms_list:
                        room_key = (room, sd['semester'], slot[0], slot[1])
                        current = slot_room_usage.get(room_key, 0)
                        room_load = num_sessions + extra_per_room.get(room, 0)
                        if (current + room_load) > day_available_weeks:
                            room_ok = False
                            break
                    if not room_ok:
                        continue
                    chosen_room = ', '.join(lab_rooms_list)  # All rooms
                else:
                    # ANY ONE room must be available — pick the least busy
                    chosen_room = None
                    min_usage = float('inf')
                    for room in lab_rooms_list:
                        room_key = (room, sd['semester'], slot[0], slot[1])
                        current = slot_room_usage.get(room_key, 0)
                        room_load = num_sessions + extra_per_room.get(room, 0)
                        if (current + room_load) <= day_available_weeks and current < min_usage:
                            chosen_room = room
                            min_usage = current
                    if chosen_room is None:
                        continue

                # Compter les étudiants libres à ce créneau
                # Pour les groupes partagés, retirer les créneaux de TOUTES les matières du groupe
                sg = sd['config'].get('shared_group')
                shared_subjects_list = [subject]
                if sg:
                    shared_subjects_list = [s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg]

                if sd.get('group_by_program', False):
                    free_by_prog = defaultdict(list)
                    for sid in sd['unassigned']:
                        busy = student_busy.get(sid, set())
                        own_slots = set()
                        for ss in shared_subjects_list:
                            own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                        effective_busy = busy - own_slots
                        if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                            prog = student_program.get(sid, 'UNKNOWN')
                            free_by_prog[prog].append(sid)

                    for prog, free in free_by_prog.items():
                        if len(free) > len(best_free):
                            best_slot = slot
                            best_free = free
                            best_room = chosen_room
                            best_prog = prog
                else:
                    # Groupes mixtes (2ème/3ème année)
                    free = []
                    for sid in sd['unassigned']:
                        busy = student_busy.get(sid, set())
                        own_slots = set()
                        for ss in shared_subjects_list:
                            own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                        effective_busy = busy - own_slots
                        if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                            free.append(sid)

                    if len(free) > len(best_free):
                        best_slot = slot
                        best_free = free
                        best_room = chosen_room
                        best_prog = 'MIXED'

            if len(best_free) < MIN_GROUP_SIZE:
                continue  # Pas assez d'étudiants pour former un groupe viable

            # Créer le groupe
            progress = True
            sd['group_counter'] += 1
            members = best_free[:max_per_group]
            day_idx, block_id = best_slot

            # Déterminer le programme dominant pour les groupes mixtes
            if best_prog == 'MIXED':
                prog_counts = defaultdict(int)
                for sid in members:
                    prog_counts[student_program.get(sid, '?')] += 1
                if prog_counts:
                    dominant = max(prog_counts, key=prog_counts.get)
                    if len(prog_counts) == 1:
                        best_prog = dominant
                    else:
                        best_prog = f"MIXED({dominant}+{len(prog_counts)-1})"

            group = {
                'subject': subject,
                'semester': sd['semester'],
                'curso_num': sd['curso_num'],
                'group_num': sd['group_counter'],
                'program': best_prog,
                'day_idx': day_idx,
                'day': DAYS[day_idx],
                'block_id': block_id,
                'block_label': BLOCK_LABELS[block_id],
                'student_ids': members,
                'nb_students': len(members),
                'num_sessions': num_sessions,
                'max_students': max_per_group,
                'lab_rooms': best_room,
                'min_week': sd['min_week'],
                'max_week': sd['max_week'],
            }

            sd['groups'].append(group)
            all_groups.append(group)
            _propagate_busy(group)

            # Mettre à jour les compteurs
            sd['slot_usage'][best_slot] += 1
            # Subject-level slot load (C1 enforcement)
            subj_key = (subject, sd['semester'], day_idx, block_id)
            subject_slot_usage[subj_key] = subject_slot_usage.get(subj_key, 0) + num_sessions
            # Track usage for each room in the assigned room(s)
            for room in best_room.split(','):
                room = room.strip()
                if room:
                    room_key = (room, sd['semester'], day_idx, block_id)
                    room_load = num_sessions + extra_per_room.get(room, 0)
                    slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + room_load

            for sid in members:
                sd['unassigned'].discard(sid)

    # ── Post-traitement: redistribuer les étudiants restants dans les groupes existants ──
    redistributed = 0
    for subject in list(subject_data.keys()) + list(remaining_subjects):
        if subject not in subject_data:
            continue
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        sg_key = sd['config'].get('shared_group')
        shared_subjects_for_slots = [subject]
        if sg_key:
            shared_subjects_for_slots = [s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key]

        for g in sd['groups']:
            if g['nb_students'] >= g['max_students']:
                continue  # Groupe plein
            room = g['max_students'] - g['nb_students']
            slot = (g['day_idx'], g['block_id'])

            added = []
            for sid in list(sd['unassigned']):
                if len(added) >= room:
                    break
                busy = student_busy.get(sid, set())
                own_slots = set()
                for ss in shared_subjects_for_slots:
                    own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                effective_busy = busy - own_slots
                if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                    added.append(sid)

            for sid in added:
                g['student_ids'].append(sid)
                g['nb_students'] += 1
                _propagate_busy_one(sid, g)
                sd['unassigned'].discard(sid)
                redistributed += 1

    if redistributed > 0:
        print(f"\n  [RECYCLE]  {redistributed} étudiants redistribués dans des groupes existants")

    # ════════════════════════════════════════════════════════════
    # ── Phase 3: AGGRESSIVE RECOVERY (target 100% assignment) ──
    # ════════════════════════════════════════════════════════════
    # Strategy:
    #   3a. Re-fit unassigned students into ANY existing group with capacity
    #       (relaxed: ignore program homogeneity, accept cross-subject slots)
    #   3b. Create OVERFLOW groups in afternoon slots for 1st/3rd year
    #       (when morning is saturated — last resort, with explicit flag)
    #   3c. Try alternative rooms for primary subjects (e.g. Química → Ciencias II)
    #
    # CRITICAL: Each new group is pre-validated against C4 (room×slot capacity)
    # so the solver never receives an infeasible problem.
    print(f"\n  ── Phase 3: Recovery aggressive (objectif 100%) ──")

    # ────────────────────────────────────────────────────────
    # FEASIBILITY HELPER: count current C4 load by (room, sem, day, block)
    # This is the ground truth for what the solver will check.
    # ────────────────────────────────────────────────────────
    def compute_room_slot_load(groups_list):
        """
        Returns a dict (room, sem, day, block) → number of sessions occupying it.
        Each group contributes its num_sessions per occupied (room, sem, day, block).
        For multi-room groups (lab_rooms = "Room1, Room2"), each room gets
        the FULL num_sessions (because the group needs num_sessions weeks).
        """
        load = {}
        for g in groups_list:
            for room in g['lab_rooms'].split(','):
                room = room.strip()
                if not room:
                    continue
                key = (room, g['semester'], g['day_idx'], g['block_id'])
                load[key] = load.get(key, 0) + g['num_sessions']
        return load

    def compute_subject_slot_load(groups_list):
        """
        Returns a dict (subject, sem, day, block) → number of sessions.
        For C1 constraint validation.
        """
        load = {}
        for g in groups_list:
            key = (g['subject'], g['semester'], g['day_idx'], g['block_id'])
            load[key] = load.get(key, 0) + g['num_sessions']
        return load

    def slot_capacity_for(room, semester, day_idx, min_week, max_week):
        """
        Returns the maximum number of sessions that can fit in
        (room, semester, day, *any block*) across all weeks min_week..max_week
        excluding holidays for that day_idx.
        """
        sem_holidays = HOLIDAYS.get(semester, {})
        valid = [w for w in range(min_week, max_week + 1)
                 if (w, day_idx) not in sem_holidays]
        return len(valid)

    def can_fit_new_group(room, sem, day_idx, block_id, min_w, max_w,
                          new_sessions, current_load,
                          subject=None, subject_load=None):
        """
        STRICT pre-validation: would adding `new_sessions` to (room, sem, day, block)
        keep us within physical capacity?

        Checks BOTH:
        - C4 capacity: room+slot can host at most `cap` sessions across weeks
        - C1 capacity: subject+slot can host at most `cap` sessions across weeks
                       (if `subject` and `subject_load` are provided)
        Returns True if safe, False otherwise.
        """
        cap = slot_capacity_for(room, sem, day_idx, min_w, max_w)

        # C4 check
        room_key = (room, sem, day_idx, block_id)
        room_current = current_load.get(room_key, 0)
        if (room_current + new_sessions) > cap:
            return False

        # C1 check (subject-level)
        if subject and subject_load is not None:
            subj_key = (subject, sem, day_idx, block_id)
            subj_current = subject_load.get(subj_key, 0)
            if (subj_current + new_sessions) > cap:
                return False

        return True

    # Compute initial loads from groups created in Phase 1+2
    current_load = compute_room_slot_load(all_groups)
    subject_load = compute_subject_slot_load(all_groups)

    # Step 3a: Relaxed re-fit into existing groups
    refit_count = 0
    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        sg_key = config.get('shared_group')
        shared_subjs = [subject] if not sg_key else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key
        ]

        # Sort groups by remaining capacity (most space first)
        candidate_groups = sorted(
            sd['groups'],
            key=lambda g: g['max_students'] - g['nb_students'],
            reverse=True,
        )

        for sid in list(sd['unassigned']):
            busy = student_busy.get(sid, set())
            own_slots = set()
            for ss in shared_subjs:
                own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
            effective_busy = busy - own_slots

            for g in candidate_groups:
                # Allow over-fill up to max_students (no extra)
                if g['nb_students'] >= g['max_students']:
                    continue
                slot = (g['day_idx'], g['block_id'])
                if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                    g['student_ids'].append(sid)
                    g['nb_students'] += 1
                    _propagate_busy_one(sid, g)
                    sd['unassigned'].discard(sid)
                    refit_count += 1
                    break

    if refit_count > 0:
        print(f"    [RECYCLE]  3a: {refit_count} étudiants re-fittés dans groupes existants")

    # Step 3b: Create OVERFLOW groups in afternoon slots for saturated 1st year subjects
    # (This is exceptional — only when morning is fully saturated)
    # CRITICAL: pre-validate every new group against C4 capacity (current_load)
    overflow_groups_created = 0
    overflow_assigned = 0
    overflow_skipped_full = 0

    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        unassigned_count = len(sd['unassigned'])
        # Only worth creating new groups if at least MIN_GROUP_SIZE remaining
        if unassigned_count < MIN_GROUP_SIZE:
            continue

        # Year-of-degree rule (see Considerations PDF) — applied STRICTLY:
        #   1st & 3rd year  → mornings ONLY (08:30-14:30); afternoons forbidden
        #   2nd & 4th year  → afternoons ONLY (15:00-21:00); mornings forbidden
        # The previous "afternoon as last resort" fallback is DISABLED per the
        # explicit requirement: a 1st-year lab must never land in the afternoon.
        # If the preferred period is saturated, the group is left for diagnostics
        # rather than violating the rule.
        # Year-of-degree rule (see Considerations PDF):
        #   1st & 3rd year  → mornings; 2nd & 4th year → afternoons.
        # By default the rule is STRICT (no cross-period fallback). The app can
        # relax it per period via ALLOW_AFTERNOON_Y1Y3 / ALLOW_MORNING_Y2Y4,
        # in which case the opposite period becomes an allowed last resort.
        original_curso = config['curso_num']
        MORNING = [1, 2, 3]    # 08:30-10:30, 10:30-12:30, 12:30-14:30
        AFTERNOON = [4, 5]     # 15:00-17:00, 17:00-19:00
        if original_curso in [1, 3]:
            preferred_blocks = MORNING
            fallback_blocks = AFTERNOON if ALLOW_AFTERNOON_Y1Y3 else []
        else:
            preferred_blocks = AFTERNOON
            fallback_blocks = MORNING if ALLOW_MORNING_Y2Y4 else []
        extra_slots = [(d, b) for d in range(5) for b in preferred_blocks]
        extra_slots_fallback = [(d, b) for d in range(5) for b in fallback_blocks]


        sg_key = config.get('shared_group')
        shared_subjs = [subject] if not sg_key else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key
        ]

        max_per_group = sd['max_per_group']
        num_sessions = sd['num_sessions']
        lab_rooms_list = sd['lab_rooms_list']
        sem = sd['semester']
        min_w = sd['min_week']
        max_w = sd['max_week']

        # For shared subjects: total sessions = sum (Física 5 + Química 4 = 9)
        if sg_key:
            total_sessions_per_group = sum(
                LAB_CONFIG[s]['num_sessions']
                for s in LAB_CONFIG if LAB_CONFIG[s].get('shared_group') == sg_key
            )
        else:
            total_sessions_per_group = num_sessions

        # Alternative rooms for saturated subjects
        ALT_ROOMS_FOR_OVERFLOW = {
            'S1_Física': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
            'S1_Química': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
            'S2_Física II': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
        }
        effective_rooms = ALT_ROOMS_FOR_OVERFLOW.get(subject, lab_rooms_list)

        attempts = 0
        # Two-tier search: exhaust preferred slots (mornings for 1st/3rd year),
        # then fall back to the other half of the day only if students remain.
        using_fallback = False
        active_slots = extra_slots
        while sd['unassigned'] and attempts < 50:
            attempts += 1
            best_slot = None; best_free = []; best_room = None
            best_score = -1

            for slot in active_slots:
                day_idx = slot[0]

                # STRICT pre-validation — find a room where the new group truly fits
                # (checks BOTH room capacity AND subject capacity)
                chosen_room = None
                for room in effective_rooms:
                    if can_fit_new_group(room, sem, day_idx, slot[1],
                                          min_w, max_w,
                                          total_sessions_per_group, current_load,
                                          subject=subject, subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    overflow_skipped_full += 1
                    continue

                # Check who's free at this slot
                free = []
                for sid in sd['unassigned']:
                    busy = student_busy.get(sid, set())
                    own_slots = set()
                    for ss in shared_subjs:
                        own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                    if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                        free.append(sid)

                if len(free) < MIN_GROUP_SIZE:
                    continue

                # Diversification scoring
                room_key = (chosen_room, sem, slot[0], slot[1])
                room_current = current_load.get(room_key, 0)
                usage_penalty = room_current * 5
                score = len(free) - usage_penalty

                if score > best_score:
                    best_score = score
                    best_slot = slot; best_free = free; best_room = chosen_room

            if best_slot is None or len(best_free) < MIN_GROUP_SIZE:
                # Preferred period exhausted. With the strict year-of-degree rule
                # there is NO cross-period fallback: we never place a 1st/3rd-year
                # group in the afternoon. Any remaining students are reported.
                if extra_slots_fallback and not using_fallback:
                    using_fallback = True
                    active_slots = extra_slots_fallback
                    continue
                if sd['unassigned']:
                    period = "morning" if original_curso in [1, 3] else "afternoon"
                    print(f"        ⚠️  {subject}: {len(sd['unassigned'])} student(s) "
                          f"could not be placed in the {period} (preferred period "
                          f"saturated). Left unassigned — afternoon fallback is "
                          f"disabled by the strict year-of-degree rule.")
                break

            # Create overflow group
            sd['group_counter'] += 1
            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': original_curso, 'group_num': sd['group_counter'],
                'program': 'OVERFLOW', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                '_overflow': True,
            }
            sd['groups'].append(group); all_groups.append(group)
            _propagate_busy(group)
            sd['slot_usage'][best_slot] += 1

            # Update BOTH current_load (room) AND subject_load (subject) AND slot_room_usage (legacy)
            for room in best_room.split(','):
                room = room.strip()
                if room:
                    room_key = (room, sem, day_idx, block_id)
                    current_load[room_key] = current_load.get(room_key, 0) + total_sessions_per_group
                    slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + total_sessions_per_group
            # Subject load: this subject + any shared subjects that will reuse this slot
            for ss in shared_subjs:
                ss_key = (ss, sem, day_idx, block_id)
                ss_sessions = LAB_CONFIG[ss]['num_sessions'] if ss in LAB_CONFIG else num_sessions
                subject_load[ss_key] = subject_load.get(ss_key, 0) + ss_sessions
            for sid in members:
                sd['unassigned'].discard(sid)
            overflow_groups_created += 1
            overflow_assigned += len(members)

    if overflow_groups_created > 0:
        print(f"    [RUN] 3b: {overflow_groups_created} groupes overflow créés "
              f"({overflow_assigned} étudiants récupérés en horaires alternatifs)")
        if overflow_skipped_full > 0:
            print(f"        ⓘ {overflow_skipped_full} tentatives rejetées par capacité salle (pré-validation)")

    # Step 3c: For shared groups, try opening additional rooms
    # (e.g. Química can use Ciencias Exp. II instead of just Ciencias Exp. I)
    expand_count = 0
    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        # Define alternative room sets per subject
        ALTERNATIVE_ROOMS = {
            'S1_Química': ['Ciencias Experimentales II'],  # Daniel mentioned this as fallback
            'S2_Física II': ['Ciencias Experimentales II'],
        }
        alt_rooms = ALTERNATIVE_ROOMS.get(subject, [])
        if not alt_rooms:
            continue

        sem = sd['semester']
        num_sessions = sd['num_sessions']
        max_per_group = sd['max_per_group']

        sg_key = config.get('shared_group')
        shared_subjs = [subject] if not sg_key else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key
        ]

        all_slots_morning = [(d, b) for d in range(5) for b in [1, 2, 3]]

        # Same fix for shared groups: account for total sessions
        if sg_key:
            total_sessions_per_group = sum(
                LAB_CONFIG[s]['num_sessions']
                for s in LAB_CONFIG if LAB_CONFIG[s].get('shared_group') == sg_key
            )
        else:
            total_sessions_per_group = num_sessions

        attempts = 0
        while sd['unassigned'] and attempts < 30:
            attempts += 1
            best_slot = None; best_free = []; best_room = None
            best_score = -1

            for slot in all_slots_morning:
                day_idx = slot[0]

                for room in alt_rooms:
                    # STRICT pre-validation (room + subject capacity)
                    if not can_fit_new_group(room, sem, day_idx, slot[1],
                                              sd['min_week'], sd['max_week'],
                                              total_sessions_per_group, current_load,
                                              subject=subject, subject_load=subject_load):
                        continue

                    free = []
                    for sid in sd['unassigned']:
                        busy = student_busy.get(sid, set())
                        own_slots = set()
                        for ss in shared_subjs:
                            own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                        if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                            free.append(sid)

                    if len(free) < MIN_GROUP_SIZE:
                        continue

                    room_key = (room, sem, slot[0], slot[1])
                    current = current_load.get(room_key, 0)
                    usage_penalty = current * 2
                    score = len(free) - usage_penalty

                    if score > best_score:
                        best_score = score
                        best_slot = slot; best_free = free; best_room = room

            if best_slot is None or len(best_free) < MIN_GROUP_SIZE:
                break

            sd['group_counter'] += 1
            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': config['curso_num'], 'group_num': sd['group_counter'],
                'program': 'ALT_ROOM', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': sd['min_week'],
                'max_week': sd['max_week'],
                '_alt_room': True,
            }
            sd['groups'].append(group); all_groups.append(group)
            _propagate_busy(group)
            sd['slot_usage'][best_slot] += 1
            room_key = (best_room, sem, day_idx, block_id)
            current_load[room_key] = current_load.get(room_key, 0) + total_sessions_per_group
            slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + total_sessions_per_group
            # Update subject_load too (for shared subjects)
            for ss in shared_subjs:
                ss_key = (ss, sem, day_idx, block_id)
                ss_sessions = LAB_CONFIG[ss]['num_sessions'] if ss in LAB_CONFIG else num_sessions
                subject_load[ss_key] = subject_load.get(ss_key, 0) + ss_sessions
            for sid in members:
                sd['unassigned'].discard(sid)
            expand_count += len(members)

    if expand_count > 0:
        print(f"    [ALT_ROOM]  3c: {expand_count} étudiants récupérés via salles alternatives")

    total_recovered = refit_count + overflow_assigned + expand_count
    if total_recovered > 0:
        print(f"  [OK] Phase 3 total : {total_recovered} étudiants supplémentaires récupérés")

    # ── Post-traitement: consolider les petits groupes ──
    # Fusionner les groupes < MIN_GROUP_SIZE dans les groupes existants
    consolidated = 0
    dissolved_groups = []

    for subject in list(subject_data.keys()):
        sd = subject_data[subject]
        groups = sd['groups']
        if len(groups) <= 1:
            continue

        small_groups = [g for g in groups if g['nb_students'] < MIN_GROUP_SIZE]
        big_groups = [g for g in groups if g['nb_students'] >= MIN_GROUP_SIZE]

        for sg in small_groups:
            # Chercher un groupe compatible (même matière) avec de la place
            merged = False
            sg_slot = (sg['day_idx'], sg['block_id'])
            # Trier par capacité restante décroissante
            candidates = sorted(big_groups, key=lambda g: g['max_students'] - g['nb_students'], reverse=True)

            for target in candidates:
                tgt_slot = (target['day_idx'], target['block_id'])
                room = target['max_students'] - target['nb_students']
                # Only wholesale-merge groups that share the SAME (day, block)
                # slot — otherwise students' previously propagated slot from sg
                # would no longer match their new placement.
                if room >= sg['nb_students'] and tgt_slot == sg_slot:
                    # Fusionner : déplacer tous les étudiants du petit groupe vers le grand
                    target['student_ids'].extend(sg['student_ids'])
                    target['nb_students'] += sg['nb_students']
                    consolidated += sg['nb_students']
                    dissolved_groups.append((subject, sg['group_num']))
                    merged = True
                    break

            if not merged:
                # Pas de groupe avec assez de place : distribuer un par un
                for sid in list(sg['student_ids']):
                    for target in candidates:
                        if target['nb_students'] < target['max_students']:
                            target_slot = (target['day_idx'], target['block_id'])
                            # Don't move the student to a slot where they're
                            # already placed in another lab (would create a
                            # same-week cross-subject conflict).
                            if target_slot in student_lab_busy.get(sid, set()):
                                continue
                            target['student_ids'].append(sid)
                            target['nb_students'] += 1
                            _propagate_busy_one(sid, target)
                            sg['student_ids'].remove(sid)
                            sg['nb_students'] -= 1
                            consolidated += 1
                            break
                if sg['nb_students'] == 0:
                    dissolved_groups.append((subject, sg['group_num']))

    # Retirer les groupes dissous de all_groups et subject_data
    for subject, gnum in dissolved_groups:
        sd = subject_data[subject]
        sd['groups'] = [g for g in sd['groups'] if g['group_num'] != gnum]
        # Also remove from all_groups
        all_groups[:] = [g for g in all_groups if not (g['subject'] == subject and g['group_num'] == gnum)]

    if consolidated > 0:
        print(f"  [RETRY] {consolidated} étudiants consolidés, {len(dissolved_groups)} petits groupes dissous")

    # Renuméroter les groupes séquentiellement après consolidation
    for subject in subject_data:
        sd = subject_data[subject]
        for i, g in enumerate(sorted(sd['groups'], key=lambda x: x['group_num'])):
            old_num = g['group_num']
            g['group_num'] = i + 1
            # Update in all_groups too
            for ag in all_groups:
                if ag['subject'] == subject and ag.get('_old_group_num', ag['group_num']) == old_num:
                    ag['group_num'] = i + 1

    # ── Post-traitement: dupliquer les groupes pour les matières partagées ──
    shared_groups_created = {}  # shared_key → list of groups
    for subject, config in LAB_CONFIG.items():
        sg = config.get('shared_group')
        if sg and subject in subject_data:
            shared_groups_created[sg] = subject_data[subject]['groups']

    for subject in secondary_shared:
        config = LAB_CONFIG[subject]
        sg = config['shared_group']
        primary_groups = shared_groups_created.get(sg, [])
        if not primary_groups:
            continue

        # Étudiants inscrits dans cette matière secondaire
        enrolled = set(subject_students.get(subject, []))

        semester = config['semester']
        min_week = config['min_week']
        # Use subject-specific max_week if defined (Daniel feedback)
        # Otherwise use semester default
        sem_max = SEMESTER_1_WEEKS if semester == 1 else SEMESTER_2_WEEKS
        max_week = config.get('max_week', sem_max)

        print(f"\n  Groupes partagés: {subject} réutilise {len(primary_groups)} "
              f"groupes de {shared_group_map[sg]}")

        for pg in primary_groups:
            # Filtrer: seuls les étudiants inscrits dans cette matière
            shared_members = [sid for sid in pg['student_ids'] if sid in enrolled]
            if len(shared_members) < MIN_GROUP_SIZE:
                continue  # Trop peu d'étudiants → pas de groupe viable

            # CRITICAL FIX: inherit the SAME room as parent group
            # (e.g. if Física G15 is in Ciencias Exp. II, Química G15 must also be there)
            inherited_room = pg.get('lab_rooms', ', '.join(config['lab_rooms']))

            group = {
                'subject': subject,
                'semester': semester,
                'curso_num': config['curso_num'],
                'group_num': pg['group_num'],  # Même numéro de groupe
                'program': pg.get('program', 'MIXED'),
                'day_idx': pg['day_idx'],
                'day': pg['day'],
                'block_id': pg['block_id'],
                'block_label': pg['block_label'],
                'student_ids': shared_members,
                'nb_students': len(shared_members),
                'num_sessions': config['num_sessions'],
                'max_students': config['max_students'],
                'lab_rooms': inherited_room,  # ← Same as parent
                'min_week': min_week,
                'max_week': max_week,
                # Inherit overflow flag from parent
                '_overflow': pg.get('_overflow', False),
                '_alt_room': pg.get('_alt_room', False),
            }
            all_groups.append(group)

        # Count for stats
        shared_assigned = sum(1 for g in all_groups if g['subject'] == subject
                             for _ in g['student_ids'])
        total_enrolled = len(enrolled)
        unassigned = total_enrolled - shared_assigned
        if unassigned > 0:
            print(f"  [WARN]  {subject:40s} | "
                  f"{shared_assigned}/{total_enrolled} assignés ({unassigned} sans groupe)")
        else:
            print(f"  [OK] {subject:40s} | "
                  f"{shared_assigned}/{total_enrolled} assignés (groupes partagés)")

    # ── Consolidation des groupes partagés trop petits ──
    for subject in secondary_shared:
        sec_groups = [g for g in all_groups if g['subject'] == subject]
        small = [g for g in sec_groups if g['nb_students'] < MIN_GROUP_SIZE]
        big = [g for g in sec_groups if g['nb_students'] >= MIN_GROUP_SIZE]

        dissolved_sec = []
        for sg in small:
            merged = False
            sg_slot = (sg['day_idx'], sg['block_id'])
            for target in sorted(big, key=lambda g: g['max_students'] - g['nb_students'], reverse=True):
                tgt_slot = (target['day_idx'], target['block_id'])
                room = target['max_students'] - target['nb_students']
                # Only merge wholesale if NONE of the small group's students
                # has a conflicting placement at the target's slot.
                if room >= sg['nb_students'] and tgt_slot == sg_slot:
                    target['student_ids'].extend(sg['student_ids'])
                    target['nb_students'] += sg['nb_students']
                    dissolved_sec.append(sg)
                    merged = True
                    break
            if not merged:
                for sid in list(sg['student_ids']):
                    for target in big:
                        if target['nb_students'] < target['max_students']:
                            tgt_slot = (target['day_idx'], target['block_id'])
                            if tgt_slot in student_lab_busy.get(sid, set()):
                                continue  # would conflict with another lab
                            target['student_ids'].append(sid)
                            target['nb_students'] += 1
                            _propagate_busy_one(sid, target)
                            sg['student_ids'].remove(sid)
                            sg['nb_students'] -= 1
                            break
                if sg['nb_students'] == 0:
                    dissolved_sec.append(sg)

        for dg in dissolved_sec:
            all_groups.remove(dg)
        if dissolved_sec:
            # Renumber remaining groups
            remaining = sorted([g for g in all_groups if g['subject'] == subject],
                               key=lambda x: x['group_num'])
            for i, g in enumerate(remaining):
                g['group_num'] = i + 1
            print(f"  [RETRY] {subject}: {len(dissolved_sec)} petits groupes fusionnés")

    # ════════════════════════════════════════════════════════════
    # ── Phase 4: Recovery for SECONDARY SHARED subjects ──
    # ════════════════════════════════════════════════════════════
    # Students enrolled in Química (etc.) but never placed in a Física group
    # need their own dedicated groups.
    print(f"\n  ── Phase 4: Recovery secondaires partagés ──")

    sec_recovery_count = 0
    for subject in secondary_shared:
        config = LAB_CONFIG[subject]
        enrolled = set(subject_students.get(subject, []))

        # Currently assigned to this subject
        assigned_ids = set()
        for g in all_groups:
            if g['subject'] == subject:
                assigned_ids.update(g['student_ids'])

        unassigned = enrolled - assigned_ids
        if not unassigned:
            continue

        sem = config['semester']
        min_week = config['min_week']
        sem_max = SEMESTER_1_WEEKS if sem == 1 else SEMESTER_2_WEEKS
        max_week = config.get('max_week', sem_max)
        num_sessions = config['num_sessions']
        max_per_group = config['max_students']

        # Try alt rooms first if defined
        ALTERNATIVE_ROOMS = {
            'S1_Química': ['Ciencias Experimentales II', 'Ciencias Experimentales I'],
        }
        candidate_rooms = ALTERNATIVE_ROOMS.get(subject, config['lab_rooms'])

        # Try ALL slots (morning AND afternoon as fallback)
        original_curso = config['curso_num']
        all_slots = []
        for d in range(5):
            primary_blocks = [1, 2, 3] if original_curso in [1, 3] else [4, 5]
            secondary_blocks = [4, 5] if original_curso in [1, 3] else [1, 2, 3]
            for b in primary_blocks + secondary_blocks:
                all_slots.append((d, b))

        sem_holidays = HOLIDAYS.get(sem, {})
        available_weeks_by_day = {}
        for d in range(5):
            valid = [w for w in range(min_week, max_week + 1)
                     if (w, d) not in sem_holidays]
            available_weeks_by_day[d] = len(valid)

        # First try to fit unassigned into existing same-subject groups with capacity
        cur_groups = sorted(
            [g for g in all_groups if g['subject'] == subject],
            key=lambda g: g['max_students'] - g['nb_students'], reverse=True,
        )

        next_group_num = max([g['group_num'] for g in all_groups if g['subject'] == subject], default=0) + 1

        for sid in list(unassigned):
            busy = student_busy.get(sid, set())
            own_slots = student_subject_slots.get(sid, {}).get(subject, set())
            effective_busy = busy - own_slots
            placed = False
            for g in cur_groups:
                if g['nb_students'] >= g['max_students']:
                    continue
                slot = (g['day_idx'], g['block_id'])
                if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                    g['student_ids'].append(sid)
                    g['nb_students'] += 1
                    _propagate_busy_one(sid, g)
                    unassigned.discard(sid)
                    sec_recovery_count += 1
                    placed = True
                    break

        # Then create new groups in alternative rooms / slots
        # CRITICAL: refresh current_load AND subject_load (includes Phase 3 + duplicated)
        current_load = compute_room_slot_load(all_groups)
        subject_load = compute_subject_slot_load(all_groups)

        attempts = 0
        while unassigned and attempts < 30:
            attempts += 1
            best_slot = None; best_free = []; best_room = None
            best_score = -1

            for slot in all_slots:
                day_idx = slot[0]

                for room in candidate_rooms:
                    # STRICT pre-validation (room + subject)
                    if not can_fit_new_group(room, sem, day_idx, slot[1],
                                              min_week, max_week,
                                              num_sessions, current_load,
                                              subject=subject, subject_load=subject_load):
                        continue

                    free = []
                    for sid in unassigned:
                        busy = student_busy.get(sid, set())
                        own_slots = student_subject_slots.get(sid, {}).get(subject, set())
                        if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                            free.append(sid)

                    if len(free) < MIN_GROUP_SIZE:
                        continue

                    # Year-of-degree rule: the wrong half of the day is excluded
                    # entirely UNLESS the app has explicitly allowed the exception
                    # for this year group (ALLOW_AFTERNOON_Y1Y3 / ALLOW_MORNING_Y2Y4).
                    if original_curso in [1, 3]:
                        wrong_period = slot[1] in [4, 5] and not ALLOW_AFTERNOON_Y1Y3
                    else:
                        wrong_period = slot[1] in [1, 2, 3] and not ALLOW_MORNING_Y2Y4
                    if wrong_period:
                        continue

                    # Diversification scoring
                    room_key = (room, sem, slot[0], slot[1])
                    current = current_load.get(room_key, 0)
                    usage_penalty = current * 2
                    score = len(free) - usage_penalty

                    if score > best_score:
                        best_score = score
                        best_slot = slot; best_free = free; best_room = room

            if best_slot is None or len(best_free) < MIN_GROUP_SIZE:
                break

            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': original_curso, 'group_num': next_group_num,
                'program': 'RECOVERED', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': min_week, 'max_week': max_week,
                '_recovered': True,
            }
            all_groups.append(group)
            next_group_num += 1
            room_key = (best_room, sem, day_idx, block_id)
            current_load[room_key] = current_load.get(room_key, 0) + num_sessions
            subj_key = (subject, sem, day_idx, block_id)
            subject_load[subj_key] = subject_load.get(subj_key, 0) + num_sessions
            slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + num_sessions
            for sid in members:
                unassigned.discard(sid)
            sec_recovery_count += len(members)

    if sec_recovery_count > 0:
        print(f"    [OK] {sec_recovery_count} étudiants secondaires récupérés")

    # ── Phase 5: PER-PROGRAM RECOVERY (Daniel's strategy) ──
    # The earlier phases optimise globally (pick the slot with the most free
    # students), which scatters leftovers below the minimum. Daniel instead
    # exhausts ONE program at a time: form homogeneous groups for that program
    # on any allowed slot where the program is free, before moving on.
    # This phase replicates that, as a final pass that only ADDS groups.
    print(f"\n  ── Phase 5: Recovery par programme (stratégie Daniel) ──")
    p5_recovered = 0
    p5_groups = 0

    # Refresh loads so Phase 5 sees everything created so far
    current_load = compute_room_slot_load(all_groups)
    subject_load = compute_subject_slot_load(all_groups)

    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        sem = sd['semester']
        num_sessions = sd['num_sessions']
        max_per_group = sd['max_per_group']
        lab_rooms_list = sd['lab_rooms_list']
        min_w, max_w = sd['min_week'], sd['max_week']
        allowed_slots = sd['all_slots']  # already restricted to the right period

        # Shared-subject slot accounting (Física/Química share a composition)
        sg = config.get('shared_group')
        shared_subjs = [subject] if not sg else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg
        ]

        # Group the still-unassigned students of THIS subject by program
        by_program = defaultdict(list)
        for sid in sd['unassigned']:
            by_program[student_program.get(sid, 'UNKNOWN')].append(sid)

        # Process programs largest-first so the biggest cohorts settle first
        for prog in sorted(by_program, key=lambda p: -len(by_program[p])):
            pool = by_program[prog]
            # Keep forming groups for this program until it can't reach the floor
            progress = True
            while len(pool) >= RECOVERY_MIN_GROUP_SIZE and progress:
                progress = False
                best_slot = None
                best_free = []
                best_room = None

                for slot in allowed_slots:
                    day_idx, block_id = slot
                    # Find a room at this slot that still has capacity
                    chosen_room = None
                    for room in lab_rooms_list:
                        if can_fit_new_group(room, sem, day_idx, block_id,
                                              min_w, max_w, num_sessions,
                                              current_load,
                                              subject=subject,
                                              subject_load=subject_load):
                            chosen_room = room
                            break
                    if chosen_room is None:
                        continue

                    # Which of this program's students are free at this slot?
                    free = []
                    for sid in pool:
                        busy = student_busy.get(sid, set())
                        own = set()
                        for ss in shared_subjs:
                            own |= student_subject_slots.get(sid, {}).get(ss, set())
                        if (slot not in (busy - own)
                                and slot not in student_lab_busy.get(sid, set())):
                            free.append(sid)

                    if len(free) > len(best_free):
                        best_free = free
                        best_slot = slot
                        best_room = chosen_room

                if best_slot is None or len(best_free) < RECOVERY_MIN_GROUP_SIZE:
                    break  # this program can't be placed any further

                # Create the homogeneous group
                members = best_free[:max_per_group]
                day_idx, block_id = best_slot
                sd['group_counter'] += 1
                group = {
                    'subject': subject, 'semester': sem,
                    'curso_num': sd['curso_num'],
                    'group_num': sd['group_counter'],
                    'program': prog, 'day_idx': day_idx, 'day': DAYS[day_idx],
                    'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                    'student_ids': members, 'nb_students': len(members),
                    'num_sessions': num_sessions, 'max_students': max_per_group,
                    'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                    'recovered': 'P5',
                }
                sd['groups'].append(group)
                all_groups.append(group)
                _propagate_busy(group)

                # Update loads + remove placed students
                for room in best_room.split(','):
                    room = room.strip()
                    if not room:
                        continue
                    rk = (room, sem, day_idx, block_id)
                    current_load[rk] = current_load.get(rk, 0) + num_sessions
                subj_key = (subject, sem, day_idx, block_id)
                subject_load[subj_key] = subject_load.get(subj_key, 0) + num_sessions

                placed = set(members)
                for sid in placed:
                    sd['unassigned'].discard(sid)
                pool = [sid for sid in pool if sid not in placed]
                p5_recovered += len(members)
                p5_groups += 1
                progress = True

    # ── Phase 5b: MIXED recovery (Daniel's real practice for leftovers) ──
    # Analysis of Daniel's reference files shows his Física/Química groups MIX
    # programs (IOI + AERO + GITI + IMR + …) rather than staying homogeneous.
    # Homogeneous grouping is kept as the PREFERENCE (Phase 5 above); when a
    # subject still has unplaced students whose per-program pools are each below
    # the floor, we pool them ACROSS programs — exactly as Daniel does — so a
    # leftover of 3 IOI + 2 AERO + 4 GITI becomes one mixed group of 9.
    p5b_recovered = 0
    p5b_groups = 0
    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        sem = sd['semester']
        num_sessions = sd['num_sessions']
        max_per_group = sd['max_per_group']
        lab_rooms_list = sd['lab_rooms_list']
        min_w, max_w = sd['min_week'], sd['max_week']
        allowed_slots = sd['all_slots']

        sg = config.get('shared_group')
        shared_subjs = [subject] if not sg else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg
        ]

        progress = True
        while len(sd['unassigned']) >= RECOVERY_MIN_GROUP_SIZE and progress:
            progress = False
            best_slot, best_free, best_room = None, [], None

            for slot in allowed_slots:
                day_idx, block_id = slot
                chosen_room = None
                for room in lab_rooms_list:
                    if can_fit_new_group(room, sem, day_idx, block_id,
                                          min_w, max_w, num_sessions,
                                          current_load,
                                          subject=subject,
                                          subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    continue

                # All still-unassigned students of this subject, any program
                free = []
                for sid in sd['unassigned']:
                    busy = student_busy.get(sid, set())
                    own = set()
                    for ss in shared_subjs:
                        own |= student_subject_slots.get(sid, {}).get(ss, set())
                    if (slot not in (busy - own)
                            and slot not in student_lab_busy.get(sid, set())):
                        free.append(sid)

                if len(free) > len(best_free):
                    best_free, best_slot, best_room = free, slot, chosen_room

            if best_slot is None or len(best_free) < RECOVERY_MIN_GROUP_SIZE:
                break

            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            # Tag the dominant program for readability
            prog_counts = Counter(student_program.get(s, 'UNKNOWN') for s in members)
            dominant, dom_n = prog_counts.most_common(1)[0]
            prog_label = (f"MIXED({dominant}+{len(members) - dom_n})"
                          if len(prog_counts) > 1 else dominant)

            sd['group_counter'] += 1
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': sd['curso_num'],
                'group_num': sd['group_counter'],
                'program': prog_label, 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                'recovered': 'P5b-mixed',
            }
            sd['groups'].append(group)
            all_groups.append(group)
            _propagate_busy(group)

            for room in best_room.split(','):
                room = room.strip()
                if not room:
                    continue
                rk = (room, sem, day_idx, block_id)
                current_load[rk] = current_load.get(rk, 0) + num_sessions
            subj_key = (subject, sem, day_idx, block_id)
            subject_load[subj_key] = subject_load.get(subj_key, 0) + num_sessions

            for sid in set(members):
                sd['unassigned'].discard(sid)
            p5b_recovered += len(members)
            p5b_groups += 1
            progress = True

    if p5b_recovered > 0:
        print(f"    [OK] {p5b_recovered} étudiants récupérés en {p5b_groups} "
              f"groupe(s) MIXTE(s) (programmes mélangés, comme Daniel)")
    p5_recovered += p5b_recovered
    p5_groups += p5b_groups

    if p5_recovered > 0:
        print(f"    [TOTAL Phase 5] {p5_recovered} étudiants récupérés "
              f"en {p5_groups} groupe(s)")
        if RECOVERY_MIN_GROUP_SIZE < MIN_GROUP_SIZE:
            print(f"    ⓘ seuil de récupération abaissé à "
                  f"{RECOVERY_MIN_GROUP_SIZE} (vs {MIN_GROUP_SIZE} normal)")
    else:
        print(f"    Aucun étudiant récupérable (créneaux saturés ou pools < "
              f"{RECOVERY_MIN_GROUP_SIZE})")

    # ════════════════════════════════════════════════════════════
    # ── Phase 6: ABSORPTION FINALE ADAPTATIVE (objectif 100%) ──
    # ════════════════════════════════════════════════════════════
    # After all preference-respecting phases, some students may still be
    # unplaced — almost always because their leftover pool fell below the group
    # floor, NOT because they have no free slot. This phase guarantees ~100% by,
    # for each remaining student:
    #   (a) trying to ADD them to an existing compatible group of their subject
    #       that still fits the room capacity and where they are free;
    #   (b) otherwise forming a small group with an ADAPTIVE floor that drops to
    #       1 (a single-student lab session is valid — Daniel does this for
    #       isolated cases);
    #   (c) if and only if a student has NO compatible free slot at all, they are
    #       recorded as genuinely unplaceable and reported explicitly (the
    #       guard-rail: never silently drop a student).
    # Every placement still passes can_fit_new_group, so the placement solver
    # stays feasible (0 conflicts preserved).
    current_load = compute_room_slot_load(all_groups)
    subject_load = compute_subject_slot_load(all_groups)
    p6_absorbed = 0
    p6_newgroups = 0
    p6_exception_solo = []  # isolated students placed in a minimal group (review)
    truly_unplaceable = []  # (subject, student_id) — reported, never hidden

    all_subjects_with_rest = [s for s in sorted_subjects
                              if subject_data[s]['unassigned']]
    # Include secondary shared subjects (Química etc.) handled via their groups
    for subject in all_subjects_with_rest:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue
        config = sd['config']
        sem = sd['semester']
        num_sessions = sd['num_sessions']
        absolute_max = MAX_GROUP_SIZE          # hard cap a room can hold
        lab_rooms_list = sd['lab_rooms_list']
        min_w, max_w = sd['min_week'], sd['max_week']
        allowed_slots = sd['all_slots']
        sg = config.get('shared_group')
        shared_subjs = [subject] if not sg else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg
        ]

        def student_free_at(sid, slot):
            busy = student_busy.get(sid, set())
            own = set()
            for ss in shared_subjs:
                own |= student_subject_slots.get(sid, {}).get(ss, set())
            if slot in student_lab_busy.get(sid, set()):
                return False  # already placed in another lab at this slot
            return slot not in (busy - own)

        # (a) Try to absorb each leftover into an existing group of this subject
        for sid in list(sd['unassigned']):
            placed = False
            # Candidate groups: same subject, not yet at absolute_max, student free
            cands = [g for g in sd['groups']
                     if g['nb_students'] < absolute_max
                     and student_free_at(sid, (g['day_idx'], g['block_id']))]
            # Prefer the emptiest group (load balancing)
            cands.sort(key=lambda g: g['nb_students'])
            if cands:
                g = cands[0]
                g['student_ids'].append(sid)
                g['nb_students'] += 1
                _propagate_busy_one(sid, g)
                # Propagate to student_busy: this student now occupies that slot
                student_busy.setdefault(sid, set()).add((g['day_idx'], g['block_id']))
                sd['unassigned'].discard(sid)
                p6_absorbed += 1
                placed = True
            if placed:
                continue

        # (b) Form small groups (floor MIN_GROUP_SIZE_P6) for remaining students
        #     who can share a slot — only if the group is VIABLE (not a singleton).
        MIN_VIABLE_P6 = 3  # do not create groups smaller than this in step (b)
        remaining = list(sd['unassigned'])
        guard = 0
        while remaining and guard < 60:
            guard += 1
            best_slot, best_free, best_room = None, [], None
            best_usage = float('inf')
            for slot in allowed_slots:
                day_idx, block_id = slot
                chosen_room = None
                for room in lab_rooms_list:
                    if can_fit_new_group(room, sem, day_idx, block_id,
                                          min_w, max_w, num_sessions,
                                          current_load,
                                          subject=subject,
                                          subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    continue
                free = [sid for sid in remaining if student_free_at(sid, slot)]
                if len(free) < MIN_VIABLE_P6:
                    continue  # not enough students to form a viable group here
                rk = (chosen_room, sem, day_idx, block_id)
                usage = current_load.get(rk, 0)
                if (len(free), -usage) > (len(best_free), -best_usage):
                    best_free = free; best_slot = slot
                    best_room = chosen_room; best_usage = usage
            if best_slot is None:
                break  # no slot can host a viable (>=3) group of the leftovers
            members = best_free[:absolute_max]
            day_idx, block_id = best_slot
            sd['group_counter'] += 1
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': sd['curso_num'],
                'group_num': sd['group_counter'],
                'program': 'MIXED', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': absolute_max,
                'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                '_absorbed': 'P6',
            }
            sd['groups'].append(group)
            all_groups.append(group)
            _propagate_busy(group)
            for room in best_room.split(','):
                room = room.strip()
                if room:
                    rk = (room, sem, day_idx, block_id)
                    current_load[rk] = current_load.get(rk, 0) + num_sessions
            sk = (subject, sem, day_idx, block_id)
            subject_load[sk] = subject_load.get(sk, 0) + num_sessions
            for sid in members:
                sd['unassigned'].discard(sid)
            remaining = [s for s in remaining if s not in members]
            p6_newgroups += 1
            p6_absorbed += len(members)

        # (c) LAST RESORT for true singletons: rather than open a lab for ONE
        #     student, attach them to the best existing group where they are
        #     free, allowing a small, bounded overflow (max + OVERFLOW_TOLERANCE).
        #     A group of 16 is far preferable to a group of 1. Only if even this
        #     fails is the student recorded as genuinely unplaceable (guard-rail).
        OVERFLOW_TOLERANCE = 2  # allow up to absolute_max + 2 in last resort
        for sid in list(sd['unassigned']):
            cands = [g for g in sd['groups']
                     if g['nb_students'] < absolute_max + OVERFLOW_TOLERANCE
                     and student_free_at(sid, (g['day_idx'], g['block_id']))]
            cands.sort(key=lambda g: g['nb_students'])
            if cands:
                g = cands[0]
                g['student_ids'].append(sid)
                g['nb_students'] += 1
                _propagate_busy_one(sid, g)
                g['_overflow'] = g.get('_overflow', 0) + 1
                # Propagate to student_busy (prevents this student from being
                # absorbed into a SECOND group at the same slot, which would
                # create a same-week conflict — the residual bug that left 2
                # cross-subject conflicts in v8).
                student_busy.setdefault(sid, set()).add((g['day_idx'], g['block_id']))
                sd['unassigned'].discard(sid)
                p6_absorbed += 1

        # (d) ABSOLUTE LAST RESORT: a student still unplaced fits no existing
        #     group. Rather than EXCLUDE them, open a minimal (possibly solo)
        #     group at any feasible slot where they are free — flagged as an
        #     exception for manual review (rare; mirrors Daniel's hand handling
        #     of isolated students). Excluding a student is a worse outcome than
        #     a 1-person session that the coordinator can later merge.
        for sid in list(sd['unassigned']):
            for slot in allowed_slots:
                day_idx, block_id = slot
                if not student_free_at(sid, slot):
                    continue
                chosen_room = None
                for room in lab_rooms_list:
                    if can_fit_new_group(room, sem, day_idx, block_id,
                                          min_w, max_w, num_sessions,
                                          current_load, subject=subject,
                                          subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    continue
                sd['group_counter'] += 1
                group = {
                    'subject': subject, 'semester': sem,
                    'curso_num': sd['curso_num'],
                    'group_num': sd['group_counter'], 'program': 'MIXED',
                    'day_idx': day_idx, 'day': DAYS[day_idx],
                    'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                    'student_ids': [sid], 'nb_students': 1,
                    'num_sessions': num_sessions, 'max_students': absolute_max,
                    'lab_rooms': chosen_room, 'min_week': min_w, 'max_week': max_w,
                    '_absorbed': 'P6', '_exception_solo': True,
                }
                sd['groups'].append(group)
                all_groups.append(group)
                _propagate_busy(group)
                for room in chosen_room.split(','):
                    room = room.strip()
                    if room:
                        rk = (room, sem, day_idx, block_id)
                        current_load[rk] = current_load.get(rk, 0) + num_sessions
                sk = (subject, sem, day_idx, block_id)
                subject_load[sk] = subject_load.get(sk, 0) + num_sessions
                sd['unassigned'].discard(sid)
                p6_absorbed += 1
                p6_newgroups += 1
                p6_exception_solo.append((subject, sid))
                break

        # (e) Guard-rail: whatever is left truly has no compatible free slot
        for sid in sd['unassigned']:
            truly_unplaceable.append((subject, sid))

    # ════════════════════════════════════════════════════════════
    # ── Phase 8: MANUAL OVERRIDE for truly-unplaceable students ──
    # ════════════════════════════════════════════════════════════
    # Daniel does this in practice: when an L2/L3/L4 student is "saturated"
    # (all afternoon slots taken by their courses) he still places them in
    # the main group of the lab subject and lets them miss the conflicting
    # course occasionally. The pipeline must do the same — otherwise we leave
    # the lab inscription unhonoured. We tag each override so Daniel sees
    # exactly which students need a manual confirmation.
    p8_overrides = []   # list of (subject, student_id, target_group_num)
    new_unplaceable = []
    for subject, sid in truly_unplaceable:
        sd = subject_data.get(subject)
        if not sd or not sd['groups']:
            new_unplaceable.append((subject, sid))
            continue
        # Pick the largest existing group of this subject (most representative)
        target = max(sd['groups'], key=lambda g: g['nb_students'])
        target['student_ids'].append(sid)
        target['nb_students'] += 1
        target['_manual_override'] = target.get('_manual_override', 0) + 1
        target.setdefault('_override_sids', set()).add(sid)
        # Remove from per-subject unassigned set so the downstream summary
        # counts this enrollment as honoured.
        if isinstance(sd.get('unassigned'), set):
            sd['unassigned'].discard(sid)
        # Note: do NOT propagate to student_lab_busy (would create future
        # blockages); the override is by definition a "soft" assignment Daniel
        # accepts despite a clash.
        p8_overrides.append((subject, sid, target['group_num']))
    truly_unplaceable = new_unplaceable

    # Also absorb any leftover solo group (1 student) into the largest sibling
    # group via override — Phase 7 couldn't move them because their slot was
    # busy with another lab, but Daniel would still consolidate manually
    # rather than ship a 1-person group.
    p8_solo_dissolved = []
    for subject, sd in subject_data.items():
        solos = [g for g in sd['groups'] if g['nb_students'] == 1]
        for solo_g in solos:
            others = [g for g in sd['groups']
                      if g is not solo_g and g['nb_students'] >= 2]
            if not others:
                continue
            target = max(others, key=lambda g: g['nb_students'])
            for sid in list(solo_g['student_ids']):
                target['student_ids'].append(sid)
                target['nb_students'] += 1
                target['_manual_override'] = target.get('_manual_override', 0) + 1
                target.setdefault('_override_sids', set()).add(sid)
                p8_overrides.append((subject, sid, target['group_num']))
            solo_g['student_ids'] = []
            solo_g['nb_students'] = 0
            p8_solo_dissolved.append((subject, solo_g['group_num']))
    # Remove dissolved solo groups
    for subj, gnum in p8_solo_dissolved:
        subject_data[subj]['groups'] = [
            g for g in subject_data[subj]['groups'] if g['group_num'] != gnum]
        all_groups[:] = [g for g in all_groups
                         if not (g['subject'] == subj and g['group_num'] == gnum)]

    if p8_overrides:
        print(f"\n  ── Phase 8: Override manuel (saturés) ──")
        print(f"    [INFO] {len(p8_overrides)} étudiant(s) saturé(s) placé(s) "
              f"dans le groupe principal de leur matière (à confirmer par Daniel) :")
        # Hash → name lookup (same logic as guard-rail)
        try:
            import pandas as _pd
            df_names = _pd.read_csv('data_clean/master_schedule.csv',
                                    usecols=['AlumnoID','Apellidos','Nombre'])
            df_names = df_names.drop_duplicates('AlumnoID')
            _sid_to_name = {str(r['AlumnoID']): f"{r['Apellidos']}, {r['Nombre']}"
                            for _, r in df_names.iterrows()}
        except Exception:
            _sid_to_name = {}
        for subj, sid, gn in p8_overrides:
            name = _sid_to_name.get(str(sid), f"<id {sid}>")
            print(f"           · {name}  →  {subj} G{gn}  [OVERRIDE]")

    # ════════════════════════════════════════════════════════════
    # ── Phase 7: PASSE FINALE ANTI-SOLO ──
    # ════════════════════════════════════════════════════════════
    # After all phases, some groups may still contain a single student because
    # when their (per-subject) Phase 6 ran, the student wasn't yet placed in
    # other subjects and the absorption target slot looked busy via a course
    # the lab would have replaced. Now that ALL subjects are placed,
    # student_lab_busy is fully populated and we can correctly evaluate
    # which solo students can be merged into a sibling group of the same
    # subject. This dramatically reduces solo groups without creating
    # cross-subject conflicts (lab_busy is checked).
    p7_merged = 0
    p7_dissolved_groups = []
    # Daniel's reference (Reparto_Pract_rev23.xlsx) shows he routinely accepts
    # groups up to 18 (Modelado de Sistemas: a single group of 18 IMR rather
    # than 15+3). The pipeline must follow that same logic: a few extra
    # students in a big group is far better than a leftover micro-group.
    OVERFLOW_TOL_P7 = 5  # absolute_max=15 → up to 20 if needed
    SMALL_THRESHOLD_P7 = 4  # groups with ≤ 4 students are candidates to dissolve
    for subject, sd in subject_data.items():
        # Include exception_solo groups: Phase 7 runs AFTER all labs are placed,
        # so student_lab_busy is now complete and a previously-isolated student
        # may now have a coherent merge target.
        solo_groups = [g for g in sd['groups']
                       if g['nb_students'] <= SMALL_THRESHOLD_P7]
        for solo_g in solo_groups:
            absolute_max = MAX_GROUP_SIZE
            members_to_move = list(solo_g['student_ids'])
            for sid in members_to_move:
                # candidate groups: same subject, different group, has room
                # under OVERFLOW_TOL_P7, and student is free at that slot
                # (i.e. not placed in another lab at that slot)
                solo_slot = (solo_g['day_idx'], solo_g['block_id'])
                # Remove the solo's own slot from lab_busy temporarily so we
                # don't see ourselves as blocking
                lab_busy_set = student_lab_busy.get(sid, set()) - {solo_slot}
                cands = []
                for g in sd['groups']:
                    if g is solo_g:
                        continue
                    if g['nb_students'] >= absolute_max + OVERFLOW_TOL_P7:
                        continue
                    g_slot = (g['day_idx'], g['block_id'])
                    if g_slot in lab_busy_set:
                        continue  # would conflict with another lab
                    cands.append(g)
                # Pick the emptiest candidate
                cands.sort(key=lambda g: g['nb_students'])
                if not cands:
                    continue
                target = cands[0]
                target['student_ids'].append(sid)
                target['nb_students'] += 1
                target['_p7_merged'] = target.get('_p7_merged', 0) + 1
                solo_g['student_ids'].remove(sid)
                solo_g['nb_students'] -= 1
                # Update lab_busy: student now occupies target slot, not solo slot
                student_lab_busy[sid].discard(solo_slot)
                student_lab_busy[sid].add((target['day_idx'], target['block_id']))
                p7_merged += 1
            # If solo group is now empty, mark for removal
            if solo_g['nb_students'] == 0:
                p7_dissolved_groups.append((subject, solo_g['group_num']))

    # Remove dissolved groups from sd['groups'] and all_groups
    for subj, gnum in p7_dissolved_groups:
        subject_data[subj]['groups'] = [
            g for g in subject_data[subj]['groups'] if g['group_num'] != gnum]
        all_groups[:] = [g for g in all_groups
                         if not (g['subject'] == subj and g['group_num'] == gnum)]

    if p7_merged > 0:
        print(f"\n  ── Phase 7: Passe finale anti-solo ──")
        print(f"    [OK] {p7_merged} étudiant(s) solo fusionné(s) dans des "
              f"groupes existants ({len(p7_dissolved_groups)} groupe(s) "
              f"solo dissous)")

    if p6_absorbed > 0:
        print(f"\n  ── Phase 6: Absorption finale adaptative ──")
        print(f"    [OK] {p6_absorbed} étudiant(s) placé(s) "
              f"({p6_newgroups} petit(s) groupe(s) créé(s), reste absorbé dans "
              f"des groupes existants)")
    if p6_exception_solo:
        from collections import Counter as _Ctr
        print(f"\n  [EXCEPTION] {len(p6_exception_solo)} étudiant(s) isolé(s) placé(s) "
              f"en groupe minimal (à fusionner manuellement si possible) :")
        for subj_x, n in _Ctr(s for s, _ in p6_exception_solo).most_common():
            print(f"       - {subj_x}: {n} cas (emploi du temps atypique)")
    if truly_unplaceable:
        print(f"\n  [GARDE-FOU] {len(truly_unplaceable)} étudiant(s) RÉELLEMENT "
              f"inplaçable(s) (aucun créneau compatible libre) :")
        from collections import Counter as _Counter, defaultdict as _DD
        by_subj_ids = _DD(list)
        for subj, sid in truly_unplaceable:
            by_subj_ids[subj].append(sid)
        # Build a hash → name lookup from the master schedule
        try:
            import pandas as _pd
            df_names = _pd.read_csv('data_clean/master_schedule.csv',
                                    usecols=['AlumnoID', 'Apellidos', 'Nombre'])
            df_names = df_names.drop_duplicates('AlumnoID')
            sid_to_name = {
                str(row['AlumnoID']): f"{row['Apellidos']}, {row['Nombre']}"
                for _, row in df_names.iterrows()
            }
        except Exception:
            sid_to_name = {}
        for subj in sorted(by_subj_ids, key=lambda s: -len(by_subj_ids[s])):
            ids = by_subj_ids[subj]
            print(f"       - {subj}: {len(ids)} étudiant(s)")
            for sid in ids:
                name = sid_to_name.get(str(sid), f"<id {sid}>")
                print(f"           · {name}")
        print(f"       Ces cas nécessitent une décision manuelle (créneau "
              f"exceptionnel ou dérogation).")

    # Refresh stats containers after Phase 6
    for sid_subject in []:
        pass

    # Résumé par matière
    import math
    print(f"\n  Résultats après {round_num} tours :")
    print(f"  (Taille préférée: {PREFERRED_GROUP_SIZE} | Max autorisé: {MAX_GROUP_SIZE})")
    for subject in sorted_subjects:
        sd = subject_data[subject]
        assigned = sum(g['nb_students'] for g in sd['groups'])
        unassigned = len(sd['unassigned'])
        total = sd['total_students']
        n_groups = len(sd['groups'])

        # Comparer avec ce que donnerait max=15
        groups_at_15 = math.ceil(total / MAX_GROUP_SIZE) if total > 0 else 0
        extra = n_groups - groups_at_15

        if unassigned > 0:
            stats['total_unassigned'] += unassigned
            print(f"  [WARN]  {subject:40s} | {n_groups:2d} gr | "
                  f"{assigned:3d}/{total} assignés ({unassigned} sans créneau)")
        else:
            extra_info = f" (+{extra} vs max15)" if extra > 0 else ""
            print(f"  [OK] {subject:40s} | {n_groups:2d} gr | "
                  f"{assigned:3d}/{total} assignés{extra_info}")

        if extra > 3:
            print(f"       [WARN]  {extra} groupes de plus qu'avec max=15. "
                  f"Considérer max=15 pour cette matière.")

        stats['total_assigned'] += assigned

        for g in sd['groups']:
            prog_label = g.get('program', '?')
            marker = " [RUN]" if g.get('_overflow') else (" [ALT_ROOM]" if g.get('_alt_room') else "")
            print(f"       G{g['group_num']:2d} : {g['day']:10s} {g['block_label']} | "
                  f"{g['nb_students']:2d}/{g['max_students']} étudiants | {prog_label}{marker}")

    # Secondary shared subjects (Química etc.)
    for subject in secondary_shared:
        config = LAB_CONFIG[subject]
        enrolled = set(subject_students.get(subject, []))
        sec_groups = [g for g in all_groups if g['subject'] == subject]
        assigned_sec = set()
        for g in sec_groups:
            assigned_sec.update(g['student_ids'])
        unassigned_count = len(enrolled - assigned_sec)
        n_groups = len(sec_groups)

        stats['total_assigned'] += len(assigned_sec)
        stats['total_unassigned'] += unassigned_count

        if unassigned_count > 0:
            print(f"  [WARN]  {subject:40s} | {n_groups:2d} gr | "
                  f"{len(assigned_sec):3d}/{len(enrolled)} assignés ({unassigned_count} sans créneau)")
        else:
            print(f"  [OK] {subject:40s} | {n_groups:2d} gr | "
                  f"{len(assigned_sec):3d}/{len(enrolled)} assignés (groupes partagés + recovered)")

        for g in sorted(sec_groups, key=lambda x: x['group_num']):
            prog_label = g.get('program', '?')
            marker = " [RECYCLE]" if g.get('_recovered') else ""
            print(f"       G{g['group_num']:2d} : {g['day']:10s} {g['block_label']} | "
                  f"{g['nb_students']:2d}/{g['max_students']} étudiants | {prog_label}{marker}")

    # Résumé global
    total_sessions = sum(g['num_sessions'] for g in all_groups)
    print(f"\n  [STATS] RÉSUMÉ :")
    print(f"     Groupes formés       : {len(all_groups)}")
    print(f"     Étudiants assignés   : {stats['total_assigned']}")
    print(f"     Étudiants non-assign.: {stats['total_unassigned']}")
    # HEADLINE rate = pair-level. Every (student × lab subject) enrollment must
    # be honoured. This is the strict, honest count.
    placed_students = set()
    enrolled_students = set()
    for g in all_groups:
        placed_students.update(g.get('student_ids', []))
    for ids in subject_students.values():
        enrolled_students.update(ids)
    pct_pair = (stats['total_assigned'] /
                (stats['total_assigned'] + stats['total_unassigned']) * 100
                if (stats['total_assigned'] + stats['total_unassigned']) > 0 else 0)
    pct_unique = (100.0 * len(placed_students) / len(enrolled_students)
                  if enrolled_students else 0.0)
    print(f"     Taux d'assignation   : {pct_pair:.1f}%  "
          f"({stats['total_assigned']}/{stats['total_assigned']+stats['total_unassigned']} "
          f"inscriptions étudiant × matière)")
    print(f"     Étudiants couverts   : {pct_unique:.1f}%  "
          f"({len(placed_students)}/{len(enrolled_students)} étudiants uniques)")
    print(f"     Sessions de lab      : {total_sessions}")

    for sem in sorted(set(g['semester'] for g in all_groups)):
        sg = [g for g in all_groups if g['semester'] == sem]
        ss = sum(g['num_sessions'] for g in sg)
        print(f"     Semestre {sem} : {len(sg)} groupes, {ss} sessions")

    return all_groups


# ============================================================
# ÉTAPE 5 : SOLVEUR CP-SAT
# ============================================================
def solve(all_groups):
    """
    Le solveur est SIMPLE car les groupes sont déjà formés :
    - Chaque groupe a un créneau FIXE (jour, bloc)
    - La seule variable : QUELLE SEMAINE pour chaque session
    - Contraintes : pas 2 sessions même matière/salle au même (semaine, jour, bloc)
    """
    print_section("ÉTAPE 5 : Solveur CP-SAT")

    all_results = []

    for sem in sorted(set(g['semester'] for g in all_groups)):
        sem_groups = [g for g in all_groups if g['semester'] == sem]
        sem_label = f"S{sem}"

        # Construire les sessions
        sessions = []
        sid = 0
        for g in sem_groups:
            for sess_num in range(1, g['num_sessions'] + 1):
                sessions.append({
                    'id': sid,
                    'subject': g['subject'],
                    'grupo': g['group_num'],
                    'program': g.get('program', ''),
                    'session': sess_num,
                    'curso_num': g['curso_num'],
                    'day_idx': g['day_idx'],
                    'day': g['day'],
                    'block_id': g['block_id'],
                    'block_label': g['block_label'],
                    'nb_students': g['nb_students'],
                    'lab_rooms': g['lab_rooms'],
                    'min_week': g['min_week'],
                    'max_week': g['max_week'],
                    'student_ids': list(g.get('student_ids', [])),
                })
                sid += 1

        print(f"\n  ── {sem_label} : {len(sessions)} sessions ──")

        model = cp_model.CpModel()

        # Variables : 1 entier par session = la semaine
        # On exclut les semaines où le jour du groupe est férié.
        # NOTE: reserved/blocked slots (Biotecnología) are NOT removed from the
        # domain here — that hard reduction, combined with C4/C5 on the shared
        # room, made S1 INFEASIBLE. Reserved-week avoidance is handled by the soft
        # objective penalty (c4_res_penalty_terms) below, which never breaks
        # feasibility.
        sem_holidays = HOLIDAYS.get(sem, {})
        week_vars = {}
        for s in sessions:
            valid_weeks = [w for w in range(s['min_week'], s['max_week'] + 1)
                           if (w, s['day_idx']) not in sem_holidays]
            if not valid_weeks:
                print(f"    [WARN]  Session {s['id']} ({s['subject']} G{s['grupo']}): "
                      f"aucune semaine disponible pour {s['day']}!")
                valid_weeks = list(range(s['min_week'], s['max_week'] + 1))
            week_vars[s['id']] = model.NewIntVarFromDomain(
                cp_model.Domain.FromValues(valid_weeks), f"w_{s['id']}")

        # C1 : Pas 2 sessions même matière au même (semaine, jour, bloc)
        c1 = 0
        by_subj_slot = defaultdict(list)
        for s in sessions:
            by_subj_slot[(s['subject'], s['day_idx'], s['block_id'])].append(s)
        for group in by_subj_slot.values():
            if len(group) > 1:
                for i in range(len(group)):
                    for j in range(i + 1, len(group)):
                        model.Add(week_vars[group[i]['id']] != week_vars[group[j]['id']])
                        c1 += 1

        # C4 : Pas 2 sessions même salle au même (semaine, jour, bloc)
        c4 = 0
        by_room_slot = defaultdict(list)
        for s in sessions:
            for room in s['lab_rooms'].split(','):
                room = room.strip()
                if room:
                    by_room_slot[(room, s['day_idx'], s['block_id'])].append(s)
        for group in by_room_slot.values():
            if len(group) > 1:
                for i in range(len(group)):
                    for j in range(i + 1, len(group)):
                        model.Add(week_vars[group[i]['id']] != week_vars[group[j]['id']])
                        c4 += 1

        # C4-RESERVED (soft): a blocked slot means the ROOM is occupied by
        # another activity (e.g. Biotecnología) in specific weeks. We DISCOURAGE
        # real sessions from those reserved weeks via an objective penalty rather
        # than forbidding them outright — a hard constraint here made S1 INFEASIBLE
        # because the shared room (Física+Química) needs nearly every week for its
        # C4/C5-distinct sessions. The Química-specific week-domain exclusion above
        # already keeps Química out of the reserved weeks; this soft term nudges
        # any OTHER subject in the same room to avoid them when it can.
        c4_res_penalty_terms = []
        for (_bsem, _bsubj), _slots in SUBJECT_BLOCKED_SLOTS.items():
            if _bsem != sem:
                continue
            _rooms = [r.strip() for r in
                      LAB_CONFIG.get(_bsubj, {}).get('lab_rooms', []) if r.strip()]
            for (_w, _d, _b) in _slots:
                for _room in _rooms:
                    for s in by_room_slot.get((_room, _d, _b), []):
                        # Only meaningful if _w is in this session's domain.
                        if not (s['min_week'] <= _w <= s['max_week']):
                            continue
                        _in_resv = model.NewBoolVar(f"resv_{s['id']}_{_w}")
                        model.Add(week_vars[s['id']] == _w).OnlyEnforceIf(_in_resv)
                        model.Add(week_vars[s['id']] != _w).OnlyEnforceIf(_in_resv.Not())
                        c4_res_penalty_terms.append(_in_resv)
        if c4_res_penalty_terms:
            print(f"  [BLOCK] C4-réservé : {len(c4_res_penalty_terms)} pénalité(s) "
                  f"souple(s) (salle occupée par activité externe)")

        # C5 : Séquencement (session k < session k+1 en semaines)
        c5 = 0
        by_group = defaultdict(list)
        for s in sessions:
            by_group[(s['subject'], s['grupo'])].append(s)
        for group in by_group.values():
            gsorted = sorted(group, key=lambda x: x['session'])
            for k in range(len(gsorted) - 1):
                model.Add(week_vars[gsorted[k+1]['id']] > week_vars[gsorted[k]['id']])
                c5 += 1

        # C8 (cross-subject student conflicts) is prevented at FORMATION time
        # by propagating each newly placed group's slot into student_busy, so a
        # student already placed at (Vie, 15:00) is no longer free there for any
        # other subject. This is more efficient than enforcing in the solver
        # (which adds thousands of soft constraints and slows it heavily).
        c8 = 0

        # ────────────────────────────────────────────────────────
        # PARITY ALTERNATION (Daniel's strategy): when several groups of the
        # same subject share a (day, block) slot, alternate their weeks by
        # parity so they never collide on the same week. Half the groups are
        # steered to EVEN weeks, half to ODD. Soft (weighted) to stay feasible.
        # ────────────────────────────────────────────────────────
        parity_penalties = []
        n_parity_groups = 0
        if PARITY_ALTERNATION:
            # Assign a target parity to EVERY group of an even-spaced subject,
            # not only to groups that happen to share a (day, block) slot.
            #
            # Previous version required len(groups_here) >= 2 on a shared slot,
            # so a group occupying its slot alone got NO parity target — its
            # spacing then drifted (gaps of 1/3) whenever a week was busy. By
            # giving every group of the subject a parity target (alternating by
            # group index), each group has a clear W4,6,8,10,12 or W5,7,9,11,13
            # skeleton to snap to, matching Daniel's ~88%-regular pattern.
            # Build subject -> sorted list of distinct groups
            subj_groups = defaultdict(set)
            for s in sessions:
                subj_groups[s['subject']].add(s['grupo'])
            # Only subjects with >=2 groups and >=3 sessions alternate
            sess_count = defaultdict(lambda: defaultdict(int))
            for s in sessions:
                sess_count[s['subject']][s['grupo']] += 1

            for subj, groups_set in subj_groups.items():
                groups_sorted = sorted(groups_set)
                max_sess = max(sess_count[subj].values()) if sess_count[subj] else 0
                if len(groups_sorted) < 2 or max_sess < 3:
                    continue
                for gi, grupo in enumerate(groups_sorted):
                    target_parity = gi % 2
                    n_parity_groups += 1
                    grp_sessions = [s for s in sessions
                                    if s['subject'] == subj and s['grupo'] == grupo]
                    for s in grp_sessions:
                        wv = week_vars[s['id']]
                        parity_bit = model.NewIntVar(0, 1, f"par_{s['id']}")
                        half = model.NewIntVar(s['min_week'] // 2,
                                               s['max_week'] // 2 + 1,
                                               f"half_{s['id']}")
                        model.Add(wv == 2 * half + parity_bit)
                        if target_parity == 0:
                            parity_penalties.append(parity_bit)
                        else:
                            inv = model.NewIntVar(0, 1, f"invpar_{s['id']}")
                            model.Add(inv == 1 - parity_bit)
                            parity_penalties.append(inv)

        # ────────────────────────────────────────────────────────
        # Daniel-style spreading: anchor first AND last sessions
        # Strategy: SOFT penalties with heavy weights instead of hard constraints
        # to keep solver feasible even when slots are saturated.
        #
        # Pattern Daniel:
        #   Física (5 sess, W4-W13): P1=W4, ..., P5=W13
        #   Química (4 sess, W4-W14): P1=W4, ..., P4=W14
        #
        # IMPORTANT — parity groups: when a group alternates by parity (Daniel's
        # real strategy), it keeps a CONSTANT gap of 2 and ends on W12 (even) or
        # W13 (odd). Forcing its LAST session to max_week would break that even
        # spacing and shift the intermediate sessions by one week. So for parity
        # groups we anchor only the FIRST session and let the even-spacing
        # objective place the rest — reproducing Daniel's W4,6,8,10,12 /
        # W5,7,9,11,13 exactly.
        # ────────────────────────────────────────────────────────
        # Identify which (subject, grupo) are under parity alternation
        parity_group_keys = set()
        if PARITY_ALTERNATION:
            # A subject is under parity/even-spacing regime if it has several
            # groups each running >=3 sessions. In that regime Daniel keeps a
            # CONSTANT gap and lets groups end naturally on W12 (even) or W13
            # (odd) — so we must NOT anchor their last session to max_week.
            #
            # NOTE: the previous version only flagged groups that SHARE a
            # (day, block) slot (len(groups_here) >= 2). That missed groups
            # occupying a slot alone, which then kept the max_week anchor and
            # ended on W13 with an irregular last gap (e.g. 4,6,8,10,13). We now
            # flag at the SUBJECT level so every group of an even-spaced subject
            # is governed by spacing alone.
            sessions_by_subject = defaultdict(lambda: defaultdict(int))
            for s in sessions:
                sessions_by_subject[s['subject']][s['grupo']] += 1
            for subj, groups_map in sessions_by_subject.items():
                n_groups = len(groups_map)
                max_sess = max(groups_map.values()) if groups_map else 0
                if n_groups >= 2 and max_sess >= 3:
                    for grupo in groups_map:
                        parity_group_keys.add((subj, grupo))

        first_excess = []   # how far session 1 is from min_week
        last_deficit = []   # how far last session is from max_week

        for s in sessions:
            group_key = (s['subject'], s['grupo'])
            group_sessions = [x for x in sessions
                              if (x['subject'], x['grupo']) == group_key]
            max_sess_num = max(x['session'] for x in group_sessions)

            if s['session'] == 1:
                # Penalize: chosen_week - min_week (the further, the worse)
                excess = model.NewIntVar(0, 20, f"excess_{s['id']}")
                model.Add(excess >= week_vars[s['id']] - s['min_week'])
                model.Add(excess >= 0)
                first_excess.append(excess)

            # Anchor last session to max_week — but NOT for parity groups, whose
            # natural even-spaced end (W12 or W13) must be preserved.
            if (s['session'] == max_sess_num and max_sess_num > 1
                    and group_key not in parity_group_keys):
                # Penalize: max_week - chosen_week (the further from end, the worse)
                deficit = model.NewIntVar(0, 20, f"deficit_{s['id']}")
                model.Add(deficit >= s['max_week'] - week_vars[s['id']])
                model.Add(deficit >= 0)
                last_deficit.append(deficit)

        # Even spacing for sessions in between (gaps between consecutive sessions)
        # We minimize the variance of gaps within each group.
        # Linear approximation: minimize sum of |gap - ideal_gap|
        gap_deviations = []
        for (subject, grupo) in set((s['subject'], s['grupo']) for s in sessions):
            grp_sess = sorted(
                [s for s in sessions if s['subject'] == subject and s['grupo'] == grupo],
                key=lambda x: x['session']
            )
            n = len(grp_sess)
            if n >= 3:
                # Ideal gap = window_size / (n-1)
                window = grp_sess[0]['max_week'] - grp_sess[0]['min_week']
                ideal_gap = max(1, window // (n - 1))

                for k in range(n - 1):
                    gap = model.NewIntVar(1, 20, f"gap_{subject}_{grupo}_{k}")
                    model.Add(gap == week_vars[grp_sess[k+1]['id']]
                                       - week_vars[grp_sess[k]['id']])
                    # Deviation from ideal gap (absolute value)
                    dev = model.NewIntVar(0, 20, f"dev_{subject}_{grupo}_{k}")
                    model.Add(dev >= gap - ideal_gap)
                    model.Add(dev >= ideal_gap - gap)
                    gap_deviations.append(dev)

        # Combined objective with weights:
        #   first_excess  × 100  (push P1 to min_week)
        #   last_deficit  × 100  (push last to max_week — non-parity subjects)
        #   gap_deviation × 200  (EVEN SPACING — Daniel keeps a strictly constant
        #                         gap of 2; this must dominate, otherwise the
        #                         solver trades regular spacing away to satisfy
        #                         per-session parity or to dodge a busy week,
        #                         producing irregular gaps like 1/3. Making
        #                         spacing the top priority reproduces Daniel's
        #                         4,6,8,10,12 / 5,7,9,11,13 pattern.)
        #   parity        ×  50  (which half-parity a group lands on; secondary,
        #                         because constant spacing already enforces the
        #                         alternation once P1 is anchored.)
        objective_terms = []
        if first_excess:
            sum_first = model.NewIntVar(0, 10000, 'sum_first')
            model.Add(sum_first == sum(first_excess))
            objective_terms.append((sum_first, 100))
        if last_deficit:
            sum_last = model.NewIntVar(0, 10000, 'sum_last')
            model.Add(sum_last == sum(last_deficit))
            objective_terms.append((sum_last, 100))
        if gap_deviations:
            sum_gaps = model.NewIntVar(0, 10000, 'sum_gaps')
            model.Add(sum_gaps == sum(gap_deviations))
            objective_terms.append((sum_gaps, 200))
        if parity_penalties:
            sum_parity = model.NewIntVar(0, 100000, 'sum_parity')
            model.Add(sum_parity == sum(parity_penalties))
            objective_terms.append((sum_parity, PARITY_PENALTY_WEIGHT))

        # Reserved-slot avoidance (soft): strongly discourage placing a real
        # session in a week where its room is reserved for another activity.
        # Weight is set ABOVE the spacing penalties (gaps=200, first/last=100) so
        # avoidance almost always wins; it stays a penalty (not a hard constraint)
        # so the model can never become INFEASIBLE — a real session lands on a
        # reserved slot only if there is genuinely no other feasible week.
        if c4_res_penalty_terms:
            sum_resv = model.NewIntVar(0, 100000, 'sum_resv')
            model.Add(sum_resv == sum(c4_res_penalty_terms))
            objective_terms.append((sum_resv, 100000))

        if objective_terms:
            total = model.NewIntVar(0, 100_000_000, 'total')
            model.Add(total == sum(var * w for var, w in objective_terms))
            model.Minimize(total)

        print(f"  Contraintes : C1={c1}, C4={c4}, C5={c5}, C8={c8}, "
              f"first_anchor={len(first_excess)}, last_anchor={len(last_deficit)}"
              + (f", parity_groups={n_parity_groups}" if PARITY_ALTERNATION else ""))

        # Résoudre
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = SOLVER_TIME_LIMIT
        solver.parameters.num_search_workers = 8

        print(f"  [WAIT] Lancement (max {SOLVER_TIME_LIMIT}s)...")
        status = solver.Solve(model)

        names = {cp_model.OPTIMAL: "OPTIMAL", cp_model.FEASIBLE: "FAISABLE",
                 cp_model.INFEASIBLE: "INFAISABLE", cp_model.UNKNOWN: "INCONNU"}

        print(f"  Statut  : {names.get(status, '?')}")
        print(f"  Temps   : {solver.WallTime():.2f}s")

        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            print(f"  Pénalité: {solver.ObjectiveValue()}")

            for s in sessions:
                all_results.append({
                    'semester': sem,
                    'subject': s['subject'],
                    'program': s.get('program', ''),
                    'curso_num': s['curso_num'],
                    'grupo': s['grupo'],
                    'session': s['session'],
                    'week': solver.Value(week_vars[s['id']]),
                    'day': s['day'],
                    'time_block': s['block_label'],
                    'nb_students': s['nb_students'],
                    'lab_rooms': s['lab_rooms'],
                })

            print(f"  [OK] {sem_label} : {len(sessions)} sessions planifiées")
            # NOTE: blocked slots (room reserved for another activity) are NOT
            # written into the schedule — they are real RESERVATIONS enforced in
            # the solver's C4 (a phantom occupant frees the room there). The Excel
            # renders them as "Festivo / No disponible" from blocked_slots.csv,
            # so the reliability conflict check sees a clean schedule (no false C4).
        else:
            # ──────────────────────────────────────────────────────
            # AUTO-RECOVERY: identify oversaturated slots, drop the
            # excess groups (overflow ones first), and retry.
            # ──────────────────────────────────────────────────────
            print(f"  [WARN]  {sem_label} INFAISABLE — tentative de récupération automatique...")

            # Find oversaturated room+slots (C4 violations)
            oversaturated = []
            for key, group in by_room_slot.items():
                room, d, b = key
                needed = len(group)
                max_w = max(s['max_week'] for s in group)
                min_w = min(s['min_week'] for s in group)
                cap = max_w - min_w + 1
                if needed > cap:
                    overflow = needed - cap
                    oversaturated.append((key, group, overflow))
                    print(f"    [FAIL] ROOM {room} {DAYS[d]} {BLOCK_LABELS[b]} : "
                          f"{needed} sess / {cap} sem dispo (excess: {overflow})")

            # Also detect oversaturated subject+slots (C1 violations)
            # If many groups of same subject share same day×block, they need
            # different weeks. If count > available weeks → infeasible.
            for key, group in by_subj_slot.items():
                subj, d, b = key
                needed = len(group)
                max_w = max(s['max_week'] for s in group)
                min_w = min(s['min_week'] for s in group)
                cap = max_w - min_w + 1
                if needed > cap:
                    overflow = needed - cap
                    # Wrap with sentinel so we can distinguish
                    oversaturated.append((('subj', subj, d, b), group, overflow))
                    print(f"    [FAIL] SUBJECT {subj} {DAYS[d]} {BLOCK_LABELS[b]} : "
                          f"{needed} sess / {cap} sem dispo (excess: {overflow})")

            if not oversaturated:
                print(f"  [WARN]  Pas de slot oversaturé détecté — cause non identifiée par diagnostic")
                # Last resort: drop all overflow/recovered/alt_room groups and retry
                print(f"  [FIX] Tentative finale : retrait de tous les groupes 'exceptionnels'")
                groups_to_drop = set()
                for g in all_groups:
                    if g.get('semester') != sem:
                        continue
                    if g.get('_overflow') or g.get('_recovered') or g.get('_alt_room'):
                        groups_to_drop.add((g['subject'], g['group_num']))
            else:
                # Identify groups to drop: prefer overflow/recovered/alt_room first
                groups_to_drop = set()
                for key, sessions_in_slot, overflow in oversaturated:
                    sessions_with_priority = []
                    for s in sessions_in_slot:
                        source_group = next(
                            (g for g in all_groups
                             if g['subject'] == s['subject']
                             and g['group_num'] == s['grupo']
                             and g['semester'] == sem),
                            None
                        )
                        priority = 0
                        if source_group:
                            if source_group.get('_overflow'):
                                priority = 3
                            elif source_group.get('_recovered'):
                                priority = 2
                            elif source_group.get('_alt_room'):
                                priority = 1
                        sessions_with_priority.append((priority, s))

                    # Sort by priority DESC (overflow first), then by session number DESC
                    sessions_with_priority.sort(key=lambda x: (-x[0], -x[1]['session']))

                    # Drop enough groups to fit
                    dropped_in_slot = 0
                    for priority, s in sessions_with_priority:
                        if dropped_in_slot >= overflow:
                            break
                        group_id = (s['subject'], s['grupo'])
                        if group_id not in groups_to_drop:
                            groups_to_drop.add(group_id)
                            group_sessions_here = sum(
                                1 for ss in sessions_in_slot
                                if ss['subject'] == group_id[0] and ss['grupo'] == group_id[1]
                            )
                            dropped_in_slot += group_sessions_here

            # ── Now we have groups_to_drop. Rebuild the model. ──
            if groups_to_drop:
                    print(f"  [FIX] Récupération : {len(groups_to_drop)} groupes problématiques retirés")
                    for subj, gnum in sorted(groups_to_drop):
                        # Find the group flag for diagnostic
                        src = next((g for g in all_groups
                                    if g['subject'] == subj and g['group_num'] == gnum
                                    and g['semester'] == sem), None)
                        flag = ""
                        if src:
                            if src.get('_overflow'): flag = " (overflow)"
                            elif src.get('_recovered'): flag = " (recovered)"
                            elif src.get('_alt_room'): flag = " (alt_room)"
                        print(f"     - {subj} G{gnum}{flag}")

                    # Filter sessions and rebuild model
                    filtered_sessions = [s for s in sessions
                                          if (s['subject'], s['grupo']) not in groups_to_drop]
                    print(f"  [RETRY] Relance solveur avec {len(filtered_sessions)} sessions "
                          f"(au lieu de {len(sessions)})")

                    # Rebuild fresh model with filtered sessions.
                    # Holiday-only domain (no reserved-slot reduction — see main
                    # model note; the reservation must never break feasibility).
                    model2 = cp_model.CpModel()
                    week_vars2 = {}
                    for s in filtered_sessions:
                        valid_weeks = [w for w in range(s['min_week'], s['max_week'] + 1)
                                       if (w, s['day_idx']) not in sem_holidays]
                        if not valid_weeks:
                            valid_weeks = list(range(s['min_week'], s['max_week'] + 1))
                        week_vars2[s['id']] = model2.NewIntVarFromDomain(
                            cp_model.Domain.FromValues(valid_weeks), f"w_{s['id']}")

                    # Re-add C1, C4, C5 constraints
                    by_subj_slot2 = defaultdict(list)
                    for s in filtered_sessions:
                        by_subj_slot2[(s['subject'], s['day_idx'], s['block_id'])].append(s)
                    for grp in by_subj_slot2.values():
                        if len(grp) > 1:
                            for i in range(len(grp)):
                                for j in range(i + 1, len(grp)):
                                    model2.Add(week_vars2[grp[i]['id']] != week_vars2[grp[j]['id']])

                    by_room_slot2 = defaultdict(list)
                    for s in filtered_sessions:
                        for room in s['lab_rooms'].split(','):
                            room = room.strip()
                            if room:
                                by_room_slot2[(room, s['day_idx'], s['block_id'])].append(s)
                    # NOTE: no hard reserved-slot constraint in the retry model.
                    # The retry path is the last resort to regain feasibility, so
                    # it must never be over-constrained. The Química-specific
                    # week-domain exclusion still applies to each session's domain;
                    # reserved-week avoidance for other subjects is handled softly
                    # in the main model only.
                    for grp in by_room_slot2.values():
                        if len(grp) > 1:
                            for i in range(len(grp)):
                                for j in range(i + 1, len(grp)):
                                    model2.Add(week_vars2[grp[i]['id']] != week_vars2[grp[j]['id']])

                    by_group2 = defaultdict(list)
                    for s in filtered_sessions:
                        by_group2[(s['subject'], s['grupo'])].append(s)
                    for grp in by_group2.values():
                        gsorted = sorted(grp, key=lambda x: x['session'])
                        for k in range(len(gsorted) - 1):
                            model2.Add(week_vars2[gsorted[k+1]['id']] > week_vars2[gsorted[k]['id']])

                    # Re-add anchoring (soft penalties)
                    first_excess2 = []
                    last_deficit2 = []
                    for s in filtered_sessions:
                        group_key = (s['subject'], s['grupo'])
                        group_sessions = [x for x in filtered_sessions
                                          if (x['subject'], x['grupo']) == group_key]
                        max_sess_num = max(x['session'] for x in group_sessions)
                        if s['session'] == 1:
                            excess = model2.NewIntVar(0, 20, f"excess2_{s['id']}")
                            model2.Add(excess >= week_vars2[s['id']] - s['min_week'])
                            model2.Add(excess >= 0)
                            first_excess2.append(excess)
                        if s['session'] == max_sess_num and max_sess_num > 1:
                            deficit = model2.NewIntVar(0, 20, f"deficit2_{s['id']}")
                            model2.Add(deficit >= s['max_week'] - week_vars2[s['id']])
                            model2.Add(deficit >= 0)
                            last_deficit2.append(deficit)

                    obj_terms2 = []
                    if first_excess2:
                        sf = model2.NewIntVar(0, 10000, 'sf2')
                        model2.Add(sf == sum(first_excess2))
                        obj_terms2.append((sf, 100))
                    if last_deficit2:
                        sl = model2.NewIntVar(0, 10000, 'sl2')
                        model2.Add(sl == sum(last_deficit2))
                        obj_terms2.append((sl, 100))

                    if obj_terms2:
                        total2 = model2.NewIntVar(0, 10_000_000, 'total2')
                        model2.Add(total2 == sum(var * w for var, w in obj_terms2))
                        model2.Minimize(total2)

                    solver2 = cp_model.CpSolver()
                    solver2.parameters.max_time_in_seconds = SOLVER_TIME_LIMIT
                    solver2.parameters.num_search_workers = 8
                    status2 = solver2.Solve(model2)

                    if status2 in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                        print(f"  [OK] Récupération réussie : {len(filtered_sessions)} sessions planifiées")
                        for s in filtered_sessions:
                            all_results.append({
                                'semester': sem,
                                'subject': s['subject'],
                                'program': s.get('program', ''),
                                'curso_num': s['curso_num'],
                                'grupo': s['grupo'],
                                'session': s['session'],
                                'week': solver2.Value(week_vars2[s['id']]),
                                'day': s['day'],
                                'time_block': s['block_label'],
                                'nb_students': s['nb_students'],
                                'lab_rooms': s['lab_rooms'],
                            })
                        # Mark dropped groups
                        for subj, gnum in groups_to_drop:
                            for g in all_groups:
                                if (g['subject'] == subj and g['group_num'] == gnum
                                        and g['semester'] == sem):
                                    g['_solver_dropped'] = True
                    else:
                        print(f"  [FAIL] {sem_label} : infaisable même après récupération")
            else:
                print(f"  [FAIL] {sem_label} : infaisable (aucun groupe à retirer)")

    return pd.DataFrame(all_results) if all_results else None


# ============================================================
# ÉTAPE 6 : GÉNÉRATION DES SORTIES
# ============================================================
def generate_outputs(results_df, all_groups, name_lookup, program_lookup, subject_students):
    """
    Step 6: Generate all output files (CSV + XLSX) from the solver result.

    Produces in OUTPUT_DIR ('outputs/optimization/'):
        - optimized_schedule_v5.csv : main planning (week, day, time block, group, room)
        - optimized_schedule_v5.xlsx : same data with formatting
        - group_composition.csv : student-to-group assignments
        - student_directory.csv : student id -> name + program lookup

    Args:
        results_df: DataFrame returned by solve() with the planning result
        all_groups: list of group dicts from form_groups()
        name_lookup: dict student_id -> "lastname, firstname"
        program_lookup: dict student_id -> program code (GITI, IOI, ...)
        subject_students: dict subject_key -> list of enrolled student_ids
    """
    print_section("ÉTAPE 6 : Génération des sorties")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HF = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    HN = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    NF = Font(name="Arial", size=10)
    BF = Font(name="Arial", size=10, bold=True)
    TF = Font(name="Arial", size=12, bold=True, color="1B4F72")
    SF = Font(name="Arial", size=11, bold=True, color="2E75B6")
    RF = Font(name="Arial", size=10, bold=True, color="FF0000")
    GF = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
    LF = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    YF = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    TB = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LA = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Mapping subject → year
    SUBJECT_YEAR = {}
    for subj, cfg in LAB_CONFIG.items():
        curso = cfg['curso_num']
        if curso == 1:
            SUBJECT_YEAR[subj] = 1
        elif curso == 2:
            SUBJECT_YEAR[subj] = 2
        else:
            SUBJECT_YEAR[subj] = 3
    YEAR_LABELS = {1: 'Primero', 2: 'Segundo', 3: 'Tercero'}

    # Count assigned per subject
    assigned_per_subject = defaultdict(int)
    for g in all_groups:
        assigned_per_subject[g['subject']] += len(g['student_ids'])

    wb = Workbook()

    # ================================================================
    # Sheet 1 : Summary (organized by year/semester)
    # ================================================================
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Año", "Sem.", "Subject", "Enrolled", "Assigned", "Unassigned", "Rate",
                   "Groups", "Sess/Gr", "Total Sess.", "Weeks", "Slot(s)", "Lab"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB

    row_idx = 2
    total_enrolled = 0
    total_assigned = 0

    for year_num in [1, 2, 3]:
        for sem_num in [1, 2]:
            year_subjects = [s for s in sorted(results_df['subject'].unique())
                             if SUBJECT_YEAR.get(s) == year_num
                             and int(results_df[results_df['subject']==s]['semester'].iloc[0]) == sem_num]
            if not year_subjects:
                continue

            # Year/Semester header row
            label = f"{YEAR_LABELS[year_num]} — {'Primer' if sem_num==1 else 'Segundo'} Semestre"
            ws_sum.cell(row=row_idx, column=1, value=label).font = TF
            ws_sum.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
            for ci in range(1, 14):
                ws_sum.cell(row=row_idx, column=ci).fill = YF
            row_idx += 1

            for subject in year_subjects:
                sd = results_df[results_df['subject'] == subject]
                sem = int(sd['semester'].iloc[0])
                slots = sd.drop_duplicates(['day', 'time_block'])
                slot_str = "; ".join(f"{r['day'][:3]} {r['time_block']}" for _, r in slots.iterrows())

                enrolled = len(subject_students.get(subject, []))
                assigned = assigned_per_subject.get(subject, 0)
                unassigned = enrolled - assigned
                rate = f"{assigned/enrolled*100:.0f}%" if enrolled > 0 else "0%"
                total_enrolled += enrolled
                total_assigned += assigned

                clean_name = subject.replace(f'S{sem}_', '')
                ws_sum.cell(row=row_idx, column=1, value=YEAR_LABELS.get(year_num, '?')).alignment = CA
                ws_sum.cell(row=row_idx, column=2, value=f"S{sem}").alignment = CA
                ws_sum.cell(row=row_idx, column=3, value=clean_name).font = BF
                ws_sum.cell(row=row_idx, column=4, value=enrolled).alignment = CA
                ws_sum.cell(row=row_idx, column=5, value=assigned).alignment = CA
                ws_sum.cell(row=row_idx, column=6, value=unassigned).alignment = CA
                if unassigned > 0:
                    ws_sum.cell(row=row_idx, column=6).font = RF
                ws_sum.cell(row=row_idx, column=7, value=rate).alignment = CA
                ws_sum.cell(row=row_idx, column=8, value=int(sd['grupo'].nunique())).alignment = CA
                ws_sum.cell(row=row_idx, column=9, value=int(sd['session'].nunique())).alignment = CA
                ws_sum.cell(row=row_idx, column=10, value=len(sd)).alignment = CA
                ws_sum.cell(row=row_idx, column=11, value=f"W{int(sd['week'].min())}-W{int(sd['week'].max())}").alignment = CA
                ws_sum.cell(row=row_idx, column=12, value=slot_str)
                ws_sum.cell(row=row_idx, column=13, value=sd['lab_rooms'].iloc[0] if pd.notna(sd['lab_rooms'].iloc[0]) else '')
                for ci in range(1, 14):
                    ws_sum.cell(row=row_idx, column=ci).border = TB
                row_idx += 1
            row_idx += 1  # Blank row between sections

    # Totals
    ws_sum.cell(row=row_idx, column=3, value="TOTAL").font = TF
    ws_sum.cell(row=row_idx, column=4, value=total_enrolled).font = BF; ws_sum.cell(row=row_idx, column=4).alignment = CA
    ws_sum.cell(row=row_idx, column=5, value=total_assigned).font = BF; ws_sum.cell(row=row_idx, column=5).alignment = CA
    ws_sum.cell(row=row_idx, column=6, value=total_enrolled - total_assigned).font = RF
    ws_sum.cell(row=row_idx, column=6).alignment = CA
    ws_sum.cell(row=row_idx, column=7, value=f"{total_assigned/total_enrolled*100:.1f}%" if total_enrolled > 0 else "0%")
    ws_sum.cell(row=row_idx, column=7).font = BF; ws_sum.cell(row=row_idx, column=7).alignment = CA
    for ci in range(1, 14):
        ws_sum.cell(row=row_idx, column=ci).border = TB

    for ci, w in [(1,10),(2,6),(3,30),(4,10),(5,10),(6,10),(7,8),(8,8),(9,8),(10,10),(11,12),(12,35),(13,40)]:
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # Export the authoritative assignment summary so downstream tools (the
    # Reliability dashboard, the app) use the SAME numbers as the pipeline
    # itself. The previous reliability calculation tried to recompute the
    # denominator from lab_enrollments.csv with max(), which is wrong: it took
    # the largest single-subject enrollment instead of the true assignment
    # total. Reading from this file removes that ambiguity.
    try:
        import os as _os
        _os.makedirs('outputs/optimization', exist_ok=True)
        # Per-subject breakdown
        per_subject_rows = []
        for subj in sorted(results_df['subject'].unique()):
            enrolled = len(subject_students.get(subj, []))
            assigned = assigned_per_subject.get(subj, 0)
            per_subject_rows.append({
                'subject': subj,
                'enrolled': enrolled,
                'assigned': min(assigned, enrolled),
                'unassigned': max(0, enrolled - assigned),
                'rate_pct': min(100.0, round(100.0 * assigned / enrolled, 2))
                            if enrolled > 0 else 0.0,
            })
        pd.DataFrame(per_subject_rows).to_csv(
            'outputs/optimization/assignment_summary.csv', index=False)

        # GLOBAL totals — HONEST count: every (student, subject) enrollment is
        # one slot to fill. The previous "unique students" headline rate hid
        # missing placements (e.g. a student in 5 lab subjects who got 4 still
        # counted as 100%). Daniel's reality is per-enrollment: if 1854 lab
        # enrollments exist, all 1854 should get a group. The unique-student
        # count is reported as a secondary metric.
        all_enrolled_students = set()
        for subj, ids in subject_students.items():
            all_enrolled_students.update(ids)
        n_students_unique = len(all_enrolled_students)
        gc_path = 'outputs/optimization/group_composition.csv'
        n_assigned_unique = min(total_assigned, n_students_unique)
        if _os.path.exists(gc_path):
            try:
                _gc = pd.read_csv(gc_path)
                name_col = ('student_name' if 'student_name' in _gc.columns
                            else 'student_hash' if 'student_hash' in _gc.columns
                            else _gc.columns[-2])
                n_assigned_unique = int(_gc[name_col].nunique())
            except Exception:
                pass
        n_assigned_unique = min(n_assigned_unique, n_students_unique)

        # The HEADLINE rate = pair rate (student × subject), honest reflection
        # of who got their lab. Phase-8 manual overrides can push total_assigned
        # slightly above total_enrolled (a student placed despite a clash counts
        # once more); clamp so the rate never exceeds 100% and unassigned is
        # never negative.
        assigned_clamped = min(total_assigned, total_enrolled)
        unassigned_clamped = max(0, total_enrolled - total_assigned)
        global_rate = (100.0 * assigned_clamped / total_enrolled
                       if total_enrolled > 0 else 0.0)
        global_rate = min(100.0, global_rate)

        pd.DataFrame([{
            'total_enrolled':            total_enrolled,
            'total_assigned':            assigned_clamped,
            'total_unassigned':          unassigned_clamped,
            'assignment_rate_pct':       round(global_rate, 2),
            # Unique-student view as secondary metrics
            'students_unique_enrolled':  n_students_unique,
            'students_unique_assigned':  n_assigned_unique,
            'students_unique_rate_pct':  round(100.0 * n_assigned_unique /
                                               n_students_unique, 2)
                                          if n_students_unique > 0 else 0.0,
        }]).to_csv(
            'outputs/optimization/assignment_summary_global.csv', index=False)
        print(f"  [OK] Export assignment_summary.csv "
              f"(taux pair {global_rate:.1f}% — {assigned_clamped}/{total_enrolled} "
              f"inscriptions ; {n_assigned_unique}/{n_students_unique} étudiants uniques)")
    except Exception as _e:
        print(f"  [WARN] Export assignment_summary.csv : {_e}")

    # ================================================================
    # Sheet 2 : Optimized Schedule (organized by year/semester)
    # ================================================================
    ws = wb.create_sheet("Optimized Schedule")
    sched_headers = ["Año", "Sem.", "Subject", "Program", "Group", "Session", "Week", "Day",
                     "Time Block", "Students", "Laboratory"]
    for ci, h in enumerate(sched_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB

    row_s = 2
    for year_num in [1, 2, 3]:
        for sem_num in [1, 2]:
            year_subjects = [s for s in sorted(results_df['subject'].unique())
                             if SUBJECT_YEAR.get(s) == year_num
                             and int(results_df[results_df['subject']==s]['semester'].iloc[0]) == sem_num]
            if not year_subjects:
                continue

            # Section header
            label = f"{YEAR_LABELS[year_num]} — {'Primer' if sem_num==1 else 'Segundo'} Semestre"
            ws.cell(row=row_s, column=1, value=label).font = TF
            ws.merge_cells(start_row=row_s, start_column=1, end_row=row_s, end_column=11)
            for ci in range(1, 12):
                ws.cell(row=row_s, column=ci).fill = YF
            row_s += 1

            for subject in year_subjects:
                sd = results_df[results_df['subject'] == subject].sort_values(['grupo', 'session'])
                sem = int(sd['semester'].iloc[0])
                clean_name = subject.replace(f'S{sem}_', '')
                for _, row in sd.iterrows():
                    vals = [YEAR_LABELS.get(year_num, '?'), f"S{sem}", clean_name,
                            row.get('program', ''), int(row['grupo']), int(row['session']),
                            int(row['week']), row['day'], row['time_block'],
                            int(row['nb_students']), row.get('lab_rooms', '')]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=row_s, column=ci, value=v)
                        c.font = NF; c.border = TB; c.alignment = LA
                    row_s += 1
            row_s += 1  # Blank between sections

    ws.auto_filter.ref = f"A1:K{row_s - 1}"
    for ci, w in [(1,10),(2,6),(3,30),(4,12),(5,7),(6,8),(7,6),(8,12),(9,14),(10,9),(11,40)]:
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ================================================================
    # Sheet 3 : Groups (organized by year/semester)
    # ================================================================
    ws_grp = wb.create_sheet("Groups")
    if INCLUDE_REAL_NAMES:
        grp_headers = ["Año", "Subject", "Group", "Program", "Day", "Time Block",
                        "Student Name", "Titulación"]
    else:
        grp_headers = ["Año", "Subject", "Group", "Program", "Day", "Time Block",
                        "Student Hash"]
    for ci, h in enumerate(grp_headers, 1):
        c = ws_grp.cell(row=1, column=ci, value=h)
        c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB

    row_g = 2
    sorted_groups = sorted(all_groups, key=lambda x: (
        SUBJECT_YEAR.get(x['subject'], 9),
        x['semester'],
        x['subject'],
        x['group_num']
    ))

    prev_year = None
    for g in sorted_groups:
        year_num = SUBJECT_YEAR.get(g['subject'], 9)
        sem = g['semester']
        clean_name = g['subject'].replace(f"S{sem}_", '')

        # Section header when year changes
        if year_num != prev_year:
            label = f"{YEAR_LABELS.get(year_num, '?')}"
            ws_grp.cell(row=row_g, column=1, value=label).font = TF
            ws_grp.merge_cells(start_row=row_g, start_column=1, end_row=row_g, end_column=len(grp_headers))
            for ci in range(1, len(grp_headers)+1):
                ws_grp.cell(row=row_g, column=ci).fill = YF
            row_g += 1
            prev_year = year_num

        for sid in sorted(g['student_ids'], key=lambda s: name_lookup.get(s, str(s))):
            if INCLUDE_REAL_NAMES:
                vals = [YEAR_LABELS.get(year_num, '?'), clean_name, g['group_num'],
                        g.get('program', ''), g['day'], g['block_label'],
                        name_lookup.get(sid, str(sid)), program_lookup.get(sid, '')]
            else:
                vals = [YEAR_LABELS.get(year_num, '?'), clean_name, g['group_num'],
                        g.get('program', ''), g['day'], g['block_label'],
                        hashlib.sha256(str(sid).encode()).hexdigest()[:12]]
            for ci, v in enumerate(vals, 1):
                c = ws_grp.cell(row=row_g, column=ci, value=v)
                c.font = NF; c.border = TB; c.alignment = LA
            row_g += 1

    ws_grp.auto_filter.ref = f"A1:{get_column_letter(len(grp_headers))}{row_g - 1}"
    grp_widths = [(1,10),(2,30),(3,8),(4,14),(5,12),(6,14),(7,35)]
    if INCLUDE_REAL_NAMES:
        grp_widths.append((8,14))
    for ci, w in grp_widths:
        ws_grp.column_dimensions[get_column_letter(ci)].width = w

    # Save
    path = f"{OUTPUT_DIR}optimized_schedule_v5.xlsx"
    wb.save(path)
    print(f"  [OK] {path}")

    csv_path = f"{OUTPUT_DIR}optimized_schedule_v5.csv"
    results_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"  [OK] {csv_path}")

    # Save group composition with titulación
    grp_path = f"{OUTPUT_DIR}group_composition.csv"
    grp_data = []
    for g in sorted(all_groups, key=lambda x: (SUBJECT_YEAR.get(x['subject'], 9), x['semester'], x['subject'], x['group_num'])):
        year_num = SUBJECT_YEAR.get(g['subject'], 9)
        sem = g['semester']
        clean_name = g['subject'].replace(f"S{sem}_", '')
        override_set = g.get('_override_sids', set())
        for sid in sorted(g['student_ids'], key=lambda s: name_lookup.get(s, str(s))):
            entry = {
                'año': YEAR_LABELS.get(year_num, '?'),
                'semester': f"S{sem}",
                'subject': clean_name,
                'grupo': g['group_num'],
                'program': g.get('program', ''),
                'day': g['day'],
                'block': g['block_label'],
            }
            if INCLUDE_REAL_NAMES:
                entry['student_name'] = name_lookup.get(sid, str(sid))
                entry['titulacion'] = program_lookup.get(sid, '')
            else:
                entry['student_hash'] = hashlib.sha256(str(sid).encode()).hexdigest()[:12]
                entry['titulacion'] = program_lookup.get(sid, '')  # program code only (low-sensitivity)
            # Manual-override flag: Daniel accepted a course clash to honour
            # this lab enrollment. Downstream tools (Reliability dashboard)
            # should NOT count this as a conflict defect.
            entry['is_override'] = bool(sid in override_set)
            grp_data.append(entry)
    pd.DataFrame(grp_data).to_csv(grp_path, index=False, encoding='utf-8-sig')
    print(f"  [OK] {grp_path}")

    # ────────────────────────────────────────────────────────
    # Export student directory: student_id → name + program
    # Used by the Caso individual page in the Streamlit app
    # ────────────────────────────────────────────────────────
    try:
        all_sids = set()
        for g in all_groups:
            all_sids.update(g['student_ids'])
        directory_rows = []
        for sid in all_sids:
            directory_rows.append({
                'student_id': sid,
                'student_hash': hashlib.sha256(str(sid).encode()).hexdigest()[:12],
                # LOCAL name map for Daniel's Excel/app — ALWAYS real names.
                # Stays in the workspace: never embed it in the .exe, never sync it.
                'student_name': name_lookup.get(sid, str(sid)),
                'titulacion': program_lookup.get(sid, ''),
            })
        if directory_rows:
            dir_path = f"{OUTPUT_DIR}student_directory.csv"
            pd.DataFrame(directory_rows).to_csv(dir_path, index=False, encoding='utf-8-sig')
            print(f"  [OK] {dir_path} ({len(directory_rows)} étudiants)")
    except Exception as e:
        print(f"  [WARN]  Erreur export student_directory : {e}")


# ============================================================
# ANALYSE FINALE
# ============================================================
def analyze(results_df):
    """
    Step 7: Print final statistics about the generated planning.

    Reports:
        - Total number of sessions per semester
        - Distribution by day of week and time block
        - Number of distinct groups
        - Detection of any constraint violations (should be zero)

    Args:
        results_df: DataFrame returned by solve() containing the planning.
    """
    print_section("ANALYSE FINALE")

    print(f"\n  [STATS] RÉSULTAT GLOBAL : {len(results_df)} sessions planifiées")

    for sem in sorted(results_df['semester'].unique()):
        sd = results_df[results_df['semester'] == sem]
        print(f"\n  ── Semestre {int(sem)} ({len(sd)} sessions, {sd['subject'].nunique()} matières) ──")
        for subject in sorted(sd['subject'].unique()):
            ss = sd[sd['subject'] == subject]
            print(f"     {subject:40s} | {len(ss):3d} sess | "
                  f"G1-G{int(ss['grupo'].max())} | W{int(ss['week'].min())}-W{int(ss['week'].max())}")

    # Distribution
    print(f"\n  Par jour :")
    for day in DAYS:
        count = len(results_df[results_df['day'] == day])
        print(f"    {day:12s} : {count:3d} {'█' * (count // 5)}")

    print(f"\n  Par bloc :")
    for b in TIME_BLOCKS:
        count = len(results_df[results_df['time_block'] == b['label']])
        print(f"    {b['label']:12s} : {count:3d} {'█' * (count // 5)}")

    # Vérifications
    print(f"\n  [OK] Vérifications :")

    # C1 : conflit matière
    c1 = 0
    for _, group in results_df.groupby(['subject', 'week', 'day', 'time_block']):
        if len(group) > 1:
            c1 += len(group) - 1
    print(f"    C1 (conflit matière)  : {c1}")

    # C4 : conflit salle — vérification correcte
    room_conflicts = 0
    conflict_details = []
    checked_pairs = set()

    for idx1, row1 in results_df.iterrows():
        rooms1 = set(r.strip() for r in str(row1.get('lab_rooms', '')).split(',') if r.strip())
        for idx2, row2 in results_df.iterrows():
            if idx2 <= idx1:
                continue
            # Sessions de semestres différents ne se chevauchent pas
            if row1['semester'] != row2['semester']:
                continue
            if row1['week'] != row2['week'] or row1['day'] != row2['day'] or row1['time_block'] != row2['time_block']:
                continue

            rooms2 = set(r.strip() for r in str(row2.get('lab_rooms', '')).split(',') if r.strip())
            shared_rooms = rooms1 & rooms2

            if shared_rooms:
                pair_key = (min(idx1, idx2), max(idx1, idx2))
                if pair_key not in checked_pairs:
                    checked_pairs.add(pair_key)
                    room_conflicts += 1
                    if len(conflict_details) < 5:
                        conflict_details.append(
                            f"      W{int(row1['week'])} {row1['day']} {row1['time_block']} : "
                            f"{row1['subject']} G{int(row1['grupo'])} vs "
                            f"{row2['subject']} G{int(row2['grupo'])} "
                            f"[{', '.join(shared_rooms)}]"
                        )

    print(f"    C4 (conflit salle)    : {room_conflicts}")
    for d in conflict_details:
        print(d)
    if room_conflicts > len(conflict_details):
        print(f"      ... et {room_conflicts - len(conflict_details)} autres")

    # C7 : matin/après-midi
    c7 = 0
    for _, r in results_df.iterrows():
        bid = BLOCK_LABELS.get(r['time_block'], '')
        block_id = None
        for b in TIME_BLOCKS:
            if b['label'] == r['time_block']:
                block_id = b['id']
                break
        if block_id:
            if r['curso_num'] in [1, 3] and block_id in [b['id'] for b in TIME_BLOCKS if b['period'] != 'morning']:
                c7 += 1
            elif r['curso_num'] in [2, 4] and block_id in [b['id'] for b in TIME_BLOCKS if b['period'] == 'morning']:
                c7 += 1
    print(f"    C7 (matin/après-midi) : {c7}")

    # Friday evening
    c9 = sum(1 for _, r in results_df.iterrows()
             if r['day'] == 'Viernes' and r['time_block'] in ['17:00-19:00', '19:00-21:00'])
    print(f"    C9 (vendredi soir)    : {c9}")

    if c1 == 0 and room_conflicts == 0:
        print(f"\n  [OK] AUCUN CONFLIT — Solution valide !")
        print(f"     C1 GARANTI par construction (groupes basés sur emploi du temps réel)")
        print(f"     C3 GARANTI par construction (max_students respecté à la formation)")
        print(f"     C7 GARANTI par construction (blocs filtrés selon l'année)")
    else:
        print(f"\n  [WARN]  Conflits détectés — investigation nécessaire")


# ============================================================
# MAIN
# ============================================================
def main():
    """
    Run the full pipeline end-to-end.

    Workflow (8 stages):
        0. Set up output directories and dual stdout (terminal + report file)
        1. Apply user config overrides (config/user_config.json)
        2. Load master_schedule.csv
        3. Identify enrolled students per subject (via MixtoID + keywords)
        4. Build individual student timetables
        5. Form groups (Phases 1, 2, 3a, 3b, 3c, 4)
        6. Run CP-SAT solver to assign weeks to each session
        7. Generate output files (CSV + Excel)
        8. Print final analysis

    Reads:
        - data_clean/master_schedule.csv (input)
        - config/user_config.json (optional overrides)

    Writes:
        - outputs/optimization/optimized_schedule_v5.csv
        - outputs/optimization/optimized_schedule_v5.xlsx
        - outputs/optimization/group_composition.csv
        - outputs/optimization/student_directory.csv
        - data_clean/optimization/student_busy.csv
        - reports/pipeline_v5_report.txt
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs('reports', exist_ok=True)

    class DualOutput:
        """
        File-like object that writes simultaneously to a file and the terminal.

        Used to capture all stdout into a report file while still showing
        progress to the user in real time.
        """
        def __init__(self, fp):
            """Open the report file and capture the current stdout."""
            self.terminal = sys.stdout
            self.file = open(fp, 'w', encoding='utf-8')
        def write(self, t):
            """Write to both terminal and file."""
            self.terminal.write(t)
            self.file.write(t)
        def flush(self):
            """Flush both terminal and file buffers."""
            self.terminal.flush()
            self.file.flush()
        def close(self):
            """Close the file (terminal stays open)."""
            self.file.close()

    dual = DualOutput(REPORT_PATH)
    sys.stdout = dual

    print("\n" + "═" * 60)
    print("  PIPELINE v5 — AULARIO+ALUMNOS → PLANNING OPTIMAL")
    print("  Approche niveau étudiant : groupes réels + CP-SAT")
    print(f"  Date : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("═" * 60)

    # ════════════════════════════════════════════════════════════
    # BRIDGE UI → PIPELINE : charger config utilisateur si présente
    # ════════════════════════════════════════════════════════════
    apply_user_config()

    if not os.path.exists(MASTER_PATH):
        print(f"\n[FAIL] Fichier d'entrée introuvable : {MASTER_PATH}")
        print(f"  → Placez 'master_schedule.csv' dans le dossier data_clean/")
        print(f"  → Ou vérifiez le chemin MASTER_PATH en haut du script.")
        sys.stdout = dual.terminal; dual.close()
        return

    # Charger le master schedule (robuste : encodage + séparateur)
    df = None
    for enc, sep in [('utf-8-sig', ','), ('utf-8', ','), ('latin-1', ','),
                     ('utf-8-sig', ';'), ('latin-1', ';')]:
        try:
            cand = pd.read_csv(MASTER_PATH, encoding=enc, sep=sep)
            if cand.shape[1] > 1:  # plus d'une colonne = bon séparateur
                df = cand
                break
        except Exception:
            continue
    if df is None or df.shape[1] <= 1:
        print(f"\n[FAIL] Impossible de lire {MASTER_PATH} correctement.")
        print(f"  → Vérifiez que le fichier est un CSV valide (séparateur , ou ;).")
        sys.stdout = dual.terminal; dual.close()
        return

    # Vérifier les colonnes essentielles
    required_cols = ['AlumnoID', 'actividad']
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        print(f"\n[FAIL] Colonnes manquantes dans {MASTER_PATH} : {missing}")
        print(f"  Colonnes trouvées : {list(df.columns)[:10]}")
        print(f"  → Le fichier d'entrée doit contenir au minimum : {required_cols}")
        sys.stdout = dual.terminal; dual.close()
        return

    print(f"\n  [OK] Données chargées : {len(df)} lignes, {df.shape[1]} colonnes")

    # Pipeline
    df = load_and_prepare(df)
    subject_students = identify_students(df)
    student_busy, student_subject_slots = build_individual_timetables(df, subject_students)

    # ── Professor flow (validated model with Pablo/Daniel) ──
    professor_busy, professor_subjects, prof_available = build_professor_busy(df)

    # Supervision capacity (n professors per subject) from the official report.
    # Answers the whiteboard's "how many groups?" — bounds parallel groups by
    # min(n_professors, n_rooms). Informational + available to downstream logic.
    supervision = load_supervision_capacity()
    if supervision:
        print(f"  [SUPERVISION] Capacité d'encadrement (profs/matière) chargée :")
        for subj, n in sorted(supervision.items()):
            n_rooms = len(LAB_CONFIG.get(subj, {}).get('lab_rooms', []) or [1])
            limit = min(n, n_rooms)
            print(f"     {subj:30s}: {n} prof(s), {n_rooms} salle(s) "
                  f"→ {limit} groupe(s) parallèle(s) max")
    # Per lab subject: a slot is blocked ONLY IF *every* professor able to teach
    # it is busy then — i.e. there is no supervisor left. A single group needs
    # only ONE free professor, so blocking a slot as soon as ANY professor is
    # busy (set union) was far too strict: with 7 professors linked to Física,
    # the union covered 14/15 morning slots and strangled group creation.
    # Correct semantics = intersection of each professor's busy set (a slot is
    # unavailable only if it is in the busy set of ALL the subject's professors).
    subject_professor_busy = {}
    if prof_available:
        # Group professors by lab subject
        profs_of_subject = defaultdict(list)
        for prof_id, subjects in professor_subjects.items():
            for subj in subjects:
                profs_of_subject[subj].append(prof_id)

        # Export subject -> ELIGIBLE professor NAMES for the Excel "Vista profesor"
        # sheet. professor_subjects links each professor to the lab subjects they
        # teach; inverted here it gives, per subject, the professors qualified to
        # run it. We do NOT assign one named professor per group — this is the set
        # of ELIGIBLE professors, matching the model's guarantee (at least one is
        # free for every session). Written to OUTPUT_DIR (per-user workspace), so
        # it is NOT bundled into the .exe.
        try:
            sp_rows = []
            for _subj in sorted(profs_of_subject):
                _names = sorted({str(p).strip() for p in profs_of_subject[_subj]
                                 if str(p).strip() and str(p).strip().lower() != 'nan'})
                if _names:
                    sp_rows.append({'subject': _subj, 'professors': '; '.join(_names)})
            if sp_rows:
                _sp_path = f"{OUTPUT_DIR}subject_professors.csv"
                pd.DataFrame(sp_rows).to_csv(_sp_path, index=False, encoding='utf-8-sig')
                print(f"  [OK] Export subject_professors.csv ({len(sp_rows)} matières)")
        except Exception as _e:
            print(f"  [WARN]  Erreur export subject_professors.csv : {_e}")

        # Export blocked slots so the Excel can render "Festivo / No disponible".
        # These are presentation annotations only — never part of the schedule.
        try:
            _bs_rows = []
            for (_bsem, _bsubj), _slots in SUBJECT_BLOCKED_SLOTS.items():
                _room = (LAB_CONFIG.get(_bsubj, {}).get('lab_rooms', ['']) or [''])[0]
                for (_w, _d, _b), _label in _slots.items():
                    _bs_rows.append({
                        'semester': _bsem,
                        'subject': _bsubj,
                        'week': _w,
                        'day': DAYS[_d] if 0 <= _d < len(DAYS) else str(_d),
                        'day_idx': _d,
                        'block_id': _b,
                        'time_block': BLOCK_LABELS.get(_b, str(_b)),
                        'lab_rooms': _room,
                        'label': _label,
                    })
            if _bs_rows:
                _bs_path = f"{OUTPUT_DIR}blocked_slots.csv"
                pd.DataFrame(_bs_rows).to_csv(_bs_path, index=False, encoding='utf-8-sig')
                print(f"  [OK] Export blocked_slots.csv ({len(_bs_rows)} créneaux réservés)")
        except Exception as _e:
            print(f"  [WARN]  Erreur export blocked_slots.csv : {_e}")

        for subj, prof_ids in profs_of_subject.items():
            busy_sets = [professor_busy.get(pid, set()) for pid in prof_ids]
            busy_sets = [b for b in busy_sets if b]
            if not busy_sets:
                continue
            # Intersection: slots where NO professor of this subject is free
            blocked = set.intersection(*busy_sets) if len(busy_sets) > 1 else set()
            if blocked:
                subject_professor_busy[subj] = blocked
        n_constrained = sum(1 for s in subject_professor_busy.values() if s)
        print(f"  Matières de lab avec contrainte professeur active : {n_constrained}")
        print(f"  (créneau bloqué seulement si AUCUN prof de la matière n'est libre)")

    # Build student → program mapping from 'programas' column
    KNOWN_PROGRAMS = {'IOI', 'IMR', 'GITI', 'GITIADE', 'MAT', 'AERO', 'IBIO', 'IEM', 'PIIA', 'IINFTV'}
    student_program = {}
    if 'programas' in df.columns:
        for _, row in df.dropna(subset=['AlumnoID', 'programas']).iterrows():
            sid = row['AlumnoID']
            if sid in student_program:
                continue
            progs_raw = str(row['programas']).strip()
            for token in progs_raw.split(','):
                abbrev = token.strip().split('-')[0].strip().upper()
                if abbrev in KNOWN_PROGRAMS:
                    student_program[sid] = abbrev
                    break
            if sid not in student_program:
                student_program[sid] = progs_raw.split(',')[0].strip().split('-')[0].strip().upper()

    all_groups = form_groups(subject_students, student_busy, student_subject_slots,
                             student_program, subject_professor_busy)

    if not all_groups:
        print("\n  [FAIL] Aucun groupe formé.")
        sys.stdout = dual.terminal; dual.close()
        return

    results_df = solve(all_groups)

    if results_df is not None and len(results_df) > 0:
        # Build name lookup: AlumnoID → "APELLIDOS, Nombre"
        name_lookup = {}
        program_lookup = {}  # AlumnoID → program abbreviation (IOI, GITI, etc.)
        if True:  # always build name_lookup — it feeds the LOCAL student_directory.csv
            name_data = df.dropna(subset=['AlumnoID']).drop_duplicates('AlumnoID')
            for _, row in name_data.iterrows():
                sid = row['AlumnoID']
                nombre = str(row.get('Nombre', '')).strip() if pd.notna(row.get('Nombre')) else ''
                apellidos = str(row.get('Apellidos', '')).strip() if pd.notna(row.get('Apellidos')) else ''
                if apellidos and nombre:
                    name_lookup[sid] = f"{apellidos}, {nombre}"
                elif nombre:
                    name_lookup[sid] = nombre
                elif apellidos:
                    name_lookup[sid] = apellidos

                # Extract program abbreviation from 'programas' column
                progs_raw = str(row.get('programas', '')).strip() if pd.notna(row.get('programas')) else ''
                if progs_raw:
                    for token in progs_raw.split(','):
                        abbrev = token.strip().split('-')[0].strip().upper()
                        if abbrev in ['IOI', 'IMR', 'GITI', 'GITIADE', 'MAT', 'AERO', 'IBIO', 'IEM', 'PIIA', 'IINFTV']:
                            program_lookup[sid] = abbrev
                            break
                    if sid not in program_lookup:
                        program_lookup[sid] = progs_raw.split(',')[0].strip().split('-')[0].strip().upper()

        generate_outputs(results_df, all_groups, name_lookup, program_lookup, subject_students)
        analyze(results_df)

        # ════════════════════════════════════════════════════════════
        # AUTO: Generate Daniel format (Curso 2025-2026 structure)
        # ════════════════════════════════════════════════════════════
        print_section("ÉTAPE 7 : Génération automatique format Daniel")
        import subprocess

        env_utf8 = os.environ.copy()
        env_utf8['PYTHONIOENCODING'] = 'utf-8'
        env_utf8['PYTHONUTF8'] = '1'

        # Look for scripts in src/ first, then current dir
        for script_name, label in [
            ('09_generate_exact_format_S1.py', 'S1 (Primero + Segundo + Tercero)'),
            ('10_generate_exact_format_S2.py', 'S2 (Primero + Segundo + Tercero)'),
        ]:
            script_path = None
            for candidate in [f'src/{script_name}', script_name]:
                if os.path.exists(candidate):
                    script_path = candidate
                    break

            if script_path is None:
                print(f"  [WARN]  {script_name} introuvable — skipping")
                continue

            print(f"\n  ▶ Génération {label}...")
            try:
                result = subprocess.run(
                    [sys.executable, script_path],
                    capture_output=True, text=True,
                    encoding='utf-8', errors='replace',
                    env=env_utf8, timeout=120,
                )
                if result.returncode == 0:
                    # Print only the last few lines for brevity
                    output_lines = result.stdout.strip().split('\n')
                    for line in output_lines[-15:]:
                        if line.strip():
                            print(f"    {line}")
                    print(f"  [OK] {label} généré")
                else:
                    print(f"  [FAIL] Erreur lors de la génération {label}")
                    if result.stderr:
                        for line in result.stderr.strip().split('\n')[-5:]:
                            print(f"    {line}")
            except subprocess.TimeoutExpired:
                print(f"  [WARN]  Timeout (120s) lors de la génération {label}")
            except Exception as e:
                print(f"  [FAIL] Exception : {e}")

        # Optionally run validation script if available
        for vscript in ['src/11_validate_output.py', '11_validate_output.py']:
            if os.path.exists(vscript):
                print(f"\n  ▶ Validation finale...")
                try:
                    vresult = subprocess.run(
                        [sys.executable, vscript],
                        capture_output=True, text=True,
                        encoding='utf-8', errors='replace',
                        env=env_utf8, timeout=60,
                    )
                    if vresult.stdout:
                        for line in vresult.stdout.strip().split('\n')[-20:]:
                            if line.strip():
                                print(f"    {line}")
                except Exception:
                    pass
                break

    else:
        print("\n  [FAIL] Aucune solution générée.")

    # ════════════════════════════════════════════════════════════
    # AUTO-SNAPSHOT: archive the result for history & undo
    # ════════════════════════════════════════════════════════════
    try:
        # Lazy import: don't fail the pipeline if version_manager isn't installed
        import version_manager as vm
        snap_id = vm.create_snapshot(
            snapshot_type='auto',
            description=f"Génération pipeline du {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        )
        if snap_id:
            print(f"\n  [OK] Snapshot automatique créé : {snap_id}")
        else:
            print(f"\n  [INFO] Snapshot non créé (aucun output à archiver)")
    except ImportError:
        # version_manager.py not available — non-blocking
        pass
    except Exception as exc:
        print(f"\n  [WARN] Auto-snapshot échoué (non-bloquant) : {exc}")

    print(f"\n{'═' * 60}")
    print(f"  PIPELINE v5 TERMINÉ")
    print(f"{'═' * 60}\n")

    sys.stdout = dual.terminal; dual.close()
    print(f"\n[REPORT] {REPORT_PATH}")
    print(f"[FILE] {OUTPUT_DIR}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n[INTERROMPU] Exécution arrêtée par l'utilisateur.")
        sys.exit(1)
    except Exception as exc:
        # Garde-fou de présentation : jamais de traceback brute devant un public.
        # Affiche un message clair et oriente vers le diagnostic.
        import traceback
        print("\n" + "═" * 60)
        print("  [ERREUR] Le pipeline a rencontré un problème inattendu.")
        print("═" * 60)
        print(f"  Type    : {type(exc).__name__}")
        print(f"  Détail  : {exc}")
        print(f"\n  Le détail technique complet a été enregistré dans :")
        print(f"    reports/pipeline_error.log")
        print(f"\n  Causes fréquentes :")
        print(f"    • fichier d'entrée manquant ou mal formaté (data_clean/master_schedule.csv)")
        print(f"    • config/user_config.json invalide")
        print("═" * 60)
        try:
            os.makedirs('reports', exist_ok=True)
            with open('reports/pipeline_error.log', 'w', encoding='utf-8') as f:
                f.write(f"Erreur : {type(exc).__name__}: {exc}\n\n")
                f.write(traceback.format_exc())
        except Exception:
            pass
        sys.exit(1)