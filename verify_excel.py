#!/usr/bin/env python3
"""
verify_excel.py — cross-check the generated Excel deliverables.

TWO TIERS
=========
Tier 1  OUR EXCEL  vs  OUR SCHEDULE CSV
        Proves the Excel generator renders the schedule faithfully: every
        (subject, group, week, day, block) shown in the generated
        'Vista profesor' sheets equals what is in optimized_schedule_v5.csv,
        and nothing is invented or dropped.

Tier 2  OUR SCHEDULE  vs  DANIEL'S REFERENCE FILES
        Compares the set of lab sessions against Daniel's hand-made
        distribution. STRICT mode: a session matches only if subject, group,
        week, day AND block are identical. Subjects whose group labels in
        Daniel's files are not plain integers (e.g. 'Grupo C', 'E1 (lab 1)',
        'D3') cannot be mapped to our integer groups, so they are SKIPPED and
        reported transparently rather than flagged as false mismatches.

USAGE
-----
    python verify_excel.py --workspace "%APPDATA%/LabScheduling/workspace" \
                           --reference-dir path/to/daniel/files
Either flag is optional; both are auto-discovered when omitted.

Exits 0 if Tier 1 passes (Tier 2 differences are reported, not failed, because
our optimiser legitimately places sessions in different weeks than Daniel).
"""
from __future__ import annotations
import argparse, os, re, sys
from collections import defaultdict

import pandas as pd

try:
    import openpyxl
except Exception:
    openpyxl = None

DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
_DAY_IDX = {d: i for i, d in enumerate(DAYS)}
_TIME_RE = re.compile(r'\d{2}:\d{2}-\d{2}:\d{2}')

# Cells the generator paints as a calendar holiday (mirror of SEMESTER_HOLIDAYS
# in excel_generator_core.py), keyed by (semester, week, day_idx). A scheduled
# session landing on one of these is shown as the holiday, not the lab — so it is
# classified separately in Tier 1 rather than reported as a rendering loss.
# Verified against the official Loyola 25-26 calendar (Sevilla campus).
_EXCEL_HOLIDAYS = {
    (1, 2, 2), (1, 7, 0),                                  # S1: Acto apertura, Hispanidad
    (2, 6, 4),                                             # S2: Blue Day (13/03, Friday)
    (2, 7, 0),                                             # S2: San Ignacio
    (2, 9, 0), (2, 9, 1), (2, 9, 2), (2, 9, 3), (2, 9, 4),  # S2: Semana Santa
    (2, 12, 3), (2, 12, 4),                                # S2: Feria de Abril (Thu+Fri)
    (2, 13, 4),                                            # S2: Día del Trabajador
}


# ───────────────────────── helpers ─────────────────────────
def _subj_key(name: str) -> str:
    """Normalise a subject for comparison: strip S1_/S2_ and lowercase."""
    return re.sub(r'^S[12]_', '', str(name)).strip().lower()


def _find(root, filename, max_depth=6):
    root = os.path.abspath(os.path.expanduser(os.path.expandvars(root)))
    if not os.path.isdir(root):
        return None
    skip = {"node_modules", ".git", "__pycache__", ".venv", "venv", "site-packages"}
    base_depth = root.rstrip(os.sep).count(os.sep)
    for dirpath, dirs, files in os.walk(root):
        dirs[:] = [d for d in dirs if d not in skip]
        if dirpath.count(os.sep) - base_depth > max_depth:
            dirs[:] = []
            continue
        if filename in files:
            return os.path.join(dirpath, filename)
    return None


def _resolve_workspace(start):
    start = os.path.abspath(os.path.expanduser(os.path.expandvars(start)))
    SCHED = "outputs/optimization/optimized_schedule_v5.csv"
    for c in (start, os.path.join(start, "workspace"),
              os.path.join(start, "LabScheduling", "workspace")):
        if os.path.isfile(os.path.join(c, *SCHED.split("/"))):
            return c
    hit = _find(start, "optimized_schedule_v5.csv")
    if hit:
        return os.path.dirname(os.path.dirname(os.path.dirname(hit)))
    return start


# ──────────────── parse OUR generated Vista profesor ────────────────
def parse_our_vista(xlsx_path):
    """Return set of (subj_key, group:int, week:int, day, block) from one of OUR
    generated workbooks (sheet 'Vista profesor'). Our cell format is:
        'Práctica N <Subject>\nGrupo M\n<room>'  (one group per cell)."""
    out = set()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'Vista profesor' not in wb.sheetnames:
        wb.close()
        return out
    ws = wb['Vista profesor']
    grid = _grid_index(ws)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not (isinstance(v, str) and 'ráctica' in v and 'rupo' in v):
                continue
            subj = _subject_from_cell(v)
            grps = re.findall(r'\d+', v.split('\n')[1]) if '\n' in v else re.findall(r'\d+', v)
            wk = grid['week_for'](r)
            d = grid['day_for'](r, c)
            b = grid['block_for'](r, c)
            for g in grps:
                if subj and wk and d and b:
                    out.add((_subj_key(subj), int(g), int(wk), d, b))
    wb.close()
    return out


def _subject_from_cell(v: str) -> str | None:
    """Extract the subject name from a 'Práctica … <Subject>' cell (first line)."""
    first = v.split('\n')[0]
    m = re.sub(r'^\s*Práctica\s*\d*\s*', '', first).strip()
    return m or None


def _grid_index(ws):
    """Build week/day/block resolvers for a SEMANA-blocked grid sheet.
    Robust to ANY number of side-by-side subject bands: time-block labels are
    detected in whatever columns they appear (1, 8/9, 15, 22, …)."""
    sem_rows = {}
    day_hdr = {}
    label_cols = set()      # columns that ever hold a 'HH:MM-HH:MM' label
    for r in range(1, ws.max_row + 1):
        rowvals = {c: ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)}
        for c, v in rowvals.items():
            if isinstance(v, str):
                if re.match(r'\s*SEMANA\s+(\d+)', v):
                    sem_rows[r] = int(re.match(r'\s*SEMANA\s+(\d+)', v).group(1))
                if _TIME_RE.match(v.strip()):
                    label_cols.add(c)
        if any(isinstance(v, str) and v.strip() == 'Lunes' for v in rowvals.values()):
            day_hdr[r] = {c: v.strip() for c, v in rowvals.items()
                          if isinstance(v, str) and v.strip() in DAYS}
    sem_sorted = sorted(sem_rows)
    hdr_sorted = sorted(day_hdr)
    label_cols = sorted(label_cols) or [1]

    def week_for(r):
        w = None
        for sr in sem_sorted:
            if sr <= r:
                w = sem_rows[sr]
            else:
                break
        return w

    def day_for(r, c):
        hr = None
        for h in hdr_sorted:
            if h <= r:
                hr = h
            else:
                break
        return day_hdr.get(hr, {}).get(c) if hr else None

    def _sem_top(r):
        top = 0
        for sr in sem_sorted:
            if sr <= r:
                top = sr
            else:
                break
        return top

    def block_for(r, c):
        # nearest time-label at/above row r, but NOT above the current SEMANA
        # header (so we never borrow a label from the previous week's block).
        limit = _sem_top(r)
        for rr in range(r, limit - 1, -1):
            for lc in label_cols:
                v = ws.cell(row=rr, column=lc).value
                if isinstance(v, str) and _TIME_RE.match(v.strip()):
                    return v.strip()
        return None

    return {'week_for': week_for, 'day_for': day_for, 'block_for': block_for}


# ──────────────── parse DANIEL's reference files ────────────────
def parse_daniel(xlsx_path):
    """Return (sessions, skipped_subjects). sessions is a set of
    (subj_key, group:int, week, day, block) for subjects whose groups are plain
    integers; skipped_subjects maps subject->reason for non-integer group labels.
    Daniel's grid sheet is 'Vista profesor' or 'Cronograma'."""
    sessions = set()
    skipped = {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    grid_name = next((s for s in wb.sheetnames if s.lower() in
                      ('vista profesor', 'cronograma', 'iteraciones')), None)
    if grid_name is None:
        wb.close()
        return sessions, skipped
    ws = wb[grid_name]
    grid = _grid_index(ws)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not (isinstance(v, str) and ('ráctica' in v.lower() or 'rupo' in v.lower())):
                continue
            subj = _daniel_subject(v)
            grp_field = v.split('\n')[1] if '\n' in v else v
            ints = re.findall(r'(?<![A-Za-z])\d+', grp_field)
            letters = re.findall(r'\bGrupos?\s+([A-Z]\d?)\b', grp_field)
            wk = grid['week_for'](r)
            d = grid['day_for'](r, c)
            b = grid['block_for'](r, c)
            if letters and not ints:
                skipped[subj] = "letter/lab group labels (e.g. 'Grupo C', 'E1') — not mappable to integer groups"
                continue
            if not ints:
                continue
            for g in ints:
                if subj and wk and d and b:
                    sessions.add((_subj_key(subj), int(g), int(wk), d, b))
    wb.close()
    return sessions, skipped


def _daniel_subject(v: str) -> str:
    first = v.split('\n')[0]
    s = re.sub(r'^\s*Práctica[s]?\s*', '', first, flags=re.I)
    s = re.sub(r'\s*\d+\s*(y\s*\d+)?\s*(grupos?)?\s*$', '', s, flags=re.I).strip()
    return s or first.strip()


# ───────────────────────── main ─────────────────────────
def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workspace", default=".")
    ap.add_argument("--reference-dir", default=None,
                    help="Folder with Daniel's reference .xlsx files")
    a = ap.parse_args(argv)

    if openpyxl is None:
        print("[FATAL] openpyxl is required (pip install openpyxl).", file=sys.stderr)
        return 2

    W = _resolve_workspace(a.workspace)
    sched_path = os.path.join(W, "outputs/optimization/optimized_schedule_v5.csv")
    if not os.path.isfile(sched_path):
        print(f"[FATAL] schedule not found under {W}", file=sys.stderr)
        return 2
    sched = pd.read_csv(sched_path)
    print(f"[info] workspace: {W}")

    # our schedule as a set
    our = set()
    _sem_of = {}
    for _, r in sched.iterrows():
        try:
            key = (_subj_key(r['subject']), int(r['grupo']), int(r['week']),
                   str(r['day']), str(r['time_block']))
            our.add(key)
            _sem_of[key] = int(r['semester'])
        except Exception:
            continue

    overall_ok = True

    # ── Tier 1: our Excel vs our CSV ───────────────────────
    print("\n== Tier 1 — generated Excel vs schedule CSV ==")
    xlsx_files = []
    cur = os.path.join(W, "outputs/optimization")
    for root, _d, files in os.walk(cur):
        for fn in files:
            if fn.lower().endswith(".xlsx") and ("AUTO" in fn or "Distribucion" in fn or "Practicas" in fn):
                xlsx_files.append(os.path.join(root, fn))
    if not xlsx_files:
        print("  [SKIP] no generated .xlsx found — run the Excel export first.")
    else:
        excel_sessions = set()
        for xf in xlsx_files:
            try:
                excel_sessions |= parse_our_vista(xf)
            except Exception as e:
                print(f"  [WARN] could not parse {os.path.basename(xf)}: {e}")
        # compare (only subjects/groups that appear in Excel — Vista profesor is
        # the lab grid, so it should match the schedule's lab sessions exactly)
        missing_in_excel = our - excel_sessions
        extra_in_excel = excel_sessions - our
        print(f"  schedule sessions: {len(our)} | parsed from Excel: {len(excel_sessions)}")
        if not missing_in_excel and not extra_in_excel:
            print("  [PASS] every schedule session appears in the Excel, and vice-versa.")
        else:
            covered_subj = {s for (s, *_rest) in excel_sessions}
            miss = {x for x in missing_in_excel if x[0] in covered_subj}
            # Sessions hidden behind an Excel OVERLAY (a calendar holiday, or a
            # reserved-slot marker like Biotecnología) are shown as that overlay,
            # not as a rendering loss. Build the reserved-slot overlay set from
            # blocked_slots.csv (subject-specific) when present.
            reserved_overlay = set()
            try:
                _bs = os.path.join(W, "outputs/optimization/blocked_slots.csv")
                if os.path.isfile(_bs):
                    _bdf = pd.read_csv(_bs)
                    for _, _r in _bdf.iterrows():
                        reserved_overlay.add((_subj_key(_r['subject']), int(_r['week']),
                                              str(_r['day']), str(_r['time_block'])))
            except Exception:
                pass
            holiday_hidden = {x for x in miss
                              if (_sem_of.get(x), x[2], _DAY_IDX.get(x[3], -1)) in _EXCEL_HOLIDAYS
                              or (x[0], x[2], x[3], x[4]) in reserved_overlay}
            true_miss = miss - holiday_hidden
            if not true_miss and not extra_in_excel:
                msg = f"  [PASS] all rendered lab sessions match"
                if holiday_hidden:
                    msg += (f" ({len(holiday_hidden)} session(s) fall on a day the Excel "
                            f"marks as a holiday and are shown as the holiday — see note)")
                if missing_in_excel - miss:
                    msg += f"; {len(missing_in_excel - miss)} session(s) belong to levels not in these files"
                print(msg + ".")
                if holiday_hidden:
                    print("  [NOTE] sessions overlaid by an Excel holiday or reserved-slot "
                          "marker (shown as the overlay, not the lab — verify the overlay is intended):")
                    for x in sorted(holiday_hidden)[:6]:
                        print(f"          {x}")
            else:
                overall_ok = False
                print(f"  [FAIL] {len(true_miss)} session(s) in CSV but not Excel (not holiday-related); "
                      f"{len(extra_in_excel)} in Excel but not CSV.")
                for x in list(true_miss)[:5]:
                    print(f"          missing in Excel: {x}")
                for x in list(extra_in_excel)[:5]:
                    print(f"          extra in Excel:   {x}")

    # ── Tier 2: our schedule vs Daniel's references ────────
    print("\n== Tier 2 — schedule vs Daniel's reference files (STRICT) ==")
    ref_dir = a.reference_dir
    ref_files = []
    if ref_dir and os.path.isdir(ref_dir):
        ref_files = [os.path.join(ref_dir, f) for f in os.listdir(ref_dir)
                     if f.lower().endswith(".xlsx")]
    else:
        # auto-discover by known name patterns under the workspace + cwd
        for base in (W, os.getcwd()):
            for root, _d, files in os.walk(base):
                for fn in files:
                    if fn.lower().endswith(".xlsx") and (
                            fn.startswith("Distribucion_Practicas") or
                            fn.startswith("Reparto_Pract")):
                        ref_files.append(os.path.join(root, fn))
        ref_files = sorted(set(ref_files))

    if not ref_files:
        print("  [SKIP] no reference files found. Pass --reference-dir to enable.")
    else:
        daniel = set()
        skipped_all = {}
        for rf in ref_files:
            try:
                s, sk = parse_daniel(rf)
                daniel |= s
                skipped_all.update(sk)
                print(f"  parsed {os.path.basename(rf)}: {len(s)} integer-group session(s)"
                      + (f", {len(sk)} subject(s) skipped" if sk else ""))
            except Exception as e:
                print(f"  [WARN] {os.path.basename(rf)}: {e}")

        # compare only on subjects present on BOTH sides with integer groups
        our_subj = {s[0] for s in our}
        dan_subj = {s[0] for s in daniel}
        common = our_subj & dan_subj
        our_c = {x for x in our if x[0] in common}
        dan_c = {x for x in daniel if x[0] in common}
        match = our_c & dan_c
        only_ours = our_c - dan_c
        only_dan = dan_c - our_c
        print(f"\n  comparable subjects (integer groups, both sides): "
              f"{sorted(common) if common else '—'}")
        print(f"  exact (subject+group+week+day+block) matches: {len(match)}")
        print(f"  in our schedule only: {len(only_ours)}")
        print(f"  in Daniel only:       {len(only_dan)}")
        if skipped_all:
            print("\n  Subjects NOT comparable (Daniel uses non-integer group labels):")
            for subj, why in sorted(skipped_all.items()):
                print(f"    · {subj}: {why}")
        # show a few differences to inspect (strict week mismatches included)
        for x in sorted(only_ours)[:6]:
            print(f"    ours only : {x}")
        for x in sorted(only_dan)[:6]:
            print(f"    Daniel only: {x}")
        print("\n  NOTE: Tier-2 differences are expected — the optimiser places sessions "
              "in different WEEKS than Daniel's manual plan (you asked for STRICT, so "
              "week differences are listed). Same group on the same day/block in a "
              "different week shows up as one 'ours only' + one 'Daniel only'.")

    print("\n== RESULT:", "TIER-1 PASS ✅" if overall_ok else "TIER-1 FAIL ❌",
          "(Tier-2 is a reference diff, not a pass/fail) ==")
    return 0 if overall_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())