# -*- coding: utf-8 -*-
"""
validation_credits.py — Rapport de validation des crédits des professeurs
=========================================================================

OBJECTIF
--------
Comparer, matière par matière (et professeur par professeur), DEUX sources :

  1. Les CRÉDITS de laboratoire (caractère « P ») assignés dans
     « Asignacion_2025-2026_v5.xlsx », feuille « Asignación docente ».
  2. Le NOMBRE RÉEL de séances de labo PLANIFIÉES dans la sortie de
     l'optimiseur (optimized_schedule_v5.csv / .xlsx).

CONVENTION (validée par le coordinateur) :  1 crédit P  =  5 séances de labo.
Donc :  Sessions attendues = (Σ crédits P) × 5.

Le rapport calcule l'ÉCART = Sessions planifiées − Sessions attendues, et
SIGNALE (sans rien bloquer) les divergences :
  • matière planifiée SANS crédit P (labo donné mais budgété en théorie « T ») ;
  • crédits P assignés à une matière NON planifiée par l'outil (hors périmètre) ;
  • écart numérique entre attendu et planifié.

SORTIE : validation_credits_professeurs.xlsx, avec trois feuilles :
  • « Résumé par matière »      : la validation de référence (écart par matière).
  • « Détail par professeur »   : colonnes demandées (Professeur, Matière, Groupe,
                                  Crédits assignés, Sessions attendues,
                                  Sessions planifiées, Écart).
  • « Méthodologie & alertes »  : hypothèses, limites et liste des alertes.

NOTE IMPORTANTE SUR L'ATTRIBUTION PAR PROFESSEUR
------------------------------------------------
L'optimiseur ne nomme PAS un professeur unique par séance : il garantit qu'au
moins un professeur habilité est libre. Le nombre de séances planifiées est donc
une donnée AU NIVEAU DE LA MATIÈRE. Dans la feuille « Détail par professeur »,
les séances planifiées de la matière sont RÉPARTIES entre les professeurs « P »
au prorata de leurs crédits, afin que l'écart par ligne reste cohérent et que la
somme des écarts par professeur égale l'écart de la matière. Ce choix est
explicité dans la feuille « Méthodologie & alertes ».
"""

from __future__ import annotations

import os
import sys
import unicodedata
import pandas as pd

import professor_credits as pc

CREDIT_TO_SESSIONS = 5  # 1 crédit P = 5 séances de labo (convention coordinateur)

DEFAULT_ASIGNACION = "Asignacion_2025-2026_v5.xlsx"
DEFAULT_OUTPUT = "validation_credits_professeurs.xlsx"

def _find_asignacion_file():
    """Localise le fichier Asignacion avec plusieurs chemins possibles."""
    candidates = [
        DEFAULT_ASIGNACION,
        '/home/ubuntu/Uploads/Asignacion_2025-2026_v5.xlsx',
        '/home/ubuntu/Shared/Uploads/Asignacion_2025-2026_v5.xlsx',
        '/home/ubuntu/lab_project/Asignacion_2025-2026_v5.xlsx',
        'data/Asignacion_2025-2026_v5.xlsx',
        'data_clean/Asignacion_2025-2026_v5.xlsx',
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    # Recherche récursive en dernier recours
    for root_dir in ['/home/ubuntu', os.getcwd()]:
        for dirpath, _, files in os.walk(root_dir):
            for f in files:
                if f.lower().startswith('asignacion') and f.endswith('.xlsx'):
                    return os.path.join(dirpath, f)
    return DEFAULT_ASIGNACION


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _norm(s):
    """Minuscule, sans accents, espaces normalisés — pour comparer les noms."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def _strip_prefix(name):
    """Retire le préfixe S1_/S2_ d'un nom de matière planifiée."""
    s = str(name)
    if s[:3] in ("S1_", "S2_"):
        return s[3:]
    return s


def _load_lab_config():
    """Récupère LAB_CONFIG depuis pipeline.py (mots-clés de mapping matière)."""
    try:
        import pipeline as P
        return P.LAB_CONFIG
    except Exception as exc:  # pragma: no cover - fallback si ortools absent
        print(f"  [WARN] Impossible d'importer pipeline.LAB_CONFIG ({exc}).")
        print("         Le mapping matière planifiée -> Asignación sera limité.")
        return {}


# --------------------------------------------------------------------------- #
# 1) Chargement du planning de sortie -> séances par matière / groupe
# --------------------------------------------------------------------------- #
def load_schedule(schedule_path):
    """Charge la sortie de l'optimiseur (CSV pipeline OU classeur xlsx).

    Renvoie un DataFrame normalisé avec au moins les colonnes :
        subject, grupo, session, day, time_block, semester
    Une ligne = une séance (groupe × numéro de séance).
    """
    if not os.path.exists(schedule_path):
        raise FileNotFoundError(f"Planning introuvable : {schedule_path}")

    ext = os.path.splitext(schedule_path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(schedule_path)
    else:
        # Classeur formaté : la feuille 'Optimized Schedule' contient le détail.
        xls = pd.ExcelFile(schedule_path)
        sheet = next((s for s in xls.sheet_names
                      if _norm(s) in ("optimized schedule", "planning", "schedule")),
                     xls.sheet_names[0])
        df = pd.read_excel(schedule_path, sheet_name=sheet)
        # Harmonise les noms de colonnes du classeur vers le schéma CSV pipeline.
        rename = {
            "Subject": "subject", "Group": "grupo", "Session": "session",
            "Day": "day", "Time Block": "time_block", "Week": "week",
            "Sem.": "semester", "Program": "program", "Students": "nb_students",
            "Laboratory": "lab_rooms",
        }
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Lignes valides uniquement (le classeur contient des lignes-titres « Año »).
    df = df[df["subject"].notna()].copy()
    df["subject"] = df["subject"].astype(str).str.strip()
    if "grupo" in df.columns:
        df["grupo"] = pd.to_numeric(df["grupo"], errors="coerce")
    return df


def sessions_by_subject(schedule_df):
    """Renvoie {matière_planifiée: {'sessions': int, 'grupos': int}}.

    'sessions' = nombre de lignes (groupe × séance) = séances de labo planifiées.
    'grupos'   = nombre de groupes distincts planifiés pour la matière.
    """
    out = {}
    for subj, grp in schedule_df.groupby("subject"):
        n_sessions = int(len(grp))
        n_grupos = int(grp["grupo"].nunique()) if "grupo" in grp.columns else 0
        out[str(subj)] = {"sessions": n_sessions, "grupos": n_grupos}
    return out


# --------------------------------------------------------------------------- #
# 2) Mapping matière planifiée  <->  matière(s) de l'Asignación
# --------------------------------------------------------------------------- #
def map_scheduled_to_asignacion(scheduled_subjects, asignacion_subjects, lab_config):
    """Pour chaque matière planifiée (ex. 'Física'), trouve la/les matière(s) de
    l'Asignación correspondante(s) via les mots-clés de LAB_CONFIG (même logique
    que le pipeline). Renvoie {matière_planifiée: [matières_asignación]}.

    Les noms planifiés sont dépouillés du préfixe S1_/S2_, on cherche donc la clé
    LAB_CONFIG dont le nom dépouillé correspond.
    """
    # Index : nom planifié dépouillé -> clé LAB_CONFIG (avec préfixe)
    stripped_to_cfg = {}
    for cfg_key in lab_config:
        stripped_to_cfg[_norm(_strip_prefix(cfg_key))] = cfg_key

    mapping = {}
    asig_list = list(asignacion_subjects)
    for sched in scheduled_subjects:
        cfg_key = stripped_to_cfg.get(_norm(_strip_prefix(sched)))
        matched = []
        if cfg_key is not None:
            cfg = lab_config[cfg_key]
            kws = [_norm(k) for k in cfg.get("keywords", [])]
            exc = [_norm(e) for e in cfg.get("keyword_exclude", [])]
            for asub in asig_list:
                na = _norm(asub)
                if any(k in na for k in kws) and not any(e in na for e in exc):
                    matched.append(asub)
        # Repli : correspondance exacte du nom dépouillé si aucun mot-clé n'a marché
        if not matched:
            target = _norm(_strip_prefix(sched))
            for asub in asig_list:
                if _norm(asub) == target:
                    matched.append(asub)
        mapping[sched] = matched
    return mapping


# --------------------------------------------------------------------------- #
# 3) Construction du rapport
# --------------------------------------------------------------------------- #
def build_report(asignacion_path=None, schedule_path=None,
                 output_path=DEFAULT_OUTPUT):
    """Génère le classeur de validation. Renvoie (summary_df, detail_df, alerts)."""
    if asignacion_path is None:
        asignacion_path = _find_asignacion_file()
    print(f"# Validation des crédits")
    print(f"  Asignación : {asignacion_path}")
    print(f"  Planning   : {schedule_path}")

    # --- Asignación docente (caractères T/P, forme longue) ---
    assign = pc.parse_assignment(asignacion_path)      # offering_id, subject, prof_code, credits, char
    budgets = pc.load_budgets(asignacion_path)         # prof_code -> prof_name
    code_to_name = dict(zip(budgets["prof_code"], budgets["prof_name"]))

    def prof_label(code):
        name = code_to_name.get(code)
        return f"{name} ({code})" if name and _norm(name) != _norm(code) else str(code)

    # --- Planning de sortie ---
    schedule_df = load_schedule(schedule_path)
    sched_info = sessions_by_subject(schedule_df)
    scheduled_subjects = sorted(sched_info.keys())

    # --- Mapping matière planifiée -> matière(s) Asignación ---
    lab_config = _load_lab_config()
    all_asig_subjects = sorted(assign["subject"].unique())
    mapping = map_scheduled_to_asignacion(scheduled_subjects, all_asig_subjects, lab_config)

    alerts = []
    summary_rows = []
    detail_rows = []
    matched_asig_subjects = set()

    # ===== A) Une entrée par matière PLANIFIÉE =====
    for sched in scheduled_subjects:
        planned_sessions = sched_info[sched]["sessions"]
        n_grupos = sched_info[sched]["grupos"]
        asig_subs = mapping.get(sched, [])
        matched_asig_subjects.update(asig_subs)

        # lignes P (labo) des matières Asignación correspondantes
        labP = assign[(assign["subject"].isin(asig_subs)) & (assign["char"] == "P")]
        theoryT = assign[(assign["subject"].isin(asig_subs)) & (assign["char"] == "T")]
        sum_p_credits = float(labP["credits"].sum())
        expected_sessions = sum_p_credits * CREDIT_TO_SESSIONS
        ecart = planned_sessions - expected_sessions

        if not asig_subs:
            estado = "⚠ Matière planifiée sans correspondance dans l'Asignación"
            alerts.append(f"{sched} : {planned_sessions} séances planifiées mais "
                          f"AUCUNE matière correspondante trouvée dans l'Asignación.")
        elif sum_p_credits == 0:
            estado = "⚠ Séances planifiées SANS crédit P (labo budgété en théorie ?)"
            alerts.append(f"{sched} : {planned_sessions} séances planifiées mais "
                          f"0 crédit P assigné (matière présente uniquement en 'T').")
        elif abs(ecart) < 1e-6:
            estado = "OK"
        else:
            estado = "⚠ Écart attendu/planifié"
            alerts.append(f"{sched} : écart de {ecart:+.0f} séances "
                          f"(attendu {expected_sessions:.0f}, planifié {planned_sessions}).")

        summary_rows.append({
            "Matière (planning)": _strip_prefix(sched),
            "Matière(s) Asignación": " / ".join(asig_subs) if asig_subs else "—",
            "Crédits P assignés (Σ)": round(sum_p_credits, 2),
            "Sessions attendues (créd×5)": round(expected_sessions, 1),
            "Groupes planifiés": n_grupos,
            "Sessions planifiées": planned_sessions,
            "Écart (planifié − attendu)": round(ecart, 1),
            "État": estado,
        })

        # ----- Détail par professeur (répartition au prorata des crédits P) -----
        if sum_p_credits > 0:
            per_prof = (labP.groupby("prof_code")["credits"].sum().reset_index())
            for _, r in per_prof.iterrows():
                cr = float(r["credits"])
                exp = cr * CREDIT_TO_SESSIONS
                share_planned = planned_sessions * (cr / sum_p_credits)
                detail_rows.append({
                    "Professeur": prof_label(r["prof_code"]),
                    "Matière": _strip_prefix(sched),
                    "Groupe": f"Tous ({n_grupos} gr.)",
                    "Crédits assignés": round(cr, 2),
                    "Sessions attendues": round(exp, 1),
                    "Sessions planifiées": round(share_planned, 1),
                    "Écart": round(share_planned - exp, 1),
                    "Type": "P",
                })
        else:
            # Aucun crédit P : on liste les professeurs 'T' pour information.
            per_prof_t = (theoryT.groupby("prof_code")["credits"].sum().reset_index())
            for _, r in per_prof_t.iterrows():
                detail_rows.append({
                    "Professeur": prof_label(r["prof_code"]),
                    "Matière": _strip_prefix(sched),
                    "Groupe": f"Tous ({n_grupos} gr.)",
                    "Crédits assignés": round(float(r["credits"]), 2),
                    "Sessions attendues": 0.0,
                    "Sessions planifiées": planned_sessions if len(per_prof_t) == 1 else "",
                    "Écart": "",
                    "Type": "T (pas de crédit P)",
                })
            if len(per_prof_t) == 0:
                detail_rows.append({
                    "Professeur": "— (non assigné)",
                    "Matière": _strip_prefix(sched),
                    "Groupe": f"Tous ({n_grupos} gr.)",
                    "Crédits assignés": 0.0,
                    "Sessions attendues": 0.0,
                    "Sessions planifiées": planned_sessions,
                    "Écart": planned_sessions,
                    "Type": "—",
                })

    # ===== B) Crédits P assignés mais matière NON planifiée (hors périmètre) =====
    labP_all = assign[assign["char"] == "P"]
    unscheduled = (labP_all[~labP_all["subject"].isin(matched_asig_subjects)]
                   .groupby("subject")["credits"].sum().reset_index())
    for _, r in unscheduled.iterrows():
        cr = float(r["credits"])
        if cr <= 0:
            continue
        expected = cr * CREDIT_TO_SESSIONS
        summary_rows.append({
            "Matière (planning)": "— (non planifié)",
            "Matière(s) Asignación": r["subject"],
            "Crédits P assignés (Σ)": round(cr, 2),
            "Sessions attendues (créd×5)": round(expected, 1),
            "Groupes planifiés": 0,
            "Sessions planifiées": 0,
            "Écart (planifié − attendu)": round(-expected, 1),
            "État": "Hors périmètre — crédits P sans planning labo (autre programme/área)",
        })
        alerts.append(f"{r['subject']} : {cr:g} crédits P assignés mais AUCUNE "
                      f"séance planifiée par l'outil (matière hors périmètre labo).")
        # détail par professeur de ces matières non planifiées
        sub_rows = labP_all[labP_all["subject"] == r["subject"]]
        for _, rr in sub_rows.groupby("prof_code")["credits"].sum().reset_index().iterrows():
            crp = float(rr["credits"])
            detail_rows.append({
                "Professeur": prof_label(rr["prof_code"]),
                "Matière": r["subject"],
                "Groupe": "—",
                "Crédits assignés": round(crp, 2),
                "Sessions attendues": round(crp * CREDIT_TO_SESSIONS, 1),
                "Sessions planifiées": 0,
                "Écart": round(-crp * CREDIT_TO_SESSIONS, 1),
                "Type": "P (non planifié)",
            })

    summary_df = pd.DataFrame(summary_rows)
    detail_df = pd.DataFrame(detail_rows)

    # Ligne de TOTAL sur le résumé (sur le périmètre planifié uniquement)
    planned_mask = summary_df["Matière (planning)"] != "— (non planifié)"
    total_row = {
        "Matière (planning)": "TOTAL (périmètre planifié)",
        "Matière(s) Asignación": "",
        "Crédits P assignés (Σ)": round(summary_df.loc[planned_mask, "Crédits P assignés (Σ)"].sum(), 2),
        "Sessions attendues (créd×5)": round(summary_df.loc[planned_mask, "Sessions attendues (créd×5)"].sum(), 1),
        "Groupes planifiés": int(summary_df.loc[planned_mask, "Groupes planifiés"].sum()),
        "Sessions planifiées": int(summary_df.loc[planned_mask, "Sessions planifiées"].sum()),
        "Écart (planifié − attendu)": round(summary_df.loc[planned_mask, "Écart (planifié − attendu)"].sum(), 1),
        "État": "",
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    _write_workbook(output_path, summary_df, detail_df, alerts)
    print(f"\n  [OK] Rapport écrit : {output_path}")
    print(f"       {len(summary_df)-1} matières analysées | {len(alerts)} alerte(s) signalée(s)")
    return summary_df, detail_df, alerts


# --------------------------------------------------------------------------- #
# 4) Écriture du classeur formaté
# --------------------------------------------------------------------------- #
def _write_workbook(output_path, summary_df, detail_df, alerts):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")
    WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
    TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)

    def _sheet_from_df(title, df, status_col=None):
        ws = wb.create_sheet(title)
        cols = list(df.columns)
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=1, column=j, value=col)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = CENTER; c.border = BORDER
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            is_total = str(row.get(cols[0], "")).startswith("TOTAL")
            status_val = str(row.get(status_col, "")) if status_col else ""
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=i, column=j, value=row[col])
                c.border = BORDER
                c.alignment = WRAP if isinstance(row[col], str) else CENTER
                if is_total:
                    c.fill = TOTAL_FILL; c.font = Font(bold=True)
                elif status_col:
                    if status_val == "OK":
                        c.fill = OK_FILL
                    elif status_val.startswith("⚠") or status_val.startswith("Hors"):
                        c.fill = WARN_FILL
        # largeurs de colonnes
        for j, col in enumerate(cols, start=1):
            maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
            ws.column_dimensions[get_column_letter(j)].width = min(48, max(12, maxlen + 2))
        ws.freeze_panes = "A2"
        return ws

    _sheet_from_df("Résumé par matière", summary_df, status_col="État")
    _sheet_from_df("Détail par professeur", detail_df, status_col="Type")

    # Feuille méthodologie & alertes
    ws = wb.create_sheet("Méthodologie & alertes")
    lines = [
        ("MÉTHODOLOGIE", True),
        ("Convention : 1 crédit P (laboratoire) = 5 séances de labo.", False),
        ("Sessions attendues = (Σ crédits P de la matière) × 5.", False),
        ("Sessions planifiées = nombre de séances (groupe × n° de séance) dans la sortie de l'optimiseur.", False),
        ("Écart = Sessions planifiées − Sessions attendues.", False),
        ("Source crédits : feuille « Asignación docente » (blocs Prof. 1..4, caractère T/P).", False),
        ("Mapping matière planifiée ↔ Asignación : mots-clés de LAB_CONFIG (même logique que le pipeline).", False),
        ("", False),
        ("ATTRIBUTION PAR PROFESSEUR", True),
        ("L'optimiseur ne nomme pas un professeur unique par séance (co-encadrement courant).", False),
        ("Dans « Détail par professeur », les séances planifiées de la matière sont réparties", False),
        ("entre les professeurs P au prorata de leurs crédits ; la somme des écarts par professeur", False),
        ("égale donc l'écart de la matière. La validation de référence est la feuille « Résumé par matière ».", False),
        ("", False),
        (f"ALERTES ({len(alerts)})", True),
    ]
    r = 1
    for text, is_head in lines:
        c = ws.cell(row=r, column=1, value=text)
        if is_head:
            c.font = Font(bold=True, color="1F4E78", size=12)
        r += 1
    for a in alerts:
        ws.cell(row=r, column=1, value="• " + a).alignment = WRAP
        r += 1
    if not alerts:
        ws.cell(row=r, column=1, value="Aucune alerte : tous les crédits P correspondent aux séances planifiées.")
    ws.column_dimensions["A"].width = 120

    wb.save(output_path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def _find_default_schedule():
    """Cherche la sortie de l'optimiseur aux emplacements habituels."""
    candidates = [
        "outputs/optimization/optimized_schedule_v5.csv",
        "optimized_schedule_v5.csv",
        "outputs/optimization/optimized_schedule_v5.xlsx",
        "optimized_schedule_v5.xlsx",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


if __name__ == "__main__":
    asig = sys.argv[1] if len(sys.argv) > 1 else _find_asignacion_file()
    sched = sys.argv[2] if len(sys.argv) > 2 else _find_default_schedule()
    out = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_OUTPUT
    if not sched:
        print("ERREUR : planning de sortie introuvable. Usage :")
        print("  python validation_credits.py [Asignacion.xlsx] [optimized_schedule_v5.csv|.xlsx] [sortie.xlsx]")
        sys.exit(1)
    build_report(asig, sched, out)
