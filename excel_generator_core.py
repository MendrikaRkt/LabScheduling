"""
excel_generator_core.py — Standalone Excel formatter (Daniel's exact layout).

Extracted from the former 09_generate_exact_format_S1.py so the app no longer
depends on external generator scripts. Contains only reusable functions and
constants — no main(), no import-time side effects. Driven by excel_export.py
for both S1 and S2.
"""

"""
09_generate_exact_format_S1
================================

Generate Excel deliverables for Semester 1 (Primero, Segundo, Tercero) using
the EXACT visual format validated by Daniel (lab scheduling coordinator).

Format requirements applied:
    - Thin borders on EVERY table cell (Daniel pixel-by-pixel review)
    - Dates centered in Iteraciones tab
    - All time blocks displayed (even empty ones)
    - All weeks displayed (even weeks with no labs)
    - Time block labels right-aligned
    - Course content centered in cells
    - Horarios: 2 programs stacked per row block

Inputs:
    outputs/optimization/optimized_schedule_v5.csv   (CP-SAT planning)
    outputs/optimization/group_composition.csv       (student-to-group mapping)
    data_clean/master_schedule.csv                   (raw timetable for Horarios)
    config/user_config.json                          (user overrides; optional)

Outputs (in outputs/optimization/Curso_2025_2026/):
    Primero/Primer semestre/Distribucion_Practicas_AUTO.xlsx
    Segundo/Primer semestre/Distribucion_Practicas_segundocurso_AUTO.xlsx
    Tercero/Primer semestre/Distribucion_Practicas_tercercurso_AUTO.xlsx
"""

# =============================================================================
# WINDOWS ENCODING FIX (must run before other imports)
# =============================================================================
import sys
import io

if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

# =============================================================================
# IMPORTS
# =============================================================================
import os
import json
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# FILE PATHS
# =============================================================================
SCHEDULE_CSV_PATH    = 'outputs/optimization/optimized_schedule_v5.csv'
GROUP_COMPOSITION_PATH = 'outputs/optimization/group_composition.csv'
MASTER_SCHEDULE_PATH = 'data_clean/master_schedule.csv'
SUPERVISION_PATH     = 'data_clean/optimization/subject_supervision.csv'
SUBJECT_PROFESSORS_PATH = 'outputs/optimization/subject_professors.csv'  # subject -> professor NAMES (workspace; not bundled)
OUTPUT_BASE_DIR      = 'outputs/optimization/Curso_2025_2026'
USER_CONFIG_PATH     = 'config/user_config.json'


# =============================================================================
# CALENDAR / SEMESTER PARAMETERS
# =============================================================================
S1_FIRST_DAY = datetime(2025, 9, 1)   # S1 week 1 (Mon of that week)
S2_FIRST_DAY = datetime(2026, 2, 2)   # S2 week 1 Monday (02 Feb 2026, per academic calendar)
DEFAULT_NUM_WEEKS = 14
NUM_WEEKS = DEFAULT_NUM_WEEKS  # Will be overridden by user config if present


def _schedule_week_span(schedule_df):
    """Return (min_week, max_week) actually present in this schedule.

    The Excel week grid must cover the real range of the sessions it renders.
    S1 labs run ~W3-W14 but S2 labs run ~W8-W20, so the old fixed cap of
    NUM_WEEKS=14 silently truncated every S2 session past week 14. We derive the
    span from the data and only fall back to 1..NUM_WEEKS when no usable 'week'
    values exist (e.g. an empty frame)."""
    try:
        import pandas as _pd
        if schedule_df is not None and 'week' in schedule_df.columns and len(schedule_df):
            w = _pd.to_numeric(schedule_df['week'], errors='coerce').dropna().astype(int)
            if len(w):
                hi = int(w.max())
                # start at week 1 so the calendar reads naturally, and never
                # render fewer weeks than the configured minimum
                return 1, max(hi, NUM_WEEKS)
    except Exception:
        pass
    return 1, NUM_WEEKS


# =============================================================================
# USER CONFIG BRIDGE (UI app -> this script)
#
# When the user changes "S1 total weeks" in the Streamlit app, the value is
# saved to config/user_config.json. We load it here so generated Excel files
# reflect the chosen number of weeks.
# =============================================================================

def load_user_config():
    """
    Load user config from disk if present.

    Returns:
        dict | None: Parsed config dictionary, or None if missing/unreadable.
    """
    if not os.path.exists(USER_CONFIG_PATH):
        return None
    try:
        with open(USER_CONFIG_PATH, 'r', encoding='utf-8') as cfg_file:
            return json.load(cfg_file)
    except Exception as exc:
        print(f"  [WARN] Cannot read {USER_CONFIG_PATH}: {exc}")
        return None


def apply_user_config_overrides():
    """
    Apply user config overrides to module-level globals.

    Currently only overrides NUM_WEEKS based on `s1_total_weeks` setting.
    Called once on module import.
    """
    global NUM_WEEKS

    user_config = load_user_config()
    if not user_config:
        return

    global_section = user_config.get('global', {})
    if 's1_total_weeks' in global_section:
        previous_value = NUM_WEEKS
        NUM_WEEKS = int(global_section['s1_total_weeks'])
        if previous_value != NUM_WEEKS:
            print(f"  [CONFIG] S1 NUM_WEEKS: {previous_value} -> {NUM_WEEKS} (from user_config.json)")


# NOTE: apply_user_config_overrides() is NOT called at import time here.
# excel_export.py calls it explicitly after patching the path constants.


# =============================================================================
# DOMAIN CONSTANTS
# =============================================================================

# Spanish weekday names used throughout Daniel's spreadsheets
DAYS_OF_WEEK = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
DAY_NAME_TO_INDEX = {day_name: index for index, day_name in enumerate(DAYS_OF_WEEK)}

# Lab room display names: keys are full names from the database,
# values are the abbreviated names Daniel validated (April 2026 feedback).
# Both the OLD long names (legacy pipeline) and the NEW short names (corrected
# pipeline, aligned with Daniel's reference files) map to the same display
# value, so this stays correct whichever name the pipeline produces.
LAB_DISPLAY_NAMES = {
    # Old long names (legacy)
    'Ciencias Experimentales I':              'Ciencias Exp. I',
    'Ciencias Experimentales II':             'Ciencias Exp. II',
    'Laboratorio de Ingeniería Telemática':   'Lab. Telemática',
    'Robótica y Automática':                  'Lab. Robótica',
    'Mecánica de Fluidos':                    'Lab. Fluidos',
    'Automoción y Resistencia de Mat.':       'Lab. Resistencia',
    'Eléctrica':                              'Lab. Eléctrica',
    'Electrónica':                            'Lab. Electrónica',
    # New short names (corrected pipeline — already Daniel-conform, idempotent)
    'Lab. Eléctrica':                         'Lab. Eléctrica',
    'Lab. Telemática':                        'Lab. Telemática',
    'Lab. Termodinámica':                     'Lab. Termodinámica',
    'Lab. Robótica':                          'Lab. Robótica',
    'Lab. Electrónica':                       'Lab. Electrónica',
}

# Time block labels (each block = 2 hours)
TIME_BLOCKS = [
    '08:30-10:30',
    '10:30-12:30',
    '12:30-14:30',
    '15:00-17:00',
    '17:00-19:00',
    '19:00-21:00',
]

# Map "minutes since midnight" -> time block label
# Used to assign courses from the raw timetable to a block
MINUTES_TO_TIME_BLOCK = {
    510:  '08:30-10:30',
    630:  '10:30-12:30',
    750:  '12:30-14:30',
    900:  '15:00-17:00',
    960:  '16:00-18:00',
    1020: '17:00-19:00',
    1140: '19:00-21:00',
}

# Known engineering program codes used at Universidad Loyola Sevilla
KNOWN_PROGRAM_CODES = {
    'IOI', 'IMR', 'GITI', 'GITIADE', 'GITIADE22', 'MAT',
    'AERO', 'IBIO', 'IINFTV', 'IEM', 'PIIA', 'MIIU',
}

# S1 academic calendar holidays: (week_number, day_index) -> event name
# Cells matching these are filled with the holiday name (no class possible)
# Holidays painted on the Excel grids, BY SEMESTER and keyed by (week, day_idx)
# in that semester's own week numbering. Verified against the official
# Universidad Loyola academic calendar 25-26 (campus Sevilla — local Córdoba
# holidays are intentionally NOT included).
#   day_idx: 0=Lunes 1=Martes 2=Miércoles 3=Jueves 4=Viernes
SEMESTER_HOLIDAYS = {
    1: {
        # Acto apertura — institutional event (kept on request), W2 Wednesday.
        (2, 2): 'Acto apertura',
        # 13/10 Día de la Hispanidad (W7 Monday).
        (7, 0): 'Día de la Hispanidad',
        # NB: Constitución (6/12) and Andalucía fall on weekends → no lab impact;
        # Inmaculada (8/12) is past the S1 lab weeks; Fuensanta/San Rafael are
        # local Córdoba holidays and are excluded for the Sevilla campus.
    },
    2: {
        # 13/03 Blue Day — institutional university holiday (W6 Friday).
        (6, 4): 'Blue Day',
        # 16/03 Canonización de San Ignacio (W7 Monday).
        (7, 0): 'Canonización de San Ignacio',
        # 30/03–05/04 Semana Santa (full W9).
        (9, 0): 'Semana Santa', (9, 1): 'Semana Santa', (9, 2): 'Semana Santa',
        (9, 3): 'Semana Santa', (9, 4): 'Semana Santa',
        # 23–25/04 Feria de Abril → 23=Thu, 24=Fri (25=Sat): W12 Thu+Fri only.
        (12, 3): 'Feria de Abril', (12, 4): 'Feria de Abril',
        # 1/05 Día del Trabajador (W13 Friday).
        (13, 4): 'Día del Trabajador',
        # NB: Feria de Córdoba (28–30/05) and Corpus Christi (4/06) are local
        # Córdoba holidays → excluded for the Sevilla campus.
    },
}
# Back-compat alias (some call sites referenced S1_HOLIDAYS directly).
S1_HOLIDAYS = SEMESTER_HOLIDAYS[1]


def _load_blocked_slots(semester):
    """Return {(week, day_idx, time_block_label): label} for blocked slots.

    Blocked slots are rooms RESERVED for another activity (e.g. Biotecnología).
    They are NOT part of the optimized schedule (so the reliability conflict
    check never sees them); they are read here purely to render
    'Festivo / No disponible' in the deliverable. Source: blocked_slots.csv,
    written by the pipeline. Searched in the workspace then data_clean/.
    """
    import os as _os
    import pandas as _pd
    out = {}
    _cands = []
    try:
        import app_paths as _ap
        for _rel in ('outputs/optimization/blocked_slots.csv',
                     'data_clean/optimization/blocked_slots.csv'):
            _p = _ap.resolve_existing(_rel)
            if _p:
                _cands.append(_p)
    except Exception:
        pass
    _cands += ['outputs/optimization/blocked_slots.csv',
               'data_clean/optimization/blocked_slots.csv']
    for _cand in _cands:
        try:
            if _cand and _os.path.exists(_cand):
                df = _pd.read_csv(_cand)
                for _, r in df.iterrows():
                    try:
                        if int(r['semester']) != int(semester):
                            continue
                    except Exception:
                        pass
                    key = (int(r['week']), int(r['day_idx']), str(r['time_block']))
                    out[key] = str(r.get('label', '') or 'Festivo / No disponible')
                if out:
                    break
        except Exception:
            continue
    return out


# =============================================================================
# LEVEL DEFINITIONS
#
# Each level (year of study) has its own Excel deliverable with:
#   - subjects: which lab courses appear
#   - programs: which engineering programs to show in Horarios
#   - file: output filename
#   - naming: 'number' (Grupo 1, 2, 3...) or 'letter' (Grupo A, B, C...)
#   - single: True = one combined Grupo tab; False = one tab per subject
# =============================================================================
LEVEL_DEFINITIONS = {
    1: {
        'label': 'Primero',
        'subjects': ['S1_Física', 'S1_Química'],
        'programs': ['IOI', 'AERO', 'IMR', 'GITI', 'MAT', 'GITIADE', 'IBIO'],
        'file': 'Distribucion_Practicas_AUTO.xlsx',
        'naming': 'number',
        'single': True,
    },
    2: {
        'label': 'Segundo',
        'subjects': ['S1_Electrotecnia', 'S1_Mecanismos', 'S1_Termodinámica'],
        'programs': ['IOI', 'IMR', 'GITI', 'GITIADE22'],
        'file': 'Distribucion_Practicas_segundocurso_AUTO.xlsx',
        'naming': 'letter',
        'single': False,
    },
    3: {
        'label': 'Tercero',
        'subjects': [
            'S1_Tecnologías de Fabricación',
            'S1_Robótica y Automatización',
            'S1_Automatización Industrial',
        ],
        'programs': ['IOI', 'IMR', 'GITI', 'GITIADE22', 'PIIA'],
        'file': 'Distribucion_Practicas_tercercurso_AUTO.xlsx',
        'naming': 'letter',
        'single': False,
    },
}


# =============================================================================
# PAIRED INTRO SESSION CONFIGURATION
#
# For these subjects, when the FIRST práctica (session 1) of two consecutive
# groups (g, g+1) lands on the SAME slot (week, day, time block), the two
# cells are merged into a single one displaying "Grupos X & Y".
#
# This matches Daniel's manual pattern observed in his rev15 reference file
# for Física: groups 1&2, 3&4, 5&6 do their first session together, each
# in a different lab (Ciencias Exp. I + II), then split for sessions 2-5.
#
# Other subjects keep one cell per group as before.
# =============================================================================
INTRO_SESSION_PAIRED_SUBJECTS = {
    'S1_Física',   # Daniel rev15 pattern: 10 paired cells in W4-W5 for Física
}


# =============================================================================
# COLOR PALETTE (subject -> hex color for cell background in Iteraciones)
# =============================================================================
SUBJECT_COLORS = {
    # One distinct, legible (black-text) colour per subject — diversified the way
    # Daniel colours his own grids. Keyed by the prefix-stripped subject name
    # (get_subject_fill strips S1_/S2_ before the lookup, so both builders match).
    'Automatic Control':               'E09882',
    'Automatización Industrial':       'F2B16D',
    'Control de Máquinas':             'F2DC8C',
    'Electrotecnia':                   'D9E064',
    'Electrónica y Automática':        'D0F28C',
    'Estructuras':                     'A1F26D',
    'Física':                          '8DE082',
    'Física II':                       '6DF280',
    'Informática y Com. Industriales': '8CF2B7',
    'Ingeniería de Control':           '64E0BA',
    'Mecanismos':                      '8CF2EE',
    'Mecánica de Fluidos':             '6DD2F2',
    'Modelado de Sistemas':            '82B0E0',
    'Métodos Numéricos':               '6D89F2',
    'Química':                         '928CF2',
    'Regulación Automática':           '8D64E0',
    'Resistencia de Materiales':       'C98CF2',
    'Robótica y Automatización':       'E16DF2',
    'Tecnología Electrónica':          'E082D2',
    'Tecnología Medio Ambiente':       'F26DBA',
    'Tecnologías de Fabricación':      'F28CAB',
    'Termodinámica':                   'E06469',
}


# =============================================================================
# OPENPYXL STYLE OBJECTS
# =============================================================================

# Border styles
THIN_BORDER_SIDE = Side(style='thin', color='000000')
FULL_THIN_BORDER = Border(
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
)

# Background fills
HEADER_BLUE_FILL  = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
HOLIDAY_PINK_FILL = PatternFill(start_color='EA7A8A', end_color='EA7A8A', fill_type='solid')

# Per-holiday colours, matching the way Daniel colours his own grids:
#   cyan  = institutional / Loyola days (Blue Day, San Ignacio)
#   red   = national / official festivos (Semana Santa, Feria, Trabajador, Hispanidad)
#   amber = internal events (Acto apertura)
HOLIDAY_COLORS = {
    'Blue Day':                      '00B0F0',
    'Canonización de San Ignacio':   '00B0F0',
    'Semana Santa':                  'DB5741',
    'Feria de Abril':                'DB5741',
    'Día del Trabajador':            'DB5741',
    'Día de la Hispanidad':          'DB5741',
    'Acto apertura':                 'FFC000',
}

def holiday_fill(name):
    """PatternFill for a holiday cell, colour-coded by holiday (falls back to pink)."""
    hexv = HOLIDAY_COLORS.get(str(name).strip(), 'EA7A8A')
    return PatternFill(start_color=hexv, end_color=hexv, fill_type='solid')


def _merge_holiday_blocks(worksheet, first_row, n_blocks, panel_day_cols,
                          sem_holidays, week_number):
    """
    Turn each holiday of the week into ONE merged vertical block per day column
    (Daniel-style: e.g. Blue Day = F57:F64), instead of repeating the name in
    every time-block row. Call once per week, AFTER the time-block rows are
    written (their day cells already painted empty with the holiday colour).

    panel_day_cols: list of day-column arrays, one per panel/program/subject.
    """
    last_row = first_row + n_blocks - 1
    centred = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for day_cols in panel_day_cols:
        for d in range(5):
            name = sem_holidays.get((week_number, d))
            if not name:
                continue
            col = day_cols[d]
            try:
                worksheet.merge_cells(start_row=first_row, start_column=col,
                                      end_row=last_row, end_column=col)
            except Exception:
                pass
            top = worksheet.cell(row=first_row, column=col)
            top.value = name
            top.font = COURSE_FONT
            top.fill = holiday_fill(name)
            top.alignment = centred
# Reserved slot (e.g. Biotecnología): a distinct PURPLE, intentionally different
# from every lab-subject colour (greens/blues/gold) and from the holiday pink, so
# a reserved room reads at a glance as "neither a normal lab nor a calendar holiday".
RESERVED_PURPLE_FILL = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
HEADER_GRAY_FILL  = PatternFill(start_color='9AABC5', end_color='9AABC5', fill_type='solid')

# Fonts
WHITE_FONT          = Font(name='Calibri', size=8,  color='FFFFFF')
PROGRAM_FONT        = Font(name='Calibri', size=10, bold=True)
WEEK_LABEL_FONT     = Font(name='Calibri', size=11, bold=True)
TIME_LABEL_FONT     = Font(name='Calibri', size=9)
COURSE_FONT         = Font(name='Calibri', size=8)
LAB_SESSION_FONT    = Font(name='Calibri', size=8, bold=True)
GROUP_HEADER_FONT   = Font(name='Calibri', size=18, bold=True)
TABLE_HEADER_FONT   = Font(name='Calibri', size=10, bold=True)
STUDENT_FONT        = Font(name='Calibri', size=9)

# Alignments
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_ALIGNMENT  = Alignment(horizontal='right',  vertical='center', wrap_text=True)
LEFT_ALIGNMENT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
WRAP_TOP_ALIGNMENT = Alignment(wrap_text=True, vertical='top')


# =============================================================================
# UTILITY HELPERS
# =============================================================================

def strip_semester_prefix(subject_name):
    """
    Strip 'S1_' or 'S2_' prefix from a subject name.

    Example:
        'S1_Física' -> 'Física'
    """
    return subject_name.replace('S1_', '').replace('S2_', '')


def get_week_dates(week_number, semester=1):
    """
    Return the 5 weekday dates for a given week of the given semester.

    Args:
        week_number: 1-based week index within the semester
        semester: 1 or 2 — selects the calendar anchor (S1 starts Sep 2025,
                  S2 starts 02 Feb 2026). Using the wrong anchor would print S1
                  dates on S2 sheets, so callers pass the sheet's semester.

    Returns:
        list[datetime]: [Monday, Tuesday, Wednesday, Thursday, Friday]
    """
    anchor = S2_FIRST_DAY if int(semester) == 2 else S1_FIRST_DAY
    monday = anchor + timedelta(weeks=week_number - 1)
    # Roll forward until we hit Monday (in case the anchor isn't a Monday)
    while monday.weekday() != 0:
        monday += timedelta(days=1)
    return [monday + timedelta(days=offset) for offset in range(5)]


def get_subject_fill(subject_name):
    """
    Return a PatternFill colored according to the subject's palette entry.

    The subject name may arrive with or without the semester prefix (one builder
    passes the raw schedule name "S1_Física", the other the cleaned "Física"), so
    we strip S1_/S2_ before the lookup. Falls back to light gray for unknowns.
    """
    key = str(subject_name).replace('S1_', '').replace('S2_', '').strip()
    color_hex = SUBJECT_COLORS.get(key, 'D9E2F3')
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')


def display_lab_name(full_lab_name):
    """
    Convert a full lab name to its abbreviated display name.

    Example:
        'Ciencias Experimentales I' -> 'Ciencias Exp. I'
    """
    cleaned = (full_lab_name or '').strip()
    return LAB_DISPLAY_NAMES.get(cleaned, cleaned)


def extract_program_code(raw_program_value):
    """
    Extract a known program code from a raw cell value.

    Splits on common separators and looks for the first token that matches
    a known program in KNOWN_PROGRAM_CODES.

    Returns:
        str | None: Program code (uppercase) or None if not found.
    """
    candidate_text = str(raw_program_value).replace(',', ' ').replace('-', ' ').replace('.', ' ')
    for token in candidate_text.split():
        normalized = token.strip().upper()
        if normalized in KNOWN_PROGRAM_CODES:
            return normalized
    return None


def sanitize_cell(value):
    """Neutralise l'injection de formule Excel/CSV.

    Excel/LibreOffice interpretent comme une FORMULE toute cellule dont le
    contenu commence par = + - @ (ou une tabulation / un retour chariot). On
    prefixe ces valeurs d'une apostrophe : elles s'affichent en texte et rien
    ne s'execute. Les valeurs legitimes (noms, matieres, salles) ne commencent
    jamais par ces caracteres.
    """
    if not isinstance(value, str) or not value:
        return value
    first = value[0]
    if first in ('=', '+', '-', '@') or (first.isspace() and first != ' '):
        return "'" + value
    return value


def write_bordered_cell(worksheet, row, col, value, font=None, fill=None, alignment=None):
    """
    Write a value to a worksheet cell with FULL_THIN_BORDER.

    Daniel's format requires borders on every table cell. This helper
    centralizes that requirement so we never forget.

    Args:
        worksheet: openpyxl worksheet
        row: 1-based row number
        col: 1-based column number
        value: cell content
        font: optional Font (defaults to inherited)
        fill: optional PatternFill (defaults to no fill)
        alignment: optional Alignment (defaults to inherited)

    Returns:
        The created cell object (for further customization if needed).
    """
    cell = worksheet.cell(row=row, column=col, value=sanitize_cell(value))
    cell.border = FULL_THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell


def format_lab_session_label(sessions_at_slot, subject, professor_lookup=None):
    """
    Build the cell label for one or more lab sessions occurring at the same slot.

    The label now includes, for each session, the four pieces of information
    needed to verify the professor flow:
        Práctica N <subject>
        Grupo M
        Sala: <lab room>
        Prof.: <assigned / eligible professor(s)>

    Args:
        sessions_at_slot: pandas DataFrame rows (sessions at the same week+day+block)
        subject: full subject name (e.g., "S1_Física")
        professor_lookup: optional dict {subject_full_name: "Prof A; Prof B"} used
            to display the professor(s) responsible for the subject. The pipeline
            does not assign a single named professor per group (it only enforces
            that at least one eligible professor is free), so this shows the
            eligible professor(s) for the subject.

    Returns:
        str: the label text to display in the cell.
    """
    subject_clean = strip_semester_prefix(subject)

    if len(sessions_at_slot) == 0:
        return ''

    def _room(row):
        r = row.get('lab_rooms', '')
        return str(r).strip() if r is not None and str(r).strip() and str(r) != 'nan' else '—'

    def _prof_names():
        if not professor_lookup:
            return []
        raw = (professor_lookup.get(subject)
               or professor_lookup.get(subject_clean)
               or professor_lookup.get(_normalize_subject_key(subject))
               or professor_lookup.get(_normalize_subject_key(subject_clean)))
        if not raw:
            return []
        return [n.strip() for n in str(raw).split(';') if n.strip()]

    def _profs(rot=0):
        """Pick ONE eligible professor to show for this cell, ROTATING across
        sessions/groups so the same person isn't shown for every session, and
        append (+N) to flag the N other professors also eligible. The pipeline
        guarantees at least one eligible professor is free per session but does
        not assign a named one, so the rotation is indicative (any of the shown
        may run it); the final assignment is the coordinator's call."""
        names = _prof_names()
        if not names:
            return None
        if len(names) == 1:
            return names[0]
        i = int(rot) % len(names)
        return f"{names[i]} (+{len(names) - 1})"

    # Single session: standard display + room + professor
    if len(sessions_at_slot) == 1:
        first = sessions_at_slot.iloc[0]
        prof_line = _profs(int(first['session']) + int(first['grupo']))
        lines = [
            f"Práctica {int(first['session'])} {subject_clean}",
            f"Grupo {int(first['grupo'])}",
            f"Sala: {_room(first)}",
        ]
        if prof_line:
            lines.append(f"Prof.: {prof_line}")
        return "\n".join(lines)

    # Multiple sessions at same slot: check for paired intro pattern
    if subject in INTRO_SESSION_PAIRED_SUBJECTS and len(sessions_at_slot) == 2:
        groups = sorted(sessions_at_slot['grupo'].astype(int).tolist())
        sessions = sessions_at_slot['session'].astype(int).tolist()

        is_session_one = all(s == 1 for s in sessions)
        is_consecutive = (groups[1] - groups[0] == 1)
        is_odd_first = (groups[0] % 2 == 1)

        if is_session_one and is_consecutive and is_odd_first:
            first = sessions_at_slot.iloc[0]
            # Rotate on the group pair so paired-intro slots vary too.
            prof_line = _profs(groups[0] + groups[1] + 1)
            lines = [
                f"Práctica 1 {subject_clean}",
                f"Grupos {groups[0]} & {groups[1]}",
                f"Sala: {_room(first)}",
            ]
            if prof_line:
                lines.append(f"Prof.: {prof_line}")
            return "\n".join(lines)

    # Fallback: more than 1 session but not a recognized pattern
    first = sessions_at_slot.iloc[0]
    prof_line = _profs(int(first['session']) + int(first['grupo']))
    lines = [
        f"Práctica {int(first['session'])} {subject_clean}",
        f"Grupo {int(first['grupo'])}",
        f"Sala: {_room(first)}",
    ]
    if prof_line:
        lines.append(f"Prof.: {prof_line}")
    return "\n".join(lines)


# =============================================================================
# BUILD: PROGRAM TIMETABLE LOOKUP (used by Horarios + Iteraciones)
# =============================================================================

def build_program_timetable(master_df, programs, year_number):
    """
    Build a lookup of regular courses for the given programs and year.

    Returns:
        dict: (program_code, time_block, day_index) -> course_activity_label
    """
    if master_df is None:
        return {}

    # Restrict to Sevilla campus and the requested academic year
    sevilla_filter = master_df['campus'].fillna('').str.contains('Sevilla', case=False)
    filtered = master_df[sevilla_filter].copy()

    if 'curso_asignatura' in filtered.columns:
        filtered = filtered[filtered['curso_asignatura'] == year_number]

    program_column = 'programas' if 'programas' in filtered.columns else 'titulacionid'

    timetable = {}
    for _, row in filtered.iterrows():
        program_code = extract_program_code(row.get(program_column, ''))
        if program_code not in programs:
            continue

        activity = str(row.get('actividad', '')).strip()
        if not activity or activity == 'nan':
            continue

        weekday_name = str(row.get('slot_jour_semaine', '')).strip()
        start_minutes = row.get('slot_hora_inicio_min')
        if pd.isna(start_minutes):
            continue

        time_block = MINUTES_TO_TIME_BLOCK.get(int(start_minutes), '')
        if not time_block:
            continue

        day_index = DAY_NAME_TO_INDEX.get(weekday_name, -1)
        if day_index < 0:
            continue

        # Cap activity label at 25 chars (Daniel's format)
        timetable_key = (program_code, time_block, day_index)
        if timetable_key not in timetable:
            timetable[timetable_key] = activity[:25].upper()

    return timetable


# =============================================================================
# BUILD: HORARIOS SHEET
# =============================================================================

def build_horarios_sheet(workbook, program_timetable, programs):
    """
    Build the Horarios tab: 2 programs per row block, all time blocks shown.

    Layout:
        Each pair of programs occupies 7 columns (1 program label + 5 day cols + 1 gap)
        on rows: header row, then 6 time-block rows, then 1 gap row.
    """
    worksheet = workbook.create_sheet('Horarios')
    current_row = 1

    for pair_start_index in range(0, len(programs), 2):
        program_pair = programs[pair_start_index:pair_start_index + 2]

        # Column layout for first and second program in the pair
        column_layout_per_program = [
            (1, [2, 3, 4, 5, 6]),       # First program: label col 1, days cols 2-6
            (8, [9, 10, 11, 12, 13]),   # Second program: label col 8, days cols 9-13
        ]

        # ---- Header row: program name + day-of-week labels
        for index, (label_col, day_cols) in enumerate(column_layout_per_program):
            if index >= len(program_pair):
                break
            worksheet.cell(row=current_row, column=label_col, value=sanitize_cell(program_pair[index])).font = PROGRAM_FONT
            worksheet.cell(row=current_row, column=label_col).border = FULL_THIN_BORDER
            for day_index, day_name in enumerate(DAYS_OF_WEEK):
                write_bordered_cell(
                    worksheet, current_row, day_cols[day_index],
                    f'{day_name} ', WHITE_FONT, HEADER_BLUE_FILL, CENTER_ALIGNMENT,
                )
        current_row += 1

        # ---- Time-block rows (one row per time block)
        for time_block in TIME_BLOCKS:
            for index, (label_col, day_cols) in enumerate(column_layout_per_program):
                if index >= len(program_pair):
                    break
                # Time block label (right-aligned)
                write_bordered_cell(
                    worksheet, current_row, label_col,
                    time_block, TIME_LABEL_FONT, None, RIGHT_ALIGNMENT,
                )
                # Course content for each day
                for day_index in range(5):
                    course_label = program_timetable.get(
                        (program_pair[index], time_block, day_index), ''
                    )
                    write_bordered_cell(
                        worksheet, current_row, day_cols[day_index],
                        course_label or '', COURSE_FONT, alignment=CENTER_ALIGNMENT,
                    )
            current_row += 1

        # Gap row between program pairs
        current_row += 1

    # ---- Column widths
    for col in [1, 8]:
        worksheet.column_dimensions[get_column_letter(col)].width = 14
    for col in [2, 3, 4, 5, 6, 9, 10, 11, 12, 13]:
        worksheet.column_dimensions[get_column_letter(col)].width = 20

    return len(program_timetable)


# =============================================================================
# BUILD: GRUPO DE PRÁCTICAS SHEET (student groups)
# =============================================================================

def build_grupos_sheet(workbook, groups_df, subjects, naming_style, is_single_tab, name_map=None):
    """
    Build the "Grupo de prácticas" tab(s).

    Args:
        workbook: openpyxl Workbook
        groups_df: DataFrame with student-group composition
        subjects: list of S1_* subject names for this level
        naming_style: 'number' (Grupo 1, 2, 3...) or 'letter' (Grupo A, B, C...)
        is_single_tab: True = one combined tab; False = one tab per subject
    """
    if len(groups_df) == 0:
        return

    # Determine which column holds the student identifier
    student_id_column = (
        'student_name' if 'student_name' in groups_df.columns else 'student_hash'
    )

    if is_single_tab:
        # One combined tab using the primary subject of this level
        primary_subject = strip_semester_prefix(subjects[0])
        primary_groups = (
            groups_df[groups_df['subject'] == primary_subject]
            .drop_duplicates(subset=['grupo', student_id_column])
            .copy()
        )
        worksheet = workbook.create_sheet('Grupo de prácticas')
        _write_groups_to_sheet(worksheet, primary_groups, student_id_column, naming_style, name_map=name_map)
    else:
        # One tab per subject
        for subject in subjects:
            subject_clean = strip_semester_prefix(subject)
            subject_groups = (
                groups_df[groups_df['subject'] == subject_clean]
                .drop_duplicates(subset=['grupo', student_id_column])
                .copy()
            )
            if len(subject_groups) == 0:
                continue
            sheet_title = f'Grupo de prácticas {subject_clean}'[:31]   # Excel 31-char limit
            worksheet = workbook.create_sheet(sheet_title)
            _write_groups_to_sheet(worksheet, subject_groups, student_id_column, naming_style, name_map=name_map)


def _write_groups_to_sheet(worksheet, groups_df, student_id_column, naming_style, name_map=None):
    """
    Write groups to a worksheet, 3 groups per row block.

    Layout:
        Columns B-C, F-G, J-K each hold one group's student list.
        Each group has a title row, then header row, then student rows.
    """
    df_local = groups_df.copy()
    name_map = name_map or {}
    def _disp(v):
        return name_map.get(v, v)
    df_local['grupo'] = df_local['grupo'].astype(int)
    group_numbers = sorted(df_local['grupo'].unique())

    # Column pairs: (Alumno col, Titulación col) for each of 3 groups per row
    GROUP_COLUMN_PAIRS = [(2, 3), (6, 7), (10, 11)]

    current_row = 2

    # Iterate in batches of 3 groups
    for batch_start in range(0, len(group_numbers), 3):
        batch_group_numbers = group_numbers[batch_start:batch_start + 3]
        current_row += 1

        # ---- Group titles row
        current_row += 1
        for batch_index, group_num in enumerate(batch_group_numbers):
            if batch_index >= len(GROUP_COLUMN_PAIRS):
                break
            student_col, _ = GROUP_COLUMN_PAIRS[batch_index]
            if naming_style == 'number':
                title_text = f'Grupo {group_num} '
            else:
                title_text = f'Grupo {chr(64 + group_num)} '   # 65=A, 66=B, ...
            worksheet.cell(row=current_row, column=student_col, value=title_text).font = GROUP_HEADER_FONT
        current_row += 1

        # ---- Column headers (Alumno/a, Titulación)
        current_row += 1
        for batch_index in range(min(len(batch_group_numbers), len(GROUP_COLUMN_PAIRS))):
            student_col, program_col = GROUP_COLUMN_PAIRS[batch_index]
            write_bordered_cell(worksheet, current_row, student_col, 'Alumno/a',     TABLE_HEADER_FONT, HEADER_GRAY_FILL)
            write_bordered_cell(worksheet, current_row, program_col, 'Titulación ', TABLE_HEADER_FONT, HEADER_GRAY_FILL)

        # ---- Student rows
        max_students_in_batch = 0
        for batch_index, group_num in enumerate(batch_group_numbers):
            if batch_index >= len(GROUP_COLUMN_PAIRS):
                break
            student_col, program_col = GROUP_COLUMN_PAIRS[batch_index]
            group_data = df_local[df_local['grupo'] == group_num].sort_values(
                by=student_id_column,
                key=lambda col: col.map(lambda v: str(_disp(v))),
            )

            for student_index, (_, student_row) in enumerate(group_data.iterrows()):
                target_row = current_row + 1 + student_index
                write_bordered_cell(worksheet, target_row, student_col, _disp(student_row.get(student_id_column, '')), STUDENT_FONT)
                titulacion = student_row.get('titulacion', '')
                if pd.notna(titulacion) and str(titulacion) != 'nan':
                    write_bordered_cell(worksheet, target_row, program_col, str(titulacion), STUDENT_FONT)
            max_students_in_batch = max(max_students_in_batch, len(group_data))

        current_row += max_students_in_batch + 1

    # ---- Column widths
    for student_col, program_col in GROUP_COLUMN_PAIRS:
        worksheet.column_dimensions[get_column_letter(student_col)].width = 35
        worksheet.column_dimensions[get_column_letter(program_col)].width = 14


# =============================================================================
# BUILD: ITERACIONES SHEET (week-by-week schedule)
# =============================================================================

def build_iteraciones_sheet(workbook, schedule_df, program_timetable, programs, subjects):
    """
    Build the Iteraciones tab: week-by-week schedule with labs overlaid on
    regular courses.

    Layout per week:
        Row 1: SEMANA N header
        Row 2: Program label + day-of-week labels
        Row 3: Empty + dates (one per day)
        Rows 4-9: Time blocks (one row per block)
    Then 1 gap row before next week.
    """
    worksheet = workbook.create_sheet('Iteraciones')

    # Blocked slots (room reserved for another activity) for this semester,
    # keyed by (week, day_idx, time_block_label). Read from blocked_slots.csv;
    # NOT part of the schedule, so the reliability conflict check is unaffected.
    try:
        _sem_it = int(str(schedule_df['semester'].iloc[0])) if len(schedule_df) else 1
    except Exception:
        _sem_it = 1
    _blocked_iter = _load_blocked_slots(_sem_it)

    # ---- Index lab sessions by (week, day_index, time_block) for fast lookup
    lab_sessions_by_slot = defaultdict(list)
    for _, schedule_row in schedule_df.iterrows():
        if schedule_row['subject'] not in subjects:
            continue

        day_index = DAY_NAME_TO_INDEX.get(schedule_row['day'], 0)
        subject_clean = strip_semester_prefix(schedule_row['subject'])

        # Normalize lab room to short display name.
        # Use the LAB_DISPLAY_NAMES table (exact match) and fall back to a
        # light generic shortening. This replaces the previous chain of
        # .replace() calls, which had a bug: it turned the real "Mecánica de
        # Fluidos" lab (used by S2_Mecánica de Fluidos) into "Lab. Termodinámica".
        raw_room = str(schedule_row.get('lab_rooms', '')).split(',')[0].strip()
        if raw_room in LAB_DISPLAY_NAMES:
            room_display = LAB_DISPLAY_NAMES[raw_room]
        else:
            room_display = (
                raw_room
                .replace('Laboratorio de Ingeniería ', 'Lab. ')
                .replace('Ciencias Experimentales',     'Ciencias Exp.')
            )

        lab_sessions_by_slot[(int(schedule_row['week']), day_index, schedule_row['time_block'])].append({
            'subject':  subject_clean,
            'grupo':    int(schedule_row['grupo']),
            'session':  int(schedule_row['session']),
            'room':     room_display,
        })

    def get_program_columns(program_index):
        """Return (label_col, [day_cols]) for the given 0-based program index."""
        pair_index = program_index // 2
        side = program_index % 2
        base = pair_index * 14
        if side == 0:
            return (base + 1, [base + 2, base + 3, base + 4, base + 5, base + 6])
        else:
            return (base + 8, [base + 9, base + 10, base + 11, base + 12, base + 13])

    current_row = 1

    # Span the weeks ACTUALLY present in this schedule, not a fixed 14. S2 labs
    # run up to week ~20, so a hard cap of 14 silently dropped S2 sessions from
    # the grid. Fall back to NUM_WEEKS only if the data has no usable weeks.
    _wmin, _wmax = _schedule_week_span(schedule_df)
    for week_number in range(_wmin, _wmax + 1):
        week_dates = get_week_dates(week_number, _sem_it)

        # ---- Row: SEMANA N header (no border, just bold)
        for program_index in range(len(programs)):
            _, day_cols = get_program_columns(program_index)
            worksheet.cell(row=current_row, column=day_cols[0], value=f'SEMANA {week_number}').font = WEEK_LABEL_FONT
            for day_col in day_cols:
                worksheet.cell(row=current_row, column=day_col).border = Border(bottom=THIN_BORDER_SIDE)
        current_row += 1

        # ---- Row: program label + day-of-week labels
        for program_index, program_code in enumerate(programs):
            label_col, day_cols = get_program_columns(program_index)
            cell = worksheet.cell(row=current_row, column=label_col, value=program_code)
            cell.font = PROGRAM_FONT
            cell.border = Border(right=THIN_BORDER_SIDE)
            cell.alignment = LEFT_ALIGNMENT
            for day_index, day_name in enumerate(DAYS_OF_WEEK):
                write_bordered_cell(
                    worksheet, current_row, day_cols[day_index],
                    f'{day_name} ', WHITE_FONT, HEADER_BLUE_FILL, CENTER_ALIGNMENT,
                )
        current_row += 1

        # ---- Row: dates (centered, blue background, borders)
        for program_index in range(len(programs)):
            label_col, day_cols = get_program_columns(program_index)
            worksheet.cell(row=current_row, column=label_col).border = Border(
                bottom=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE
            )
            for day_index in range(5):
                cell = write_bordered_cell(
                    worksheet, current_row, day_cols[day_index],
                    week_dates[day_index], WHITE_FONT, HEADER_BLUE_FILL, CENTER_ALIGNMENT,
                )
                cell.number_format = 'YYYY-MM-DD'
        current_row += 1

        # ---- Time block rows (all 6 blocks, even if empty)
        for time_block in TIME_BLOCKS:
            for program_index, program_code in enumerate(programs):
                label_col, day_cols = get_program_columns(program_index)
                # Time block label (right-aligned, blue background)
                write_bordered_cell(
                    worksheet, current_row, label_col,
                    time_block, TIME_LABEL_FONT, HEADER_BLUE_FILL, RIGHT_ALIGNMENT,
                )
                for day_index in range(5):
                    # Holiday: paint the cell filled (no text here). The name is
                    # written once and merged into a vertical block per day column
                    # by _merge_holiday_blocks() after this loop (Daniel style).
                    holiday = SEMESTER_HOLIDAYS.get(_sem_it, {}).get((week_number, day_index))
                    if holiday:
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            '', COURSE_FONT, holiday_fill(holiday), CENTER_ALIGNMENT,
                        )
                        continue

                    # Blocked slot (room reserved for another activity) takes
                    # priority. Rendered in a DISTINCT purple (not the lab colours,
                    # not the holiday pink). Read from blocked_slots.csv (not part
                    # of the schedule, so reliability is unaffected).
                    _blk_lbl = _blocked_iter.get((week_number, day_index, time_block))
                    if _blk_lbl:
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            _blk_lbl, WHITE_FONT, RESERVED_PURPLE_FILL, CENTER_ALIGNMENT,
                        )
                        continue

                    # Lab session takes priority over regular course
                    lab_at_slot = lab_sessions_by_slot.get((week_number, day_index, time_block), [])
                    if lab_at_slot:
                        first_lab = lab_at_slot[0]
                        cell_text = (
                            f"Práctica {first_lab['session']} {first_lab['subject']}\n"
                            f"Grupo {first_lab['grupo']}\n"
                            f"{first_lab['room']}"
                        )
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            cell_text, LAB_SESSION_FONT, get_subject_fill(first_lab['subject']),
                            WRAP_TOP_ALIGNMENT,
                        )
                    else:
                        # Regular course (if any)
                        course_label = program_timetable.get(
                            (program_code, time_block, day_index), ''
                        )
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            course_label, COURSE_FONT, alignment=CENTER_ALIGNMENT,
                        )
            current_row += 1

        # ---- Holidays: one merged vertical block per day column (Daniel style)
        _merge_holiday_blocks(
            worksheet, current_row - len(TIME_BLOCKS), len(TIME_BLOCKS),
            [get_program_columns(pi)[1] for pi in range(len(programs))],
            SEMESTER_HOLIDAYS.get(_sem_it, {}), week_number,
        )

        # Gap row between weeks
        current_row += 1

    # ---- Column widths
    if programs:
        max_col = get_program_columns(len(programs) - 1)[1][-1]
    else:
        max_col = 13
    for col in range(1, max_col + 1):
        # First and last column of each pair = narrower
        worksheet.column_dimensions[get_column_letter(col)].width = (
            14 if (col - 1) % 7 in [0, 6] else 18
        )


# =============================================================================
# BUILD: VISTA PROFESOR SHEET (subject-by-subject schedule)
# =============================================================================

def _normalize_subject_key(name):
    """Lowercase, strip the S1_/S2_ prefix and accents, collapse spaces — so
    'S1_Física', 'Física' and 'fisica' all map to a single matching key."""
    import unicodedata
    s = strip_semester_prefix(str(name)).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return ' '.join(s.split())


def _build_professor_lookup():
    """Return {subject_variant: 'Prof A; Prof B'} for the Vista profesor sheet.

    Priority:
      1) outputs/optimization/subject_professors.csv — subject -> professor NAMES,
         written by the pipeline from the master schedule. Authoritative source,
         lives in the per-user workspace (never bundled).
      2) a professor-NAMES column inside subject_supervision.csv, if one exists
         (several plausible column names are tried). That file may instead carry
         only 'n_professors' (a COUNT) — in which case it yields no names, which
         is exactly why the professor line used to be missing.

    Each subject is registered under raw, prefix-stripped and accent/case
    normalized keys, so the cell formatter matches however the schedule spells it.
    """
    import os as _os
    import pandas as _pd
    lookup = {}

    def _register(subj, names):
        names = str(names).strip()
        if not str(subj).strip() or not names or names.lower() == 'nan':
            return
        for key in (str(subj), strip_semester_prefix(str(subj)), _normalize_subject_key(subj)):
            if key:
                lookup[key] = names

    # Search order (first existing wins):
    #   1) the re-pointed module constant (set by excel_export when frozen),
    #   2) data_clean/optimization/ — the AUTHORITATIVE file derived from the
    #      official enrolment report (informeDetalleGruposPorCurso) by
    #      build_subject_professors.py; shipped as reference data,
    #   3) outputs/optimization/ — a fallback the pipeline may write from the
    #      master schedule's keyword deduction.
    _candidates = []
    _g = globals().get('SUBJECT_PROFESSORS_PATH')
    if _g:
        _candidates.append(_g)
    # Correct locations first, then a known common TYPO ('optimizarion'), so a
    # misplaced file is still found instead of silently showing "N/A".
    _candidates += ['data_clean/optimization/subject_professors.csv',
                    'outputs/optimization/subject_professors.csv',
                    'data_clean/optimizarion/subject_professors.csv',
                    'outputs/optimizarion/subject_professors.csv']
    # Resolve each via app_paths when frozen, keeping the raw relative as fallback.
    _resolved = []
    try:
        import app_paths as _ap_sp
        for _c in _candidates:
            _r = _ap_sp.resolve_existing(_c)
            if _r:
                _resolved.append(_r)
    except Exception:
        pass
    _candidates = _resolved + _candidates
    # Last resort: search the workspace tree for ANY subject_professors.csv
    # (covers misspelled/relocated folders we didn't anticipate).
    try:
        import app_paths as _ap_sp2
        _ws = getattr(_ap_sp2, 'WORKSPACE', None) or _ap_sp2.workspace_path()
        _ws = _os.path.dirname(_ws) if _os.path.splitext(str(_ws))[1] else str(_ws)
        for _root, _dirs, _files in _os.walk(_ws):
            if 'subject_professors.csv' in _files:
                _candidates.append(_os.path.join(_root, 'subject_professors.csv'))
    except Exception:
        pass
    for _cand in _candidates:
        try:
            if _cand and _os.path.exists(_cand):
                df = _pd.read_csv(_cand)
                if 'subject' in df.columns and 'professors' in df.columns:
                    for _, r in df.iterrows():
                        _register(r['subject'], r.get('professors', ''))
                    if lookup:
                        break
        except Exception:
            continue

    if not lookup:
        sup_path = globals().get('SUPERVISION_PATH',
                                 'data_clean/optimization/subject_supervision.csv')
        try:
            if sup_path and _os.path.exists(sup_path):
                df = _pd.read_csv(sup_path)
                name_col = next((c for c in ('professors', 'profesores', 'professor_names',
                                             'docentes', 'profesorado')
                                 if c in df.columns), None)
                if 'subject' in df.columns and name_col:
                    for _, r in df.iterrows():
                        _register(r['subject'], r.get(name_col, ''))
        except Exception:
            pass

    return lookup


def build_vista_profesor_sheet(workbook, schedule_df, subjects, professor_lookup=None):
    """
    Build the Vista profesor tab: same layout as Iteraciones but
    grouped by subject instead of by program.

    professor_lookup: optional {subject: "Prof A; Prof B"} so each session cell
    can show the responsible professor(s). If None, it is loaded from
    subject_supervision.csv via the module-level SUPERVISION_PATH constant.
    """
    worksheet = workbook.create_sheet('Vista profesor')

    if len(schedule_df) == 0:
        return

    # Build professor lookup (subject -> eligible professor names) if not provided.
    if professor_lookup is None:
        professor_lookup = _build_professor_lookup()

    # Only include subjects that have at least one lab session
    lab_subjects = sorted(set(s for s in subjects if s in schedule_df['subject'].unique()))
    if not lab_subjects:
        return

    # Blocked slots (room reserved for another activity) for this semester,
    # grouped by subject -> {(week, day_idx, time_block_label)}. Rendered as
    # "Festivo / No disponible" in that subject's column band. Not part of the
    # schedule, so the reliability conflict check is unaffected.
    try:
        _sem_vp = int(str(schedule_df['semester'].iloc[0]))
    except Exception:
        _sem_vp = 1
    _blocked_all = _load_blocked_slots(_sem_vp)
    _blocked_by_subject = {}
    _bs_cands = []
    try:
        import app_paths as _ap2
        for _rel in ('outputs/optimization/blocked_slots.csv',
                     'data_clean/optimization/blocked_slots.csv'):
            _p = _ap2.resolve_existing(_rel)
            if _p:
                _bs_cands.append(_p)
    except Exception:
        pass
    _bs_cands += ['outputs/optimization/blocked_slots.csv',
                  'data_clean/optimization/blocked_slots.csv']
    for _cand in _bs_cands:
        try:
            import os as _os, pandas as _pd
            if _os.path.exists(_cand):
                _bdf = _pd.read_csv(_cand)
                for _, _r in _bdf.iterrows():
                    try:
                        if int(_r['semester']) != _sem_vp:
                            continue
                    except Exception:
                        pass
                    _subj = str(_r['subject'])
                    _blocked_by_subject.setdefault(_subj, {})[
                        (int(_r['week']), int(_r['day_idx']), str(_r['time_block']))
                    ] = str(_r.get('label', '') or 'Festivo / No disponible')
                if _blocked_by_subject:
                    break
        except Exception:
            continue

    def get_subject_columns(subject_index):
        """Return (label_col, [day_cols]) for the given 0-based subject index."""
        pair_index = subject_index // 2
        side = subject_index % 2
        base = pair_index * 14
        if side == 0:
            return (base + 1, [base + 2, base + 3, base + 4, base + 5, base + 6])
        else:
            return (base + 8, [base + 9, base + 10, base + 11, base + 12, base + 13])

    # ── Legend (explains the "Prof.: Nombre (+N)" notation) ─────────────────
    # Written at the top, above the week grid, so the meaning of (+N) and the
    # rotation is documented inside the deliverable itself.
    _legend_last_col = get_subject_columns(min(1, len(lab_subjects) - 1))[1][-1]
    try:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6, _legend_last_col))
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(6, _legend_last_col))
    except Exception:
        pass
    _lc = worksheet.cell(row=1, column=1, value="Leyenda — Profesores")
    _lc.font = PROGRAM_FONT
    _le = worksheet.cell(
        row=2, column=1,
        value=("«Prof.: Nombre (+N)» = profesor indicado para la sesión y N profesor(es) "
               "adicional(es) habilitado(s) para la asignatura. El nombre ROTA entre sesiones/grupos: "
               "todos los mostrados pueden impartirla; la asignación final la decide el coordinador. "
               "La lista completa está en subject_professors.csv."),
    )
    _le.alignment = WRAP_TOP_ALIGNMENT
    worksheet.row_dimensions[2].height = 44

    current_row = 4  # leave the legend (rows 1-2) + a blank row 3 above the grid

    _wmin, _wmax = _schedule_week_span(schedule_df)
    for week_number in range(_wmin, _wmax + 1):
        week_dates = get_week_dates(week_number, _sem_vp)

        # ---- SEMANA N header
        for subject_index in range(len(lab_subjects)):
            _, day_cols = get_subject_columns(subject_index)
            worksheet.cell(row=current_row, column=day_cols[0], value=f'SEMANA {week_number}').font = WEEK_LABEL_FONT
            for day_col in day_cols:
                worksheet.cell(row=current_row, column=day_col).border = Border(bottom=THIN_BORDER_SIDE)
        current_row += 1

        # ---- Subject label + day-of-week labels
        for subject_index, subject in enumerate(lab_subjects):
            label_col, day_cols = get_subject_columns(subject_index)
            cell = worksheet.cell(row=current_row, column=label_col, value=sanitize_cell(strip_semester_prefix(subject)))
            cell.font = PROGRAM_FONT
            cell.border = Border(right=THIN_BORDER_SIDE)
            cell.alignment = LEFT_ALIGNMENT
            for day_index, day_name in enumerate(DAYS_OF_WEEK):
                write_bordered_cell(
                    worksheet, current_row, day_cols[day_index],
                    f'{day_name} ', WHITE_FONT, HEADER_BLUE_FILL, CENTER_ALIGNMENT,
                )
        current_row += 1

        # ---- Dates row
        for subject_index in range(len(lab_subjects)):
            label_col, day_cols = get_subject_columns(subject_index)
            worksheet.cell(row=current_row, column=label_col).border = Border(
                bottom=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE
            )
            for day_index in range(5):
                cell = write_bordered_cell(
                    worksheet, current_row, day_cols[day_index],
                    week_dates[day_index], WHITE_FONT, HEADER_BLUE_FILL, CENTER_ALIGNMENT,
                )
                cell.number_format = 'YYYY-MM-DD'
        current_row += 1

        # ---- Time block rows
        for time_block in TIME_BLOCKS:
            # Taller rows so the 4-line cell (Práctica / Grupo / Sala / Prof.)
            # is fully visible without manual resizing, including longer
            # professor names.
            worksheet.row_dimensions[current_row].height = 74
            for subject_index, subject in enumerate(lab_subjects):
                label_col, day_cols = get_subject_columns(subject_index)
                # Time block label
                write_bordered_cell(
                    worksheet, current_row, label_col,
                    time_block, TIME_LABEL_FONT, HEADER_BLUE_FILL, RIGHT_ALIGNMENT,
                )

                # Filter sessions for this subject + week + time block
                subject_filter = (
                    (schedule_df['subject']    == subject)
                    & (schedule_df['week']       == week_number)
                    & (schedule_df['time_block'] == time_block)
                )
                sessions_at_slot = schedule_df[subject_filter]

                for day_index in range(5):
                    # Holiday: paint the cell filled (no text here). The name is
                    # written once and merged into a vertical block per day column
                    # by _merge_holiday_blocks() after this loop (Daniel style).
                    holiday = SEMESTER_HOLIDAYS.get(_sem_vp, {}).get((week_number, day_index))
                    if holiday:
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            '', COURSE_FONT, holiday_fill(holiday), CENTER_ALIGNMENT,
                        )
                        continue

                    # Blocked slot (room reserved for another activity) for THIS
                    # subject takes priority over an empty cell. Rendered in a
                    # DISTINCT purple (not a lab colour, not the holiday pink),
                    # with no professor / no práctica.
                    _blk_lbl = _blocked_by_subject.get(subject, {}).get(
                        (week_number, day_index, time_block))
                    if _blk_lbl:
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            _blk_lbl, WHITE_FONT, RESERVED_PURPLE_FILL, CENTER_ALIGNMENT,
                        )
                        continue

                    # Look for sessions matching this day
                    sessions_today = sessions_at_slot[sessions_at_slot['day'] == DAYS_OF_WEEK[day_index]]
                    if len(sessions_today) > 0:
                        subject_clean = strip_semester_prefix(subject)
                        # format_lab_session_label handles single OR paired-intro display
                        cell_text = format_lab_session_label(sessions_today, subject, professor_lookup)
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            cell_text, LAB_SESSION_FONT, get_subject_fill(subject_clean),
                            WRAP_TOP_ALIGNMENT,
                        )
                    else:
                        write_bordered_cell(
                            worksheet, current_row, day_cols[day_index],
                            '', alignment=CENTER_ALIGNMENT,
                        )
            current_row += 1

        # ---- Holidays: one merged vertical block per day column (Daniel style)
        _merge_holiday_blocks(
            worksheet, current_row - len(TIME_BLOCKS), len(TIME_BLOCKS),
            [get_subject_columns(si)[1] for si in range(len(lab_subjects))],
            SEMESTER_HOLIDAYS.get(_sem_vp, {}), week_number,
        )

        # Gap row between weeks
        current_row += 1

    # ---- Column widths
    if lab_subjects:
        max_col = get_subject_columns(len(lab_subjects) - 1)[1][-1]
    else:
        max_col = 13
    for col in range(1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = (
            16 if (col - 1) % 7 in [0, 6] else 23
        )


# =============================================================================
# VUE PROFESSEUR (consolidée) — Partie 1
# =============================================================================
#
# Cette section ajoute une NOUVELLE feuille « Vue Professeur » à chaque classeur
# de niveau (Primero / Segundo / Tercero, S1 et S2). Contrairement à la feuille
# « Vista profesor » existante (une GRILLE horaire par MATIÈRE), cette feuille est
# CENTRÉE SUR LE PROFESSEUR : une ligne par (professeur, matière) avec :
#   • le nom du professeur,
#   • la/les matière(s) et le(s) groupe(s) qu'il encadre,
#   • les crédits assignés par matière (lus dans « Asignación docente »),
#   • le nombre de séances de labo planifiées par matière,
#   • l'horaire détaillé des séances (jours / heures / salle).
#
# Tout est en FRANÇAIS (demande explicite). La convention 1 crédit P = 5 séances
# est rappelée dans l'en-tête de la feuille.
# =============================================================================

CREDIT_TO_SESSIONS = 5  # 1 crédit de laboratoire (P) = 5 séances (convention coordinateur)

# En-têtes et styles dédiés à la Vue Professeur (palette sobre, lisible)
VP_HEADER_FILL  = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
VP_HEADER_FONT  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
VP_TITLE_FONT   = Font(name='Calibri', size=13, bold=True, color='1F4E78')
VP_NOTE_FONT    = Font(name='Calibri', size=9, italic=True, color='595959')
VP_PROF_FONT    = Font(name='Calibri', size=10, bold=True)
VP_CELL_FONT    = Font(name='Calibri', size=9)
VP_OK_FILL      = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
VP_WARN_FILL    = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
VP_BAND_FILL    = PatternFill(start_color='F2F6FC', end_color='F2F6FC', fill_type='solid')


def _find_asignacion_file():
    """Localise le classeur « Asignacion_*.xlsx » (source des crédits T/P).

    Cherche, dans l'ordre :
      1) une éventuelle constante de module ASIGNACION_PATH (re-pointée par
         excel_export, comme les autres chemins, si elle existe) ;
      2) les noms usuels via app_paths.resolve_existing ;
      3) une recherche récursive dans l'arborescence de travail.

    Renvoie un chemin existant ou None (la feuille affichera alors « N/D »).
    """
    common = [
        'Asignacion_2025-2026_v5.xlsx',
        'Asignacion_2025-2026_v5.xlsx'.replace('-', '_'),
        'data_clean/Asignacion_2025-2026_v5.xlsx',
        'data/Asignacion_2025-2026_v5.xlsx',
        '/home/ubuntu/Uploads/Asignacion_2025-2026_v5.xlsx',
        '/home/ubuntu/Shared/Uploads/Asignacion_2025-2026_v5.xlsx',
        '/home/ubuntu/lab_project/Asignacion_2025-2026_v5.xlsx',
    ]
    g = globals().get('ASIGNACION_PATH')
    if g and os.path.exists(g):
        return g
    # via app_paths
    try:
        import app_paths as _ap
        for rel in ([g] if g else []) + common:
            if not rel:
                continue
            r = _ap.resolve_existing(rel)
            if r and os.path.exists(r):
                return r
    except Exception:
        pass
    # chemins relatifs bruts
    for rel in common:
        if os.path.exists(rel):
            return rel
    # recherche récursive : tout fichier commençant par « Asignacion » et .xlsx
    search_roots = []
    try:
        import app_paths as _ap2
        ws = getattr(_ap2, 'WORKSPACE', None) or _ap2.workspace_path()
        ws = os.path.dirname(ws) if os.path.splitext(str(ws))[1] else str(ws)
        search_roots.append(ws)
    except Exception:
        pass
    search_roots.append(os.getcwd())
    seen = set()
    for root in search_roots:
        if not root or root in seen or not os.path.isdir(root):
            continue
        seen.add(root)
        for dirpath, _dirs, files in os.walk(root):
            for f in files:
                fl = f.lower()
                if fl.startswith('asignacion') and fl.endswith('.xlsx'):
                    return os.path.join(dirpath, f)
    return None


def _get_lab_config():
    """Renvoie le LAB_CONFIG du pipeline (mots-clés de mapping matière→Asignación).

    Importé paresseusement : dans l'application, pipeline.py est déjà chargé.
    En cas d'échec (ortools absent dans un contexte isolé), on renvoie {} et le
    mapping se rabat sur une correspondance par nom dépouillé du préfixe S1_/S2_.
    """
    try:
        import pipeline as _P
        return getattr(_P, 'LAB_CONFIG', {}) or {}
    except Exception:
        return {}


def _load_professor_lab_credits():
    """Charge les crédits de LABORATOIRE (caractère « P ») par professeur et matière.

    Renvoie un dict :
        { matière_Asignación : { prof_label : crédits_P (float) } }
    et un dict de mapping inverse pratique :
        { matière_Asignación : "Prof A; Prof B" }  (noms des encadrants P)

    En cas d'absence du fichier Asignación ou d'erreur, renvoie ({}, {}) : la
    feuille s'affichera alors avec « N/D » pour les crédits (jamais d'exception).
    """
    fp = _find_asignacion_file()
    if not fp:
        print('    [WARN] Teacher View: Asignación file not found — credits = N/A')
        return {}, {}
    try:
        import professor_credits as _pc
        assign = _pc.parse_assignment(fp)      # offering_id, subject, prof_code, credits, char
        budgets = _pc.load_budgets(fp)         # prof_code -> prof_name
        code_to_name = dict(zip(budgets['prof_code'], budgets['prof_name']))
    except Exception as exc:
        print(f'    [WARN] Teacher View: cannot read Asignación ({exc}) — credits = N/A')
        return {}, {}

    def _label(code):
        name = code_to_name.get(code)
        if name and str(name).strip() and str(name).strip().lower() != str(code).strip().lower():
            return f"{name} ({code})"
        return str(code)

    labP = assign[assign['char'] == 'P']
    credits_by_subject = {}
    names_by_subject = {}
    for subj, grp in labP.groupby('subject'):
        per_prof = grp.groupby('prof_code')['credits'].sum()
        d = {}
        names = []
        for code, cr in per_prof.items():
            lab = _label(code)
            d[lab] = float(cr)
            names.append(lab)
        credits_by_subject[str(subj)] = d
        names_by_subject[str(subj)] = '; '.join(names)
    return credits_by_subject, names_by_subject


def _map_sched_to_asignacion(sched_subject, asignacion_subjects, lab_config):
    """Pour une matière planifiée (ex. « S1_Física »), renvoie la liste des
    matières de l'Asignación correspondantes, via les mots-clés de LAB_CONFIG.

    Repli : correspondance exacte sur le nom dépouillé du préfixe si aucun
    mot-clé ne correspond (ou si LAB_CONFIG est indisponible).
    """
    target = _normalize_subject_key(sched_subject)
    # index nom dépouillé -> clé LAB_CONFIG
    stripped_to_cfg = {}
    for k in lab_config:
        stripped_to_cfg[_normalize_subject_key(k)] = k
    cfg_key = stripped_to_cfg.get(target)
    matched = []
    if cfg_key is not None:
        cfg = lab_config[cfg_key]
        kws = [_normalize_subject_key(k) for k in cfg.get('keywords', [])]
        exc = [_normalize_subject_key(e) for e in cfg.get('keyword_exclude', [])]
        for asub in asignacion_subjects:
            na = _normalize_subject_key(asub)
            if any(k and k in na for k in kws) and not any(e and e in na for e in exc):
                matched.append(asub)
    if not matched:
        for asub in asignacion_subjects:
            if _normalize_subject_key(asub) == target:
                matched.append(asub)
    return matched


def _format_session_timetable(subject_sessions):
    """Construit l'horaire détaillé des séances d'une matière, par groupe.

    subject_sessions : DataFrame des lignes de planning pour UNE matière.
    Renvoie une chaîne lisible, p.ex. :
        « G1 : Lunes 12:30-14:30 (Ciencias Exp. I) — 12 séances ; G2 : Martes ... »

    On regroupe par (groupe, jour, créneau, salle) pour ne pas répéter chaque
    semaine, et on indique le nombre de séances de ce créneau récurrent.
    """
    if len(subject_sessions) == 0:
        return ''
    parts = []
    # tri par numéro de groupe
    try:
        groups = sorted(subject_sessions['grupo'].dropna().unique(),
                        key=lambda x: (float(x) if str(x).replace('.', '', 1).isdigit() else 1e9, str(x)))
    except Exception:
        groups = list(subject_sessions['grupo'].dropna().unique())
    for g in groups:
        gs = subject_sessions[subject_sessions['grupo'] == g]
        slot_parts = []
        # créneaux récurrents distincts pour ce groupe
        grouping_cols = [c for c in ('day', 'time_block', 'lab_rooms') if c in gs.columns]
        if grouping_cols:
            agg = gs.groupby(grouping_cols).size().reset_index(name='n')
            for _, row in agg.iterrows():
                day = str(row.get('day', '')).strip()
                tb = str(row.get('time_block', '')).strip()
                room = str(row.get('lab_rooms', '')).strip()
                room_disp = display_lab_name(room) if room and room.lower() != 'nan' else ''
                n = int(row['n'])
                seg = f"{day} {tb}".strip()
                if room_disp:
                    seg += f" ({room_disp})"
                seg += f" — {n} session{'s' if n > 1 else ''}"
                slot_parts.append(seg)
        try:
            g_disp = int(float(g))
        except Exception:
            g_disp = g
        parts.append(f"G{g_disp} : " + " / ".join(slot_parts))
    return " ; ".join(parts)


def build_vue_professeur_consolidada_sheet(workbook, schedule_df, subjects,
                                           credits_by_subject=None,
                                           names_by_subject=None):
    """Construit la feuille « Vue Professeur » (consolidée, centrée professeur).

    Une ligne par (professeur, matière) pour les matières de CE niveau, avec :
        Professeur | Matière | Crédits assignés (P) | Sessions attendues (créd×5)
        | Nº groupes | Sessions planifiées | Horaire des séances

    Arguments :
        workbook      : classeur openpyxl en cours de construction.
        schedule_df   : planning du niveau (déjà filtré sur ses matières).
        subjects      : liste des noms de matières du niveau (préfixés S1_/S2_).
        credits_by_subject / names_by_subject : pré-chargés une seule fois par
            excel_export (sinon chargés ici). Voir _load_professor_lab_credits().

    Robustesse : si l'Asignación est absente, les crédits affichent « N/D » et
    aucune exception n'est levée (la feuille reste générée).
    """
    worksheet = workbook.create_sheet('Teacher View')

    # ---- Per-group professor assignment (recommendation #1) -------------
    # We link every scheduled lab group to ONE responsible professor, derived
    # from the official P credits (1 P credit = 5 sessions). This replaces the
    # previous subject-level "shared co-supervision" display by a precise,
    # per-professor breakdown. Falls back gracefully to the legacy credit map
    # (or N/D) if the Asignación source / module is unavailable.
    sgmap = {}            # {(lpa_subject_key, group_int): prof_name}  actual schedule
    expected_map = {}     # {(lpa_subject_key, prof_name): sessions_expected}
    lpa_ok = False
    try:
        import lab_professor_assignment as _lpa
        _fp = _find_asignacion_file()
        if _fp:
            # subjects actually scheduled at this level -> their planned groups
            subject_to_groups = {}
            for _s in subjects:
                if _s not in schedule_df['subject'].unique():
                    continue
                _key = _lpa._norm(strip_semester_prefix(_s))
                _grps = []
                for _g in schedule_df[schedule_df['subject'] == _s]['grupo'].dropna().unique():
                    try:
                        _grps.append(int(float(_g)))
                    except Exception:
                        pass
                if _grps:
                    subject_to_groups[_key] = sorted(set(_grps))
            if subject_to_groups:
                sgmap = _lpa.assign_schedule_groups(_fp, subject_to_groups)
                _exp = _lpa.expected_sessions(_fp)
                for _, _r in _exp.iterrows():
                    expected_map[(_r['subject_clean'], _r['prof_name'])] = \
                        float(_r['sessions_expected'])
                lpa_ok = bool(sgmap)
    except Exception as _exc:
        print(f'    [WARN] Teacher View: per-group assignment unavailable ({_exc})')

    # Legacy credit map (subject-level) kept as a fallback only.
    if credits_by_subject is None or names_by_subject is None:
        credits_by_subject, names_by_subject = _load_professor_lab_credits()
    lab_config = _get_lab_config()
    asignacion_subjects = list(credits_by_subject.keys())

    # ---- Title + methodology note (English) ----
    title = worksheet.cell(row=1, column=1,
                           value='Teacher View — lab session breakdown by professor')
    title.font = VP_TITLE_FONT
    if lpa_ok:
        note_txt = ("Convention: 1 lab credit (P) = 5 sessions. Each lab group is "
                    "assigned to ONE responsible professor, allocated in proportion "
                    "to the official P credits (source: Asignación docente). "
                    "\"Expected sessions\" = P credits x 5; \"Planned sessions\" counts "
                    "the sessions actually scheduled for that professor's groups. "
                    "Volume gaps are flagged, never blocking.")
    else:
        note_txt = ("Convention: 1 lab credit (P) = 5 sessions. Sessions and schedule "
                    "are shown at SUBJECT level (the Asignación source was not found, "
                    "so per-professor allocation is unavailable). Credits \"N/A\" = "
                    "Asignación file not available.")
    note = worksheet.cell(row=2, column=1, value=note_txt)
    note.font = VP_NOTE_FONT

    if schedule_df is None or len(schedule_df) == 0:
        worksheet.cell(row=4, column=1,
                       value='No sessions scheduled for this level.').font = VP_CELL_FONT
        worksheet.column_dimensions['A'].width = 60
        return

    # ---- Table headers (English) ----
    headers = [
        'Professor', 'Subject', 'Lab credits (P)', 'Expected sessions (cred x5)',
        'Assigned groups', 'Planned sessions (professor)',
        'Schedule (day / time / room)',
    ]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        write_bordered_cell(worksheet, header_row, j, h,
                            VP_HEADER_FONT, VP_HEADER_FILL, CENTER_ALIGNMENT)

    # ---- Build rows ----
    # rows: (prof, subject, credits, expected, groups_str, planned, horaire, state)
    rows = []
    scheduled_present = sorted(set(s for s in subjects
                                   if s in schedule_df['subject'].unique()))
    for sched in scheduled_present:
        subj_sessions = schedule_df[schedule_df['subject'] == sched]
        planned_total = int(len(subj_sessions))
        n_groups = int(subj_sessions['grupo'].nunique()) if 'grupo' in subj_sessions.columns else 0
        horaire = _format_session_timetable(subj_sessions)
        matiere_disp = strip_semester_prefix(sched)
        lpa_key = None
        try:
            import lab_professor_assignment as _lpa2
            lpa_key = _lpa2._norm(matiere_disp)
        except Exception:
            lpa_key = None

        # group -> prof for this subject's actually scheduled groups
        prof_groups = defaultdict(list)
        if lpa_ok and lpa_key is not None:
            for _g in subj_sessions['grupo'].dropna().unique():
                try:
                    gi = int(float(_g))
                except Exception:
                    continue
                pname = sgmap.get((lpa_key, gi))
                if pname:
                    prof_groups[pname].append(gi)

        if prof_groups:
            # One row per responsible professor with their precise figures.
            for pname in sorted(prof_groups):
                grps = sorted(set(prof_groups[pname]))
                grp_str = ', '.join(f'G{g}' for g in grps)
                planned_prof = int(len(subj_sessions[subj_sessions['grupo'].apply(
                    lambda x: _safe_int(x) in set(grps))]))
                expected = expected_map.get((lpa_key, pname))
                if expected is None:
                    expected = planned_prof  # no source target -> neutral
                    credits = round(expected / CREDIT_TO_SESSIONS, 2)
                    state = ''
                else:
                    credits = round(expected / CREDIT_TO_SESSIONS, 2)
                    state = 'OK' if abs(planned_prof - expected) < 1e-6 else 'Gap'
                rows.append((pname, matiere_disp, credits, round(expected, 1),
                             grp_str, planned_prof, horaire, state))
        elif credits_by_subject:
            # Fallback: subject-level credit map (legacy display).
            asig_subs = _map_sched_to_asignacion(sched, asignacion_subjects, lab_config)
            prof_credits = defaultdict(float)
            for asub in asig_subs:
                for prof_label, cr in credits_by_subject.get(asub, {}).items():
                    prof_credits[prof_label] += cr
            subj_expected = sum(prof_credits.values()) * CREDIT_TO_SESSIONS
            if prof_credits:
                subj_state = 'OK' if abs(planned_total - subj_expected) < 1e-6 else 'Gap'
                for prof_label in sorted(prof_credits):
                    cr = prof_credits[prof_label]
                    rows.append((prof_label, matiere_disp, round(cr, 2),
                                 round(cr * CREDIT_TO_SESSIONS, 1),
                                 f'{n_groups} group(s)', planned_total, horaire,
                                 subj_state))
            else:
                rows.append(('- (no P credit assigned)', matiere_disp, 0.0, 0.0,
                             f'{n_groups} group(s)', planned_total, horaire,
                             'No P credit'))
        else:
            # Asignación unavailable: credits N/A
            rows.append(('N/A', matiere_disp, 'N/A', 'N/A',
                         f'{n_groups} group(s)', planned_total, horaire, ''))

    rows.sort(key=lambda r: (str(r[0]), str(r[1])))

    # ---- Write rows with per-professor banding ----
    current_row = header_row + 1
    prev_prof = None
    band = False
    for (prof_label, matiere_disp, cr, expected, grp_str, planned, horaire, state) in rows:
        if prof_label != prev_prof:
            band = not band
            prev_prof = prof_label
        row_fill = VP_BAND_FILL if band else None
        state_fill = None
        if state == 'OK':
            state_fill = VP_OK_FILL
        elif state in ('Gap', 'No P credit'):
            state_fill = VP_WARN_FILL

        values = [prof_label, matiere_disp, cr, expected, grp_str, planned, horaire]
        for j, val in enumerate(values, start=1):
            font = VP_PROF_FONT if j == 1 else VP_CELL_FONT
            align = WRAP_TOP_ALIGNMENT if j == 7 else (LEFT_ALIGNMENT if j <= 2 or j == 5 else CENTER_ALIGNMENT)
            write_bordered_cell(worksheet, current_row, j, val, font, row_fill, align)
        if state_fill:
            for j in (3, 4, 6):
                worksheet.cell(row=current_row, column=j).fill = state_fill
        worksheet.row_dimensions[current_row].height = 30
        current_row += 1

    # ---- Column widths + freeze headers ----
    widths = [30, 26, 14, 18, 22, 16, 64]
    for j, w in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(j)].width = w
    worksheet.freeze_panes = f'A{header_row + 1}'


def _safe_int(x):
    """Best-effort int conversion for group numbers (returns None on failure)."""
    try:
        return int(float(x))
    except Exception:
        return None


# =============================================================================
# MAIN
# =============================================================================