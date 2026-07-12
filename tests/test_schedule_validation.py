"""
Tests for schedule_validation and validation_sheet.

These are read-only, additive validation utilities. The tests exercise the
report structure, the hard/indicator classification, the proportional scoring,
and the Excel sheet builder — without touching any generation logic.
"""

import os
import sys

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import schedule_validation as sv  # noqa: E402
import validation_sheet as vs  # noqa: E402


@pytest.fixture(scope="module")
def report():
    return sv.validate_schedule()


def test_report_has_expected_keys(report):
    for key in ("status", "reliability_score", "checks", "counts",
                "teacher_load", "solver", "reliability_formula"):
        assert key in report


def test_status_is_valid(report):
    assert report["status"] in ("PASS", "WARN", "FAIL", "NO_DATA")


def test_score_in_range(report):
    assert 0.0 <= report["reliability_score"] <= 100.0


def test_hard_and_indicator_checks_present(report):
    checks = report["checks"]
    for hard in ("room_conflicts", "student_theory_conflicts",
                 "student_lab_double_booking"):
        assert hard in checks
        assert checks[hard].get("kind") == "hard"
    for indic in ("professor_double_booking", "professor_busy"):
        assert indic in checks
        assert checks[indic].get("kind") == "indicator"


def test_indicator_never_forces_fail(report):
    # If only indicator checks fail (no hard violation), status must not be FAIL.
    checks = report["checks"]
    hard_fail = any(
        not checks[n].get("passed", True)
        for n in ("room_conflicts", "student_theory_conflicts",
                  "student_lab_double_booking")
        if checks[n].get("checkable", True)
    )
    if not hard_fail:
        assert report["status"] != "FAIL"


def test_room_conflicts_have_totals(report):
    chk = report["checks"]["room_conflicts"]
    assert "affected" in chk and "total" in chk
    assert chk["affected"] <= chk["total"] or chk["total"] == 0


def test_build_validation_sheet_creates_tab(report):
    wb = Workbook()
    wb.remove(wb.active)
    ws = vs.build_validation_sheet(wb, report)
    assert ws.title == "Validation"
    assert ws["A1"].value == "Schedule validation and reliability"
    assert ws.max_row > 10


def test_build_validation_sheet_custom_title(report):
    wb = Workbook()
    wb.remove(wb.active)
    ws = vs.build_validation_sheet(wb, report, sheet_title="Fiabilidad")
    assert ws.title == "Fiabilidad"


def test_effective_busy_context_available():
    ctx = sv.build_effective_busy_context()
    # Context should either build successfully or degrade gracefully.
    assert "available" in ctx
    if ctx["available"]:
        assert 1 in ctx["student_busy_sem"]
        assert 2 in ctx["student_busy_sem"]



# ---------------------------------------------------------------------------
# Breakdown coherence — the per-slice figures must reconcile with the global
# totals for the ADDITIVE metrics, proving the Validation-sheet numbers are
# authentic. Distinct-student head-counts are deliberately non-additive.
# ---------------------------------------------------------------------------
_ADDITIVE = ("sessions", "groups", "subjects", "enrollments")


def test_breakdown_keys_present(report):
    for key in ("counts_by_semester", "counts_by_level",
                "counts_by_level_semester", "counts_by_titulacion",
                "counts_by_subject"):
        assert key in report, f"missing breakdown key: {key}"


def test_global_counts_have_enrollments(report):
    counts = report["counts"]
    assert "enrollments" in counts
    assert "avg_sessions_per_group" in counts
    # Enrollments >= distinct students (a student can register in many groups).
    if counts.get("students"):
        assert counts["enrollments"] >= counts["students"]


@pytest.mark.parametrize("bkey", ["counts_by_semester", "counts_by_level",
                                  "counts_by_level_semester"])
def test_additive_metrics_reconcile(report, bkey):
    slices = report.get(bkey) or []
    if not slices:
        pytest.skip(f"no data for {bkey}")
    counts = report["counts"]
    for metric in _ADDITIVE:
        gtotal = int(counts.get(metric, 0) or 0)
        if gtotal == 0:
            continue
        ssum = sum(int(s.get(metric, 0) or 0) for s in slices)
        assert ssum == gtotal, (
            f"{bkey}: {metric} sums to {ssum}, expected global {gtotal}")


def test_titulacion_students_partition_exactly(report):
    by_tit = report.get("counts_by_titulacion") or []
    if not by_tit:
        pytest.skip("no titulación data")
    g_students = int(report["counts"].get("students", 0) or 0)
    tsum = sum(int(t.get("students", 0) or 0) for t in by_tit)
    # Each student has exactly one degree -> distinct students partition.
    assert tsum == g_students
    # Enrollments across degrees also reconcile with the global total.
    g_enrol = int(report["counts"].get("enrollments", 0) or 0)
    esum = sum(int(t.get("enrollments", 0) or 0) for t in by_tit)
    assert esum == g_enrol


def test_by_subject_totals_reconcile(report):
    by_subj = report.get("counts_by_subject") or []
    if not by_subj:
        pytest.skip("no subject data")
    counts = report["counts"]
    assert sum(int(s.get("sessions", 0) or 0) for s in by_subj) == \
        int(counts.get("sessions", 0) or 0)
    assert sum(int(s.get("groups", 0) or 0) for s in by_subj) == \
        int(counts.get("groups", 0) or 0)
    assert sum(int(s.get("enrollments", 0) or 0) for s in by_subj) == \
        int(counts.get("enrollments", 0) or 0)
    # One row per subject, matching the global subject count.
    assert len(by_subj) == int(counts.get("subjects", 0) or 0)
