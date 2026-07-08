"""Tests for the Phase 3 export format & template manager (export_manager.py).

Covers format resolution, preference loading, the 'standard' backward-compat
delegation path, and the template engine (discovery, validation, placeholder
extraction and rendering).
"""

import os

import pytest
from openpyxl import Workbook

import export_manager as m
from excel_export_enhanced import ExportOptions


# ────────────────────────────────────────────────────────────
# Format presets & resolution
# ────────────────────────────────────────────────────────────
def test_all_formats_present():
    assert set(m.FORMAT_PRESETS) == {"standard", "summary", "enhanced", "detailed"}
    assert m.VALID_FORMATS == tuple(m.FORMAT_PRESETS)


def test_resolve_options_returns_export_options():
    o = m.resolve_options("enhanced")
    assert isinstance(o, ExportOptions)
    assert o.quality_metrics is True


def test_resolve_options_summary_disables_rooms():
    o = m.resolve_options("summary")
    assert o.room_utilization is False
    assert o.quality_metrics is True


def test_resolve_options_detailed_protects():
    o = m.resolve_options("detailed")
    assert o.protect_sheets is True


def test_resolve_options_overrides_win():
    o = m.resolve_options("enhanced", {"legend": False})
    assert o.legend is False


def test_resolve_options_invalid_format_raises():
    with pytest.raises(ValueError):
        m.resolve_options("nonsense")


def test_standard_preset_minimal():
    o = m.FORMAT_PRESETS["standard"]
    assert o.conditional_formatting is False
    assert o.legend is False


# ────────────────────────────────────────────────────────────
# Preferences
# ────────────────────────────────────────────────────────────
def test_load_preferences_has_defaults():
    prefs = m.load_preferences()
    assert prefs["default_format"] in m.VALID_FORMATS
    assert prefs["default_color_scheme"] in ("loyola", "default", "monochrome")
    assert "formats" in prefs


# ────────────────────────────────────────────────────────────
# export_with_format
# ────────────────────────────────────────────────────────────
def test_export_with_format_unknown_returns_error():
    res = m.export_with_format("bogus", semester=1)
    assert res["ok"] is False
    assert "Unknown format" in res["error"]


def test_export_with_format_enhanced_writes(tmp_path):
    out = tmp_path / "enh.xlsx"
    res = m.export_with_format("enhanced", semester=1, out_path=str(out),
                               color_scheme="loyola")
    assert res["ok"] is True
    assert res["format"] == "enhanced"
    assert out.exists()


def test_export_with_format_summary(tmp_path):
    out = tmp_path / "sum.xlsx"
    res = m.export_with_format("summary", semester=2, out_path=str(out))
    assert res["ok"] is True
    assert res["format"] == "summary"


def test_export_with_format_standard_delegates():
    # 'standard' routes to the legacy exporter; it should not raise and should
    # report the standard format tag regardless of data availability.
    res = m.export_with_format("standard", semester=1)
    assert res["format"] == "standard"
    assert "files" in res


# ────────────────────────────────────────────────────────────
# Template engine
# ────────────────────────────────────────────────────────────
def _template_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Loyola - {{TITLE}}"
    ws["A2"] = "Group {{GROUP}} / {{PROFESSOR}} / {{ROOM}}"
    ws["A3"] = "Mystery {{WHATEVER}}"
    path = tmp_path / "tpl.xlsx"
    wb.save(path)
    return path


def test_list_templates_returns_loyola():
    # The repo ships templates/loyola_schedule_template.xlsx.
    names = m.list_templates()
    assert any("loyola" in n.lower() for n in names)


def test_extract_placeholders(tmp_path):
    path = _template_workbook(tmp_path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    found = m.extract_placeholders(wb)
    assert {"TITLE", "GROUP", "PROFESSOR", "ROOM", "WHATEVER"} <= found


def test_validate_template_reports_unknown(tmp_path):
    path = _template_workbook(tmp_path)
    report = m.validate_template(str(path))
    assert report["ok"] is True
    assert "WHATEVER" in report["unknown"]
    assert "TITLE" in report["known"]


def test_validate_template_missing_file():
    report = m.validate_template("does_not_exist.xlsx")
    assert report["ok"] is False
    assert report["error"]


def test_load_template_missing_raises():
    with pytest.raises(FileNotFoundError):
        m.load_template("nope_not_here.xlsx")


def test_render_template_substitutes(tmp_path):
    path = _template_workbook(tmp_path)
    out = tmp_path / "rendered.xlsx"
    res = m.render_template(str(path), {"TITLE": "T", "GROUP": "G1",
                                        "PROFESSOR": "P", "ROOM": "R"},
                            out_path=str(out))
    assert res["ok"] is True
    assert set(res["substituted"]) >= {"TITLE", "GROUP", "PROFESSOR", "ROOM"}
    # WHATEVER has no context value -> reported missing, left intact.
    assert "WHATEVER" in res["missing"]
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.active["A1"].value == "Loyola - T"
    assert "{{WHATEVER}}" in wb.active["A3"].value


def test_render_template_defaults_date(tmp_path):
    wb = Workbook()
    wb.active["A1"] = "Date: {{DATE}}"
    path = tmp_path / "d.xlsx"
    wb.save(path)
    out = tmp_path / "d_out.xlsx"
    res = m.render_template(str(path), {}, out_path=str(out))
    assert res["ok"] is True
    from openpyxl import load_workbook
    val = load_workbook(out).active["A1"].value
    assert "{{DATE}}" not in val


def test_render_loyola_template(tmp_path):
    out = tmp_path / "loy.xlsx"
    res = m.render_template("loyola_schedule_template.xlsx",
                            {"TITLE": "Practicas", "SEMESTER": "1",
                             "SUBJECT": "Fisica", "GROUP": "G1",
                             "PROFESSOR": "Prof", "ROOM": "Lab",
                             "DAY": "Lunes", "BLOCK": "12:30-14:30"},
                            out_path=str(out))
    assert res["ok"] is True
    assert out.exists()
    assert not res["missing"]  # loyola template placeholders all supplied


def test_safe_name_sanitises():
    import excel_export_enhanced as e
    assert e._safe_name("Room Utilization") == "Room_Utilization"
    assert e._safe_name("1abc")[0] == "_"
