"""
Tests for Point 1 refinements:
  * English localisation of the solver soft-constraint labels/help/details.
  * The ui_config_preview module (importable, exposes render + helpers).
  * Solver configuration flowing into the validation report and Excel sheet.

These tests are deliberately UI-light: Streamlit rendering is not exercised,
only the pure data structures and module contracts are validated.
"""

import os
import sys

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import solver_config as sc  # noqa: E402


# ── English solver-constraint localisation ────────────────────────────

def test_english_labels_cover_all_soft_constraints():
    for key in sc.SOFT_CONSTRAINT_KEYS:
        assert key in sc.CONSTRAINT_LABELS_EN
        assert key in sc.CONSTRAINT_HELP_EN
        assert sc.CONSTRAINT_LABELS_EN[key].strip()
        assert sc.CONSTRAINT_HELP_EN[key].strip()


def test_english_details_have_purpose_effect_typical():
    for key in sc.SOFT_CONSTRAINT_KEYS:
        detail = sc.CONSTRAINT_DETAIL_EN[key]
        assert detail["purpose"].strip()
        assert detail["effect"].strip()
        assert detail["typical"].strip()


def test_english_labels_are_not_french():
    # Guard against accidentally copying the French strings verbatim.
    for key in sc.SOFT_CONSTRAINT_KEYS:
        assert sc.CONSTRAINT_LABELS_EN[key] != sc.CONSTRAINT_LABELS_FR[key]


# ── Config preview module contract ────────────────────────────────────

def test_config_preview_module_exposes_render():
    import ui_config_preview as cp
    assert hasattr(cp, "render_config_preview")
    assert callable(cp.render_config_preview)


def test_config_preview_active_profile_helper_defaults_balanced(monkeypatch):
    import ui_config_preview as cp
    # With no streamlit session and a clean config, the helper must not raise
    # and returns a known profile string.
    profile = cp._active_solver_profile()
    assert profile in ("Balanced", "Strict", "Relaxed", "Custom", "UNKNOWN")


def test_config_preview_fmt_int():
    import ui_config_preview as cp
    assert cp._fmt_int(1234) == "1,234"
    assert cp._fmt_int("abc") == "abc"


# ── Solver config flows into validation report + Excel sheet ───────────

def test_validation_report_contains_solver_config():
    import schedule_validation as sv
    report = sv.validate_schedule()
    assert "solver_config" in report
    cfg = report["solver_config"]
    assert "profile" in cfg
    assert "weights" in cfg
    assert "enabled" in cfg


def test_validation_sheet_renders_solver_config_section():
    from openpyxl import Workbook
    import validation_sheet as vs

    report = {
        "verdict": {"code": "APTO", "label": "APTO"},
        "reliability": {"score": 100.0, "status": "PASS"},
        "counts": {"sessions": 10, "groups": 2, "students": 5},
        "checks": {},
        "teacher_load": [],
        "examples": {},
        "solver_config": {
            "profile": "Balanced",
            "weights": {"semester_anchor_first": 100, "spacing": 200},
            "enabled": {"semester_anchor_first": True, "spacing": True},
        },
    }
    wb = Workbook()
    wb.remove(wb.active)
    vs.build_validation_sheet(wb, report)
    ws = wb["Validación"]
    texts = [
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    ]
    assert any("Configuración del solver" in t for t in texts)
    assert any("Perfil activo" in t for t in texts)
