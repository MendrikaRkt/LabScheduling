"""Tests for the Phase 3 enhanced Excel export engine (excel_export_enhanced.py).

Tests build workbooks from small synthetic DataFrames so they never depend on
real optimisation artifacts, then assert on the resulting openpyxl objects
(sheets, colours, conditional formatting, charts, comments, freeze panes,
auto-filters). Colour-blind-safe palette invariants are checked explicitly.
"""

import pandas as pd
import pytest
from openpyxl import Workbook

import excel_export_enhanced as e
from excel_export_enhanced import ExportOptions


# ────────────────────────────────────────────────────────────
# Fixtures / helpers
# ────────────────────────────────────────────────────────────
def _schedule():
    """Synthetic schedule with one optimal, one over, one under group."""
    rows = []
    # Optimal group: 12 students, 5 sessions.
    for s in range(1, 6):
        rows.append(dict(semester=1, subject="Fisica", program="GITI", curso_num=1,
                         grupo=1, session=s, week=s, day="Lunes",
                         time_block="12:30-14:30", nb_students=12,
                         lab_rooms="Ciencias Exp. I", professor="Prof A"))
    # Over-subscribed group: 20 students.
    for s in range(1, 6):
        rows.append(dict(semester=1, subject="Fisica", program="GITI", curso_num=1,
                         grupo=2, session=s, week=s, day="Martes",
                         time_block="10:30-12:30", nb_students=20,
                         lab_rooms="Ciencias Exp. II", professor="Prof B"))
    # Under-utilised group: 4 students.
    for s in range(1, 4):
        rows.append(dict(semester=1, subject="Quimica", program="IMR", curso_num=1,
                         grupo=1, session=s, week=s, day="Viernes",
                         time_block="08:30-10:30", nb_students=4,
                         lab_rooms="Lab. Electronica", professor="Prof C"))
    return pd.DataFrame(rows)


def _kpi():
    return {
        "groups": {"total": 3, "overflow": 1, "size_min": 4, "size_max": 20,
                   "size_mean": 12.0},
        "placement": {"enrolled": 100, "placed": 98, "unplaced": 2,
                      "placement_pct": 98.0},
        "total_sessions": 13,
    }


def _solver_stats():
    return [
        {"semester": 1, "label": "S1", "status": "OPTIMAL", "n_sessions": 13,
         "wall_time_s": 1.2, "objective": 100.0, "best_bound": 100.0, "gap": 0.0},
    ]


# ────────────────────────────────────────────────────────────
# Colour helpers
# ────────────────────────────────────────────────────────────
def test_group_color_stable_for_index():
    assert e.group_color(0) == e.group_color(0)
    assert e.group_color(0) != e.group_color(1)


def test_group_color_extends_beyond_base_palette():
    n = len(e._BASE_GROUP_PALETTE)
    c1 = e.group_color(n + 3)
    c2 = e.group_color(n + 3)
    assert c1 == c2                    # deterministic
    assert c1.startswith("FF") and len(c1) == 8


def test_build_group_color_map_consistent_and_unique():
    keys = [(1, "A", 1), (1, "A", 2), (1, "B", 1)]
    cmap = e.build_group_color_map(keys)
    assert set(cmap.keys()) == set(keys)
    assert len(set(cmap.values())) == 3


def test_group_status_thresholds():
    assert e.group_status(e.MAX_GROUP_SIZE + 1) == "over"
    assert e.group_status(e.MIN_GROUP_SIZE - 1) == "under"
    assert e.group_status(e.PREFERRED_GROUP_SIZE) == "optimal"
    assert e.group_status(e.MIN_GROUP_SIZE) == "optimal"
    assert e.group_status(e.MAX_GROUP_SIZE) == "optimal"


def test_status_palette_is_colorblind_safe_distinct():
    # Okabe-Ito derived hues must be distinct from each other.
    cols = {e.STATUS_OK, e.STATUS_WARN, e.STATUS_CRIT}
    assert len(cols) == 3


def test_readable_fg_contrast():
    assert e._readable_fg(e.LOYOLA_NAVY) == e.WHITE
    assert e._readable_fg(e.LOYOLA_GOLD) == e.BLACK


def test_heat_argb_monotonic_endpoints():
    light = e._heat_argb(0.0)
    dark = e._heat_argb(1.0)
    assert light != dark
    assert dark == e.LOYOLA_NAVY


# ────────────────────────────────────────────────────────────
# ExportOptions
# ────────────────────────────────────────────────────────────
def test_export_options_defaults_all_true():
    o = ExportOptions()
    assert o.color_coded_groups and o.quality_metrics and o.legend


def test_export_options_from_dict_ignores_unknown():
    o = ExportOptions.from_dict({"legend": False, "bogus": 123})
    assert o.legend is False
    assert not hasattr(o, "bogus")


# ────────────────────────────────────────────────────────────
# Group sheet
# ────────────────────────────────────────────────────────────
def test_groups_sheet_rows_and_header():
    wb = Workbook(); wb.remove(wb.active)
    e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Groups"]
    assert ws.cell(row=1, column=1).value == "Semester"
    # 3 groups -> 3 data rows.
    assert ws.max_row == 4


def test_groups_sheet_group_cell_colored():
    wb = Workbook(); wb.remove(wb.active)
    cmap = e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Groups"]
    gcell = ws.cell(row=2, column=4)
    assert gcell.fill.fill_type == "solid"
    assert len(cmap) == 3


def test_groups_sheet_conditional_formatting_present():
    wb = Workbook(); wb.remove(wb.active)
    e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Groups"]
    assert len(list(ws.conditional_formatting)) >= 1


def test_groups_sheet_freeze_and_filter():
    wb = Workbook(); wb.remove(wb.active)
    e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Groups"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_groups_sheet_comments_added():
    wb = Workbook(); wb.remove(wb.active)
    e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Groups"]
    # professor comment on column 11, room comment on column 10
    assert ws.cell(row=2, column=11).comment is not None


def test_groups_sheet_respects_disabled_formatting():
    opts = ExportOptions(conditional_formatting=False, freeze_panes=False,
                         auto_filter=False, cell_comments=False)
    wb = Workbook(); wb.remove(wb.active)
    e.build_color_coded_groups_sheet(wb, _schedule(), options=opts)
    ws = wb["Groups"]
    assert ws.freeze_panes is None
    assert ws.auto_filter.ref is None
    assert len(list(ws.conditional_formatting)) == 0
    assert ws.cell(row=2, column=11).comment is None


# ────────────────────────────────────────────────────────────
# Legend
# ────────────────────────────────────────────────────────────
def test_legend_sheet_lists_all_groups():
    wb = Workbook(); wb.remove(wb.active)
    cmap = e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    e.build_legend_sheet(wb, _schedule(), cmap)
    ws = wb["Legend"]
    labels = [ws.cell(row=r, column=1).value for r in range(2, 5)]
    assert all(lbl and lbl.startswith("S1") for lbl in labels)


# ────────────────────────────────────────────────────────────
# Analysis sheets
# ────────────────────────────────────────────────────────────
def test_room_utilization_sheet_has_chart_and_rules():
    wb = Workbook(); wb.remove(wb.active)
    e.build_room_utilization_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Room Utilization"]
    assert ws.max_row >= 2
    assert len(ws._charts) == 1
    assert len(list(ws.conditional_formatting)) >= 1


def test_professor_workload_sheet_aggregates():
    wb = Workbook(); wb.remove(wb.active)
    e.build_professor_workload_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Professor Workload"]
    profs = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert {"Prof A", "Prof B", "Prof C"} <= profs
    assert len(ws._charts) == 1


def test_student_placement_sheet_status():
    wb = Workbook(); wb.remove(wb.active)
    e.build_student_placement_sheet(wb, _schedule(), _kpi(), options=ExportOptions())
    ws = wb["Student Placement"]
    # header at row 7; data starts row 8.
    statuses = {ws.cell(row=r, column=7).value for r in range(8, ws.max_row + 1)}
    assert "Over-subscribed" in statuses or "Under-utilized" in statuses


def test_time_slot_analysis_heatmap_built():
    wb = Workbook(); wb.remove(wb.active)
    e.build_time_slot_analysis_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Time Slot Analysis"]
    assert ws.cell(row=2, column=1).value == "Time Block"


def test_quality_metrics_sheet_values():
    wb = Workbook(); wb.remove(wb.active)
    e.build_quality_metrics_sheet(wb, _kpi(), _solver_stats(), options=ExportOptions())
    ws = wb["Quality Metrics"]
    # Find "Placement %" row.
    found = False
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Placement %":
            assert ws.cell(row=r, column=2).value == 98.0
            found = True
    assert found


# ────────────────────────────────────────────────────────────
# Room capacity proxy
# ────────────────────────────────────────────────────────────
def test_derive_room_capacity_uses_max():
    caps = e.derive_room_capacity(_schedule())
    assert caps["Ciencias Exp. II"] == 20
    assert caps["Ciencias Exp. I"] == 12


def test_derive_room_capacity_empty():
    assert e.derive_room_capacity(pd.DataFrame()) == {}


# ────────────────────────────────────────────────────────────
# Full workbook + edge cases
# ────────────────────────────────────────────────────────────
def test_build_enhanced_workbook_all_sheets():
    wb = e.build_enhanced_workbook(_schedule(), kpi=_kpi(),
                                   solver_stats=_solver_stats())
    for name in ("Overview", "Groups", "Legend", "Room Utilization",
                 "Professor Workload", "Student Placement",
                 "Time Slot Analysis", "Quality Metrics"):
        assert name in wb.sheetnames


def test_build_enhanced_workbook_empty_schedule_ok():
    wb = e.build_enhanced_workbook(pd.DataFrame(), kpi={}, solver_stats=[])
    assert "Overview" in wb.sheetnames
    assert len(wb.sheetnames) >= 1


def test_build_enhanced_workbook_selective_options():
    opts = ExportOptions(room_utilization=False, professor_workload=False,
                         time_slot_analysis=False, legend=False,
                         quality_metrics=False, student_placement=False)
    wb = e.build_enhanced_workbook(_schedule(), kpi=_kpi(),
                                   solver_stats=_solver_stats(), options=opts)
    assert "Room Utilization" not in wb.sheetnames
    assert "Groups" in wb.sheetnames


@pytest.mark.parametrize("scheme", ["loyola", "default", "monochrome"])
def test_color_schemes_apply(scheme):
    wb = e.build_enhanced_workbook(_schedule(), kpi=_kpi(),
                                   solver_stats=_solver_stats(),
                                   color_scheme=scheme)
    ws = wb["Groups"]
    expected_bg = e.COLOR_SCHEMES[scheme]["header_bg"]
    assert ws.cell(row=1, column=1).fill.start_color.rgb == expected_bg


def test_export_enhanced_writes_file(tmp_path):
    out = tmp_path / "enh.xlsx"
    res = e.export_enhanced(semester=1, out_path=str(out),
                            options=ExportOptions())
    # Uses real data if present; otherwise still writes a valid (near-empty) wb.
    assert res["ok"] is True
    assert out.exists()


def test_data_validation_and_named_ranges():
    wb = Workbook(); wb.remove(wb.active)
    e.build_color_coded_groups_sheet(wb, _schedule(), options=ExportOptions())
    ws = wb["Groups"]
    e.add_dropdown_validations(ws, _schedule())
    assert len(ws.data_validations.dataValidation) >= 1
    e.add_named_ranges(wb, "Groups", e._GROUP_HEADERS, ws.max_row)
    assert any(name.startswith("Groups_") for name in wb.defined_names)
