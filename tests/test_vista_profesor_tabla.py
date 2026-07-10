"""Tests for the colour-coded, filterable "Vista profesor (tabla)" sheet.

Point 2: the enhanced Excel features (per-group / per-program colours and
native column filters) are integrated into the default deliverables through a
new companion sheet that lives next to the validated "Vista profesor" grid.
The grid itself is never modified (additive-only rule).
"""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook

import excel_generator_core as core


def _make_schedule():
    """Small synthetic schedule with two subjects, two programs, two groups."""
    rows = [
        # semester, subject, program, curso_num, grupo, session, week, day,
        # time_block, nb_students, lab_rooms, professor
        (1, "S1-Fisica", "GITI", 1, 1, 1, 1, "Lunes", "09:00-11:00", 20, "Lab A", "Prof One"),
        (1, "S1-Fisica", "GITI", 1, 2, 1, 1, "Martes", "11:00-13:00", 18, "Lab A", "Prof Two"),
        (1, "S1-Quimica", "GIM", 1, 1, 1, 2, "Miercoles", "09:00-11:00", 22, "Lab B", "Prof Three"),
        (1, "S1-Quimica", "GIM", 1, 3, 2, 2, "Jueves", "15:00-17:00", 15, "Lab B", "Prof One"),
    ]
    cols = ["semester", "subject", "program", "curso_num", "grupo", "session",
            "week", "day", "time_block", "nb_students", "lab_rooms", "professor"]
    return pd.DataFrame(rows, columns=cols)


def test_color_helpers_are_stable_and_return_fills():
    f1 = core.get_program_fill("GITI")
    f2 = core.get_program_fill("GITI")
    assert f1.fill_type == "solid"
    # Deterministic: same program -> same colour.
    assert f1.start_color.rgb == f2.start_color.rgb
    # Unknown program falls back to the default fill (no crash).
    assert core.get_program_fill("MIXED(GITI+2)").fill_type == "solid"

    g1 = core.get_group_fill(1)
    g2 = core.get_group_fill(1)
    assert g1.start_color.rgb == g2.start_color.rgb
    # Non-numeric group is handled gracefully.
    assert core.get_group_fill("").fill_type == "solid"


def test_tabla_sheet_created_with_headers_filter_and_freeze():
    wb = Workbook()
    schedule = _make_schedule()
    subjects = ["S1-Fisica", "S1-Quimica"]

    core.build_vista_profesor_tabla_sheet(wb, schedule, subjects)

    assert "Vista profesor (tabla)" in wb.sheetnames
    ws = wb["Vista profesor (tabla)"]

    expected_headers = ["Week", "Date", "Day", "Time slot", "Subject",
                        "Program", "Group", "Practice", "Room", "Students",
                        "Professor"]
    actual = [ws.cell(row=4, column=c).value for c in range(1, len(expected_headers) + 1)]
    assert actual == expected_headers

    # Native column filters on the header row.
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A4")
    # Frozen header row so filters stay visible while scrolling.
    assert ws.freeze_panes == "A5"

    # One data row per session (header row 4, data from row 5).
    data_rows = [r for r in range(5, ws.max_row + 1)
                 if ws.cell(row=r, column=1).value not in (None, "")]
    assert len(data_rows) == len(schedule)


def test_tabla_sheet_applies_colour_coding():
    wb = Workbook()
    schedule = _make_schedule()
    subjects = ["S1-Fisica", "S1-Quimica"]

    core.build_vista_profesor_tabla_sheet(wb, schedule, subjects)
    ws = wb["Vista profesor (tabla)"]

    # Subject (col 5), program (col 6) and group (col 7) cells are filled.
    for r in range(5, 5 + len(schedule)):
        assert ws.cell(row=r, column=5).fill.fill_type == "solid"
        assert ws.cell(row=r, column=6).fill.fill_type == "solid"
        assert ws.cell(row=r, column=7).fill.fill_type == "solid"


def test_tabla_sheet_empty_schedule_is_safe():
    wb = Workbook()
    empty = pd.DataFrame(columns=["semester", "subject", "program", "curso_num",
                                  "grupo", "session", "week", "day", "time_block",
                                  "nb_students", "lab_rooms", "professor"])
    # Must not raise on an empty schedule.
    core.build_vista_profesor_tabla_sheet(wb, empty, ["S1-Fisica"])
    assert "Vista profesor (tabla)" in wb.sheetnames
