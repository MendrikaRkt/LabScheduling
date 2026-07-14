"""
diagnostic_sheet.py
===================

Rendu de la feuille Excel « Diagnostic & Remèdes ».

Cette feuille matérialise, pour le jury et l'établissement, la thèse centrale
du projet : un statut solveur « OPTIMAL » ne prouve PAS que la solution est
correcte. Le pré-traitement peut absorber une infaisabilité en déformant la
solution (groupes minuscules/solo, séances hors-période, sur-souscription,
surcharge professeur). Cette feuille LISTE ces anomalies détectées par
``diagnostics.audit_schedule``, les compte par NIVEAU × SEMESTRE, et affiche
pour chacune un REMÈDE PROPOSÉ et chiffré — jamais appliqué automatiquement.

Conception
----------
* Purement additive : n'altère aucune autre feuille ni la logique de génération.
* Read-only : consomme le dict produit par ``diagnostics.audit_schedule``.
* Contenu en FRANÇAIS (demande de l'utilisateur).
* Identité visuelle Loyola (marine #1B3A6F, or #FFCC00), pas de rouge/vert vif —
  le sens est toujours renforcé par un libellé texte explicite.

Utilisation
-----------
    from openpyxl import Workbook
    import diagnostics, diagnostic_sheet

    report = diagnostics.audit_schedule(rows)
    wb = Workbook()
    diagnostic_sheet.build_diagnostic_sheet(wb, report)
    wb.save("out.xlsx")
"""

from __future__ import annotations

from typing import Any, Dict, List

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# On réutilise la palette et les briques de style de validation_sheet pour
# garantir une cohérence visuelle parfaite entre les deux feuilles d'analyse.
from validation_sheet import (
    BAD, BAD_BG, BORDER, GOOD, GOOD_BG, GOLD, GREY_BG, NAVY, NAVY_DEEP,
    WARN, WARN_BG, WHITE, _plain_table, _section, _title,
)

import diagnostics as _dg


# Libellés FR par type d'anomalie.
_TYPE_LABELS = {
    "tiny_group": "Groupe sous-dimensionné",
    "wrong_period": "Séance hors-période",
    "oversubscription": "Matière sur-souscrite",
    "bottleneck": "Goulot d'étranglement",
    "credit_overload": "Surcharge professeur",
}

# Libellés FR par sévérité + couleurs (texte, fond).
_SEV_LABELS = {
    _dg.SEV_CRITICAL: "Critique",
    _dg.SEV_WARNING: "Avertissement",
    _dg.SEV_INFO: "Info",
}
_SEV_COLORS = {
    _dg.SEV_CRITICAL: (BAD, BAD_BG),
    _dg.SEV_WARNING: (WARN, WARN_BG),
    _dg.SEV_INFO: (GOOD, GOOD_BG),
}


def _kpi_band(ws, row, report: Dict[str, Any]) -> int:
    """Bandeau de synthèse : total, critiques, groupes analysés, verdict."""
    n_total = report.get("n_total", 0)
    n_crit = report.get("n_critical", 0)
    n_groups = report.get("n_groups_analyzed", 0)
    healthy = report.get("healthy", n_total == 0)

    cells = [
        ("Anomalies détectées", n_total,
         (GOOD if n_total == 0 else BAD), (GOOD_BG if n_total == 0 else BAD_BG)),
        ("Dont critiques", n_crit,
         (GOOD if n_crit == 0 else BAD), (GOOD_BG if n_crit == 0 else BAD_BG)),
        ("Groupes analysés", n_groups, NAVY_DEEP, GREY_BG),
        ("Verdict", "CONFORME" if healthy else "À CORRIGER",
         (GOOD if healthy else BAD), (GOOD_BG if healthy else BAD_BG)),
    ]
    for i, (label, value, fg, bg) in enumerate(cells):
        col = 1 + i
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = Font(bold=True, size=9, color=NAVY_DEEP)
        lc.fill = PatternFill("solid", fgColor=GREY_BG)
        lc.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        lc.border = BORDER
        vc = ws.cell(row=row + 1, column=col, value=value)
        vc.font = Font(bold=True, size=18, color=fg)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = BORDER
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row + 1].height = 34
    return row + 2


def build_diagnostic_sheet(workbook, report: Dict[str, Any],
                           sheet_title="Diagnostic & Remèdes"):
    """Construit la feuille « Diagnostic & Remèdes » dans *workbook*.

    Args:
        workbook: classeur openpyxl cible.
        report: dict retourné par ``diagnostics.audit_schedule``.
        sheet_title: nom de l'onglet.
    """
    # Nettoie la feuille par défaut vide si présente.
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
        ws = workbook["Sheet"]
        ws.title = sheet_title
    else:
        ws = workbook.create_sheet(title=sheet_title)

    # Largeurs de colonnes (7 colonnes utiles).
    widths = [22, 10, 10, 26, 14, 40, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = _title(ws, row, "Diagnostic & Remèdes — Audit métier de la solution",
                 ncols=7)

    # Encart explicatif (la thèse du projet).
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    intro = ws.cell(row=row, column=1, value=(
        "Un statut solveur « OPTIMAL » ne garantit pas une solution CONFORME. "
        "Le pré-traitement peut absorber une infaisabilité en déformant la "
        "solution. Cette feuille liste les anomalies métier détectées et "
        "propose pour chacune un remède chiffré (à décider par l'utilisateur, "
        "jamais appliqué automatiquement)."))
    intro.font = Font(size=10, italic=True, color=NAVY_DEEP)
    intro.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 46
    row += 1
    row += 1

    # Bandeau KPI.
    row = _kpi_band(ws, row, report)
    row += 1

    # Cas sain : message et sortie anticipée.
    if report.get("healthy", report.get("n_total", 0) == 0):
        row = _section(ws, row, "Résultat", ncols=7)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=(
            "Aucune anomalie métier détectée : tous les groupes respectent la "
            "taille minimale et la période attendue pour leur niveau. La "
            "solution est conforme aux règles vérifiées."))
        c.font = Font(size=11, color=NAVY_DEEP)
        c.fill = PatternFill("solid", fgColor=GOOD_BG)
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 40
        ws.freeze_panes = "A2"
        return ws

    # Synthèse par type d'anomalie.
    row = _section(ws, row, "Synthèse par type d'anomalie", ncols=7)
    by_type = report.get("by_type", {})
    type_rows = [
        [_TYPE_LABELS.get(k, k), v]
        for k, v in sorted(by_type.items(), key=lambda kv: -kv[1])
    ]
    row = _plain_table(ws, row, ["Type d'anomalie", "Nombre"], type_rows,
                       col_widths=[30, 12])
    row += 1

    # Réconciliation par niveau × semestre.
    row = _section(ws, row, "Répartition par niveau × semestre", ncols=7)
    bls = report.get("by_level_semester", {})
    ls_rows = []
    for key in sorted(bls.keys()):
        d = bls[key]
        ls_rows.append([
            key,
            d.get("tiny_group", 0),
            d.get("wrong_period", 0),
            d.get("oversubscription", 0)
            + d.get("bottleneck", 0) + d.get("credit_overload", 0),
            d.get("total", 0),
        ])
    row = _plain_table(
        ws, row,
        ["Niveau · Semestre", "Groupes\nsous-dim.", "Hors-\npériode",
         "Autres", "Total"],
        ls_rows, col_widths=[24, 12, 12, 12, 10])
    row += 1

    # Détail des anomalies + remèdes proposés.
    row = _section(ws, row, "Détail des anomalies et remèdes proposés",
                   ncols=7)

    headers = ["Niveau", "Sem.", "Grp.", "Matière", "Sévérité",
               "Anomalie", "Remède proposé (chiffré)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26
    row += 1

    for a in report.get("anomalies", []):
        sev = a.get("severity", _dg.SEV_INFO)
        fg, bg = _SEV_COLORS.get(sev, (NAVY_DEEP, GREY_BG))
        remedy = a.get("remedy", {}) or {}
        values = [
            a.get("level", ""),
            a.get("semester", ""),
            str(a.get("grupo", "")),
            a.get("subject", ""),
            _SEV_LABELS.get(sev, sev),
            a.get("detail", ""),
            remedy.get("text", ""),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       indent=1)
            if c == 5:  # colonne sévérité colorée
                cell.font = Font(bold=True, color=fg)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
            elif c in (1, 4):
                cell.font = Font(bold=True, color=NAVY_DEEP)
        ws.row_dimensions[row].height = 42
        row += 1

    ws.freeze_panes = "A2"
    return ws
