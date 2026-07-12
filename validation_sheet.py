"""
validation_sheet.py
====================

Render the final reliability report (from ``schedule_validation``) into a
styled Excel worksheet — the "Validation" tab — that is appended to every
generated workbook (Daniel format) and can also be exported on its own.

Design goals
------------
* Purely additive: it never touches the existing sheets or their logic; it
  only *adds* a "Validation" sheet to a workbook it is handed.
* Read-only with respect to the schedule: it consumes the report dict produced
  by :func:`schedule_validation.validate_schedule`.
* Content is in ENGLISH (explicit request): every label, verdict, formula and
  remediation hint is written in English so the sheet reads consistently with
  the analytical tabs (Teacher View, Vista profesor (tabla)).
* Loyola visual identity (navy #1B3A6F, gold #FFCC00).

Usage
-----
    from openpyxl import Workbook
    from schedule_validation import validate_schedule
    import validation_sheet

    wb = Workbook()
    report = validate_schedule()
    validation_sheet.build_validation_sheet(wb, report)
    wb.save("out.xlsx")
"""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Loyola palette ─────────────────────────────────────────────────────────
# Status colors follow the Universidad Loyola brand (blue/white). We avoid
# bright green/red: positive = Loyola blue, attention = gold, problem = muted
# wine. Meaning is always reinforced by an explicit text label in each cell.
NAVY = "1B3A6F"
NAVY_DEEP = "0F2344"
GOLD = "FFCC00"
GOOD = "2E86AB"
GOOD_BG = "E4F0F6"
WARN = "B8860B"
WARN_BG = "FBF3DC"
BAD = "B26575"
BAD_BG = "F3E5E9"
GREY_BG = "F2F5FA"
WHITE = "FFFFFF"

_THIN = Side(style="thin", color="C7D2E4")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# English labels for the five checks + group sizing
CHECK_LABELS_EN = {
    "room_conflicts": "Room conflicts",
    "student_theory_conflicts": "Student theory clash",
    "student_lab_double_booking": "Student lab double-booking",
    "professor_double_booking": "Professor double-booking",
    "professor_busy": "Professor busy (teaching slot)",
    "group_sizing": "Group sizing",
}
CHECK_KIND_EN = {"hard": "Hard constraint", "indicator": "Indicator"}
# English detail text (the validation module stores French detail; we localise
# it here for the Excel output).
CHECK_DETAILS_EN = {
    "room_conflicts": "A physical room hosts a single group per (semester, week, day, time slot).",
    "student_theory_conflicts": "No student in a lab on a busy theory-class slot (same semester).",
    "student_lab_double_booking": "No student with two lab sessions in the same (semester, week, day, slot).",
    "professor_double_booking": "Indicator (assignment done after optimisation, not constrained by the solver): no professor with two sessions in the same (semester, week, day, slot).",
    "professor_busy": "Indicator (post-optimisation assignment): session on the professor's theory-class slot (same semester, except the replaced hour).",
    "group_sizing": "Group size within policy (min 7 / preferred 12 / max 15).",
}
# Likely CAUSE of a residual conflict, per check — shown in the examples table.
CHECK_CAUSE_EN = {
    "room_conflicts": "Two groups placed in the same physical room during the same (semester, week, day, slot).",
    "student_theory_conflicts": "A lab session overlaps a compulsory theory class attended by at least one enrolled student.",
    "student_lab_double_booking": "A student belongs to two lab groups scheduled in the same (semester, week, day, slot).",
    "professor_double_booking": "A professor is assigned to two sessions in the same slot (assignment happens after optimisation, outside the CP-SAT model).",
    "professor_busy": "A professor's assigned session falls on one of their own theory-class slots (post-optimisation assignment).",
    "group_sizing": "One or more groups fall outside the size policy (min 7 / preferred 12 / max 15).",
}
# Proposed REMEDY, per check — actionable and tied to the app's tools.
CHECK_REMEDY_EN = {
    "room_conflicts": "Add room capacity for that slot, or move one group to a free room/slot (Simulator > add resources).",
    "student_theory_conflicts": "Reschedule the lab to a slot free of theory classes for those students, or split the affected group.",
    "student_lab_double_booking": "Move one of the two lab sessions to a different slot, or rebalance group membership.",
    "professor_double_booking": "Re-assign one session to another eligible professor, or shift it to a free slot.",
    "professor_busy": "Assign an eligible substitute professor for that slot, or move the session.",
    "group_sizing": "Merge under-sized groups or split over-sized ones; adjust the target size in Configuration.",
}
STATUS_EN = {
    "PASS": "PASS",
    "WARN": "PASS WITH WARNINGS",
    "FAIL": "FAIL",
    "NO_DATA": "NO DATA",
}


def _title(ws, row, text, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30
    return row + 1


def _section(ws, row, text, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color=NAVY_DEEP)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1


def _kv(ws, row, key, value, key_col=1, val_col=2, val_span=5):
    kc = ws.cell(row=row, column=key_col, value=key)
    kc.font = Font(bold=True, color=NAVY_DEEP)
    kc.fill = PatternFill("solid", fgColor=GREY_BG)
    kc.alignment = Alignment(vertical="center", indent=1)
    kc.border = BORDER
    if val_span > 1:
        ws.merge_cells(start_row=row, start_column=val_col,
                       end_row=row, end_column=val_col + val_span - 1)
    vc = ws.cell(row=row, column=val_col, value=value)
    vc.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    vc.border = BORDER
    for col in range(val_col, val_col + val_span):
        ws.cell(row=row, column=col).border = BORDER
    return row + 1


def _status_colors(status):
    return {
        "PASS": (GOOD, GOOD_BG),
        "WARN": (WARN, WARN_BG),
        "FAIL": (BAD, BAD_BG),
        "NO_DATA": (WARN, WARN_BG),
    }.get(status, (WARN, WARN_BG))


def _breakdown_table(ws, row, first_header, slice_labels, metrics, slices,
                     global_counts):
    """Render a compact breakdown grid: one row per metric, one column per
    slice (semester or level), plus a final 'Global (Total)' column.

    A ✓ mark next to the total confirms that the per-slice figures sum exactly
    to the global total — the visual proof of authenticity requested.
    Columns used: A (metric) .. then one per slice .. then Total (2 cols).
    """
    n_slices = len(slice_labels)
    # Header row -------------------------------------------------------------
    hc = ws.cell(row=row, column=1, value=first_header)
    hc.font = Font(bold=True, color=WHITE)
    hc.fill = PatternFill("solid", fgColor=NAVY)
    hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hc.border = BORDER
    col = 2
    for lbl in slice_labels:
        c = ws.cell(row=row, column=col, value=lbl)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        col += 1
    # 'Global (Total)' spanning the remaining columns (up to col 6).
    gcol = col
    tc = ws.cell(row=row, column=gcol, value="Global (Total)")
    tc.font = Font(bold=True, color=NAVY_DEEP)
    tc.fill = PatternFill("solid", fgColor=GOLD)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    last_col = max(gcol, 6)
    if last_col > gcol:
        ws.merge_cells(start_row=row, start_column=gcol,
                       end_row=row, end_column=last_col)
    for cc in range(gcol, last_col + 1):
        ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=GOLD)
        ws.cell(row=row, column=cc).border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    # Metric rows ------------------------------------------------------------
    for label, key in metrics:
        mc = ws.cell(row=row, column=1, value=label)
        mc.font = Font(bold=True, color=NAVY_DEEP)
        mc.fill = PatternFill("solid", fgColor=GREY_BG)
        mc.alignment = Alignment(vertical="center", indent=1)
        mc.border = BORDER
        col = 2
        slice_sum = 0
        for s in slices:
            val = int(s.get(key, 0) or 0)
            slice_sum += val
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            col += 1
        gtotal = int(global_counts.get(key, 0) or 0)
        # For 'subjects' and 'students' a slice sum may exceed the global total
        # because a subject/student can appear in more than one slice; only the
        # additive metrics (sessions, groups) are expected to reconcile exactly.
        reconciles = key in ("sessions", "groups") and slice_sum == gtotal
        mark = "  ✓" if reconciles else ""
        gv = ws.cell(row=row, column=gcol, value=f"{gtotal}{mark}")
        gv.font = Font(bold=True, color=NAVY_DEEP)
        gv.alignment = Alignment(horizontal="center", vertical="center")
        if last_col > gcol:
            ws.merge_cells(start_row=row, start_column=gcol,
                           end_row=row, end_column=last_col)
        for cc in range(gcol, last_col + 1):
            ws.cell(row=row, column=cc).border = BORDER
        row += 1

    # Footnote ---------------------------------------------------------------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    fn = ws.cell(
        row=row, column=1,
        value=("✓ = per-slice figures sum exactly to the global total "
               "(Lab sessions, Practice groups). Subjects and Students may "
               "recur across slices, so their per-slice values can overlap."),
    )
    fn.font = Font(size=9, italic=True, color=NAVY_DEEP)
    fn.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 26
    row += 1
    return row


def build_validation_sheet(workbook, report, sheet_title="Validation"):
    """Append a styled 'Validation' sheet built from a validation *report*.

    Parameters
    ----------
    workbook : openpyxl.Workbook
    report   : dict returned by schedule_validation.validate_schedule()
    sheet_title : str
    """
    ws = workbook.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False
    # Column widths tuned so the wide "Detail" column (F) fits its text without
    # overlapping neighbouring rows; the narrow numeric columns stay compact.
    widths = [34, 20, 14, 12, 18, 52]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = _title(ws, row, "Schedule validation and reliability")
    row += 2

    # ── Global verdict ─────────────────────────────────────────────────
    status = report.get("status", "NO_DATA")
    fg, bg = _status_colors(status)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=2)
    v = ws.cell(row=row, column=1, value=STATUS_EN.get(status, status))
    v.font = Font(size=18, bold=True, color=WHITE)
    v.fill = PatternFill("solid", fgColor=fg)
    v.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rr in (row, row + 1):
        for cc in (1, 2):
            ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=fg)

    score = report.get("reliability_score", 0.0)
    ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=4)
    sc = ws.cell(row=row, column=3, value=f"{score:.1f} / 100")
    sc.font = Font(size=20, bold=True, color=NAVY_DEEP)
    sc.fill = PatternFill("solid", fgColor=bg)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (row, row + 1):
        for cc in (3, 4):
            ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=bg)
    lbl = ws.cell(row=row, column=5, value="Reliability index")
    lbl.font = Font(size=10, italic=True, color=NAVY_DEEP)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    solver = report.get("solver", {})
    slv = ws.cell(row=row + 1, column=5,
                  value=f"Solver: {solver.get('status', 'N/A')}")
    slv.font = Font(size=10, bold=True, color=NAVY_DEEP)
    ws.merge_cells(start_row=row + 1, start_column=5, end_row=row + 1, end_column=6)
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row + 1].height = 24
    row += 3

    # ── Key counts ─────────────────────────────────────────────────────
    row = _section(ws, row, "Schedule summary")
    counts = report.get("counts", {})
    row = _kv(ws, row, "Lab sessions", counts.get("sessions", 0))
    row = _kv(ws, row, "Practice groups", counts.get("groups", 0))
    row = _kv(ws, row, "Subjects", counts.get("subjects", 0))
    row = _kv(ws, row, "Semesters", counts.get("semesters", 0))
    row = _kv(ws, row, "Students involved", counts.get("students", 0))
    row += 2

    # ── Breakdown by semester (authenticity: per-slice reconciles to total)
    by_sem = report.get("counts_by_semester") or []
    if by_sem:
        row = _section(ws, row, "Breakdown by semester")
        row = _breakdown_table(
            ws, row,
            first_header="Metric",
            slice_labels=[s.get("semester", "?") for s in by_sem],
            metrics=[
                ("Lab sessions", "sessions"),
                ("Practice groups", "groups"),
                ("Subjects", "subjects"),
                ("Students involved", "students"),
            ],
            slices=by_sem,
            global_counts=counts,
        )
        row += 2

    # ── Breakdown by level / course ────────────────────────────────────
    by_lvl = report.get("counts_by_level") or []
    if by_lvl:
        row = _section(ws, row, "Breakdown by level (course)")
        row = _breakdown_table(
            ws, row,
            first_header="Metric",
            slice_labels=["Course " + str(s.get("level", "?")) for s in by_lvl],
            metrics=[
                ("Lab sessions", "sessions"),
                ("Practice groups", "groups"),
                ("Subjects", "subjects"),
                ("Students involved", "students"),
            ],
            slices=by_lvl,
            global_counts=counts,
        )
        row += 2

    # ── Checks table ───────────────────────────────────────────────────
    row = _section(ws, row, "Conflict checks")
    headers = ["Check", "Type", "Status", "Conflicts",
               "Affected / Total", "Detail"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 24
    row += 1

    checks = report.get("checks", {})
    order = ["room_conflicts", "student_theory_conflicts",
             "student_lab_double_booking", "professor_double_booking",
             "professor_busy", "group_sizing"]
    for name in order:
        chk = checks.get(name)
        if not chk:
            continue
        checkable = chk.get("checkable", True)
        passed = chk.get("passed", True)
        kind = chk.get("kind") or ("hard" if name != "group_sizing" else "indicator")
        if not checkable:
            state, sfg, sbg = "N/A", WARN, WARN_BG
        elif passed:
            state, sfg, sbg = "OK", GOOD, GOOD_BG
        elif kind == "indicator":
            state, sfg, sbg = "WARNING", WARN, WARN_BG
        else:
            state, sfg, sbg = "CONFLICT", BAD, BAD_BG
        affected = chk.get("affected")
        total = chk.get("total")
        aff_total = (f"{affected} / {total}"
                     if affected is not None and total is not None else "—")
        values = [
            CHECK_LABELS_EN.get(name, name),
            CHECK_KIND_EN.get(kind, kind),
            state,
            chk.get("count", 0),
            aff_total,
            CHECK_DETAILS_EN.get(name, chk.get("detail", "")),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if c in (2, 3, 4, 5) else "left",
                wrap_text=(c == 6), indent=0 if c in (2, 3, 4, 5) else 1)
            if c == 3:
                cell.font = Font(bold=True, color=sfg)
                cell.fill = PatternFill("solid", fgColor=sbg)
        ws.row_dimensions[row].height = 44
        row += 1
    row += 2

    # ── Reliability formula (laid out for readability) ─────────────────
    row = _section(ws, row, "Reliability index formula")
    # The formula on its own line, in a monospace-like style.
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    fcell = ws.cell(
        row=row, column=1,
        value=("index = 100 x Σ[ weight x (1 - conflicting_entities / "
               "total_entities) ] / Σ weights  x  solver_factor"),
    )
    fcell.font = Font(name="Consolas", size=11, bold=True, color=NAVY_DEEP)
    fcell.fill = PatternFill("solid", fgColor=GREY_BG)
    fcell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    for cc in range(1, 7):
        ws.cell(row=row, column=cc).border = BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    # Weights sub-table (hard constraints only).
    whdr = ["Term (hard constraint)", "Weight"]
    ws.cell(row=row, column=1, value=whdr[0]).font = Font(bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=row, column=1).border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    for cc in range(1, 5):
        ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=row, column=cc).border = BORDER
    wc = ws.cell(row=row, column=5, value=whdr[1])
    wc.font = Font(bold=True, color=WHITE)
    wc.fill = PatternFill("solid", fgColor=NAVY)
    wc.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    for cc in range(5, 7):
        ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=row, column=cc).border = BORDER
    row += 1
    for term, weight in (("Room conflicts", 40),
                         ("Student theory clash", 30),
                         ("Student lab double-booking", 30)):
        tc = ws.cell(row=row, column=1, value=term)
        tc.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        for cc in range(1, 5):
            ws.cell(row=row, column=cc).border = BORDER
        wv = ws.cell(row=row, column=5, value=weight)
        wv.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        for cc in range(5, 7):
            ws.cell(row=row, column=cc).border = BORDER
        row += 1

    # Legend / interpretation of the factors, as bullet lines.
    legend_lines = [
        "solver_factor = 1.0 if the solver status is OPTIMAL/FEASIBLE, else 0.9.",
        "Status = PASS if there is no hard violation; PASS WITH WARNINGS if the "
        "residual is <= 1% of entities OR indicators are flagged; FAIL if > 1%.",
        "Professor checks (double-booking, busy) are quality INDICATORS: the "
        "teaching assignment is done AFTER optimisation, outside the CP-SAT "
        "model, so they lower the status to WARNING but never to FAIL.",
    ]
    for line in legend_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        lc = ws.cell(row=row, column=1, value="•  " + line)
        lc.font = Font(size=10, color=NAVY_DEEP)
        lc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        for cc in range(1, 7):
            ws.cell(row=row, column=cc).border = BORDER
        ws.row_dimensions[row].height = 28
        row += 1
    row += 1

    # ── Solver configuration applied (traceability) ────────────────────
    solver_cfg = report.get("solver_config") or {}
    if solver_cfg:
        row = _section(ws, row, "Solver configuration (applied parameters)")
        # Header for the constraints sub-table.
        cfg_hdr = ["Soft constraint", "Weight", "State"]
        ws.cell(row=row, column=1, value=cfg_hdr[0]).font = Font(bold=True, color=WHITE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        wcol = ws.cell(row=row, column=4, value=cfg_hdr[1])
        wcol.font = Font(bold=True, color=WHITE)
        wcol.alignment = Alignment(horizontal="center")
        scol = ws.cell(row=row, column=5, value=cfg_hdr[2])
        scol.font = Font(bold=True, color=WHITE)
        scol.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        for cc in range(1, 7):
            ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(row=row, column=cc).border = BORDER
        row += 1
        # Active profile line.
        row = _kv(ws, row, "Active profile", solver_cfg.get("profile", "-"))
        _labels_en = {
            "semester_anchor_first": "Anchor first session to semester start",
            "semester_anchor_last": "Anchor last session to semester end",
            "spacing": "Regular spacing between sessions",
            "parity": "Parity alternation between parallel groups",
        }
        weights = solver_cfg.get("weights", {}) or {}
        enabled = solver_cfg.get("enabled", {}) or {}
        for key in ("semester_anchor_first", "semester_anchor_last",
                    "spacing", "parity"):
            if key not in weights and key not in enabled:
                continue
            state = "enabled" if enabled.get(key, False) else "disabled"
            nc = ws.cell(row=row, column=1, value=_labels_en.get(key, key))
            nc.alignment = Alignment(vertical="center", indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            wv = ws.cell(row=row, column=4, value=weights.get(key, 0))
            wv.alignment = Alignment(horizontal="center")
            sv = ws.cell(row=row, column=5, value=state)
            sv.alignment = Alignment(horizontal="center")
            sv.font = Font(color=(GOOD if state == "enabled" else WARN), bold=True)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            for cc in range(1, 7):
                ws.cell(row=row, column=cc).border = BORDER
            row += 1
        row += 1

    # ── Teacher load (top entries) ─────────────────────────────────────
    teacher_load = report.get("teacher_load", [])
    if teacher_load:
        row = _section(ws, row, "Teaching load (sessions per professor)")
        thdr = ["Professor", "Sessions", "Groups", "Subjects"]
        for c, h in enumerate(thdr, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        for cc in range(4, 7):
            ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(row=row, column=cc).border = BORDER
        row += 1
        for entry in teacher_load[:25]:
            pc = ws.cell(row=row, column=1, value=entry.get("professor", ""))
            pc.border = BORDER
            pc.alignment = Alignment(vertical="center", indent=1)
            for col, key in ((2, "sessions"), (3, "groups")):
                cc = ws.cell(row=row, column=col, value=entry.get(key, 0))
                cc.alignment = Alignment(horizontal="center")
                cc.border = BORDER
            subj = entry.get("subjects", 0)
            if isinstance(subj, list):
                subj = ", ".join(subj)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            scell = ws.cell(row=row, column=4, value=subj)
            scell.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
            for cc in range(4, 7):
                ws.cell(row=row, column=cc).border = BORDER
            row += 1
        row += 1

    # ── Examples of residual conflicts (with cause & remedy) ───────────
    problem_checks = [
        (name, chk) for name, chk in checks.items()
        if chk.get("examples") and not chk.get("passed", True)
    ]
    if problem_checks:
        row = _section(ws, row, "Residual conflicts — examples, causes and remedies")
        for name, chk in problem_checks:
            # Check name banner.
            hc = ws.cell(row=row, column=1, value=CHECK_LABELS_EN.get(name, name))
            hc.font = Font(bold=True, color=WHITE)
            hc.fill = PatternFill("solid", fgColor=NAVY)
            hc.alignment = Alignment(vertical="center", indent=1)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            for cc in range(1, 7):
                ws.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=NAVY)
                ws.cell(row=row, column=cc).border = BORDER
            row += 1
            # Cause line.
            row = _kv(ws, row, "Likely cause",
                      CHECK_CAUSE_EN.get(name, "—"))
            # Remedy line.
            row = _kv(ws, row, "Proposed remedy",
                      CHECK_REMEDY_EN.get(name, "—"))
            # Example instances.
            ex_hdr = ws.cell(row=row, column=1, value="Example instances")
            ex_hdr.font = Font(bold=True, italic=True, color=NAVY_DEEP)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            for ex in chk.get("examples", [])[:10]:
                txt = "; ".join(f"{k}={v}" for k, v in ex.items())
                ec = ws.cell(row=row, column=1, value="- " + txt)
                ec.font = Font(size=9, color="333333")
                ec.alignment = Alignment(wrap_text=True, indent=1)
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=6)
                row += 1
            row += 1

    ws.sheet_view.showGridLines = False
    # Freeze the report title so section banners stay in context while scrolling
    # through the (potentially long) checks / examples tables.
    ws.freeze_panes = "A2"
    return ws
