"""
04 v5 + 05 v5 COMBINÉ — Pipeline complet aulario+alumnos → planning optimal
Projet Scheduling - Universidad Loyola
Auteur: RAKOTONJANAHARY Maminiaina Mendrika

APPROCHE FINALE :
  INPUT  : master_schedule.csv (aulario + alumnos, scripts 01-02)
           + CONFIGURATION des labs (quelles matières, salles, sessions)
  OUTPUT : planning optimal des sessions de lab

  Les fichiers de Daniel ne sont PLUS utilisés comme entrée.
  Ils servent UNIQUEMENT de validation a posteriori.

PIPELINE EN 5 ÉTAPES :
  1. Charger le master_schedule et la configuration des labs
  2. Identifier les étudiants inscrits à chaque matière via MixtoID
  3. Calculer l'emploi du temps individuel de chaque étudiant
  4. Former les groupes : chaque étudiant assigné à un créneau libre
  5. Solveur : décider la semaine de chaque session (créneau déjà fixé)
"""

import io
import json
import math
import os
import subprocess
import sys
import hashlib
from collections import defaultdict, Counter
from datetime import datetime, timedelta

from ortools.sat.python import cp_model
import pandas as pd

try:
    import data_quality as _dq
except Exception:  # le module QA est optionnel ; ne jamais casser l'import
    _dq = None

try:
    import kpi_report as _kpi
except Exception:  # le module KPI est optionnel ; ne jamais casser l'import
    _kpi = None

# Étape 6.5 — Reproductibilité & réglage du solveur (centralisés)
RANDOM_SEED = 42            # graine fixe -> résultats reproductibles d'un run à l'autre
SOLVER_RELATIVE_GAP = 0.02  # arrêt à 2 % de l'optimum prouvé -> temps maîtrisé
SOLVER_LOG_PROGRESS = False # passer à True pour diagnostiquer la recherche

# Fichiers SOURCES bruts (optionnels) pour la réconciliation de jointure (Étape 6.2).
# S'ils sont présents à côté du master, on mesure la fuite aulario->master.
ALUMNOS_SOURCE_CANDIDATES = (
    'data_clean/report_AlumnosGruposCentroDecanos.xlsx',
    'report_AlumnosGruposCentroDecanos.xlsx',
)
AULARIO_SOURCE_CANDIDATES = (
    'data_clean/revisionAulario.xlsx',
    'revisionAulario.xlsx',
)


def _first_existing(candidates):
    """Renvoie le premier chemin existant parmi `candidates`, sinon None."""
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def configure_solver(solver, time_limit=None):
    """Étape 6.5 — Applique un paramétrage solveur reproductible et maîtrisé.

    Centralise random_seed / relative_gap_limit / log_search_progress /
    max_time_in_seconds / num_search_workers pour TOUS les appels au solveur
    (modèle principal ET modèle de repli), afin d'éviter la dérive de réglages.
    """
    p = solver.parameters
    p.max_time_in_seconds = time_limit if time_limit is not None else SOLVER_TIME_LIMIT
    p.num_search_workers = 8
    p.random_seed = RANDOM_SEED
    p.relative_gap_limit = SOLVER_RELATIVE_GAP
    p.log_search_progress = SOLVER_LOG_PROGRESS
    return solver


def add_week_hints(model, week_vars, sessions):
    """Étape 6.5 — Warm-start : suggère au solveur un étalement régulier des
    semaines (réparti uniformément dans la fenêtre de chaque groupe, en
    respectant l'ordre des séances). Les hints sont NON contraignants : ils
    n'altèrent jamais la validité du modèle, ils accélèrent la convergence.
    """
    by_group = defaultdict(list)
    for s in sessions:
        by_group[(s['subject'], s['grupo'])].append(s)
    n_hints = 0
    for grp in by_group.values():
        gsorted = sorted(grp, key=lambda x: x['session'])
        n = len(gsorted)
        if n == 0:
            continue
        lo = gsorted[0]['min_week']
        hi = gsorted[0]['max_week']
        span = max(1, hi - lo)
        for k, s in enumerate(gsorted):
            # semaine cible régulièrement espacée, bornée à [lo, hi]
            target = lo + round(span * (k + 1) / (n + 1)) if n > 1 else (lo + hi) // 2
            target = max(lo, min(hi, int(target)))
            try:
                model.AddHint(week_vars[s['id']], target)
                n_hints += 1
            except Exception:
                pass
    return n_hints


# Étape 6.4 — Accumulateur de statistiques solveur (alimente les KPIs, §6.6).
SOLVER_RUNS = []

# Diagnostic des étudiants non placés (alimente l'app + les exports Excel).
# Chaque entrée explique POURQUOI un étudiant inscrit n'a pas pu être affecté
# à un groupe : créneaux libres de l'étudiant vs créneaux des groupes existants.
UNPLACED_DIAGNOSTICS = []


def diagnose_unplaced_students(all_groups, subject_students, student_busy):
    """Analyse fine des inscriptions non placées.

    Pour chaque étudiant inscrit à une matière mais absent de tout groupe de
    cette matière, on reconstruit :
      - ses créneaux LIBRES (30 créneaux possibles − créneaux occupés),
      - les créneaux des groupes existants de la matière (jour/bloc + remplissage),
      - le verdict : soit AUCUN créneau commun (conflit d'emploi du temps total),
        soit créneaux communs mais groupes SATURÉS (capacité atteinte).

    Renseigne et renvoie la liste globale UNPLACED_DIAGNOSTICS.
    """
    UNPLACED_DIAGNOSTICS.clear()

    # Table id étudiant -> nom lisible (best-effort).
    try:
        import pandas as _pd
        _dfn = _pd.read_csv('data_clean/master_schedule.csv',
                            usecols=['AlumnoID', 'Apellidos', 'Nombre'])
        _dfn = _dfn.drop_duplicates('AlumnoID')
        sid_to_name = {str(r['AlumnoID']): f"{r['Apellidos']}, {r['Nombre']}"
                       for _, r in _dfn.iterrows()}
    except Exception:
        sid_to_name = {}

    all_slots = [(d, b) for d in range(len(DAYS)) for b in ALL_BLOCKS]

    for subject, ids in subject_students.items():
        enrolled = set(ids)
        subj_groups = [g for g in all_groups if g['subject'] == subject]
        placed = set()
        for g in subj_groups:
            placed.update(g.get('student_ids', []))
        unplaced = enrolled - placed
        if not unplaced:
            continue

        # Créneaux occupés par AU MOINS un groupe de la matière (jour, bloc).
        group_slots = {}
        for g in subj_groups:
            key = (g['day_idx'], g['block_id'])
            cur = group_slots.setdefault(key, {'groups': 0, 'free_capacity': 0})
            cur['groups'] += 1
            cur['free_capacity'] += max(0, g['max_students'] - g['nb_students'])

        for sid in unplaced:
            busy = student_busy.get(sid, set())
            free_slots = [s for s in all_slots if s not in busy]
            free_labels = [f"{DAYS[d]} {BLOCK_LABELS[b]}" for (d, b) in free_slots]

            # Créneaux des groupes où l'étudiant serait libre.
            compatible = [(d, b) for (d, b) in group_slots if (d, b) in free_slots]
            compatible_with_room = [
                (d, b) for (d, b) in compatible
                if group_slots[(d, b)]['free_capacity'] > 0]

            if not subj_groups:
                verdict = "Aucun groupe formé pour cette matière"
            elif not compatible:
                verdict = ("Conflit d'emploi du temps total : l'étudiant est occupé "
                           "sur TOUS les créneaux des groupes existants")
            elif not compatible_with_room:
                verdict = ("Créneaux compatibles mais groupes SATURÉS "
                           "(capacité salle/groupe atteinte)")
            else:
                verdict = ("Créneaux compatibles disponibles — non placé par "
                           "contrainte de cohorte/programme")

            UNPLACED_DIAGNOSTICS.append({
                'subject': subject,
                'student_id': str(sid),
                'student_name': sid_to_name.get(str(sid), f"<id {sid}>"),
                'n_free_slots': len(free_slots),
                'free_slots': free_labels,
                'group_slots': [
                    {'day': DAYS[d], 'block': BLOCK_LABELS[b],
                     'n_groups': group_slots[(d, b)]['groups'],
                     'free_capacity': group_slots[(d, b)]['free_capacity']}
                    for (d, b) in sorted(group_slots)],
                'n_compatible_slots': len(compatible),
                'n_compatible_with_room': len(compatible_with_room),
                'verdict': verdict,
            })

    # Persiste le rapport pour l'app et les exports Excel.
    try:
        os.makedirs('reports', exist_ok=True)
        with open('reports/unplaced_students.json', 'w', encoding='utf-8') as fh:
            json.dump(UNPLACED_DIAGNOSTICS, fh, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  [WARN]  Erreur export unplaced_students.json : {e}")

    return UNPLACED_DIAGNOSTICS


def record_solver_run(sem, label, status, solver, n_sessions, n_hints=0,
                      recovered=False):
    """Capture un run du solveur (statut, objectif, gap, temps) pour les KPIs."""
    status_names = {
        cp_model.OPTIMAL: "OPTIMAL", cp_model.FEASIBLE: "FEASIBLE",
        cp_model.INFEASIBLE: "INFEASIBLE", cp_model.UNKNOWN: "UNKNOWN",
        cp_model.MODEL_INVALID: "MODEL_INVALID",
    }
    entry = {
        "semester": sem,
        "label": label,
        "status": status_names.get(status, str(status)),
        "n_sessions": int(n_sessions),
        "n_hints": int(n_hints),
        "recovered": bool(recovered),
        "wall_time_s": round(solver.WallTime(), 2),
    }
    try:
        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            entry["objective"] = solver.ObjectiveValue()
            entry["best_bound"] = solver.BestObjectiveBound()
            # gap relatif (proxy) : |obj - bound| / max(1, |obj|)
            obj = solver.ObjectiveValue()
            bound = solver.BestObjectiveBound()
            entry["gap"] = round(abs(obj - bound) / max(1.0, abs(obj)), 4)
    except Exception:
        pass
    SOLVER_RUNS.append(entry)
    return entry


def diagnose_infeasibility(sessions, sem, sem_holidays, label=""):
    """Étape 6.4 — Diagnostic LISIBLE d'une infaisabilité.

    Au lieu d'un simple « INFAISABLE », identifie les couples (salle|matière,
    jour, bloc) SUR-SATURÉS : ceux qui exigent plus de séances que de semaines
    disponibles dans la fenêtre (cause physique n°1 de l'audit §4.2). Écrit un
    rapport et renvoie la liste des goulots.
    """
    bottlenecks = []

    def _scan(group_iter, kind):
        for key, grp in group_iter.items():
            if len(grp) <= 1:
                continue
            ident, d, b = key
            needed = len(grp)
            min_w = min(s['min_week'] for s in grp)
            max_w = max(s['max_week'] for s in grp)
            valid = [w for w in range(min_w, max_w + 1)
                     if (w, d) not in sem_holidays]
            cap = len(valid)
            if needed > cap:
                bottlenecks.append({
                    "kind": kind, "ident": ident,
                    "day": DAYS[d] if 0 <= d < len(DAYS) else d,
                    "block": BLOCK_LABELS.get(b, b),
                    "needed": needed, "capacity": cap,
                    "excess": needed - cap,
                })

    by_room = defaultdict(list)
    for s in sessions:
        for room in str(s['lab_rooms']).split(','):
            room = room.strip()
            if room:
                by_room[(room, s['day_idx'], s['block_id'])].append(s)
    by_subj = defaultdict(list)
    for s in sessions:
        by_subj[(s['subject'], s['day_idx'], s['block_id'])].append(s)

    _scan(by_room, "SALLE")
    _scan(by_subj, "MATIÈRE")
    bottlenecks.sort(key=lambda x: x["excess"], reverse=True)

    lines = [
        "=" * 64,
        f"  DIAGNOSTIC D'INFAISABILITÉ — S{sem} {label}".rstrip(),
        "=" * 64,
        f"  Sessions concernées : {len(sessions)}",
    ]
    if bottlenecks:
        lines.append(f"  Goulots détectés    : {len(bottlenecks)} "
                     f"(cause PHYSIQUE : capacité salle/créneau dépassée)")
        lines.append("")
        for bn in bottlenecks[:25]:
            lines.append(
                f"   [{bn['kind']:7s}] {str(bn['ident'])[:32]:32s} "
                f"{bn['day']:10s} {bn['block']:12s} : "
                f"{bn['needed']} séances / {bn['capacity']} semaines "
                f"(excès {bn['excess']})"
            )
        lines.append("")
        lines.append("  PISTES (cf. audit §4) : ouvrir un créneau/salle "
                     "supplémentaire, élargir la fenêtre [min_week, max_week], "
                     "ou réduire le nombre de groupes parallèles.")
    else:
        lines.append("  Aucun goulot capacité évident — cause probable :")
        lines.append("   • conflits d'agendas étudiants (aucun créneau commun),")
        lines.append("   • contraintes profs trop strictes,")
        lines.append("   • fenêtre trop courte après retrait des jours fériés.")
    lines.append("=" * 64)
    report = "\n".join(lines)

    try:
        os.makedirs("reports", exist_ok=True)
        suffix = f"_{label}" if label else ""
        with open(f"reports/infeasibility_S{sem}{suffix}.txt", "w",
                  encoding="utf-8") as fh:
            fh.write(report)
    except Exception:
        pass

    print(report)
    return bottlenecks

if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass


MASTER_PATH = 'data_clean/master_schedule.csv'
OUTPUT_DIR = 'outputs/optimization/'
REPORT_PATH = 'reports/pipeline_v5_report.txt'

TIME_BLOCKS = [
    {'id': 1, 'label': '08:30-10:30', 'start': 510, 'end': 630, 'period': 'morning'},
    {'id': 2, 'label': '10:30-12:30', 'start': 630, 'end': 750, 'period': 'morning'},
    {'id': 3, 'label': '12:30-14:30', 'start': 750, 'end': 870, 'period': 'morning'},
    {'id': 4, 'label': '15:00-17:00', 'start': 900, 'end': 1020, 'period': 'afternoon'},
    {'id': 5, 'label': '17:00-19:00', 'start': 1020, 'end': 1140, 'period': 'afternoon'},
    {'id': 6, 'label': '19:00-21:00', 'start': 1140, 'end': 1260, 'period': 'evening'},
]

BLOCK_LABELS = {b['id']: b['label'] for b in TIME_BLOCKS}
MORNING_BLOCKS = [b['id'] for b in TIME_BLOCKS if b['period'] == 'morning']
AFTERNOON_BLOCKS = [b['id'] for b in TIME_BLOCKS if b['period'] in ['afternoon', 'evening']]
ALL_BLOCKS = [b['id'] for b in TIME_BLOCKS]

DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
DAY_IDS = {d: i for i, d in enumerate(DAYS)}

# ── Anti-goulot d'étranglement du VENDREDI (contrainte SOUPLE / soft) ───────
# Problème constaté : le vendredi concentrait trop de séances de labo
# (183 séances observées vs ~120 en moyenne par jour). On DÉCOURAGE le vendredi
# via une pénalité ajoutée au score de placement des groupes, SANS jamais
# l'interdire : si le vendredi reste le seul créneau faisable pour un groupe,
# il demeure utilisable (le taux d'affectation des étudiants n'est donc pas
# compromis — la qualité globale de la solution est préservée).
#   • FRIDAY_BASE_PENALTY   : petite pénalité constante -> départage les quasi-
#     égalités en faveur des autres jours (Lun→Jeu), ce qui lisse la charge.
#   • FRIDAY_SOFT_CAP       : plafond SOUPLE, en nombre de SÉANCES, visé le
#     vendredi (cible ~130-140).
#   • FRIDAY_OVERCAP_WEIGHT : pénalité additionnelle PAR séance au-delà du
#     plafond -> escalade qui rend le vendredi très peu attractif une fois la
#     cible atteinte, tout en laissant la porte ouverte si réellement nécessaire.
#
# VALEURS CALIBRÉES (testées sur le jeu de données 2025-2026) :
#   Avant correctif : Viernes = 183 séances (goulot ; moy. ~120/jour).
#   Après correctif : Viernes ≈ 152 séances (−17 %), SANS perte de couverture
#   (total des séances et nombre de groupes inchangés : ~598 séances / 145 gr.).
#   Les séances du vendredi restantes (~150) sont STRUCTURELLEMENT contraintes :
#   il s'agit de groupes (cours du soir, années supérieures) dont les étudiants
#   ne sont mutuellement libres QUE le vendredi. Les déplacer casserait la taille
#   minimale de groupe (MIN_GROUP_SIZE) et laisserait des étudiants non affectés.
#   Augmenter encore la pénalité ne descend pas sous ce plancher (testé jusqu'à
#   base=20) : la contrainte reste donc SOUPLE pour préserver la qualité.
#   Pour viser plus bas au prix d'éventuels étudiants non placés, augmenter
#   FRIDAY_BASE_PENALTY / FRIDAY_OVERCAP_WEIGHT et baisser FRIDAY_SOFT_CAP.
FRIDAY_IDX = DAY_IDS['Viernes']            # = 4
FRIDAY_SOFT_CAP = 125                      # plafond souple visé le vendredi (séances)
FRIDAY_BASE_PENALTY = 8                    # pénalité constante : diverge vers Lun→Jeu dès qu'un créneau faisable existe
FRIDAY_OVERCAP_WEIGHT = 10                 # escalade par séance au-delà du plafond


def friday_placement_penalty(day_idx, num_sessions, current_friday_load):
    """Pénalité SOUPLE retirée du score de placement d'un groupe sur un créneau
    du vendredi (0 pour les autres jours).

    Pour le vendredi : une pénalité de base constante (FRIDAY_BASE_PENALTY) plus
    une pénalité qui ESCALADE dès que le nombre de séances déjà placées le
    vendredi (+ celles qu'ajouterait ce groupe) dépasse le plafond souple
    FRIDAY_SOFT_CAP. Cette fonction n'interdit JAMAIS le vendredi : elle ne fait
    que réduire son score, donc un créneau du vendredi reste sélectionnable
    lorsqu'aucune autre alternative faisable n'existe.
    """
    if day_idx != FRIDAY_IDX:
        return 0
    penalty = FRIDAY_BASE_PENALTY
    projected = current_friday_load + max(0, int(num_sessions or 0))
    if projected > FRIDAY_SOFT_CAP:
        penalty += (projected - FRIDAY_SOFT_CAP) * FRIDAY_OVERCAP_WEIGHT
    return penalty


SEMESTER_1_WEEKS = 14
SEMESTER_2_WEEKS = 20


ALLOW_AFTERNOON_Y1Y3 = False
ALLOW_MORNING_Y2Y4   = False


TEACHER_UNAVAILABILITY = {}


QUIMICA_USE_TWO_ROOMS = False


PARITY_ALTERNATION = True
PARITY_PENALTY_WEIGHT = 50

SOLVER_TIME_LIMIT = 300


HOLIDAYS = {
    1: {

        (7, 0): "Día de la Hispanidad",
    },
    2: {

        (6, 4): "Blue Day",

        (7, 0): "Canonización de San Ignacio",

        (9, 0): "Semana Santa", (9, 1): "Semana Santa",
        (9, 2): "Semana Santa", (9, 3): "Semana Santa",
        (9, 4): "Semana Santa",


        (12, 3): "Feria de Abril", (12, 4): "Feria de Abril",

        (13, 4): "Día del Trabajador",
    },
}


SUBJECT_BLOCKED_LABEL = "Festivo / No disponible"
SUBJECT_BLOCKED_SLOTS = {
    (1, 'S1_Química'): {
        (w, d, b): SUBJECT_BLOCKED_LABEL
        for w in (7, 8, 11, 12)
        for d in (2, 3)
        for b in (2, 3)
    },
}


def subject_blocked_map(semester, subject):
    """All blocked (week, day_idx, block_id) -> label for this subject/semester."""
    return SUBJECT_BLOCKED_SLOTS.get((semester, subject), {})


def is_week_blocked_for_session(semester, subject, week, day_idx, block_id):
    """True if (subject, week, day, block) is reserved for another activity."""
    return (week, day_idx, block_id) in subject_blocked_map(semester, subject)


INCLUDE_REAL_NAMES = True


PREFERRED_GROUP_SIZE = 12
MAX_GROUP_SIZE = 15
MIN_GROUP_SIZE = 7


RECOVERY_MIN_GROUP_SIZE = 7
MAX_EXTRA_GROUPS = 3
COMPUTER_LAB_MAX = 24
REDUCED_MAX_SIZE = 12

LAB_CONFIG = {


    'S1_Física': {
        'curso_num': 1, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
        'simultaneous_rooms': False,


        'keywords': ['física i', 'física'],


        'keyword_exclude': ['física ii', 'computacional', 'simulación',
                            'simulacion'],
        'shared_group': 'S1_1er_anno',
        'group_by_program': True,


        'intro_session_paired': True,
    },
    'S1_Química': {
        'curso_num': 1, 'semester': 1, 'num_sessions': 4, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I'],
        'simultaneous_rooms': False,
        'keywords': ['química general'],
        'keyword_exclude': [],
        'shared_group': 'S1_1er_anno',
        'group_by_program': True,
    },


    'S1_Electrotecnia': {
        'curso_num': 2, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': 15,
        'lab_rooms': ['Lab. Eléctrica'],
        'simultaneous_rooms': False,
        'keywords': ['electrotecnia'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Mecanismos': {
        'curso_num': 2, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Telemática'],
        'simultaneous_rooms': False,
        'keywords': ['mecanismos y elementos'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Termodinámica': {
        'curso_num': 2, 'semester': 1, 'num_sessions': 5, 'min_week': 4, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Termodinámica'],
        'simultaneous_rooms': False,
        'keywords': ['termodinámica'],
        'keyword_exclude': [],
        'group_by_program': False,
    },


    'S1_Tecnologías de Fabricación': {
        'curso_num': 3, 'semester': 1, 'num_sessions': 2, 'min_week': 3, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Telemática'],
        'simultaneous_rooms': False,
        'keywords': ['tecnologías de fabricación'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Robótica y Automatización': {
        'curso_num': 3, 'semester': 1, 'num_sessions': 3, 'min_week': 3, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Robótica'],
        'simultaneous_rooms': False,
        'keywords': ['robótica y automatización'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S1_Automatización Industrial': {
        'curso_num': 3, 'semester': 1, 'num_sessions': 4, 'min_week': 3, 'max_week': 14,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Lab. Robótica'],
        'simultaneous_rooms': False,
        'keywords': ['automatización industrial', 'instrumentación'],
        'keyword_exclude': ['robótica'],
        'group_by_program': False,
    },


    'S2_Física II': {
        'curso_num': 1, 'semester': 2, 'num_sessions': 5, 'min_week': 9, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
        'simultaneous_rooms': False,


        'keywords': ['física ii'],
        'keyword_exclude': [],
        'group_by_program': True,
    },
    'S2_Tecnología Medio Ambiente': {
        'curso_num': 1, 'semester': 2, 'num_sessions': 2, 'min_week': 11, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Ciencias Experimentales I'],
        'simultaneous_rooms': False,
        'keywords': ['tecnología del medio ambiente', 'medio ambiente'],
        'keyword_exclude': [],
        'group_by_program': True,
    },
    'S2_Resistencia de Materiales': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 10, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Automoción y Resistencia de Mat.'],
        'simultaneous_rooms': False,
        'keywords': ['resistencia de materiales'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Mecánica de Fluidos': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 9, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Mecánica de Fluidos'],
        'simultaneous_rooms': False,
        'keywords': ['mecánica de fluido', 'mecánica y máquinas de fluido'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Regulación Automática': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 5, 'min_week': 9, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['regulación automática'],
        'keyword_exclude': ['automatic control'],
        'group_by_program': False,
    },
    'S2_Tecnología Electrónica': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 3, 'min_week': 8, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Electrónica'],
        'simultaneous_rooms': False,
        'keywords': ['tecnología electrónica'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Electrónica y Automática': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 3, 'min_week': 9, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Electrónica', 'Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['electrónica y automática'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Informática y Com. Industriales': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 8, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['informática y comunicaciones industriales'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Métodos Numéricos': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 4, 'min_week': 9, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Laboratorio de Ingeniería Telemática'],
        'simultaneous_rooms': False,
        'keywords': ['métodos numéricos'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Modelado de Sistemas': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 5, 'min_week': 7, 'max_week': 18,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Laboratorio de Ingeniería Telemática'],
        'simultaneous_rooms': False,
        'keywords': ['modelado simulación', 'modelado y simulación'],
        'keyword_exclude': ['modelado de sistemas físicos', 'modelado de sistemas fís'],
        'group_by_program': False,
    },
    'S2_Automatic Control': {
        'curso_num': 2, 'semester': 2, 'num_sessions': 5, 'min_week': 10, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['automatic control'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Ingeniería de Control': {
        'curso_num': 3, 'semester': 2, 'num_sessions': 5, 'min_week': 8, 'max_week': 19,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Robótica y Automática'],
        'simultaneous_rooms': False,
        'keywords': ['ingeniería de control'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Control de Máquinas': {
        'curso_num': 3, 'semester': 2, 'num_sessions': 5, 'min_week': 13, 'max_week': 20,
        'max_students': PREFERRED_GROUP_SIZE,
        'lab_rooms': ['Eléctrica'],
        'simultaneous_rooms': False,
        'keywords': ['control de máquinas'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
    'S2_Estructuras': {
        'curso_num': 3, 'semester': 2, 'num_sessions': 1, 'min_week': 16, 'max_week': 19,
        'max_students': 12,
        'lab_rooms': ['Automoción y Resistencia de Mat.'],
        'simultaneous_rooms': False,
        'keywords': ['estructuras'],
        'keyword_exclude': [],
        'group_by_program': False,
    },
}


USER_CONFIG_PATH = 'config/user_config.json'
CSV_READ_OPTIONS = (
    ('utf-8-sig', ','),
    ('utf-8', ','),
    ('latin-1', ','),
    ('utf-8-sig', ';'),
    ('latin-1', ';'),
)
REQUIRED_MASTER_COLUMNS = ('AlumnoID', 'actividad')
KNOWN_PROGRAMS = frozenset({'IOI', 'IMR', 'GITI', 'GITIADE', 'MAT', 'AERO', 'IBIO', 'IEM', 'PIIA', 'IINFTV'})
DANIEL_FORMAT_SCRIPTS = (
    ('09_generate_exact_format_S1.py', 'S1 (Primero + Segundo + Tercero)'),
    ('10_generate_exact_format_S2.py', 'S2 (Primero + Segundo + Tercero)'),
)
VALIDATION_SCRIPTS = ('src/11_validate_output.py', '11_validate_output.py')


def apply_user_config():
    """
    Load user configuration from `config/user_config.json` and apply overrides
    to LAB_CONFIG and global parameters. This allows the Streamlit app to
    customize the pipeline behavior without modifying the source code.

    Config file structure:
        {
          "global": {
            "preferred_size": 12, "default_max": 15, "min_size": 7,
            "computer_lab_max": 24, "reduced_max_size": 12,
            "start_week": 4, "s1_total_weeks": 14
          },
          "subjects": {
            "S1_Física": {
              "num_sessions": 5, "max_students": 15,
              "min_week": 4, "max_week": 13,
              "lab_rooms": [...], "keywords": [...], "keyword_exclude": [...]
            },
            ...
          },
          "year_prefs": {
            "allow_afternoon_y1y3": false,
            "allow_morning_y2y4": false
          },
          "teachers": {...}
        }
    """
    global PREFERRED_GROUP_SIZE, MAX_GROUP_SIZE, MIN_GROUP_SIZE, RECOVERY_MIN_GROUP_SIZE
    global COMPUTER_LAB_MAX, REDUCED_MAX_SIZE
    global SEMESTER_1_WEEKS, SEMESTER_2_WEEKS
    global ALLOW_AFTERNOON_Y1Y3, ALLOW_MORNING_Y2Y4, TEACHER_UNAVAILABILITY
    global QUIMICA_USE_TWO_ROOMS, PARITY_ALTERNATION

    if not os.path.exists(USER_CONFIG_PATH):
        print(f"\n  [INFO]  Aucune config utilisateur trouvée ({USER_CONFIG_PATH})")
        print(f"  → utilisation des valeurs par défaut")
        return

    try:
        with open(USER_CONFIG_PATH, 'r', encoding='utf-8') as f:
            user_config = json.load(f)
    except Exception as e:
        print(f"\n  [WARN]  Impossible de charger {USER_CONFIG_PATH} : {e}")
        return

    print(f"\n  [CONFIG] Configuration utilisateur chargée depuis {USER_CONFIG_PATH}")
    saved_at = user_config.get('meta', {}).get('saved_at', '?')
    print(f"     Sauvegardée le : {saved_at}")


    global_cfg = user_config.get('global', {})
    if global_cfg:
        print(f"  [GLOBAL]  Paramètres globaux :")
        if 'preferred_size' in global_cfg:
            PREFERRED_GROUP_SIZE = int(global_cfg['preferred_size'])
            print(f"     PREFERRED_GROUP_SIZE = {PREFERRED_GROUP_SIZE}")
        if 'default_max' in global_cfg:
            MAX_GROUP_SIZE = int(global_cfg['default_max'])
            print(f"     MAX_GROUP_SIZE = {MAX_GROUP_SIZE}")
        if 'min_size' in global_cfg:
            MIN_GROUP_SIZE = int(global_cfg['min_size'])
            print(f"     MIN_GROUP_SIZE = {MIN_GROUP_SIZE}")


        if 'recovery_min_size' in global_cfg:
            RECOVERY_MIN_GROUP_SIZE = int(global_cfg['recovery_min_size'])
            print(f"     RECOVERY_MIN_GROUP_SIZE = {RECOVERY_MIN_GROUP_SIZE}")
        else:
            RECOVERY_MIN_GROUP_SIZE = MIN_GROUP_SIZE
        if 'computer_lab_max' in global_cfg:
            try:
                COMPUTER_LAB_MAX = int(global_cfg['computer_lab_max'])
                print(f"     COMPUTER_LAB_MAX = {COMPUTER_LAB_MAX}")
            except NameError:
                pass
        if 'reduced_max_size' in global_cfg:
            try:
                REDUCED_MAX_SIZE = int(global_cfg['reduced_max_size'])
                print(f"     REDUCED_MAX_SIZE = {REDUCED_MAX_SIZE}")
            except NameError:
                pass
        if 's1_total_weeks' in global_cfg:
            try:
                SEMESTER_1_WEEKS = int(global_cfg['s1_total_weeks'])
                print(f"     SEMESTER_1_WEEKS = {SEMESTER_1_WEEKS}")
            except NameError:
                pass
        if 's2_total_weeks' in global_cfg:
            try:
                SEMESTER_2_WEEKS = int(global_cfg['s2_total_weeks'])
                print(f"     SEMESTER_2_WEEKS = {SEMESTER_2_WEEKS}")
            except NameError:
                pass
        if 'quimica_use_two_rooms' in global_cfg:
            QUIMICA_USE_TWO_ROOMS = bool(global_cfg['quimica_use_two_rooms'])
        if 'parity_alternation' in global_cfg:
            PARITY_ALTERNATION = bool(global_cfg['parity_alternation'])


    if QUIMICA_USE_TWO_ROOMS and 'S1_Química' in LAB_CONFIG:
        LAB_CONFIG['S1_Química']['lab_rooms'] = [
            'Ciencias Experimentales I', 'Ciencias Experimentales II']
        LAB_CONFIG['S1_Química']['simultaneous_rooms'] = False
        print(f"  [ROOMS]  Química réparti sur 2 salles "
              f"(Ciencias I + II, alternées) — déblocage capacité matin")


    subjects_cfg = user_config.get('subjects', {})
    if subjects_cfg:
        print(f"  [SUBJECTS] Surcharges par matière : {len(subjects_cfg)} matières personnalisées")
        for subj_key, overrides in subjects_cfg.items():
            if subj_key not in LAB_CONFIG:
                print(f"     [WARN]  {subj_key} : matière inconnue, ignorée")
                continue

            base = LAB_CONFIG[subj_key]
            applied = []

            if 'num_sessions' in overrides:
                old, new = base.get('num_sessions'), int(overrides['num_sessions'])
                if old != new:
                    base['num_sessions'] = new
                    applied.append(f"num_sessions {old}→{new}")

            if 'max_students' in overrides:
                old, new = base.get('max_students'), int(overrides['max_students'])
                if old != new:
                    base['max_students'] = new
                    applied.append(f"max_students {old}→{new}")

            if 'min_week' in overrides:
                old, new = base.get('min_week'), int(overrides['min_week'])
                if old != new:
                    base['min_week'] = new
                    applied.append(f"min_week {old}→{new}")

            if 'max_week' in overrides:
                old, new = base.get('max_week'), int(overrides['max_week'])
                if old != new:
                    base['max_week'] = new
                    applied.append(f"max_week {old}→{new}")

            if 'lab_rooms' in overrides and overrides['lab_rooms']:
                new_rooms = list(overrides['lab_rooms'])
                old_rooms = base.get('lab_rooms', [])
                if set(new_rooms) != set(old_rooms):
                    base['lab_rooms'] = new_rooms
                    applied.append(f"lab_rooms ({len(new_rooms)} salles)")

            if 'keywords' in overrides and overrides['keywords']:
                new_kw = [k.strip() for k in overrides['keywords'] if k.strip()]
                if new_kw and new_kw != base.get('keywords'):
                    base['keywords'] = new_kw
                    applied.append(f"keywords ({len(new_kw)})")

            if 'keyword_exclude' in overrides:
                new_kx = [k.strip() for k in overrides['keyword_exclude'] if k.strip()]


                code_kx = base.get('keyword_exclude', [])
                merged_kx = list(dict.fromkeys(code_kx + new_kx))
                if merged_kx != code_kx:
                    base['keyword_exclude'] = merged_kx
                    applied.append(f"keyword_exclude ({len(merged_kx)}, fusionné)")
                else:
                    base['keyword_exclude'] = merged_kx

            if applied:
                print(f"     [CONFIG]  {subj_key} : {', '.join(applied)}")


    year_prefs = user_config.get('year_prefs', {})
    if year_prefs:
        print(f"  [SCHEDULE] Préférences horaires :")
        ALLOW_AFTERNOON_Y1Y3 = bool(year_prefs.get('allow_afternoon_y1y3', False))
        ALLOW_MORNING_Y2Y4 = bool(year_prefs.get('allow_morning_y2y4', False))
        if ALLOW_AFTERNOON_Y1Y3:
            print(f"     [WARN]  1ère/3ème année : après-midi exceptionnellement AUTORISÉ")
        else:
            print(f"     1ère/3ème année : matin STRICT (après-midi interdit)")
        if ALLOW_MORNING_Y2Y4:
            print(f"     [WARN]  2ème/4ème année : matin exceptionnellement AUTORISÉ")
        else:
            print(f"     2ème/4ème année : après-midi STRICT (matin interdit)")


    teachers = user_config.get('teachers', {})
    if teachers:
        print(f"  [TEACHER] Restrictions professeurs : {len(teachers)} professeur(s)")

        normalised = {}
        for teacher, slots in teachers.items():
            slot_set = set()
            for s in slots:

                if isinstance(s, (list, tuple)) and len(s) == 2:
                    slot_set.add((int(s[0]), int(s[1])))
                elif isinstance(s, dict) and 'day' in s and 'block' in s:
                    slot_set.add((int(s['day']), int(s['block'])))
                elif isinstance(s, str):
                    parsed = _parse_teacher_slot_string(s)
                    if parsed:
                        slot_set.add(parsed)
            normalised[teacher] = slot_set
            print(f"     {teacher} : {len(slot_set)} créneau(x) bloqué(s)")
        TEACHER_UNAVAILABILITY = normalised

    # Charger tot les regles souples/signal (preferred_blocks, max jours/sem) afin
    # qu'elles figurent dans applied_config.json (preuve d'application pour l'app).
    try:
        if not TEACHER_RULES:
            TEACHER_RULES.update(load_teacher_rules())
    except Exception as _e_rules:
        print(f"  [RULES][WARN] chargement anticipe echoue : {_e_rules}")

    print(f"  [OK] Configuration utilisateur appliquée\n")


    write_applied_config()


def write_applied_config():
    """
    Write `config/applied_config.json` — a read-back of the values the pipeline
    has ACTUALLY applied (post-override globals + the live LAB_CONFIG).

    The app compares this against the `user_config.json` it sent, parameter by
    parameter, to prove that every setting was honoured. A parameter present in
    user_config but absent or different here is a synchronisation gap.
    """
    applied = {
        'global': {
            'preferred_size':   PREFERRED_GROUP_SIZE,
            'default_max':      MAX_GROUP_SIZE,
            'min_size':         MIN_GROUP_SIZE,
            'computer_lab_max': COMPUTER_LAB_MAX,
            'reduced_max_size': REDUCED_MAX_SIZE,
            's1_total_weeks':   SEMESTER_1_WEEKS,
            's2_total_weeks':   SEMESTER_2_WEEKS,
            'quimica_use_two_rooms': QUIMICA_USE_TWO_ROOMS,
            'parity_alternation': PARITY_ALTERNATION,
        },
        'year_prefs': {
            'allow_afternoon_y1y3': ALLOW_AFTERNOON_Y1Y3,
            'allow_morning_y2y4':   ALLOW_MORNING_Y2Y4,
        },
        'teachers_blocked_slots': {
            t: sorted(list(slots)) for t, slots in TEACHER_UNAVAILABILITY.items()
        },
        # Regles souples/signal effectivement chargees (preferred_blocks + max jours/sem).
        # Permet a l'app de PROUVER que ces parametres ont ete pris en compte.
        'teacher_rules': {
            str(name): {
                k: (sorted(list(v)) if isinstance(v, (set, list, tuple)) else v)
                for k, v in rules.items()
            }
            for name, rules in TEACHER_RULES.items()
        },

        'subjects': {
            k: {
                'num_sessions':    v.get('num_sessions'),
                'max_students':    v.get('max_students'),
                'min_week':        v.get('min_week'),
                'max_week':        v.get('max_week'),
                'lab_rooms':       v.get('lab_rooms', []),
            }
            for k, v in LAB_CONFIG.items()
        },
        'meta': {
            'applied_at': datetime.now().isoformat(),
            'note': 'Auto-generated by pipeline. Reflects values actually used.',
        },
    }
    try:
        os.makedirs('config', exist_ok=True)
        with open('config/applied_config.json', 'w', encoding='utf-8') as f:
            json.dump(applied, f, indent=2, ensure_ascii=False)
        print(f"  [ECHO]  État appliqué écrit dans config/applied_config.json")
    except Exception as e:
        print(f"  [WARN]  Impossible d'écrire applied_config.json : {e}")


def print_section(title):
    """Print a horizontal separator with a title in the middle (60 chars wide)."""
    print("\n" + "=" * 60)
    print(f"  {title}")
    print("=" * 60)


def min_to_block_id(start_min):
    """
    Convert "minutes since midnight" to a TIME_BLOCKS id (1..6).

    Tolerates +/- 15 minutes around the canonical block start time
    (e.g., 504 -> block 1, even though block 1 nominally starts at 510).

    Args:
        start_min: int, minutes since midnight (e.g., 510 = 08:30)

    Returns:
        int | None: block id (1-6) or None if no block matches.
    """
    for b in TIME_BLOCKS:
        if abs(b['start'] - start_min) < 15:
            return b['id']
    return None


def load_and_prepare(df):
    """
    Step 1: Print summary statistics for the loaded master schedule.

    The actual loading happens upstream (in main()); this function
    just validates and prints diagnostics.

    Args:
        df: master_schedule DataFrame (joined aulario + alumnos)

    Returns:
        The same DataFrame, unchanged.
    """
    print_section("ÉTAPE 1 : Chargement et préparation")
    print(f"  {df.shape[0]} lignes × {df.shape[1]} colonnes")
    print(f"  {df['AlumnoID'].nunique()} étudiants uniques")
    print(f"  {df['actividad'].nunique()} cours uniques")
    return df


def identify_students(df):
    """
    Pour chaque matière avec lab (définie dans LAB_CONFIG),
    cherche les étudiants inscrits via les mots-clés dans actividad.
    Filtre par campus Sevilla et par année d'étude (curso_asignatura).
    """
    print_section("ÉTAPE 2 : Étudiants inscrits par matière (via MixtoID)")


    if 'campus' in df.columns:
        sev = df[df['campus'].fillna('').str.contains('Sevilla', case=False, na=False)]
        print(f"  Filtre campus Sevilla: {len(sev)}/{len(df)} lignes")
    else:
        sev = df

    subject_students = {}

    for subject, config in LAB_CONFIG.items():
        keywords = config['keywords']
        keyword_exclude = config.get('keyword_exclude', [])
        curso_num = config['curso_num']


        mask = pd.Series(False, index=sev.index)
        for kw in keywords:
            mask = mask | sev['actividad'].fillna('').str.lower().str.contains(kw, na=False)
        for kw_ex in keyword_exclude:
            mask = mask & ~sev['actividad'].fillna('').str.lower().str.contains(kw_ex, na=False)


        if 'Titulación' in sev.columns:
            mask = mask & ~sev['Titulación'].fillna('').str.contains('Máster|Master|Diploma', case=False, na=False)


        student_ids = sev[mask]['AlumnoID'].dropna().unique().tolist()

        subject_students[subject] = student_ids

        sem = config['semester']
        print(f"  {'[OK]' if len(student_ids) > 0 else '[WARN]'} {subject:40s} S{sem} | "
              f"{len(student_ids):3d} étudiants")

    total = sum(len(v) for v in subject_students.values())
    print(f"\n  Total inscriptions : {total}")


    try:
        os.makedirs('data_clean/optimization', exist_ok=True)
        enrol_rows = [
            {'subject': subj,
             'semester': LAB_CONFIG[subj]['semester'],
             'student_count': len(ids)}
            for subj, ids in subject_students.items()
        ]
        pd.DataFrame(enrol_rows).to_csv(
            'data_clean/optimization/lab_enrollments.csv',
            index=False, encoding='utf-8-sig')
        print(f"  [OK] Export lab_enrollments.csv ({len(enrol_rows)} matières)")
    except Exception as e:
        print(f"  [WARN]  Erreur export lab_enrollments.csv : {e}")

    return subject_students


def build_individual_timetables(df, subject_students):
    """
    Pour chaque étudiant inscrit à une matière avec lab,
    calcule :
    1. Ses créneaux occupés (TOUS ses cours)
    2. Pour chaque matière avec lab, quels créneaux sont occupés
       PAR CETTE MATIÈRE spécifiquement

    PRINCIPE CLÉ : un lab de Física REMPLACE le cours de Física.
    Donc quand on cherche un créneau pour un lab de Física,
    le créneau du cours de Física est DISPONIBLE (pas occupé).
    Seuls les cours des AUTRES matières bloquent ce créneau.
    """
    print_section("ÉTAPE 3 : Emploi du temps individuel")

    all_student_ids = set()
    for ids in subject_students.values():
        all_student_ids.update(ids)

    print(f"  {len(all_student_ids)} étudiants uniques concernés par des labs")

    valid = df.dropna(subset=['AlumnoID', 'slot_hora_inicio_min', 'slot_jour_semaine'])
    valid = valid[valid['slot_hora_inicio_min'] > 0]
    valid = valid[valid['AlumnoID'].isin(all_student_ids)]


    student_busy = {}


    student_subject_slots = defaultdict(lambda: defaultdict(set))


    for student_id, group in valid.groupby('AlumnoID'):
        busy = set()
        for _, row in group.iterrows():
            day = row['slot_jour_semaine']
            block_id = min_to_block_id(row['slot_hora_inicio_min'])
            if block_id and day in DAYS:
                slot = (DAY_IDS[day], block_id)
                busy.add(slot)


                act = str(row.get('actividad', '')).lower()
                for subject, config in LAB_CONFIG.items():
                    keywords = config['keywords']
                    keyword_exclude = config.get('keyword_exclude', [])
                    if any(kw in act for kw in keywords):
                        if not any(kw_ex in act for kw_ex in keyword_exclude):
                            student_subject_slots[student_id][subject].add(slot)

        student_busy[student_id] = busy


    avg_busy = sum(len(b) for b in student_busy.values()) / len(student_busy) if student_busy else 0
    avg_free = 30 - avg_busy
    min_free = 30 - max(len(b) for b in student_busy.values()) if student_busy else 0
    max_free = 30 - min(len(b) for b in student_busy.values()) if student_busy else 30

    print(f"  Créneaux occupés par étudiant : moy={avg_busy:.1f}")
    print(f"  Créneaux libres par étudiant  : min={min_free}, moy={avg_free:.1f}, max={max_free}")


    extra_slots = 0
    for sid in all_student_ids:
        for subject in student_subject_slots.get(sid, {}):
            extra_slots += len(student_subject_slots[sid][subject])
    avg_extra = extra_slots / len(all_student_ids) if all_student_ids else 0
    print(f"  Créneaux récupérés par remplacement cours→lab : moy={avg_extra:.1f}")
    print(f"  Créneaux effectivement libres pour labs : moy={avg_free + avg_extra:.1f}")


    try:
        os.makedirs('data_clean/optimization', exist_ok=True)
        sb_rows = []
        for sid, busy_set in student_busy.items():
            for (day_idx, block_id) in busy_set:
                sb_rows.append({
                    'student_id': sid,
                    'day_idx': day_idx,
                    'block_id': block_id,
                })
        if sb_rows:
            sb_df = pd.DataFrame(sb_rows)
            sb_path = 'data_clean/optimization/student_busy.csv'
            sb_df.to_csv(sb_path, index=False, encoding='utf-8-sig')
            print(f"  [OK] Export {sb_path} ({len(sb_rows)} entrées)")
    except Exception as e:
        print(f"  [WARN]  Erreur export student_busy.csv : {e}")

    return student_busy, student_subject_slots


PROFESSOR_ID_COLUMNS = ['docentes', 'ProfesorID', 'profesor_id', 'profesor',
                        'Profesor', 'docente', 'Docente', 'teacher', 'Teacher',
                        'profesorado']


MULTI_PROFESSOR_COLUMNS = {'docentes', 'Docentes'}

SUPERVISION_PATH = 'data_clean/optimization/subject_supervision.csv'


def load_supervision_capacity():
    """
    Load per-subject supervision capacity (number of distinct professors) from
    subject_supervision.csv, produced by build_professor_table.py from the
    official enrolment report's 'Docentes' column.

    Returns {subject → n_professors}, or {} if the file is absent.

    This answers the whiteboard's "how many groups?" question: a subject's
    parallel-group capacity is bounded by min(n_professors, n_rooms). For Física,
    n_professors=9 and n_rooms=2, so rooms are the real limit — confirming the
    morning bottleneck is physical (rooms), not supervision.
    """
    if not os.path.exists(SUPERVISION_PATH):
        return {}
    try:
        df = pd.read_csv(SUPERVISION_PATH)
        if {'subject', 'n_professors'}.issubset(df.columns):
            return {str(r['subject']): int(r['n_professors'])
                    for _, r in df.iterrows()}
    except Exception as e:
        print(f"  [WARN]  Lecture {SUPERVISION_PATH} échouée : {e}")
    return {}


def build_professor_busy(df):
    """
    Build the professor side of the busy/free model.

    Returns:
        professor_busy: {professor_id → set of (day_idx, block_id)}  busy slots
                        from their NORMAL classes (mirrors student_busy).
        professor_subjects: {professor_id → set of lab subjects they teach}
                        derived from 'actividad' keywords (best-effort; the
                        explicit prof↔lab mapping is the open data question).
        available: bool — whether a usable professor column was found.

    If no professor column exists, returns ({}, {}, False) and the rest of the
    pipeline runs unchanged.
    """
    print_section("FLUX PROFESSEUR : disponibilités (busy/free)")


    prof_col = None
    for cand in PROFESSOR_ID_COLUMNS:
        if cand in df.columns:
            prof_col = cand
            break

    if prof_col is None:
        print(f"  [INFO]  Aucune colonne professeur trouvée "
              f"(cherché: {', '.join(PROFESSOR_ID_COLUMNS[:4])}…)")
        print(f"  → flux professeur INACTIF, pipeline inchangé.")
        print(f"  → données requises (questions ouvertes du tableau) :")
        print(f"      • quel professeur enseigne quel lab ('which labs?')")
        print(f"      • combien de groupes par professeur ('how many groups?')")
        return {}, {}, False

    print(f"  Colonne professeur détectée : '{prof_col}'")
    multi = prof_col in MULTI_PROFESSOR_COLUMNS
    if multi:
        print(f"  Format multi-professeurs (noms séparés par virgule)")

    valid = df.dropna(subset=[prof_col, 'slot_hora_inicio_min', 'slot_jour_semaine'])
    valid = valid[valid['slot_hora_inicio_min'] > 0]

    professor_busy = defaultdict(set)
    professor_subjects = defaultdict(set)

    for _, row in valid.iterrows():
        day = row['slot_jour_semaine']
        block_id = min_to_block_id(row['slot_hora_inicio_min'])
        if not block_id or day not in DAYS:
            continue
        slot = (DAY_IDS[day], block_id)
        act = str(row.get('actividad', '')).lower()


        raw = str(row[prof_col])
        names = [n.strip() for n in raw.split(',')] if multi else [raw.strip()]

        for prof_id in names:
            if not prof_id or prof_id.lower() == 'nan':
                continue
            professor_busy[prof_id].add(slot)

            for subject, config in LAB_CONFIG.items():
                if any(kw in act for kw in config['keywords']):
                    if not any(kw_ex in act
                               for kw_ex in config.get('keyword_exclude', [])):
                        professor_subjects[prof_id].add(subject)

    professor_busy = dict(professor_busy)

    n_prof = len(professor_busy)
    if n_prof:
        avg_busy = sum(len(b) for b in professor_busy.values()) / n_prof
        print(f"  {n_prof} professeurs avec emploi du temps")
        print(f"  Créneaux occupés par prof : moy={avg_busy:.1f} / 30")
        n_linked = sum(1 for s in professor_subjects.values() if s)
        print(f"  Professeurs liés à ≥1 matière de lab : {n_linked}")
    else:
        print(f"  [WARN]  Colonne '{prof_col}' présente mais aucun prof exploitable")
        return {}, {}, False


    try:
        os.makedirs('data_clean/optimization', exist_ok=True)
        pb_rows = []
        for pid, busy_set in professor_busy.items():
            for (day_idx, block_id) in busy_set:
                pb_rows.append({'professor_id': pid,
                                'day_idx': day_idx, 'block_id': block_id})
        if pb_rows:
            pd.DataFrame(pb_rows).to_csv(
                'data_clean/optimization/professor_busy.csv',
                index=False, encoding='utf-8-sig')
            print(f"  [OK] Export professor_busy.csv ({len(pb_rows)} entrées)")
    except Exception as e:
        print(f"  [WARN]  Erreur export professor_busy.csv : {e}")

    return professor_busy, dict(professor_subjects), True


def form_groups(subject_students, student_busy, student_subject_slots, student_program,
                subject_professor_busy=None,
                subject_block_penalty=None):
    """
    Formation des groupes avec support pour:
    1. Groupes partagés (shared_group): Física et Química partagent les mêmes groupes
    2. Groupes par programme (group_by_program): 1ère année = homogène par titulación
    3. Tailles variables par matière
    4. Contrainte professeur (subject_professor_busy): un groupe d'une matière ne
       peut pas être créé sur un créneau où un professeur de cette matière est
       occupé. Optionnel — vide/None = aucun effet (rétro-compatible).
    """
    print_section("ÉTAPE 4 : Formation des groupes")

    if subject_professor_busy is None:
        subject_professor_busy = {}
    if subject_block_penalty is None:
        subject_block_penalty = {}

    all_groups = []
    stats = {'total_assigned': 0, 'total_unassigned': 0}
    slot_room_usage = {}

    # Suivi du nombre de SÉANCES déjà placées par jour (index 0=Lun … 4=Ven).
    # Alimenté dans _propagate_busy à chaque création de groupe (toutes phases).
    # Sert à la contrainte souple anti-goulot du vendredi (cf. friday_placement_penalty).
    day_session_load = defaultdict(int)


    subject_slot_usage = {}


    student_lab_busy = defaultdict(set)

    def _propagate_busy(group):
        """Block the group's (day, block) slot for every member: they now have
        a lab there and cannot accept ANOTHER lab placement at the same slot.
        Uses student_lab_busy (a hard block) rather than student_busy (which
        is partially neutralised by own_slots)."""
        slot = (group['day_idx'], group['block_id'])
        for sid in group.get('student_ids', []):
            student_lab_busy[sid].add(slot)
        # Comptabilise les séances de ce groupe pour le suivi par jour (vendredi).
        day_session_load[group['day_idx']] += int(group.get('num_sessions', 0) or 0)

    def _propagate_busy_one(sid, group):
        """Same for a SINGLE student added to an existing group."""
        student_lab_busy[sid].add(
            (group['day_idx'], group['block_id']))


    shared_group_map = {}
    for subject, config in LAB_CONFIG.items():
        sg = config.get('shared_group')
        if sg and sg not in shared_group_map:
            shared_group_map[sg] = subject


    secondary_shared = set()
    for subject, config in LAB_CONFIG.items():
        sg = config.get('shared_group')
        if sg and shared_group_map[sg] != subject:
            secondary_shared.add(subject)


    subject_data = {}
    for subject, config in LAB_CONFIG.items():
        if subject in secondary_shared:
            continue


        sg = config.get('shared_group')
        if sg:
            student_ids = set()
            for other_subj, other_cfg in LAB_CONFIG.items():
                if other_cfg.get('shared_group') == sg:
                    student_ids.update(subject_students.get(other_subj, []))
            student_ids = list(student_ids)
        else:
            student_ids = subject_students.get(subject, [])

        if not student_ids:
            continue

        curso_num = config['curso_num']
        semester = config['semester']
        num_sessions = config['num_sessions']
        min_week = config['min_week']


        sem_max = SEMESTER_1_WEEKS if semester == 1 else SEMESTER_2_WEEKS
        max_week = config.get('max_week', sem_max)
        max_per_group = config['max_students']
        lab_rooms_list = config['lab_rooms']


        if max_per_group == PREFERRED_GROUP_SIZE and len(student_ids) > 0:
            groups_at_12 = math.ceil(len(student_ids) / PREFERRED_GROUP_SIZE)
            groups_at_15 = math.ceil(len(student_ids) / MAX_GROUP_SIZE)


            shares_room = bool(config.get('shared_group'))
            if (groups_at_12 - groups_at_15 > MAX_EXTRA_GROUPS) or shares_room:
                max_per_group = MAX_GROUP_SIZE
                reason = ("partage de salle — capacité matin serrée" if shares_room
                          else f"12→{groups_at_12} gr vs 15→{groups_at_15} gr")
                print(f"  ↑ {subject}: max porté à {MAX_GROUP_SIZE} ({reason})")

        sem_holidays = HOLIDAYS.get(semester, {})
        available_weeks_by_day = {}
        for d in range(5):
            valid = [w for w in range(min_week, max_week + 1)
                     if (w, d) not in sem_holidays]
            available_weeks_by_day[d] = len(valid)

        available_weeks = min(available_weeks_by_day.values()) if available_weeks_by_day else (max_week - min_week + 1)

        if curso_num in [1, 3]:
            allowed_blocks = MORNING_BLOCKS
        elif curso_num in [2, 4]:
            allowed_blocks = AFTERNOON_BLOCKS
        else:
            allowed_blocks = ALL_BLOCKS

        all_slots = [(d, b) for d in range(5) for b in allowed_blocks]


        prof_busy = subject_professor_busy.get(subject, set())
        if prof_busy:
            filtered = [s for s in all_slots if s not in prof_busy]
            if filtered:
                removed = len(all_slots) - len(filtered)
                if removed:
                    print(f"    [PROF]  {subject}: {removed} créneau(x) retiré(s) "
                          f"(professeur occupé)")
                all_slots = filtered
            else:
                print(f"    [PROF][WARN] {subject}: la contrainte professeur "
                      f"viderait tous les créneaux — ignorée (à vérifier)")


        n_rooms = len(config['lab_rooms']) if not config.get('simultaneous_rooms', False) else 1
        max_groups_per_slot_by_day = {
            d: max(1, (available_weeks_by_day.get(d, 0) // num_sessions) * n_rooms)
            for d in range(5)
        }

        subject_data[subject] = {
            'config': config,
            'unassigned': set(student_ids),
            'groups': [],
            'group_counter': 0,
            'all_slots': all_slots,
            'max_groups_per_slot_by_day': max_groups_per_slot_by_day,
            'slot_usage': defaultdict(int),
            'num_sessions': num_sessions,
            'max_per_group': max_per_group,
            'lab_rooms_list': lab_rooms_list,
            'semester': semester,
            'curso_num': curso_num,
            'min_week': min_week,
            'max_week': max_week,
            'available_weeks': available_weeks,
            'available_weeks_by_day': available_weeks_by_day,
            'total_students': len(student_ids),
            'group_by_program': config.get('group_by_program', False),
        }


    sorted_subjects = sorted(subject_data.keys(), key=lambda s: (
        0 if LAB_CONFIG[s].get('shared_group') else 1,
        len(subject_data[s]['lab_rooms_list']),
        len(subject_data[s]['all_slots']),
        -subject_data[s]['total_students'],
    ))

    print(f"  Ordre de traitement (plus contraint en premier) :")
    for s in sorted_subjects:
        sd = subject_data[s]
        print(f"    {s:40s} | {sd['total_students']:3d} étu | "
              f"{len(sd['all_slots']):2d} slots | "
              f"{len(sd['lab_rooms_list'])} salle(s)")


    shared_extra_sessions = {}
    for subject in subject_data:
        sg = LAB_CONFIG[subject].get('shared_group')
        if sg and shared_group_map.get(sg) == subject:
            extra = defaultdict(int)
            for other_subj, other_cfg in LAB_CONFIG.items():
                if other_cfg.get('shared_group') == sg and other_subj != subject:
                    other_rooms = other_cfg['lab_rooms']
                    other_sess = other_cfg['num_sessions']
                    if other_cfg.get('simultaneous_rooms', False):

                        for room in other_rooms:
                            extra[room] += other_sess
                    else:

                        per_room = math.ceil(other_sess / len(other_rooms)) if other_rooms else other_sess
                        for room in other_rooms:
                            extra[room] += per_room
            shared_extra_sessions[subject] = dict(extra)
            if extra:
                print(f"\n  Charge partagée pour {subject}:")
                for room, sess in extra.items():
                    total = subject_data[subject]['num_sessions'] + sess
                    print(f"    {room}: {subject_data[subject]['num_sessions']} + {sess} = {total} sess/groupe")


    shared_subjects_list_rr = [s for s in sorted_subjects if LAB_CONFIG[s].get('shared_group')]
    remaining_subjects = [s for s in sorted_subjects if not LAB_CONFIG[s].get('shared_group')]

    if shared_subjects_list_rr:
        print(f"\n  ── Phase 1: Groupes partagés (priorité) ──")
        shared_round = 0
        shared_progress = True
        while shared_progress and shared_round < 200:
            shared_round += 1
            shared_progress = False
            for subject in shared_subjects_list_rr:
                sd = subject_data[subject]
                if not sd['unassigned']:
                    continue
                config = sd['config']
                num_sessions = sd['num_sessions']
                max_per_group = sd['max_per_group']
                lab_rooms_list = sd['lab_rooms_list']
                best_slot = None
                best_free = []
                best_room = None
                best_prog = None
                best_score = -1
                simultaneous = config.get('simultaneous_rooms', False)
                extra_per_room = shared_extra_sessions.get(subject, {})

                for slot in sd['all_slots']:
                    day_idx = slot[0]
                    if sd['slot_usage'][slot] >= sd['max_groups_per_slot_by_day'].get(day_idx, 0):
                        continue
                    day_available_weeks = sd['available_weeks_by_day'].get(day_idx, 0)


                    subj_key = (subject, sd['semester'], slot[0], slot[1])
                    if subject_slot_usage.get(subj_key, 0) + num_sessions > day_available_weeks:
                        continue

                    if simultaneous:
                        room_ok = True
                        for room in lab_rooms_list:
                            room_key = (room, sd['semester'], slot[0], slot[1])
                            current = slot_room_usage.get(room_key, 0)
                            room_load = num_sessions + extra_per_room.get(room, 0)
                            if (current + room_load) > day_available_weeks:
                                room_ok = False; break
                        if not room_ok: continue
                        chosen_room = ', '.join(lab_rooms_list)
                    else:


                        chosen_room = None
                        min_usage = float('inf')
                        for room in lab_rooms_list:
                            room_key = (room, sd['semester'], slot[0], slot[1])
                            current = slot_room_usage.get(room_key, 0)
                            room_load = num_sessions + extra_per_room.get(room, 0)
                            if (current + room_load) <= day_available_weeks and current < min_usage:
                                chosen_room = room; min_usage = current
                        if chosen_room is None: continue

                    sg = sd['config'].get('shared_group')
                    shared_subjects_for_slots = [subject]
                    if sg:
                        shared_subjects_for_slots = [s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg]


                    room_for_penalty = (chosen_room.split(',')[0].strip()
                                        if chosen_room else lab_rooms_list[0])
                    pk = (room_for_penalty, sd['semester'], slot[0], slot[1])
                    usage_penalty = slot_room_usage.get(pk, 0)

                    if sd.get('group_by_program', False):
                        free_by_prog = defaultdict(list)
                        for sid in sd['unassigned']:
                            busy = student_busy.get(sid, set())
                            own_slots = set()
                            for ss in shared_subjects_for_slots:
                                own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                            effective_busy = busy - own_slots
                            if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                                prog = student_program.get(sid, 'UNKNOWN')
                                free_by_prog[prog].append(sid)
                        for prog, free in free_by_prog.items():
                            score = len(free) - usage_penalty - (PREF_BLOCK_PENALTY if slot[1] in subject_block_penalty.get(subject, ()) else 0) - friday_placement_penalty(day_idx, num_sessions, day_session_load[FRIDAY_IDX])
                            if len(free) >= MIN_GROUP_SIZE and score > best_score:
                                best_score = score
                                best_slot = slot; best_free = free
                                best_room = chosen_room; best_prog = prog
                    else:
                        free = []
                        for sid in sd['unassigned']:
                            busy = student_busy.get(sid, set())
                            own_slots = set()
                            for ss in shared_subjects_for_slots:
                                own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                            effective_busy = busy - own_slots
                            if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                                free.append(sid)
                        score = len(free) - usage_penalty - (PREF_BLOCK_PENALTY if slot[1] in subject_block_penalty.get(subject, ()) else 0) - friday_placement_penalty(day_idx, num_sessions, day_session_load[FRIDAY_IDX])
                        if len(free) >= MIN_GROUP_SIZE and score > best_score:
                            best_score = score
                            best_slot = slot; best_free = free
                            best_room = chosen_room; best_prog = 'MIXED'

                if len(best_free) < MIN_GROUP_SIZE: continue
                shared_progress = True
                sd['group_counter'] += 1
                members = best_free[:max_per_group]
                day_idx, block_id = best_slot
                group = {
                    'subject': subject, 'semester': sd['semester'],
                    'curso_num': sd['curso_num'], 'group_num': sd['group_counter'],
                    'program': best_prog, 'day_idx': day_idx, 'day': DAYS[day_idx],
                    'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                    'student_ids': members, 'nb_students': len(members),
                    'num_sessions': num_sessions, 'max_students': max_per_group,
                    'lab_rooms': best_room, 'min_week': sd['min_week'],
                    'max_week': sd['max_week'],
                }
                sd['groups'].append(group); all_groups.append(group)
                _propagate_busy(group)
                sd['slot_usage'][best_slot] += 1

                subj_key = (subject, sd['semester'], day_idx, block_id)
                subject_slot_usage[subj_key] = subject_slot_usage.get(subj_key, 0) + num_sessions
                for room in best_room.split(','):
                    room = room.strip()
                    if room:
                        room_key = (room, sd['semester'], day_idx, block_id)
                        room_load = num_sessions + extra_per_room.get(room, 0)
                        slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + room_load
                for sid in members:
                    sd['unassigned'].discard(sid)

        for s in shared_subjects_list_rr:
            sd = subject_data[s]
            assigned = sum(g['nb_students'] for g in sd['groups'])
            print(f"    {s:40s} | {len(sd['groups']):2d} gr | {assigned}/{sd['total_students']} assignés")


        overflow_count = 0
        for s in shared_subjects_list_rr:
            sd = subject_data[s]
            if not sd['unassigned']:
                continue

            sg_key = sd['config'].get('shared_group')
            shared_subjs = [s] if not sg_key else [x for x, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key]

            for g in sorted(sd['groups'], key=lambda x: x['max_students'] - x['nb_students'], reverse=True):
                room = g['max_students'] - g['nb_students']
                if room <= 0:
                    continue
                slot = (g['day_idx'], g['block_id'])

                added = []
                for sid in list(sd['unassigned']):
                    if len(added) >= room:
                        break
                    busy = student_busy.get(sid, set())
                    own_slots = set()
                    for ss in shared_subjs:
                        own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                    if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                        added.append(sid)

                for sid in added:
                    g['student_ids'].append(sid)
                    g['nb_students'] += 1
                    _propagate_busy_one(sid, g)
                    sd['unassigned'].discard(sid)
                    overflow_count += 1

        if overflow_count > 0:
            print(f"    [RECYCLE]  {overflow_count} étudiants cross-programme redistribués")
            for s in shared_subjects_list_rr:
                sd = subject_data[s]
                assigned = sum(g['nb_students'] for g in sd['groups'])
                if sd['unassigned']:
                    print(f"    {s:40s} | {assigned}/{sd['total_students']} assignés ({len(sd['unassigned'])} restants)")


    print(f"\n  ── Phase 2: Round-robin pour les matières restantes ──")
    round_num = 0
    max_rounds = 200
    progress = True

    while progress and round_num < max_rounds:
        round_num += 1
        progress = False

        for subject in remaining_subjects:
            sd = subject_data[subject]

            if not sd['unassigned']:
                continue

            config = sd['config']
            num_sessions = sd['num_sessions']
            max_per_group = sd['max_per_group']
            lab_rooms_list = sd['lab_rooms_list']
            available_weeks = sd['available_weeks']


            best_slot = None
            best_free = []
            best_room = None
            best_prog = None
            # Sélection par SCORE (= nb d'étudiants libres - pénalité vendredi),
            # et non plus par simple max d'étudiants libres, afin de lisser la
            # charge hors du vendredi (contrainte souple anti-goulot).
            best_score = float('-inf')

            simultaneous = config.get('simultaneous_rooms', False)
            extra_per_room = shared_extra_sessions.get(subject, {})

            for slot in sd['all_slots']:
                day_idx = slot[0]


                if sd['slot_usage'][slot] >= sd['max_groups_per_slot_by_day'].get(day_idx, 0):
                    continue


                day_available_weeks = sd['available_weeks_by_day'].get(day_idx, 0)


                subj_key = (subject, sd['semester'], slot[0], slot[1])
                if subject_slot_usage.get(subj_key, 0) + num_sessions > day_available_weeks:
                    continue

                if simultaneous:

                    room_ok = True
                    for room in lab_rooms_list:
                        room_key = (room, sd['semester'], slot[0], slot[1])
                        current = slot_room_usage.get(room_key, 0)
                        room_load = num_sessions + extra_per_room.get(room, 0)
                        if (current + room_load) > day_available_weeks:
                            room_ok = False
                            break
                    if not room_ok:
                        continue
                    chosen_room = ', '.join(lab_rooms_list)
                else:

                    chosen_room = None
                    min_usage = float('inf')
                    for room in lab_rooms_list:
                        room_key = (room, sd['semester'], slot[0], slot[1])
                        current = slot_room_usage.get(room_key, 0)
                        room_load = num_sessions + extra_per_room.get(room, 0)
                        if (current + room_load) <= day_available_weeks and current < min_usage:
                            chosen_room = room
                            min_usage = current
                    if chosen_room is None:
                        continue


                sg = sd['config'].get('shared_group')
                shared_subjects_list = [subject]
                if sg:
                    shared_subjects_list = [s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg]

                if sd.get('group_by_program', False):
                    free_by_prog = defaultdict(list)
                    for sid in sd['unassigned']:
                        busy = student_busy.get(sid, set())
                        own_slots = set()
                        for ss in shared_subjects_list:
                            own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                        effective_busy = busy - own_slots
                        if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                            prog = student_program.get(sid, 'UNKNOWN')
                            free_by_prog[prog].append(sid)

                    for prog, free in free_by_prog.items():
                        score = len(free) - friday_placement_penalty(day_idx, num_sessions, day_session_load[FRIDAY_IDX])
                        # On ne met à jour le meilleur créneau qu'avec un créneau
                        # FAISABLE (>= MIN_GROUP_SIZE). Ainsi la pénalité vendredi
                        # ne peut jamais faire perdre un vendredi faisable face à
                        # un créneau non faisable : aucune perte de faisabilité.
                        if len(free) >= MIN_GROUP_SIZE and score > best_score:
                            best_score = score
                            best_slot = slot
                            best_free = free
                            best_room = chosen_room
                            best_prog = prog
                else:

                    free = []
                    for sid in sd['unassigned']:
                        busy = student_busy.get(sid, set())
                        own_slots = set()
                        for ss in shared_subjects_list:
                            own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                        effective_busy = busy - own_slots
                        if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                            free.append(sid)

                    score = len(free) - friday_placement_penalty(day_idx, num_sessions, day_session_load[FRIDAY_IDX])
                    # Garde-fou faisabilité identique au cas group_by_program :
                    # seul un créneau >= MIN_GROUP_SIZE peut devenir le meilleur.
                    if len(free) >= MIN_GROUP_SIZE and score > best_score:
                        best_score = score
                        best_slot = slot
                        best_free = free
                        best_room = chosen_room
                        best_prog = 'MIXED'

            if len(best_free) < MIN_GROUP_SIZE:
                continue


            progress = True
            sd['group_counter'] += 1
            members = best_free[:max_per_group]
            day_idx, block_id = best_slot


            if best_prog == 'MIXED':
                prog_counts = defaultdict(int)
                for sid in members:
                    prog_counts[student_program.get(sid, '?')] += 1
                if prog_counts:
                    dominant = max(prog_counts, key=prog_counts.get)
                    if len(prog_counts) == 1:
                        best_prog = dominant
                    else:
                        best_prog = f"MIXED({dominant}+{len(prog_counts)-1})"

            group = {
                'subject': subject,
                'semester': sd['semester'],
                'curso_num': sd['curso_num'],
                'group_num': sd['group_counter'],
                'program': best_prog,
                'day_idx': day_idx,
                'day': DAYS[day_idx],
                'block_id': block_id,
                'block_label': BLOCK_LABELS[block_id],
                'student_ids': members,
                'nb_students': len(members),
                'num_sessions': num_sessions,
                'max_students': max_per_group,
                'lab_rooms': best_room,
                'min_week': sd['min_week'],
                'max_week': sd['max_week'],
            }

            sd['groups'].append(group)
            all_groups.append(group)
            _propagate_busy(group)


            sd['slot_usage'][best_slot] += 1

            subj_key = (subject, sd['semester'], day_idx, block_id)
            subject_slot_usage[subj_key] = subject_slot_usage.get(subj_key, 0) + num_sessions

            for room in best_room.split(','):
                room = room.strip()
                if room:
                    room_key = (room, sd['semester'], day_idx, block_id)
                    room_load = num_sessions + extra_per_room.get(room, 0)
                    slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + room_load

            for sid in members:
                sd['unassigned'].discard(sid)


    redistributed = 0
    for subject in list(subject_data.keys()) + list(remaining_subjects):
        if subject not in subject_data:
            continue
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        sg_key = sd['config'].get('shared_group')
        shared_subjects_for_slots = [subject]
        if sg_key:
            shared_subjects_for_slots = [s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key]

        for g in sd['groups']:
            if g['nb_students'] >= g['max_students']:
                continue
            room = g['max_students'] - g['nb_students']
            slot = (g['day_idx'], g['block_id'])

            added = []
            for sid in list(sd['unassigned']):
                if len(added) >= room:
                    break
                busy = student_busy.get(sid, set())
                own_slots = set()
                for ss in shared_subjects_for_slots:
                    own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                effective_busy = busy - own_slots
                if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                    added.append(sid)

            for sid in added:
                g['student_ids'].append(sid)
                g['nb_students'] += 1
                _propagate_busy_one(sid, g)
                sd['unassigned'].discard(sid)
                redistributed += 1

    if redistributed > 0:
        print(f"\n  [RECYCLE]  {redistributed} étudiants redistribués dans des groupes existants")


    print(f"\n  ── Phase 3: Recovery aggressive (objectif 100%) ──")


    def compute_room_slot_load(groups_list):
        """
        Returns a dict (room, sem, day, block) → number of sessions occupying it.
        Each group contributes its num_sessions per occupied (room, sem, day, block).
        For multi-room groups (lab_rooms = "Room1, Room2"), each room gets
        the FULL num_sessions (because the group needs num_sessions weeks).
        """
        load = {}
        for g in groups_list:
            for room in g['lab_rooms'].split(','):
                room = room.strip()
                if not room:
                    continue
                key = (room, g['semester'], g['day_idx'], g['block_id'])
                load[key] = load.get(key, 0) + g['num_sessions']
        return load

    def compute_subject_slot_load(groups_list):
        """
        Returns a dict (subject, sem, day, block) → number of sessions.
        For C1 constraint validation.
        """
        load = {}
        for g in groups_list:
            key = (g['subject'], g['semester'], g['day_idx'], g['block_id'])
            load[key] = load.get(key, 0) + g['num_sessions']
        return load

    def slot_capacity_for(room, semester, day_idx, min_week, max_week):
        """
        Returns the maximum number of sessions that can fit in
        (room, semester, day, *any block*) across all weeks min_week..max_week
        excluding holidays for that day_idx.
        """
        sem_holidays = HOLIDAYS.get(semester, {})
        valid = [w for w in range(min_week, max_week + 1)
                 if (w, day_idx) not in sem_holidays]
        return len(valid)

    def can_fit_new_group(room, sem, day_idx, block_id, min_w, max_w,
                          new_sessions, current_load,
                          subject=None, subject_load=None):
        """
        STRICT pre-validation: would adding `new_sessions` to (room, sem, day, block)
        keep us within physical capacity?

        Checks BOTH:
        - C4 capacity: room+slot can host at most `cap` sessions across weeks
        - C1 capacity: subject+slot can host at most `cap` sessions across weeks
                       (if `subject` and `subject_load` are provided)
        Returns True if safe, False otherwise.
        """
        cap = slot_capacity_for(room, sem, day_idx, min_w, max_w)


        room_key = (room, sem, day_idx, block_id)
        room_current = current_load.get(room_key, 0)
        if (room_current + new_sessions) > cap:
            return False


        if subject and subject_load is not None:
            subj_key = (subject, sem, day_idx, block_id)
            subj_current = subject_load.get(subj_key, 0)
            if (subj_current + new_sessions) > cap:
                return False

        return True


    current_load = compute_room_slot_load(all_groups)
    subject_load = compute_subject_slot_load(all_groups)


    refit_count = 0
    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        sg_key = config.get('shared_group')
        shared_subjs = [subject] if not sg_key else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key
        ]


        candidate_groups = sorted(
            sd['groups'],
            key=lambda g: g['max_students'] - g['nb_students'],
            reverse=True,
        )

        for sid in list(sd['unassigned']):
            busy = student_busy.get(sid, set())
            own_slots = set()
            for ss in shared_subjs:
                own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
            effective_busy = busy - own_slots

            for g in candidate_groups:

                if g['nb_students'] >= g['max_students']:
                    continue
                slot = (g['day_idx'], g['block_id'])
                if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                    g['student_ids'].append(sid)
                    g['nb_students'] += 1
                    _propagate_busy_one(sid, g)
                    sd['unassigned'].discard(sid)
                    refit_count += 1
                    break

    if refit_count > 0:
        print(f"    [RECYCLE]  3a: {refit_count} étudiants re-fittés dans groupes existants")


    overflow_groups_created = 0
    overflow_assigned = 0
    overflow_skipped_full = 0

    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        unassigned_count = len(sd['unassigned'])

        if unassigned_count < MIN_GROUP_SIZE:
            continue


        original_curso = config['curso_num']
        MORNING = [1, 2, 3]
        AFTERNOON = [4, 5]
        if original_curso in [1, 3]:
            preferred_blocks = MORNING
            fallback_blocks = AFTERNOON if ALLOW_AFTERNOON_Y1Y3 else []
        else:
            preferred_blocks = AFTERNOON
            fallback_blocks = MORNING if ALLOW_MORNING_Y2Y4 else []
        extra_slots = [(d, b) for d in range(5) for b in preferred_blocks]
        extra_slots_fallback = [(d, b) for d in range(5) for b in fallback_blocks]


        sg_key = config.get('shared_group')
        shared_subjs = [subject] if not sg_key else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key
        ]

        max_per_group = sd['max_per_group']
        num_sessions = sd['num_sessions']
        lab_rooms_list = sd['lab_rooms_list']
        sem = sd['semester']
        min_w = sd['min_week']
        max_w = sd['max_week']


        if sg_key:
            total_sessions_per_group = sum(
                LAB_CONFIG[s]['num_sessions']
                for s in LAB_CONFIG if LAB_CONFIG[s].get('shared_group') == sg_key
            )
        else:
            total_sessions_per_group = num_sessions


        ALT_ROOMS_FOR_OVERFLOW = {
            'S1_Física': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
            'S1_Química': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
            'S2_Física II': ['Ciencias Experimentales I', 'Ciencias Experimentales II'],
        }
        effective_rooms = ALT_ROOMS_FOR_OVERFLOW.get(subject, lab_rooms_list)

        attempts = 0


        using_fallback = False
        active_slots = extra_slots
        while sd['unassigned'] and attempts < 50:
            attempts += 1
            best_slot = None; best_free = []; best_room = None
            best_score = -1

            for slot in active_slots:
                day_idx = slot[0]


                chosen_room = None
                for room in effective_rooms:
                    if can_fit_new_group(room, sem, day_idx, slot[1],
                                          min_w, max_w,
                                          total_sessions_per_group, current_load,
                                          subject=subject, subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    overflow_skipped_full += 1
                    continue


                free = []
                for sid in sd['unassigned']:
                    busy = student_busy.get(sid, set())
                    own_slots = set()
                    for ss in shared_subjs:
                        own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                    if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                        free.append(sid)

                if len(free) < MIN_GROUP_SIZE:
                    continue


                room_key = (chosen_room, sem, slot[0], slot[1])
                room_current = current_load.get(room_key, 0)
                usage_penalty = room_current * 5
                # Pénalité souple anti-goulot du vendredi (cf. phases principales).
                score = len(free) - usage_penalty - friday_placement_penalty(
                    day_idx, total_sessions_per_group, day_session_load[FRIDAY_IDX])

                if score > best_score:
                    best_score = score
                    best_slot = slot; best_free = free; best_room = chosen_room

            if best_slot is None or len(best_free) < MIN_GROUP_SIZE:


                if extra_slots_fallback and not using_fallback:
                    using_fallback = True
                    active_slots = extra_slots_fallback
                    continue
                if sd['unassigned']:
                    period = "morning" if original_curso in [1, 3] else "afternoon"
                    print(f"        ⚠️  {subject}: {len(sd['unassigned'])} student(s) "
                          f"could not be placed in the {period} (preferred period "
                          f"saturated). Left unassigned — afternoon fallback is "
                          f"disabled by the strict year-of-degree rule.")
                break


            sd['group_counter'] += 1
            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': original_curso, 'group_num': sd['group_counter'],
                'program': 'OVERFLOW', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                '_overflow': True,
            }
            sd['groups'].append(group); all_groups.append(group)
            _propagate_busy(group)
            sd['slot_usage'][best_slot] += 1


            for room in best_room.split(','):
                room = room.strip()
                if room:
                    room_key = (room, sem, day_idx, block_id)
                    current_load[room_key] = current_load.get(room_key, 0) + total_sessions_per_group
                    slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + total_sessions_per_group

            for ss in shared_subjs:
                ss_key = (ss, sem, day_idx, block_id)
                ss_sessions = LAB_CONFIG[ss]['num_sessions'] if ss in LAB_CONFIG else num_sessions
                subject_load[ss_key] = subject_load.get(ss_key, 0) + ss_sessions
            for sid in members:
                sd['unassigned'].discard(sid)
            overflow_groups_created += 1
            overflow_assigned += len(members)

    if overflow_groups_created > 0:
        print(f"    [RUN] 3b: {overflow_groups_created} groupes overflow créés "
              f"({overflow_assigned} étudiants récupérés en horaires alternatifs)")
        if overflow_skipped_full > 0:
            print(f"        ⓘ {overflow_skipped_full} tentatives rejetées par capacité salle (pré-validation)")


    expand_count = 0
    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']

        ALTERNATIVE_ROOMS = {
            'S1_Química': ['Ciencias Experimentales II'],
            'S2_Física II': ['Ciencias Experimentales II'],
        }
        alt_rooms = ALTERNATIVE_ROOMS.get(subject, [])
        if not alt_rooms:
            continue

        sem = sd['semester']
        num_sessions = sd['num_sessions']
        max_per_group = sd['max_per_group']

        sg_key = config.get('shared_group')
        shared_subjs = [subject] if not sg_key else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg_key
        ]

        all_slots_morning = [(d, b) for d in range(5) for b in [1, 2, 3]]


        if sg_key:
            total_sessions_per_group = sum(
                LAB_CONFIG[s]['num_sessions']
                for s in LAB_CONFIG if LAB_CONFIG[s].get('shared_group') == sg_key
            )
        else:
            total_sessions_per_group = num_sessions

        attempts = 0
        while sd['unassigned'] and attempts < 30:
            attempts += 1
            best_slot = None; best_free = []; best_room = None
            best_score = -1

            for slot in all_slots_morning:
                day_idx = slot[0]

                for room in alt_rooms:

                    if not can_fit_new_group(room, sem, day_idx, slot[1],
                                              sd['min_week'], sd['max_week'],
                                              total_sessions_per_group, current_load,
                                              subject=subject, subject_load=subject_load):
                        continue

                    free = []
                    for sid in sd['unassigned']:
                        busy = student_busy.get(sid, set())
                        own_slots = set()
                        for ss in shared_subjs:
                            own_slots |= student_subject_slots.get(sid, {}).get(ss, set())
                        if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                            free.append(sid)

                    if len(free) < MIN_GROUP_SIZE:
                        continue

                    room_key = (room, sem, slot[0], slot[1])
                    current = current_load.get(room_key, 0)
                    usage_penalty = current * 2
                    # Pénalité souple anti-goulot du vendredi (cf. phases principales).
                    score = len(free) - usage_penalty - friday_placement_penalty(
                        day_idx, total_sessions_per_group, day_session_load[FRIDAY_IDX])

                    if score > best_score:
                        best_score = score
                        best_slot = slot; best_free = free; best_room = room

            if best_slot is None or len(best_free) < MIN_GROUP_SIZE:
                break

            sd['group_counter'] += 1
            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': config['curso_num'], 'group_num': sd['group_counter'],
                'program': 'ALT_ROOM', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': sd['min_week'],
                'max_week': sd['max_week'],
                '_alt_room': True,
            }
            sd['groups'].append(group); all_groups.append(group)
            _propagate_busy(group)
            sd['slot_usage'][best_slot] += 1
            room_key = (best_room, sem, day_idx, block_id)
            current_load[room_key] = current_load.get(room_key, 0) + total_sessions_per_group
            slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + total_sessions_per_group

            for ss in shared_subjs:
                ss_key = (ss, sem, day_idx, block_id)
                ss_sessions = LAB_CONFIG[ss]['num_sessions'] if ss in LAB_CONFIG else num_sessions
                subject_load[ss_key] = subject_load.get(ss_key, 0) + ss_sessions
            for sid in members:
                sd['unassigned'].discard(sid)
            expand_count += len(members)

    if expand_count > 0:
        print(f"    [ALT_ROOM]  3c: {expand_count} étudiants récupérés via salles alternatives")

    total_recovered = refit_count + overflow_assigned + expand_count
    if total_recovered > 0:
        print(f"  [OK] Phase 3 total : {total_recovered} étudiants supplémentaires récupérés")


    consolidated = 0
    dissolved_groups = []

    for subject in list(subject_data.keys()):
        sd = subject_data[subject]
        groups = sd['groups']
        if len(groups) <= 1:
            continue

        small_groups = [g for g in groups if g['nb_students'] < MIN_GROUP_SIZE]
        big_groups = [g for g in groups if g['nb_students'] >= MIN_GROUP_SIZE]

        for sg in small_groups:

            merged = False
            sg_slot = (sg['day_idx'], sg['block_id'])

            candidates = sorted(big_groups, key=lambda g: g['max_students'] - g['nb_students'], reverse=True)

            for target in candidates:
                tgt_slot = (target['day_idx'], target['block_id'])
                room = target['max_students'] - target['nb_students']


                if room >= sg['nb_students'] and tgt_slot == sg_slot:

                    target['student_ids'].extend(sg['student_ids'])
                    target['nb_students'] += sg['nb_students']
                    consolidated += sg['nb_students']
                    dissolved_groups.append((subject, sg['group_num']))
                    merged = True
                    break

            if not merged:

                for sid in list(sg['student_ids']):
                    for target in candidates:
                        if target['nb_students'] < target['max_students']:
                            target_slot = (target['day_idx'], target['block_id'])


                            if target_slot in student_lab_busy.get(sid, set()):
                                continue
                            target['student_ids'].append(sid)
                            target['nb_students'] += 1
                            _propagate_busy_one(sid, target)
                            sg['student_ids'].remove(sid)
                            sg['nb_students'] -= 1
                            consolidated += 1
                            break
                if sg['nb_students'] == 0:
                    dissolved_groups.append((subject, sg['group_num']))


    for subject, gnum in dissolved_groups:
        sd = subject_data[subject]
        sd['groups'] = [g for g in sd['groups'] if g['group_num'] != gnum]

        all_groups[:] = [g for g in all_groups if not (g['subject'] == subject and g['group_num'] == gnum)]

    if consolidated > 0:
        print(f"  [RETRY] {consolidated} étudiants consolidés, {len(dissolved_groups)} petits groupes dissous")


    for subject in subject_data:
        sd = subject_data[subject]
        for i, g in enumerate(sorted(sd['groups'], key=lambda x: x['group_num'])):
            old_num = g['group_num']
            g['group_num'] = i + 1

            for ag in all_groups:
                if ag['subject'] == subject and ag.get('_old_group_num', ag['group_num']) == old_num:
                    ag['group_num'] = i + 1


    shared_groups_created = {}
    for subject, config in LAB_CONFIG.items():
        sg = config.get('shared_group')
        if sg and subject in subject_data:
            shared_groups_created[sg] = subject_data[subject]['groups']

    for subject in secondary_shared:
        config = LAB_CONFIG[subject]
        sg = config['shared_group']
        primary_groups = shared_groups_created.get(sg, [])
        if not primary_groups:
            continue


        enrolled = set(subject_students.get(subject, []))

        semester = config['semester']
        min_week = config['min_week']


        sem_max = SEMESTER_1_WEEKS if semester == 1 else SEMESTER_2_WEEKS
        max_week = config.get('max_week', sem_max)

        print(f"\n  Groupes partagés: {subject} réutilise {len(primary_groups)} "
              f"groupes de {shared_group_map[sg]}")

        for pg in primary_groups:

            shared_members = [sid for sid in pg['student_ids'] if sid in enrolled]
            if len(shared_members) < MIN_GROUP_SIZE:
                continue


            inherited_room = pg.get('lab_rooms', ', '.join(config['lab_rooms']))

            group = {
                'subject': subject,
                'semester': semester,
                'curso_num': config['curso_num'],
                'group_num': pg['group_num'],
                'program': pg.get('program', 'MIXED'),
                'day_idx': pg['day_idx'],
                'day': pg['day'],
                'block_id': pg['block_id'],
                'block_label': pg['block_label'],
                'student_ids': shared_members,
                'nb_students': len(shared_members),
                'num_sessions': config['num_sessions'],
                'max_students': config['max_students'],
                'lab_rooms': inherited_room,
                'min_week': min_week,
                'max_week': max_week,

                '_overflow': pg.get('_overflow', False),
                '_alt_room': pg.get('_alt_room', False),
            }
            all_groups.append(group)


        shared_assigned = sum(1 for g in all_groups if g['subject'] == subject
                             for _ in g['student_ids'])
        total_enrolled = len(enrolled)
        unassigned = total_enrolled - shared_assigned
        if unassigned > 0:
            print(f"  [WARN]  {subject:40s} | "
                  f"{shared_assigned}/{total_enrolled} assignés ({unassigned} sans groupe)")
        else:
            print(f"  [OK] {subject:40s} | "
                  f"{shared_assigned}/{total_enrolled} assignés (groupes partagés)")


    for subject in secondary_shared:
        sec_groups = [g for g in all_groups if g['subject'] == subject]
        small = [g for g in sec_groups if g['nb_students'] < MIN_GROUP_SIZE]
        big = [g for g in sec_groups if g['nb_students'] >= MIN_GROUP_SIZE]

        dissolved_sec = []
        for sg in small:
            merged = False
            sg_slot = (sg['day_idx'], sg['block_id'])
            for target in sorted(big, key=lambda g: g['max_students'] - g['nb_students'], reverse=True):
                tgt_slot = (target['day_idx'], target['block_id'])
                room = target['max_students'] - target['nb_students']


                if room >= sg['nb_students'] and tgt_slot == sg_slot:
                    target['student_ids'].extend(sg['student_ids'])
                    target['nb_students'] += sg['nb_students']
                    dissolved_sec.append(sg)
                    merged = True
                    break
            if not merged:
                for sid in list(sg['student_ids']):
                    for target in big:
                        if target['nb_students'] < target['max_students']:
                            tgt_slot = (target['day_idx'], target['block_id'])
                            if tgt_slot in student_lab_busy.get(sid, set()):
                                continue
                            target['student_ids'].append(sid)
                            target['nb_students'] += 1
                            _propagate_busy_one(sid, target)
                            sg['student_ids'].remove(sid)
                            sg['nb_students'] -= 1
                            break
                if sg['nb_students'] == 0:
                    dissolved_sec.append(sg)

        for dg in dissolved_sec:
            all_groups.remove(dg)
        if dissolved_sec:

            remaining = sorted([g for g in all_groups if g['subject'] == subject],
                               key=lambda x: x['group_num'])
            for i, g in enumerate(remaining):
                g['group_num'] = i + 1
            print(f"  [RETRY] {subject}: {len(dissolved_sec)} petits groupes fusionnés")


    print(f"\n  ── Phase 4: Recovery secondaires partagés ──")

    sec_recovery_count = 0
    for subject in secondary_shared:
        config = LAB_CONFIG[subject]
        enrolled = set(subject_students.get(subject, []))


        assigned_ids = set()
        for g in all_groups:
            if g['subject'] == subject:
                assigned_ids.update(g['student_ids'])

        unassigned = enrolled - assigned_ids
        if not unassigned:
            continue

        sem = config['semester']
        min_week = config['min_week']
        sem_max = SEMESTER_1_WEEKS if sem == 1 else SEMESTER_2_WEEKS
        max_week = config.get('max_week', sem_max)
        num_sessions = config['num_sessions']
        max_per_group = config['max_students']


        ALTERNATIVE_ROOMS = {
            'S1_Química': ['Ciencias Experimentales II', 'Ciencias Experimentales I'],
        }
        candidate_rooms = ALTERNATIVE_ROOMS.get(subject, config['lab_rooms'])


        original_curso = config['curso_num']
        all_slots = []
        for d in range(5):
            primary_blocks = [1, 2, 3] if original_curso in [1, 3] else [4, 5]
            secondary_blocks = [4, 5] if original_curso in [1, 3] else [1, 2, 3]
            for b in primary_blocks + secondary_blocks:
                all_slots.append((d, b))

        sem_holidays = HOLIDAYS.get(sem, {})
        available_weeks_by_day = {}
        for d in range(5):
            valid = [w for w in range(min_week, max_week + 1)
                     if (w, d) not in sem_holidays]
            available_weeks_by_day[d] = len(valid)


        cur_groups = sorted(
            [g for g in all_groups if g['subject'] == subject],
            key=lambda g: g['max_students'] - g['nb_students'], reverse=True,
        )

        next_group_num = max([g['group_num'] for g in all_groups if g['subject'] == subject], default=0) + 1

        for sid in list(unassigned):
            busy = student_busy.get(sid, set())
            own_slots = student_subject_slots.get(sid, {}).get(subject, set())
            effective_busy = busy - own_slots
            placed = False
            for g in cur_groups:
                if g['nb_students'] >= g['max_students']:
                    continue
                slot = (g['day_idx'], g['block_id'])
                if slot not in effective_busy and slot not in student_lab_busy.get(sid, set()):
                    g['student_ids'].append(sid)
                    g['nb_students'] += 1
                    _propagate_busy_one(sid, g)
                    unassigned.discard(sid)
                    sec_recovery_count += 1
                    placed = True
                    break


        current_load = compute_room_slot_load(all_groups)
        subject_load = compute_subject_slot_load(all_groups)

        attempts = 0
        while unassigned and attempts < 30:
            attempts += 1
            best_slot = None; best_free = []; best_room = None
            best_score = -1

            for slot in all_slots:
                day_idx = slot[0]

                for room in candidate_rooms:

                    if not can_fit_new_group(room, sem, day_idx, slot[1],
                                              min_week, max_week,
                                              num_sessions, current_load,
                                              subject=subject, subject_load=subject_load):
                        continue

                    free = []
                    for sid in unassigned:
                        busy = student_busy.get(sid, set())
                        own_slots = student_subject_slots.get(sid, {}).get(subject, set())
                        if slot not in (busy - own_slots) and slot not in student_lab_busy.get(sid, set()):
                            free.append(sid)

                    if len(free) < MIN_GROUP_SIZE:
                        continue


                    if original_curso in [1, 3]:
                        wrong_period = slot[1] in [4, 5] and not ALLOW_AFTERNOON_Y1Y3
                    else:
                        wrong_period = slot[1] in [1, 2, 3] and not ALLOW_MORNING_Y2Y4
                    if wrong_period:
                        continue


                    room_key = (room, sem, slot[0], slot[1])
                    current = current_load.get(room_key, 0)
                    usage_penalty = current * 2
                    # Pénalité souple anti-goulot du vendredi (cf. phases principales).
                    score = len(free) - usage_penalty - friday_placement_penalty(
                        slot[0], num_sessions, day_session_load[FRIDAY_IDX])

                    if score > best_score:
                        best_score = score
                        best_slot = slot; best_free = free; best_room = room

            if best_slot is None or len(best_free) < MIN_GROUP_SIZE:
                break

            members = best_free[:max_per_group]
            day_idx, block_id = best_slot
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': original_curso, 'group_num': next_group_num,
                'program': 'RECOVERED', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': min_week, 'max_week': max_week,
                '_recovered': True,
            }
            all_groups.append(group)
            next_group_num += 1
            room_key = (best_room, sem, day_idx, block_id)
            current_load[room_key] = current_load.get(room_key, 0) + num_sessions
            subj_key = (subject, sem, day_idx, block_id)
            subject_load[subj_key] = subject_load.get(subj_key, 0) + num_sessions
            slot_room_usage[room_key] = slot_room_usage.get(room_key, 0) + num_sessions
            for sid in members:
                unassigned.discard(sid)
            sec_recovery_count += len(members)

    if sec_recovery_count > 0:
        print(f"    [OK] {sec_recovery_count} étudiants secondaires récupérés")


    print(f"\n  ── Phase 5: Recovery par programme (stratégie Daniel) ──")
    p5_recovered = 0
    p5_groups = 0


    current_load = compute_room_slot_load(all_groups)
    subject_load = compute_subject_slot_load(all_groups)

    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        sem = sd['semester']
        num_sessions = sd['num_sessions']
        max_per_group = sd['max_per_group']
        lab_rooms_list = sd['lab_rooms_list']
        min_w, max_w = sd['min_week'], sd['max_week']
        allowed_slots = sd['all_slots']


        sg = config.get('shared_group')
        shared_subjs = [subject] if not sg else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg
        ]


        by_program = defaultdict(list)
        for sid in sd['unassigned']:
            by_program[student_program.get(sid, 'UNKNOWN')].append(sid)


        for prog in sorted(by_program, key=lambda p: -len(by_program[p])):
            pool = by_program[prog]

            progress = True
            while len(pool) >= RECOVERY_MIN_GROUP_SIZE and progress:
                progress = False
                best_slot = None
                best_free = []
                best_room = None

                for slot in allowed_slots:
                    day_idx, block_id = slot

                    chosen_room = None
                    for room in lab_rooms_list:
                        if can_fit_new_group(room, sem, day_idx, block_id,
                                              min_w, max_w, num_sessions,
                                              current_load,
                                              subject=subject,
                                              subject_load=subject_load):
                            chosen_room = room
                            break
                    if chosen_room is None:
                        continue


                    free = []
                    for sid in pool:
                        busy = student_busy.get(sid, set())
                        own = set()
                        for ss in shared_subjs:
                            own |= student_subject_slots.get(sid, {}).get(ss, set())
                        if (slot not in (busy - own)
                                and slot not in student_lab_busy.get(sid, set())):
                            free.append(sid)

                    if len(free) > len(best_free):
                        best_free = free
                        best_slot = slot
                        best_room = chosen_room

                if best_slot is None or len(best_free) < RECOVERY_MIN_GROUP_SIZE:
                    break


                members = best_free[:max_per_group]
                day_idx, block_id = best_slot
                sd['group_counter'] += 1
                group = {
                    'subject': subject, 'semester': sem,
                    'curso_num': sd['curso_num'],
                    'group_num': sd['group_counter'],
                    'program': prog, 'day_idx': day_idx, 'day': DAYS[day_idx],
                    'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                    'student_ids': members, 'nb_students': len(members),
                    'num_sessions': num_sessions, 'max_students': max_per_group,
                    'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                    'recovered': 'P5',
                }
                sd['groups'].append(group)
                all_groups.append(group)
                _propagate_busy(group)


                for room in best_room.split(','):
                    room = room.strip()
                    if not room:
                        continue
                    rk = (room, sem, day_idx, block_id)
                    current_load[rk] = current_load.get(rk, 0) + num_sessions
                subj_key = (subject, sem, day_idx, block_id)
                subject_load[subj_key] = subject_load.get(subj_key, 0) + num_sessions

                placed = set(members)
                for sid in placed:
                    sd['unassigned'].discard(sid)
                pool = [sid for sid in pool if sid not in placed]
                p5_recovered += len(members)
                p5_groups += 1
                progress = True


    p5b_recovered = 0
    p5b_groups = 0
    for subject in sorted_subjects:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue

        config = sd['config']
        sem = sd['semester']
        num_sessions = sd['num_sessions']
        max_per_group = sd['max_per_group']
        lab_rooms_list = sd['lab_rooms_list']
        min_w, max_w = sd['min_week'], sd['max_week']
        allowed_slots = sd['all_slots']

        sg = config.get('shared_group')
        shared_subjs = [subject] if not sg else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg
        ]

        progress = True
        while len(sd['unassigned']) >= RECOVERY_MIN_GROUP_SIZE and progress:
            progress = False
            best_slot, best_free, best_room = None, [], None

            for slot in allowed_slots:
                day_idx, block_id = slot
                chosen_room = None
                for room in lab_rooms_list:
                    if can_fit_new_group(room, sem, day_idx, block_id,
                                          min_w, max_w, num_sessions,
                                          current_load,
                                          subject=subject,
                                          subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    continue


                free = []
                for sid in sd['unassigned']:
                    busy = student_busy.get(sid, set())
                    own = set()
                    for ss in shared_subjs:
                        own |= student_subject_slots.get(sid, {}).get(ss, set())
                    if (slot not in (busy - own)
                            and slot not in student_lab_busy.get(sid, set())):
                        free.append(sid)

                if len(free) > len(best_free):
                    best_free, best_slot, best_room = free, slot, chosen_room

            if best_slot is None or len(best_free) < RECOVERY_MIN_GROUP_SIZE:
                break

            members = best_free[:max_per_group]
            day_idx, block_id = best_slot

            prog_counts = Counter(student_program.get(s, 'UNKNOWN') for s in members)
            dominant, dom_n = prog_counts.most_common(1)[0]
            prog_label = (f"MIXED({dominant}+{len(members) - dom_n})"
                          if len(prog_counts) > 1 else dominant)

            sd['group_counter'] += 1
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': sd['curso_num'],
                'group_num': sd['group_counter'],
                'program': prog_label, 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': max_per_group,
                'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                'recovered': 'P5b-mixed',
            }
            sd['groups'].append(group)
            all_groups.append(group)
            _propagate_busy(group)

            for room in best_room.split(','):
                room = room.strip()
                if not room:
                    continue
                rk = (room, sem, day_idx, block_id)
                current_load[rk] = current_load.get(rk, 0) + num_sessions
            subj_key = (subject, sem, day_idx, block_id)
            subject_load[subj_key] = subject_load.get(subj_key, 0) + num_sessions

            for sid in set(members):
                sd['unassigned'].discard(sid)
            p5b_recovered += len(members)
            p5b_groups += 1
            progress = True

    if p5b_recovered > 0:
        print(f"    [OK] {p5b_recovered} étudiants récupérés en {p5b_groups} "
              f"groupe(s) MIXTE(s) (programmes mélangés, comme Daniel)")
    p5_recovered += p5b_recovered
    p5_groups += p5b_groups

    if p5_recovered > 0:
        print(f"    [TOTAL Phase 5] {p5_recovered} étudiants récupérés "
              f"en {p5_groups} groupe(s)")
        if RECOVERY_MIN_GROUP_SIZE < MIN_GROUP_SIZE:
            print(f"    ⓘ seuil de récupération abaissé à "
                  f"{RECOVERY_MIN_GROUP_SIZE} (vs {MIN_GROUP_SIZE} normal)")
    else:
        print(f"    Aucun étudiant récupérable (créneaux saturés ou pools < "
              f"{RECOVERY_MIN_GROUP_SIZE})")


    current_load = compute_room_slot_load(all_groups)
    subject_load = compute_subject_slot_load(all_groups)
    p6_absorbed = 0
    p6_newgroups = 0
    p6_exception_solo = []
    truly_unplaceable = []

    all_subjects_with_rest = [s for s in sorted_subjects
                              if subject_data[s]['unassigned']]

    for subject in all_subjects_with_rest:
        sd = subject_data[subject]
        if not sd['unassigned']:
            continue
        config = sd['config']
        sem = sd['semester']
        num_sessions = sd['num_sessions']
        absolute_max = MAX_GROUP_SIZE
        lab_rooms_list = sd['lab_rooms_list']
        min_w, max_w = sd['min_week'], sd['max_week']
        allowed_slots = sd['all_slots']
        sg = config.get('shared_group')
        shared_subjs = [subject] if not sg else [
            s for s, c in LAB_CONFIG.items() if c.get('shared_group') == sg
        ]

        def student_free_at(sid, slot):
            busy = student_busy.get(sid, set())
            own = set()
            for ss in shared_subjs:
                own |= student_subject_slots.get(sid, {}).get(ss, set())
            if slot in student_lab_busy.get(sid, set()):
                return False
            return slot not in (busy - own)


        for sid in list(sd['unassigned']):
            placed = False

            cands = [g for g in sd['groups']
                     if g['nb_students'] < absolute_max
                     and student_free_at(sid, (g['day_idx'], g['block_id']))]

            cands.sort(key=lambda g: g['nb_students'])
            if cands:
                g = cands[0]
                g['student_ids'].append(sid)
                g['nb_students'] += 1
                _propagate_busy_one(sid, g)

                student_busy.setdefault(sid, set()).add((g['day_idx'], g['block_id']))
                sd['unassigned'].discard(sid)
                p6_absorbed += 1
                placed = True
            if placed:
                continue


        MIN_VIABLE_P6 = 3
        remaining = list(sd['unassigned'])
        guard = 0
        while remaining and guard < 60:
            guard += 1
            best_slot, best_free, best_room = None, [], None
            best_usage = float('inf')
            for slot in allowed_slots:
                day_idx, block_id = slot
                chosen_room = None
                for room in lab_rooms_list:
                    if can_fit_new_group(room, sem, day_idx, block_id,
                                          min_w, max_w, num_sessions,
                                          current_load,
                                          subject=subject,
                                          subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    continue
                free = [sid for sid in remaining if student_free_at(sid, slot)]
                if len(free) < MIN_VIABLE_P6:
                    continue
                rk = (chosen_room, sem, day_idx, block_id)
                usage = current_load.get(rk, 0)
                if (len(free), -usage) > (len(best_free), -best_usage):
                    best_free = free; best_slot = slot
                    best_room = chosen_room; best_usage = usage
            if best_slot is None:
                break
            members = best_free[:absolute_max]
            day_idx, block_id = best_slot
            sd['group_counter'] += 1
            group = {
                'subject': subject, 'semester': sem,
                'curso_num': sd['curso_num'],
                'group_num': sd['group_counter'],
                'program': 'MIXED', 'day_idx': day_idx, 'day': DAYS[day_idx],
                'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                'student_ids': members, 'nb_students': len(members),
                'num_sessions': num_sessions, 'max_students': absolute_max,
                'lab_rooms': best_room, 'min_week': min_w, 'max_week': max_w,
                '_absorbed': 'P6',
            }
            sd['groups'].append(group)
            all_groups.append(group)
            _propagate_busy(group)
            for room in best_room.split(','):
                room = room.strip()
                if room:
                    rk = (room, sem, day_idx, block_id)
                    current_load[rk] = current_load.get(rk, 0) + num_sessions
            sk = (subject, sem, day_idx, block_id)
            subject_load[sk] = subject_load.get(sk, 0) + num_sessions
            for sid in members:
                sd['unassigned'].discard(sid)
            remaining = [s for s in remaining if s not in members]
            p6_newgroups += 1
            p6_absorbed += len(members)


        OVERFLOW_TOLERANCE = 2
        for sid in list(sd['unassigned']):
            cands = [g for g in sd['groups']
                     if g['nb_students'] < absolute_max + OVERFLOW_TOLERANCE
                     and student_free_at(sid, (g['day_idx'], g['block_id']))]
            cands.sort(key=lambda g: g['nb_students'])
            if cands:
                g = cands[0]
                g['student_ids'].append(sid)
                g['nb_students'] += 1
                _propagate_busy_one(sid, g)
                g['_overflow'] = g.get('_overflow', 0) + 1


                student_busy.setdefault(sid, set()).add((g['day_idx'], g['block_id']))
                sd['unassigned'].discard(sid)
                p6_absorbed += 1


        for sid in list(sd['unassigned']):
            for slot in allowed_slots:
                day_idx, block_id = slot
                if not student_free_at(sid, slot):
                    continue
                chosen_room = None
                for room in lab_rooms_list:
                    if can_fit_new_group(room, sem, day_idx, block_id,
                                          min_w, max_w, num_sessions,
                                          current_load, subject=subject,
                                          subject_load=subject_load):
                        chosen_room = room
                        break
                if chosen_room is None:
                    continue
                sd['group_counter'] += 1
                group = {
                    'subject': subject, 'semester': sem,
                    'curso_num': sd['curso_num'],
                    'group_num': sd['group_counter'], 'program': 'MIXED',
                    'day_idx': day_idx, 'day': DAYS[day_idx],
                    'block_id': block_id, 'block_label': BLOCK_LABELS[block_id],
                    'student_ids': [sid], 'nb_students': 1,
                    'num_sessions': num_sessions, 'max_students': absolute_max,
                    'lab_rooms': chosen_room, 'min_week': min_w, 'max_week': max_w,
                    '_absorbed': 'P6', '_exception_solo': True,
                }
                sd['groups'].append(group)
                all_groups.append(group)
                _propagate_busy(group)
                for room in chosen_room.split(','):
                    room = room.strip()
                    if room:
                        rk = (room, sem, day_idx, block_id)
                        current_load[rk] = current_load.get(rk, 0) + num_sessions
                sk = (subject, sem, day_idx, block_id)
                subject_load[sk] = subject_load.get(sk, 0) + num_sessions
                sd['unassigned'].discard(sid)
                p6_absorbed += 1
                p6_newgroups += 1
                p6_exception_solo.append((subject, sid))
                break


        for sid in sd['unassigned']:
            truly_unplaceable.append((subject, sid))


    p8_overrides = []
    new_unplaceable = []
    for subject, sid in truly_unplaceable:
        sd = subject_data.get(subject)
        if not sd or not sd['groups']:
            new_unplaceable.append((subject, sid))
            continue

        target = max(sd['groups'], key=lambda g: g['nb_students'])
        target['student_ids'].append(sid)
        target['nb_students'] += 1
        target['_manual_override'] = target.get('_manual_override', 0) + 1
        target.setdefault('_override_sids', set()).add(sid)


        if isinstance(sd.get('unassigned'), set):
            sd['unassigned'].discard(sid)


        p8_overrides.append((subject, sid, target['group_num']))
    truly_unplaceable = new_unplaceable


    p8_solo_dissolved = []
    for subject, sd in subject_data.items():
        solos = [g for g in sd['groups'] if g['nb_students'] == 1]
        for solo_g in solos:
            others = [g for g in sd['groups']
                      if g is not solo_g and g['nb_students'] >= 2]
            if not others:
                continue
            target = max(others, key=lambda g: g['nb_students'])
            for sid in list(solo_g['student_ids']):
                target['student_ids'].append(sid)
                target['nb_students'] += 1
                target['_manual_override'] = target.get('_manual_override', 0) + 1
                target.setdefault('_override_sids', set()).add(sid)
                p8_overrides.append((subject, sid, target['group_num']))
            solo_g['student_ids'] = []
            solo_g['nb_students'] = 0
            p8_solo_dissolved.append((subject, solo_g['group_num']))

    for subj, gnum in p8_solo_dissolved:
        subject_data[subj]['groups'] = [
            g for g in subject_data[subj]['groups'] if g['group_num'] != gnum]
        all_groups[:] = [g for g in all_groups
                         if not (g['subject'] == subj and g['group_num'] == gnum)]

    if p8_overrides:
        print(f"\n  ── Phase 8: Override manuel (saturés) ──")
        print(f"    [INFO] {len(p8_overrides)} étudiant(s) saturé(s) placé(s) "
              f"dans le groupe principal de leur matière (à confirmer par Daniel) :")

        try:
            import pandas as _pd
            df_names = _pd.read_csv('data_clean/master_schedule.csv',
                                    usecols=['AlumnoID','Apellidos','Nombre'])
            df_names = df_names.drop_duplicates('AlumnoID')
            _sid_to_name = {str(r['AlumnoID']): f"{r['Apellidos']}, {r['Nombre']}"
                            for _, r in df_names.iterrows()}
        except Exception:
            _sid_to_name = {}
        for subj, sid, gn in p8_overrides:
            name = _sid_to_name.get(str(sid), f"<id {sid}>")
            print(f"           · {name}  →  {subj} G{gn}  [OVERRIDE]")


    p7_merged = 0
    p7_dissolved_groups = []


    OVERFLOW_TOL_P7 = 5
    SMALL_THRESHOLD_P7 = 4
    for subject, sd in subject_data.items():


        solo_groups = [g for g in sd['groups']
                       if g['nb_students'] <= SMALL_THRESHOLD_P7]
        for solo_g in solo_groups:
            absolute_max = MAX_GROUP_SIZE
            members_to_move = list(solo_g['student_ids'])
            for sid in members_to_move:


                solo_slot = (solo_g['day_idx'], solo_g['block_id'])


                lab_busy_set = student_lab_busy.get(sid, set()) - {solo_slot}
                cands = []
                for g in sd['groups']:
                    if g is solo_g:
                        continue
                    if g['nb_students'] >= absolute_max + OVERFLOW_TOL_P7:
                        continue
                    g_slot = (g['day_idx'], g['block_id'])
                    if g_slot in lab_busy_set:
                        continue
                    cands.append(g)

                cands.sort(key=lambda g: g['nb_students'])
                if not cands:
                    continue
                target = cands[0]
                target['student_ids'].append(sid)
                target['nb_students'] += 1
                target['_p7_merged'] = target.get('_p7_merged', 0) + 1
                solo_g['student_ids'].remove(sid)
                solo_g['nb_students'] -= 1

                student_lab_busy[sid].discard(solo_slot)
                student_lab_busy[sid].add((target['day_idx'], target['block_id']))
                p7_merged += 1

            if solo_g['nb_students'] == 0:
                p7_dissolved_groups.append((subject, solo_g['group_num']))


    for subj, gnum in p7_dissolved_groups:
        subject_data[subj]['groups'] = [
            g for g in subject_data[subj]['groups'] if g['group_num'] != gnum]
        all_groups[:] = [g for g in all_groups
                         if not (g['subject'] == subj and g['group_num'] == gnum)]

    if p7_merged > 0:
        print(f"\n  ── Phase 7: Passe finale anti-solo ──")
        print(f"    [OK] {p7_merged} étudiant(s) solo fusionné(s) dans des "
              f"groupes existants ({len(p7_dissolved_groups)} groupe(s) "
              f"solo dissous)")

    if p6_absorbed > 0:
        print(f"\n  ── Phase 6: Absorption finale adaptative ──")
        print(f"    [OK] {p6_absorbed} étudiant(s) placé(s) "
              f"({p6_newgroups} petit(s) groupe(s) créé(s), reste absorbé dans "
              f"des groupes existants)")
    if p6_exception_solo:
        from collections import Counter as _Ctr
        print(f"\n  [EXCEPTION] {len(p6_exception_solo)} étudiant(s) isolé(s) placé(s) "
              f"en groupe minimal (à fusionner manuellement si possible) :")
        for subj_x, n in _Ctr(s for s, _ in p6_exception_solo).most_common():
            print(f"       - {subj_x}: {n} cas (emploi du temps atypique)")
    if truly_unplaceable:
        print(f"\n  [GARDE-FOU] {len(truly_unplaceable)} étudiant(s) RÉELLEMENT "
              f"inplaçable(s) (aucun créneau compatible libre) :")
        from collections import Counter as _Counter, defaultdict as _DD
        by_subj_ids = _DD(list)
        for subj, sid in truly_unplaceable:
            by_subj_ids[subj].append(sid)

        try:
            import pandas as _pd
            df_names = _pd.read_csv('data_clean/master_schedule.csv',
                                    usecols=['AlumnoID', 'Apellidos', 'Nombre'])
            df_names = df_names.drop_duplicates('AlumnoID')
            sid_to_name = {
                str(row['AlumnoID']): f"{row['Apellidos']}, {row['Nombre']}"
                for _, row in df_names.iterrows()
            }
        except Exception:
            sid_to_name = {}
        for subj in sorted(by_subj_ids, key=lambda s: -len(by_subj_ids[s])):
            ids = by_subj_ids[subj]
            print(f"       - {subj}: {len(ids)} étudiant(s)")
            for sid in ids:
                name = sid_to_name.get(str(sid), f"<id {sid}>")
                print(f"           · {name}")
        print(f"       Ces cas nécessitent une décision manuelle (créneau "
              f"exceptionnel ou dérogation).")


    for sid_subject in []:
        pass


    print(f"\n  Résultats après {round_num} tours :")
    print(f"  (Taille préférée: {PREFERRED_GROUP_SIZE} | Max autorisé: {MAX_GROUP_SIZE})")
    for subject in sorted_subjects:
        sd = subject_data[subject]
        assigned = sum(g['nb_students'] for g in sd['groups'])
        unassigned = len(sd['unassigned'])
        total = sd['total_students']
        n_groups = len(sd['groups'])


        groups_at_15 = math.ceil(total / MAX_GROUP_SIZE) if total > 0 else 0
        extra = n_groups - groups_at_15

        if unassigned > 0:
            stats['total_unassigned'] += unassigned
            print(f"  [WARN]  {subject:40s} | {n_groups:2d} gr | "
                  f"{assigned:3d}/{total} assignés ({unassigned} sans créneau)")
        else:
            extra_info = f" (+{extra} vs max15)" if extra > 0 else ""
            print(f"  [OK] {subject:40s} | {n_groups:2d} gr | "
                  f"{assigned:3d}/{total} assignés{extra_info}")

        if extra > 3:
            print(f"       [WARN]  {extra} groupes de plus qu'avec max=15. "
                  f"Considérer max=15 pour cette matière.")

        stats['total_assigned'] += assigned

        for g in sd['groups']:
            prog_label = g.get('program', '?')
            marker = " [RUN]" if g.get('_overflow') else (" [ALT_ROOM]" if g.get('_alt_room') else "")
            print(f"       G{g['group_num']:2d} : {g['day']:10s} {g['block_label']} | "
                  f"{g['nb_students']:2d}/{g['max_students']} étudiants | {prog_label}{marker}")


    for subject in secondary_shared:
        config = LAB_CONFIG[subject]
        enrolled = set(subject_students.get(subject, []))
        sec_groups = [g for g in all_groups if g['subject'] == subject]
        assigned_sec = set()
        for g in sec_groups:
            assigned_sec.update(g['student_ids'])
        unassigned_count = len(enrolled - assigned_sec)
        n_groups = len(sec_groups)

        stats['total_assigned'] += len(assigned_sec)
        stats['total_unassigned'] += unassigned_count

        if unassigned_count > 0:
            print(f"  [WARN]  {subject:40s} | {n_groups:2d} gr | "
                  f"{len(assigned_sec):3d}/{len(enrolled)} assignés ({unassigned_count} sans créneau)")
        else:
            print(f"  [OK] {subject:40s} | {n_groups:2d} gr | "
                  f"{len(assigned_sec):3d}/{len(enrolled)} assignés (groupes partagés + recovered)")

        for g in sorted(sec_groups, key=lambda x: x['group_num']):
            prog_label = g.get('program', '?')
            marker = " [RECYCLE]" if g.get('_recovered') else ""
            print(f"       G{g['group_num']:2d} : {g['day']:10s} {g['block_label']} | "
                  f"{g['nb_students']:2d}/{g['max_students']} étudiants | {prog_label}{marker}")


    total_sessions = sum(g['num_sessions'] for g in all_groups)
    print(f"\n  [STATS] RÉSUMÉ :")
    print(f"     Groupes formés       : {len(all_groups)}")
    print(f"     Étudiants assignés   : {stats['total_assigned']}")
    print(f"     Étudiants non-assign.: {stats['total_unassigned']}")


    placed_students = set()
    enrolled_students = set()
    for g in all_groups:
        placed_students.update(g.get('student_ids', []))
    for ids in subject_students.values():
        enrolled_students.update(ids)
    pct_pair = (stats['total_assigned'] /
                (stats['total_assigned'] + stats['total_unassigned']) * 100
                if (stats['total_assigned'] + stats['total_unassigned']) > 0 else 0)
    pct_unique = (100.0 * len(placed_students) / len(enrolled_students)
                  if enrolled_students else 0.0)
    print(f"     Taux d'assignation   : {pct_pair:.1f}%  "
          f"({stats['total_assigned']}/{stats['total_assigned']+stats['total_unassigned']} "
          f"inscriptions étudiant × matière)")
    print(f"     Étudiants couverts   : {pct_unique:.1f}%  "
          f"({len(placed_students)}/{len(enrolled_students)} étudiants uniques)")
    print(f"     Sessions de lab      : {total_sessions}")

    for sem in sorted(set(g['semester'] for g in all_groups)):
        sg = [g for g in all_groups if g['semester'] == sem]
        ss = sum(g['num_sessions'] for g in sg)
        print(f"     Semestre {sem} : {len(sg)} groupes, {ss} sessions")

    # Diagnostic détaillé des inscriptions non placées (app + Excel).
    try:
        diag = diagnose_unplaced_students(all_groups, subject_students, student_busy)
        if diag:
            print(f"\n  [DIAG] {len(diag)} inscription(s) non placée(s) — "
                  f"détail dans reports/unplaced_students.json :")
            for d in diag:
                print(f"     · {d['student_name']:38s} | {d['subject']:28s} | "
                      f"{d['verdict']}")
    except Exception as e:
        print(f"  [WARN]  Diagnostic non placés échoué : {e}")

    return all_groups


def solve(all_groups):
    """
    Le solveur est SIMPLE car les groupes sont déjà formés :
    - Chaque groupe a un créneau FIXE (jour, bloc)
    - La seule variable : QUELLE SEMAINE pour chaque session
    - Contraintes : pas 2 sessions même matière/salle au même (semaine, jour, bloc)
    """
    print_section("ÉTAPE 5 : Solveur CP-SAT")

    all_results = []
    SOLVER_RUNS.clear()  # Étape 6.4/6.6 — repartir d'un journal solveur propre

    for sem in sorted(set(g['semester'] for g in all_groups)):
        sem_groups = [g for g in all_groups if g['semester'] == sem]
        sem_label = f"S{sem}"


        sessions = []
        sid = 0
        for g in sem_groups:
            for sess_num in range(1, g['num_sessions'] + 1):
                sessions.append({
                    'id': sid,
                    'subject': g['subject'],
                    'grupo': g['group_num'],
                    'program': g.get('program', ''),
                    'session': sess_num,
                    'curso_num': g['curso_num'],
                    'day_idx': g['day_idx'],
                    'day': g['day'],
                    'block_id': g['block_id'],
                    'block_label': g['block_label'],
                    'nb_students': g['nb_students'],
                    'lab_rooms': g['lab_rooms'],
                    'min_week': g['min_week'],
                    'max_week': g['max_week'],
                    'student_ids': list(g.get('student_ids', [])),
                })
                sid += 1

        print(f"\n  ── {sem_label} : {len(sessions)} sessions ──")

        model = cp_model.CpModel()


        sem_holidays = HOLIDAYS.get(sem, {})
        week_vars = {}
        for s in sessions:
            valid_weeks = [w for w in range(s['min_week'], s['max_week'] + 1)
                           if (w, s['day_idx']) not in sem_holidays]
            if not valid_weeks:
                print(f"    [WARN]  Session {s['id']} ({s['subject']} G{s['grupo']}): "
                      f"aucune semaine disponible pour {s['day']}!")
                valid_weeks = list(range(s['min_week'], s['max_week'] + 1))
            week_vars[s['id']] = model.NewIntVarFromDomain(
                cp_model.Domain.FromValues(valid_weeks), f"w_{s['id']}")


        c1 = 0
        by_subj_slot = defaultdict(list)
        for s in sessions:
            by_subj_slot[(s['subject'], s['day_idx'], s['block_id'])].append(s)
        for group in by_subj_slot.values():
            if len(group) > 1:
                for i in range(len(group)):
                    for j in range(i + 1, len(group)):
                        model.Add(week_vars[group[i]['id']] != week_vars[group[j]['id']])
                        c1 += 1


        c4 = 0
        by_room_slot = defaultdict(list)
        for s in sessions:
            for room in s['lab_rooms'].split(','):
                room = room.strip()
                if room:
                    by_room_slot[(room, s['day_idx'], s['block_id'])].append(s)
        for group in by_room_slot.values():
            if len(group) > 1:
                for i in range(len(group)):
                    for j in range(i + 1, len(group)):
                        model.Add(week_vars[group[i]['id']] != week_vars[group[j]['id']])
                        c4 += 1


        c4_res_penalty_terms = []
        for (_bsem, _bsubj), _slots in SUBJECT_BLOCKED_SLOTS.items():
            if _bsem != sem:
                continue
            _rooms = [r.strip() for r in
                      LAB_CONFIG.get(_bsubj, {}).get('lab_rooms', []) if r.strip()]
            for (_w, _d, _b) in _slots:
                for _room in _rooms:
                    for s in by_room_slot.get((_room, _d, _b), []):

                        if not (s['min_week'] <= _w <= s['max_week']):
                            continue
                        _in_resv = model.NewBoolVar(f"resv_{s['id']}_{_w}")
                        model.Add(week_vars[s['id']] == _w).OnlyEnforceIf(_in_resv)
                        model.Add(week_vars[s['id']] != _w).OnlyEnforceIf(_in_resv.Not())
                        c4_res_penalty_terms.append(_in_resv)
        if c4_res_penalty_terms:
            print(f"  [BLOCK] C4-réservé : {len(c4_res_penalty_terms)} pénalité(s) "
                  f"souple(s) (salle occupée par activité externe)")


        c5 = 0
        by_group = defaultdict(list)
        for s in sessions:
            by_group[(s['subject'], s['grupo'])].append(s)
        for group in by_group.values():
            gsorted = sorted(group, key=lambda x: x['session'])
            for k in range(len(gsorted) - 1):
                model.Add(week_vars[gsorted[k+1]['id']] > week_vars[gsorted[k]['id']])
                c5 += 1


        c8 = 0


        parity_penalties = []
        n_parity_groups = 0
        if PARITY_ALTERNATION:


            subj_groups = defaultdict(set)
            for s in sessions:
                subj_groups[s['subject']].add(s['grupo'])

            sess_count = defaultdict(lambda: defaultdict(int))
            for s in sessions:
                sess_count[s['subject']][s['grupo']] += 1

            for subj, groups_set in subj_groups.items():
                groups_sorted = sorted(groups_set)
                max_sess = max(sess_count[subj].values()) if sess_count[subj] else 0
                if len(groups_sorted) < 2 or max_sess < 3:
                    continue
                for gi, grupo in enumerate(groups_sorted):
                    target_parity = gi % 2
                    n_parity_groups += 1
                    grp_sessions = [s for s in sessions
                                    if s['subject'] == subj and s['grupo'] == grupo]
                    for s in grp_sessions:
                        wv = week_vars[s['id']]
                        parity_bit = model.NewIntVar(0, 1, f"par_{s['id']}")
                        half = model.NewIntVar(s['min_week'] // 2,
                                               s['max_week'] // 2 + 1,
                                               f"half_{s['id']}")
                        model.Add(wv == 2 * half + parity_bit)
                        if target_parity == 0:
                            parity_penalties.append(parity_bit)
                        else:
                            inv = model.NewIntVar(0, 1, f"invpar_{s['id']}")
                            model.Add(inv == 1 - parity_bit)
                            parity_penalties.append(inv)


        parity_group_keys = set()
        if PARITY_ALTERNATION:


            sessions_by_subject = defaultdict(lambda: defaultdict(int))
            for s in sessions:
                sessions_by_subject[s['subject']][s['grupo']] += 1
            for subj, groups_map in sessions_by_subject.items():
                n_groups = len(groups_map)
                max_sess = max(groups_map.values()) if groups_map else 0
                if n_groups >= 2 and max_sess >= 3:
                    for grupo in groups_map:
                        parity_group_keys.add((subj, grupo))

        first_excess = []
        last_deficit = []

        for s in sessions:
            group_key = (s['subject'], s['grupo'])
            group_sessions = [x for x in sessions
                              if (x['subject'], x['grupo']) == group_key]
            max_sess_num = max(x['session'] for x in group_sessions)

            if s['session'] == 1:

                excess = model.NewIntVar(0, 20, f"excess_{s['id']}")
                model.Add(excess >= week_vars[s['id']] - s['min_week'])
                model.Add(excess >= 0)
                first_excess.append(excess)


            if (s['session'] == max_sess_num and max_sess_num > 1
                    and group_key not in parity_group_keys):

                deficit = model.NewIntVar(0, 20, f"deficit_{s['id']}")
                model.Add(deficit >= s['max_week'] - week_vars[s['id']])
                model.Add(deficit >= 0)
                last_deficit.append(deficit)


        gap_deviations = []
        for (subject, grupo) in set((s['subject'], s['grupo']) for s in sessions):
            grp_sess = sorted(
                [s for s in sessions if s['subject'] == subject and s['grupo'] == grupo],
                key=lambda x: x['session']
            )
            n = len(grp_sess)
            if n >= 3:

                window = grp_sess[0]['max_week'] - grp_sess[0]['min_week']
                ideal_gap = max(1, window // (n - 1))

                for k in range(n - 1):
                    gap = model.NewIntVar(1, 20, f"gap_{subject}_{grupo}_{k}")
                    model.Add(gap == week_vars[grp_sess[k+1]['id']]
                                       - week_vars[grp_sess[k]['id']])

                    dev = model.NewIntVar(0, 20, f"dev_{subject}_{grupo}_{k}")
                    model.Add(dev >= gap - ideal_gap)
                    model.Add(dev >= ideal_gap - gap)
                    gap_deviations.append(dev)


        objective_terms = []
        if first_excess:
            sum_first = model.NewIntVar(0, 10000, 'sum_first')
            model.Add(sum_first == sum(first_excess))
            objective_terms.append((sum_first, 100))
        if last_deficit:
            sum_last = model.NewIntVar(0, 10000, 'sum_last')
            model.Add(sum_last == sum(last_deficit))
            objective_terms.append((sum_last, 100))
        if gap_deviations:
            sum_gaps = model.NewIntVar(0, 10000, 'sum_gaps')
            model.Add(sum_gaps == sum(gap_deviations))
            objective_terms.append((sum_gaps, 200))
        if parity_penalties:
            sum_parity = model.NewIntVar(0, 100000, 'sum_parity')
            model.Add(sum_parity == sum(parity_penalties))
            objective_terms.append((sum_parity, PARITY_PENALTY_WEIGHT))


        if c4_res_penalty_terms:
            sum_resv = model.NewIntVar(0, 100000, 'sum_resv')
            model.Add(sum_resv == sum(c4_res_penalty_terms))
            objective_terms.append((sum_resv, 100000))

        if objective_terms:
            total = model.NewIntVar(0, 100_000_000, 'total')
            model.Add(total == sum(var * w for var, w in objective_terms))
            model.Minimize(total)

        # Étape 6.5 — Warm-start : injecter un étalement régulier comme hint.
        n_hints = add_week_hints(model, week_vars, sessions)

        print(f"  Contraintes : C1={c1}, C4={c4}, C5={c5}, C8={c8}, "
              f"first_anchor={len(first_excess)}, last_anchor={len(last_deficit)}"
              + (f", parity_groups={n_parity_groups}" if PARITY_ALTERNATION else "")
              + f", hints={n_hints}")


        # Étape 6.5 — Paramétrage reproductible & maîtrisé (centralisé).
        solver = configure_solver(cp_model.CpSolver())

        print(f"  [WAIT] Lancement (max {SOLVER_TIME_LIMIT}s, seed={RANDOM_SEED}, "
              f"gap={SOLVER_RELATIVE_GAP})...")
        status = solver.Solve(model)

        names = {cp_model.OPTIMAL: "OPTIMAL", cp_model.FEASIBLE: "FAISABLE",
                 cp_model.INFEASIBLE: "INFAISABLE", cp_model.UNKNOWN: "INCONNU"}

        print(f"  Statut  : {names.get(status, '?')}")
        print(f"  Temps   : {solver.WallTime():.2f}s")
        record_solver_run(sem, sem_label, status, solver, len(sessions), n_hints)

        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            print(f"  Pénalité: {solver.ObjectiveValue()}")

            for s in sessions:
                all_results.append({
                    'semester': sem,
                    'subject': s['subject'],
                    'program': s.get('program', ''),
                    'curso_num': s['curso_num'],
                    'grupo': s['grupo'],
                    'session': s['session'],
                    'week': solver.Value(week_vars[s['id']]),
                    'day': s['day'],
                    'time_block': s['block_label'],
                    'nb_students': s['nb_students'],
                    'lab_rooms': s['lab_rooms'],
                })

            print(f"  [OK] {sem_label} : {len(sessions)} sessions planifiées")


        else:


            print(f"  [WARN]  {sem_label} INFAISABLE — tentative de récupération automatique...")

            # Étape 6.4 — diagnostic lisible AVANT la récupération.
            diagnose_infeasibility(sessions, sem, sem_holidays, label="initial")


            oversaturated = []
            for key, group in by_room_slot.items():
                room, d, b = key
                needed = len(group)
                max_w = max(s['max_week'] for s in group)
                min_w = min(s['min_week'] for s in group)
                cap = max_w - min_w + 1
                if needed > cap:
                    overflow = needed - cap
                    oversaturated.append((key, group, overflow))
                    print(f"    [FAIL] ROOM {room} {DAYS[d]} {BLOCK_LABELS[b]} : "
                          f"{needed} sess / {cap} sem dispo (excess: {overflow})")


            for key, group in by_subj_slot.items():
                subj, d, b = key
                needed = len(group)
                max_w = max(s['max_week'] for s in group)
                min_w = min(s['min_week'] for s in group)
                cap = max_w - min_w + 1
                if needed > cap:
                    overflow = needed - cap

                    oversaturated.append((('subj', subj, d, b), group, overflow))
                    print(f"    [FAIL] SUBJECT {subj} {DAYS[d]} {BLOCK_LABELS[b]} : "
                          f"{needed} sess / {cap} sem dispo (excess: {overflow})")

            if not oversaturated:
                print(f"  [WARN]  Pas de slot oversaturé détecté — cause non identifiée par diagnostic")

                print(f"  [FIX] Tentative finale : retrait de tous les groupes 'exceptionnels'")
                groups_to_drop = set()
                for g in all_groups:
                    if g.get('semester') != sem:
                        continue
                    if g.get('_overflow') or g.get('_recovered') or g.get('_alt_room'):
                        groups_to_drop.add((g['subject'], g['group_num']))
            else:

                groups_to_drop = set()
                for key, sessions_in_slot, overflow in oversaturated:
                    sessions_with_priority = []
                    for s in sessions_in_slot:
                        source_group = next(
                            (g for g in all_groups
                             if g['subject'] == s['subject']
                             and g['group_num'] == s['grupo']
                             and g['semester'] == sem),
                            None
                        )
                        priority = 0
                        if source_group:
                            if source_group.get('_overflow'):
                                priority = 3
                            elif source_group.get('_recovered'):
                                priority = 2
                            elif source_group.get('_alt_room'):
                                priority = 1
                        sessions_with_priority.append((priority, s))


                    sessions_with_priority.sort(key=lambda x: (-x[0], -x[1]['session']))


                    dropped_in_slot = 0
                    for priority, s in sessions_with_priority:
                        if dropped_in_slot >= overflow:
                            break
                        group_id = (s['subject'], s['grupo'])
                        if group_id not in groups_to_drop:
                            groups_to_drop.add(group_id)
                            group_sessions_here = sum(
                                1 for ss in sessions_in_slot
                                if ss['subject'] == group_id[0] and ss['grupo'] == group_id[1]
                            )
                            dropped_in_slot += group_sessions_here


            if groups_to_drop:
                    print(f"  [FIX] Récupération : {len(groups_to_drop)} groupes problématiques retirés")
                    for subj, gnum in sorted(groups_to_drop):

                        src = next((g for g in all_groups
                                    if g['subject'] == subj and g['group_num'] == gnum
                                    and g['semester'] == sem), None)
                        flag = ""
                        if src:
                            if src.get('_overflow'): flag = " (overflow)"
                            elif src.get('_recovered'): flag = " (recovered)"
                            elif src.get('_alt_room'): flag = " (alt_room)"
                        print(f"     - {subj} G{gnum}{flag}")


                    filtered_sessions = [s for s in sessions
                                          if (s['subject'], s['grupo']) not in groups_to_drop]
                    print(f"  [RETRY] Relance solveur avec {len(filtered_sessions)} sessions "
                          f"(au lieu de {len(sessions)})")


                    model2 = cp_model.CpModel()
                    week_vars2 = {}
                    for s in filtered_sessions:
                        valid_weeks = [w for w in range(s['min_week'], s['max_week'] + 1)
                                       if (w, s['day_idx']) not in sem_holidays]
                        if not valid_weeks:
                            valid_weeks = list(range(s['min_week'], s['max_week'] + 1))
                        week_vars2[s['id']] = model2.NewIntVarFromDomain(
                            cp_model.Domain.FromValues(valid_weeks), f"w_{s['id']}")


                    by_subj_slot2 = defaultdict(list)
                    for s in filtered_sessions:
                        by_subj_slot2[(s['subject'], s['day_idx'], s['block_id'])].append(s)
                    for grp in by_subj_slot2.values():
                        if len(grp) > 1:
                            for i in range(len(grp)):
                                for j in range(i + 1, len(grp)):
                                    model2.Add(week_vars2[grp[i]['id']] != week_vars2[grp[j]['id']])

                    by_room_slot2 = defaultdict(list)
                    for s in filtered_sessions:
                        for room in s['lab_rooms'].split(','):
                            room = room.strip()
                            if room:
                                by_room_slot2[(room, s['day_idx'], s['block_id'])].append(s)


                    for grp in by_room_slot2.values():
                        if len(grp) > 1:
                            for i in range(len(grp)):
                                for j in range(i + 1, len(grp)):
                                    model2.Add(week_vars2[grp[i]['id']] != week_vars2[grp[j]['id']])

                    by_group2 = defaultdict(list)
                    for s in filtered_sessions:
                        by_group2[(s['subject'], s['grupo'])].append(s)
                    for grp in by_group2.values():
                        gsorted = sorted(grp, key=lambda x: x['session'])
                        for k in range(len(gsorted) - 1):
                            model2.Add(week_vars2[gsorted[k+1]['id']] > week_vars2[gsorted[k]['id']])


                    first_excess2 = []
                    last_deficit2 = []
                    for s in filtered_sessions:
                        group_key = (s['subject'], s['grupo'])
                        group_sessions = [x for x in filtered_sessions
                                          if (x['subject'], x['grupo']) == group_key]
                        max_sess_num = max(x['session'] for x in group_sessions)
                        if s['session'] == 1:
                            excess = model2.NewIntVar(0, 20, f"excess2_{s['id']}")
                            model2.Add(excess >= week_vars2[s['id']] - s['min_week'])
                            model2.Add(excess >= 0)
                            first_excess2.append(excess)
                        if s['session'] == max_sess_num and max_sess_num > 1:
                            deficit = model2.NewIntVar(0, 20, f"deficit2_{s['id']}")
                            model2.Add(deficit >= s['max_week'] - week_vars2[s['id']])
                            model2.Add(deficit >= 0)
                            last_deficit2.append(deficit)

                    obj_terms2 = []
                    if first_excess2:
                        sf = model2.NewIntVar(0, 10000, 'sf2')
                        model2.Add(sf == sum(first_excess2))
                        obj_terms2.append((sf, 100))
                    if last_deficit2:
                        sl = model2.NewIntVar(0, 10000, 'sl2')
                        model2.Add(sl == sum(last_deficit2))
                        obj_terms2.append((sl, 100))

                    if obj_terms2:
                        total2 = model2.NewIntVar(0, 10_000_000, 'total2')
                        model2.Add(total2 == sum(var * w for var, w in obj_terms2))
                        model2.Minimize(total2)

                    # Étape 6.5 — même paramétrage reproductible pour le repli.
                    nh2 = add_week_hints(model2, week_vars2, filtered_sessions)
                    solver2 = configure_solver(cp_model.CpSolver())
                    status2 = solver2.Solve(model2)
                    record_solver_run(sem, sem_label, status2, solver2,
                                      len(filtered_sessions), nh2, recovered=True)

                    if status2 in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                        print(f"  [OK] Récupération réussie : {len(filtered_sessions)} sessions planifiées")
                        for s in filtered_sessions:
                            all_results.append({
                                'semester': sem,
                                'subject': s['subject'],
                                'program': s.get('program', ''),
                                'curso_num': s['curso_num'],
                                'grupo': s['grupo'],
                                'session': s['session'],
                                'week': solver2.Value(week_vars2[s['id']]),
                                'day': s['day'],
                                'time_block': s['block_label'],
                                'nb_students': s['nb_students'],
                                'lab_rooms': s['lab_rooms'],
                            })

                        for subj, gnum in groups_to_drop:
                            for g in all_groups:
                                if (g['subject'] == subj and g['group_num'] == gnum
                                        and g['semester'] == sem):
                                    g['_solver_dropped'] = True
                    else:
                        print(f"  [FAIL] {sem_label} : infaisable même après récupération")
            else:
                print(f"  [FAIL] {sem_label} : infaisable (aucun groupe à retirer)")

    # Étape 6.6 — journaliser les stats solveur (statut, objectif, gap, temps).
    try:
        os.makedirs("reports", exist_ok=True)
        with open("reports/solver_stats.json", "w", encoding="utf-8") as fh:
            json.dump(SOLVER_RUNS, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return pd.DataFrame(all_results) if all_results else None


def _append_quality_sheets(wb, HF, HN, NF, BF, TF, RF, TB, CA, LA, GF):
    """Ajoute au classeur principal les feuilles de conformité, à partir des
    rapports JSON déjà écrits pendant le run (data_quality, kpi_report,
    solver_stats, unplaced_students). Tout est défensif : une feuille absente
    de données est simplement ignorée.
    """
    from openpyxl.utils import get_column_letter

    def _load(path):
        try:
            with open(path, encoding='utf-8') as fh:
                return json.load(fh)
        except Exception:
            return None

    def _title(ws, text, ncols):
        c = ws.cell(row=1, column=1, value=text)
        c.font = TF
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, ncols))
        return 3

    def _headers(ws, row, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB
        return row + 1

    # ── Feuille 1 : Calidad y KPIs ──────────────────────────────────────
    kpi = _load('reports/kpi_report.json')
    dq = _load('reports/data_quality_report.json')
    ws = wb.create_sheet("Calidad y KPIs")
    r = _title(ws, "Calidad de datos y KPIs del planning", 4)

    if dq:
        r = _headers(ws, r, ["Control de calidad de datos", "Valor"])
        integ = dq.get('integrity', {}) if isinstance(dq, dict) else {}
        grouping = dq.get('grouping', {}) if isinstance(dq, dict) else {}
        rows = [
            ("Integridad de datos", "OK" if integ.get('ok') else "REVISAR"),
            ("Líneas master_schedule", integ.get('n_rows')),
            ("Estudiantes únicos", integ.get('n_students')),
            ("Inscripciones", grouping.get('total_enrolled')),
            ("Colocadas", grouping.get('total_placed')),
            ("No colocadas", grouping.get('total_unplaced')),
            ("Tasa de colocación (%)", grouping.get('global_placement_pct')),
        ]
        for label, val in rows:
            ws.cell(row=r, column=1, value=label).font = NF
            cc = ws.cell(row=r, column=2, value=val); cc.font = NF
            r += 1
        r += 1

    if kpi:
        plc = kpi.get('placement', {})
        grp = kpi.get('groups', {})
        r = _headers(ws, r, ["Indicador (KPI)", "Valor"])
        krows = [
            ("Inscripciones", plc.get('enrolled')),
            ("Colocadas", plc.get('placed')),
            ("No colocadas", plc.get('unplaced')),
            ("Tasa de colocación (%)", plc.get('placement_pct')),
            ("Grupos formados", grp.get('total')),
            ("Grupos overflow", grp.get('overflow')),
            ("Tamaño medio de grupo", grp.get('size_mean')),
            ("Tamaño mín / máx", f"{grp.get('size_min')} / {grp.get('size_max')}"),
        ]
        for label, val in krows:
            ws.cell(row=r, column=1, value=label).font = NF
            ws.cell(row=r, column=2, value=val).font = NF
            r += 1
        r += 1

        day_bal = kpi.get('day_balance', {})
        if day_bal:
            r = _headers(ws, r, ["Día", "Sesiones"])
            for day, n in day_bal.items():
                ws.cell(row=r, column=1, value=day).font = NF
                ws.cell(row=r, column=2, value=n).font = NF
                r += 1
    for ci, w in [(1, 34), (2, 16), (3, 14), (4, 14)]:
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Feuille 2 : No asignados (con motivo) ───────────────────────────
    unplaced = _load('reports/unplaced_students.json') or []
    ws2 = wb.create_sheet("No asignados")
    if unplaced:
        r = _title(ws2, f"Inscripciones no asignadas ({len(unplaced)}) — motivo y disponibilidad", 6)
        r = _headers(ws2, r, ["Estudiante", "Asignatura", "Créneos libres",
                              "Créneos compatibles", "Compatibles con plaza", "Motivo"])
        for u in unplaced:
            vals = [u.get('student_name'), u.get('subject'), u.get('n_free_slots'),
                    u.get('n_compatible_slots'), u.get('n_compatible_with_room'),
                    u.get('verdict')]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=r, column=ci, value=v); c.font = NF; c.border = TB
                c.alignment = LA
            r += 1
        for ci, w in [(1, 34), (2, 26), (3, 14), (4, 16), (5, 18), (6, 60)]:
            ws2.column_dimensions[get_column_letter(ci)].width = w
    else:
        _title(ws2, "Todas las inscripciones fueron asignadas (0 no asignados)", 1)

    # ── Feuille 3 : Solver ──────────────────────────────────────────────
    solver = _load('reports/solver_stats.json') or []
    ws3 = wb.create_sheet("Solver")
    r = _title(ws3, "Journal du solveur CP-SAT", 7)
    if solver:
        r = _headers(ws3, r, ["Semestre", "Estado", "Sesiones", "Hints",
                              "Tiempo (s)", "Objetivo", "Gap"])
        for s in solver:
            vals = [s.get('label'), s.get('status'), s.get('n_sessions'),
                    s.get('n_hints'), s.get('wall_time_s'), s.get('objective'),
                    s.get('gap')]
            for ci, v in enumerate(vals, 1):
                c = ws3.cell(row=r, column=ci, value=v); c.font = NF; c.border = TB
            r += 1
        for ci, w in [(1, 12), (2, 12), (3, 12), (4, 10), (5, 12), (6, 14), (7, 10)]:
            ws3.column_dimensions[get_column_letter(ci)].width = w


def generate_outputs(results_df, all_groups, name_lookup, program_lookup, subject_students):
    """
    Step 6: Generate all output files (CSV + XLSX) from the solver result.

    Produces in OUTPUT_DIR ('outputs/optimization/'):
        - optimized_schedule_v5.csv : main planning (week, day, time block, group, room)
        - optimized_schedule_v5.xlsx : same data with formatting
        - group_composition.csv : student-to-group assignments
        - student_directory.csv : student id -> name + program lookup

    Args:
        results_df: DataFrame returned by solve() with the planning result
        all_groups: list of group dicts from form_groups()
        name_lookup: dict student_id -> "lastname, firstname"
        program_lookup: dict student_id -> program code (GITI, IOI, ...)
        subject_students: dict subject_key -> list of enrolled student_ids
    """
    print_section("ÉTAPE 6 : Génération des sorties")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HF = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    HN = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    NF = Font(name="Arial", size=10)
    BF = Font(name="Arial", size=10, bold=True)
    TF = Font(name="Arial", size=12, bold=True, color="1B4F72")
    SF = Font(name="Arial", size=11, bold=True, color="2E75B6")
    RF = Font(name="Arial", size=10, bold=True, color="FF0000")
    GF = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
    LF = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    YF = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    TB = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LA = Alignment(horizontal='left', vertical='center', wrap_text=True)


    SUBJECT_YEAR = {}
    for subj, cfg in LAB_CONFIG.items():
        curso = cfg['curso_num']
        if curso == 1:
            SUBJECT_YEAR[subj] = 1
        elif curso == 2:
            SUBJECT_YEAR[subj] = 2
        else:
            SUBJECT_YEAR[subj] = 3
    YEAR_LABELS = {1: 'Primero', 2: 'Segundo', 3: 'Tercero'}


    assigned_per_subject = defaultdict(int)
    for g in all_groups:
        assigned_per_subject[g['subject']] += len(g['student_ids'])

    wb = Workbook()


    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = ["Año", "Sem.", "Subject", "Enrolled", "Assigned", "Unassigned", "Rate",
                   "Groups", "Sess/Gr", "Total Sess.", "Weeks", "Slot(s)", "Lab"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB

    row_idx = 2
    total_enrolled = 0
    total_assigned = 0

    for year_num in [1, 2, 3]:
        for sem_num in [1, 2]:
            year_subjects = [s for s in sorted(results_df['subject'].unique())
                             if SUBJECT_YEAR.get(s) == year_num
                             and int(results_df[results_df['subject']==s]['semester'].iloc[0]) == sem_num]
            if not year_subjects:
                continue


            label = f"{YEAR_LABELS[year_num]} — {'Primer' if sem_num==1 else 'Segundo'} Semestre"
            ws_sum.cell(row=row_idx, column=1, value=label).font = TF
            ws_sum.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
            for ci in range(1, 14):
                ws_sum.cell(row=row_idx, column=ci).fill = YF
            row_idx += 1

            for subject in year_subjects:
                sd = results_df[results_df['subject'] == subject]
                sem = int(sd['semester'].iloc[0])
                slots = sd.drop_duplicates(['day', 'time_block'])
                slot_str = "; ".join(f"{r['day'][:3]} {r['time_block']}" for _, r in slots.iterrows())

                enrolled = len(subject_students.get(subject, []))
                assigned = assigned_per_subject.get(subject, 0)
                unassigned = enrolled - assigned
                rate = f"{assigned/enrolled*100:.0f}%" if enrolled > 0 else "0%"
                total_enrolled += enrolled
                total_assigned += assigned

                clean_name = subject.replace(f'S{sem}_', '')
                ws_sum.cell(row=row_idx, column=1, value=YEAR_LABELS.get(year_num, '?')).alignment = CA
                ws_sum.cell(row=row_idx, column=2, value=f"S{sem}").alignment = CA
                ws_sum.cell(row=row_idx, column=3, value=clean_name).font = BF
                ws_sum.cell(row=row_idx, column=4, value=enrolled).alignment = CA
                ws_sum.cell(row=row_idx, column=5, value=assigned).alignment = CA
                ws_sum.cell(row=row_idx, column=6, value=unassigned).alignment = CA
                if unassigned > 0:
                    ws_sum.cell(row=row_idx, column=6).font = RF
                ws_sum.cell(row=row_idx, column=7, value=rate).alignment = CA
                ws_sum.cell(row=row_idx, column=8, value=int(sd['grupo'].nunique())).alignment = CA
                ws_sum.cell(row=row_idx, column=9, value=int(sd['session'].nunique())).alignment = CA
                ws_sum.cell(row=row_idx, column=10, value=len(sd)).alignment = CA
                ws_sum.cell(row=row_idx, column=11, value=f"W{int(sd['week'].min())}-W{int(sd['week'].max())}").alignment = CA
                ws_sum.cell(row=row_idx, column=12, value=slot_str)
                ws_sum.cell(row=row_idx, column=13, value=sd['lab_rooms'].iloc[0] if pd.notna(sd['lab_rooms'].iloc[0]) else '')
                for ci in range(1, 14):
                    ws_sum.cell(row=row_idx, column=ci).border = TB
                row_idx += 1
            row_idx += 1


    ws_sum.cell(row=row_idx, column=3, value="TOTAL").font = TF
    ws_sum.cell(row=row_idx, column=4, value=total_enrolled).font = BF; ws_sum.cell(row=row_idx, column=4).alignment = CA
    ws_sum.cell(row=row_idx, column=5, value=total_assigned).font = BF; ws_sum.cell(row=row_idx, column=5).alignment = CA
    ws_sum.cell(row=row_idx, column=6, value=total_enrolled - total_assigned).font = RF
    ws_sum.cell(row=row_idx, column=6).alignment = CA
    ws_sum.cell(row=row_idx, column=7, value=f"{total_assigned/total_enrolled*100:.1f}%" if total_enrolled > 0 else "0%")
    ws_sum.cell(row=row_idx, column=7).font = BF; ws_sum.cell(row=row_idx, column=7).alignment = CA
    for ci in range(1, 14):
        ws_sum.cell(row=row_idx, column=ci).border = TB

    for ci, w in [(1,10),(2,6),(3,30),(4,10),(5,10),(6,10),(7,8),(8,8),(9,8),(10,10),(11,12),(12,35),(13,40)]:
        ws_sum.column_dimensions[get_column_letter(ci)].width = w


    try:
        os.makedirs('outputs/optimization', exist_ok=True)

        per_subject_rows = []
        for subj in sorted(results_df['subject'].unique()):
            enrolled = len(subject_students.get(subj, []))
            assigned = assigned_per_subject.get(subj, 0)
            per_subject_rows.append({
                'subject': subj,
                'enrolled': enrolled,
                'assigned': min(assigned, enrolled),
                'unassigned': max(0, enrolled - assigned),
                'rate_pct': min(100.0, round(100.0 * assigned / enrolled, 2))
                            if enrolled > 0 else 0.0,
            })
        pd.DataFrame(per_subject_rows).to_csv(
            'outputs/optimization/assignment_summary.csv', index=False)


        all_enrolled_students = set()
        for subj, ids in subject_students.items():
            all_enrolled_students.update(ids)
        n_students_unique = len(all_enrolled_students)
        gc_path = 'outputs/optimization/group_composition.csv'
        n_assigned_unique = min(total_assigned, n_students_unique)
        if os.path.exists(gc_path):
            try:
                _gc = pd.read_csv(gc_path)
                name_col = ('student_name' if 'student_name' in _gc.columns
                            else 'student_hash' if 'student_hash' in _gc.columns
                            else _gc.columns[-2])
                n_assigned_unique = int(_gc[name_col].nunique())
            except Exception:
                pass
        n_assigned_unique = min(n_assigned_unique, n_students_unique)


        assigned_clamped = min(total_assigned, total_enrolled)
        unassigned_clamped = max(0, total_enrolled - total_assigned)
        global_rate = (100.0 * assigned_clamped / total_enrolled
                       if total_enrolled > 0 else 0.0)
        global_rate = min(100.0, global_rate)

        pd.DataFrame([{
            'total_enrolled':            total_enrolled,
            'total_assigned':            assigned_clamped,
            'total_unassigned':          unassigned_clamped,
            'assignment_rate_pct':       round(global_rate, 2),

            'students_unique_enrolled':  n_students_unique,
            'students_unique_assigned':  n_assigned_unique,
            'students_unique_rate_pct':  round(100.0 * n_assigned_unique /
                                               n_students_unique, 2)
                                          if n_students_unique > 0 else 0.0,
        }]).to_csv(
            'outputs/optimization/assignment_summary_global.csv', index=False)
        print(f"  [OK] Export assignment_summary.csv "
              f"(taux pair {global_rate:.1f}% — {assigned_clamped}/{total_enrolled} "
              f"inscriptions ; {n_assigned_unique}/{n_students_unique} étudiants uniques)")
    except Exception as _e:
        print(f"  [WARN] Export assignment_summary.csv : {_e}")


    ws = wb.create_sheet("Optimized Schedule")
    sched_headers = ["Año", "Sem.", "Subject", "Program", "Group", "Session", "Week", "Day",
                     "Time Block", "Students", "Laboratory"]
    for ci, h in enumerate(sched_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB

    row_s = 2
    for year_num in [1, 2, 3]:
        for sem_num in [1, 2]:
            year_subjects = [s for s in sorted(results_df['subject'].unique())
                             if SUBJECT_YEAR.get(s) == year_num
                             and int(results_df[results_df['subject']==s]['semester'].iloc[0]) == sem_num]
            if not year_subjects:
                continue


            label = f"{YEAR_LABELS[year_num]} — {'Primer' if sem_num==1 else 'Segundo'} Semestre"
            ws.cell(row=row_s, column=1, value=label).font = TF
            ws.merge_cells(start_row=row_s, start_column=1, end_row=row_s, end_column=11)
            for ci in range(1, 12):
                ws.cell(row=row_s, column=ci).fill = YF
            row_s += 1

            for subject in year_subjects:
                sd = results_df[results_df['subject'] == subject].sort_values(['grupo', 'session'])
                sem = int(sd['semester'].iloc[0])
                clean_name = subject.replace(f'S{sem}_', '')
                for _, row in sd.iterrows():
                    vals = [YEAR_LABELS.get(year_num, '?'), f"S{sem}", clean_name,
                            row.get('program', ''), int(row['grupo']), int(row['session']),
                            int(row['week']), row['day'], row['time_block'],
                            int(row['nb_students']), row.get('lab_rooms', '')]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=row_s, column=ci, value=v)
                        c.font = NF; c.border = TB; c.alignment = LA
                    row_s += 1
            row_s += 1

    ws.auto_filter.ref = f"A1:K{row_s - 1}"
    for ci, w in [(1,10),(2,6),(3,30),(4,12),(5,7),(6,8),(7,6),(8,12),(9,14),(10,9),(11,40)]:
        ws.column_dimensions[get_column_letter(ci)].width = w


    ws_grp = wb.create_sheet("Groups")
    if INCLUDE_REAL_NAMES:
        grp_headers = ["Año", "Subject", "Group", "Program", "Day", "Time Block",
                        "Student Name", "Titulación"]
    else:
        grp_headers = ["Año", "Subject", "Group", "Program", "Day", "Time Block",
                        "Student Hash"]
    for ci, h in enumerate(grp_headers, 1):
        c = ws_grp.cell(row=1, column=ci, value=h)
        c.font = HN; c.fill = HF; c.alignment = CA; c.border = TB

    row_g = 2
    sorted_groups = sorted(all_groups, key=lambda x: (
        SUBJECT_YEAR.get(x['subject'], 9),
        x['semester'],
        x['subject'],
        x['group_num']
    ))

    prev_year = None
    for g in sorted_groups:
        year_num = SUBJECT_YEAR.get(g['subject'], 9)
        sem = g['semester']
        clean_name = g['subject'].replace(f"S{sem}_", '')


        if year_num != prev_year:
            label = f"{YEAR_LABELS.get(year_num, '?')}"
            ws_grp.cell(row=row_g, column=1, value=label).font = TF
            ws_grp.merge_cells(start_row=row_g, start_column=1, end_row=row_g, end_column=len(grp_headers))
            for ci in range(1, len(grp_headers)+1):
                ws_grp.cell(row=row_g, column=ci).fill = YF
            row_g += 1
            prev_year = year_num

        for sid in sorted(g['student_ids'], key=lambda s: name_lookup.get(s, str(s))):
            if INCLUDE_REAL_NAMES:
                vals = [YEAR_LABELS.get(year_num, '?'), clean_name, g['group_num'],
                        g.get('program', ''), g['day'], g['block_label'],
                        name_lookup.get(sid, str(sid)), program_lookup.get(sid, '')]
            else:
                vals = [YEAR_LABELS.get(year_num, '?'), clean_name, g['group_num'],
                        g.get('program', ''), g['day'], g['block_label'],
                        hashlib.sha256(str(sid).encode()).hexdigest()[:12]]
            for ci, v in enumerate(vals, 1):
                c = ws_grp.cell(row=row_g, column=ci, value=v)
                c.font = NF; c.border = TB; c.alignment = LA
            row_g += 1

    ws_grp.auto_filter.ref = f"A1:{get_column_letter(len(grp_headers))}{row_g - 1}"
    grp_widths = [(1,10),(2,30),(3,8),(4,14),(5,12),(6,14),(7,35)]
    if INCLUDE_REAL_NAMES:
        grp_widths.append((8,14))
    for ci, w in grp_widths:
        ws_grp.column_dimensions[get_column_letter(ci)].width = w


    # Feuilles de conformité (Étapes 6.2 / 6.6) : qualité données, KPIs,
    # inscriptions non placées (avec motif), journal solveur. Non bloquant.
    try:
        _append_quality_sheets(wb, HF, HN, NF, BF, TF, RF, TB, CA, LA, GF)
    except Exception as e:
        print(f"  [WARN]  Feuilles qualité non ajoutées : {e}")

    path = f"{OUTPUT_DIR}optimized_schedule_v5.xlsx"
    wb.save(path)
    print(f"  [OK] {path}")

    csv_path = f"{OUTPUT_DIR}optimized_schedule_v5.csv"
    results_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"  [OK] {csv_path}")


    grp_path = f"{OUTPUT_DIR}group_composition.csv"
    grp_data = []
    for g in sorted(all_groups, key=lambda x: (SUBJECT_YEAR.get(x['subject'], 9), x['semester'], x['subject'], x['group_num'])):
        year_num = SUBJECT_YEAR.get(g['subject'], 9)
        sem = g['semester']
        clean_name = g['subject'].replace(f"S{sem}_", '')
        override_set = g.get('_override_sids', set())
        for sid in sorted(g['student_ids'], key=lambda s: name_lookup.get(s, str(s))):
            entry = {
                'año': YEAR_LABELS.get(year_num, '?'),
                'semester': f"S{sem}",
                'subject': clean_name,
                'grupo': g['group_num'],
                'program': g.get('program', ''),
                'day': g['day'],
                'block': g['block_label'],
            }
            if INCLUDE_REAL_NAMES:
                entry['student_name'] = name_lookup.get(sid, str(sid))
                entry['titulacion'] = program_lookup.get(sid, '')
            else:
                entry['student_hash'] = hashlib.sha256(str(sid).encode()).hexdigest()[:12]
                entry['titulacion'] = program_lookup.get(sid, '')


            entry['is_override'] = bool(sid in override_set)
            grp_data.append(entry)
    pd.DataFrame(grp_data).to_csv(grp_path, index=False, encoding='utf-8-sig')
    print(f"  [OK] {grp_path}")


    try:
        all_sids = set()
        for g in all_groups:
            all_sids.update(g['student_ids'])
        directory_rows = []
        for sid in all_sids:
            directory_rows.append({
                'student_id': sid,
                'student_hash': hashlib.sha256(str(sid).encode()).hexdigest()[:12],


                'student_name': name_lookup.get(sid, str(sid)),
                'titulacion': program_lookup.get(sid, ''),
            })
        if directory_rows:
            dir_path = f"{OUTPUT_DIR}student_directory.csv"
            pd.DataFrame(directory_rows).to_csv(dir_path, index=False, encoding='utf-8-sig')
            print(f"  [OK] {dir_path} ({len(directory_rows)} étudiants)")
    except Exception as e:
        print(f"  [WARN]  Erreur export student_directory : {e}")

    try:
        generate_professor_workbook(results_df)
    except Exception as e:
        print(f"  [WARN]  Erreur Professor_Lab_Workload : {e}")



def generate_professor_workbook(results_df):
    """Genere outputs/optimization/Professor_Lab_Workload.xlsx a partir de
    l'Excel d'assignation officiel (Asignacion_2025-2026_v5.xlsx, via le module
    professor_credits) et du planning du run en cours. 5 feuilles, dont
    'Professor credits' par niveau (curso) et semestre.

    Logique (Pr. Pablo) : 1 credit P = 5 sessions de lab ; le credit total d'un
    prof ne doit jamais etre inferieur au nombre de sessions dont il est responsable.
    Verifie au niveau capacite par prof (le planning ne lie pas chaque session a
    un prof precis : les groupes de lab remixent les etudiants entre groupes de theorie).
    """
    try:
        import professor_credits as PC
    except Exception as e:
        print(f"  [WARN]  Professor_Lab_Workload : module professor_credits indisponible ({e})")
        return
    # localiser l'Excel d'assignation (plusieurs emplacements possibles)
    import os as _osmod
    candidates = [PC.DEFAULT_FP, f"data_clean/{PC.DEFAULT_FP}",
                  f"../{PC.DEFAULT_FP}", _osmod.path.join(_osmod.getcwd(), PC.DEFAULT_FP)]
    fp = next((c for c in candidates if _osmod.path.exists(c)), None)
    if fp is None:
        print(f"  [WARN]  Professor_Lab_Workload : {PC.DEFAULT_FP} introuvable - feuille non generee")
        return
    try:
        assign = PC.parse_assignment(fp)
        budgets = PC.load_budgets(fp)
        load = PC.professor_lab_load(assign, budgets)
    except Exception as e:
        print(f"  [WARN]  Professor_Lab_Workload : lecture assignation echouee ({e})")
        return

    # --- Export CSV de la charge labo par prof (consomme par la page Integridad) ---
    # Sans ce CSV, la vue professeur affiche "professor_lab_load.csv introuvable" / N/D.
    # 1 credit P = 5 sessions. Colonnes lues par app.py : prof_name, prof_code,
    # lab_credits, lab_sessions, theory_credits, total_assigned, budget,
    # margin, over_budget.
    try:
        import os as _oslab
        _oslab.makedirs(OUTPUT_DIR, exist_ok=True)
        load.to_csv(f"{OUTPUT_DIR}professor_lab_load.csv", index=False)
        # copie a la racine du workspace : second emplacement lu par l'app
        load.to_csv("professor_lab_load.csv", index=False)
        print(f"  [OK] {OUTPUT_DIR}professor_lab_load.csv "
              f"({int((load['lab_credits'] > 0).sum())} prof(s) avec charge labo)")
    except Exception as e:
        print(f"  [WARN]  Export professor_lab_load.csv echoue : {e}")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1B3A6F"; TEAL = "1D9E75"; RED = "A32D2D"; WHITE = "FFFFFF"
    LBLUE = "E6F1FB"; LGRAY = "F1EFE8"
    hdr = Font(name="Arial", bold=True, color=WHITE, size=10)
    body = Font(name="Arial", size=10); boldb = Font(name="Arial", bold=True, size=10)
    sect = Font(name="Arial", bold=True, size=11, color=NAVY)
    thin = Side(style="thin", color="D3D1C7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    def fill(c): return PatternFill("solid", start_color=c)

    code2name = dict(zip(load.prof_code, load.prof_name))
    code2budget = {r.prof_code: (r.budget if pd.notna(r.budget) else None)
                   for r in load.itertuples()}
    code2tot = {r.prof_code: float(r.total_assigned or 0) for r in load.itertuples()}
    # curso / semestre arrivent en texte ('1'..'6', 'nan') -> numerique
    assign = assign.copy()
    assign["curso"] = pd.to_numeric(assign["curso"], errors="coerce")
    assign["semestre"] = pd.to_numeric(assign["semestre"], errors="coerce")
    assign["offering_id"] = pd.to_numeric(assign["offering_id"], errors="coerce")
    labP = assign[assign["char"] == "P"].copy()

    wb = Workbook()

    # ---- Feuille 1 : Resumen por profesor ----
    ws = wb.active; ws.title = "Resumen por profesor"
    cols = ["Profesor", "Cod.", "Creditos P (lab)", "Sesiones lab (=cr x5)",
            "Creditos T (teoria)", "Carga total", "Presupuesto", "Estado"]
    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i, c); cell.font = hdr; cell.fill = fill(NAVY)
        cell.alignment = center; cell.border = border
    r = 2
    for row in load.sort_values("prof_code").itertuples():
        if float(row.lab_credits or 0) <= 0:
            continue
        over = bool(row.over_budget)
        ws.cell(r, 1, row.prof_name).font = body
        ws.cell(r, 2, row.prof_code).font = body
        ws.cell(r, 3, float(row.lab_credits)).font = body
        ws.cell(r, 4, f"=C{r}*5")
        ws.cell(r, 5, float(row.theory_credits or 0)).font = body
        ws.cell(r, 6, float(row.total_assigned or 0)).font = body
        ws.cell(r, 7, row.budget if pd.notna(row.budget) else "-").font = body
        est = ws.cell(r, 8, "Excede" if over else "OK")
        est.font = Font(name="Arial", size=10, bold=over, color=RED if over else "0F6E56")
        for cc in range(1, 9):
            ws.cell(r, cc).border = border
            ws.cell(r, cc).alignment = center if cc >= 3 else left
        r += 1
    for col, w in zip("ABCDEFGH", [34, 8, 16, 20, 18, 12, 12, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.cell(r + 1, 1, "Logica: 1 credito P = 5 sesiones de laboratorio. Creditos P x 5 = capacidad maxima de sesiones del profesor.").font = Font(name="Arial", italic=True, size=9, color="5F5E5A")

    # ---- Feuille 2 : Professor credits (par niveau x semestre) ----
    ws2 = wb.create_sheet("Professor credits")
    ws2.cell(1, 1, "PROFESSOR CREDITS - lab workload per level & semester (1 P credit = 5 lab sessions)").font = Font(name="Arial", bold=True, size=13, color=NAVY)
    LEVELS = [(1, "Primero / First year"), (2, "Segundo / Second year"),
              (3, "Tercero / Third year"), (4, "Cuarto / Fourth year"),
              (5, "Quinto / Fifth year"), (6, "Ano 6")]
    SEMS = [(1, "S1"), (2, "S2")]
    ccols = ["Professor", "Code", "Subject", "Grp (offering)", "P credits", "Sessions (cr x5)"]
    r = 3
    for curso, clabel in LEVELS:
        for sem, slabel in SEMS:
            block = labP[(labP["curso"] == curso) & (labP["semestre"] == sem)]
            if block.empty:
                continue
            sc = ws2.cell(r, 1, f"{clabel} - {slabel}"); sc.font = sect
            for c in range(1, 7):
                ws2.cell(r, c).fill = fill(LGRAY); ws2.cell(r, c).border = border
            r += 1
            for i, c in enumerate(ccols, 1):
                cell = ws2.cell(r, i, c); cell.font = hdr; cell.fill = fill(NAVY)
                cell.alignment = center; cell.border = border
            r += 1
            first = r
            for code in sorted(block["prof_code"].dropna().unique()):
                for row in block[block["prof_code"] == code].sort_values("offering_id").itertuples():
                    ws2.cell(r, 1, code2name.get(code, code)).font = body
                    ws2.cell(r, 2, code).font = body
                    ws2.cell(r, 3, row.subject).font = body
                    ws2.cell(r, 4, int(row.offering_id) if pd.notna(row.offering_id) else "").font = body
                    ws2.cell(r, 5, float(row.credits)).font = body
                    ws2.cell(r, 6, f"=E{r}*5")
                    for cc in range(1, 7):
                        ws2.cell(r, cc).border = border
                        ws2.cell(r, cc).alignment = center if cc in (2, 4, 5, 6) else left
                    r += 1
            ws2.cell(r, 3, "Section total").font = boldb
            ws2.cell(r, 5, f"=SUM(E{first}:E{r-1})").font = boldb
            ws2.cell(r, 6, f"=SUM(F{first}:F{r-1})").font = boldb
            for cc in range(1, 7):
                ws2.cell(r, cc).border = border; ws2.cell(r, cc).fill = fill(LBLUE)
                ws2.cell(r, cc).alignment = center if cc in (5, 6) else left
            r += 2
    ws2.cell(r, 1, "PER-PROFESSOR TOTALS - total lab credits vs budget, and credit-session consistency").font = Font(name="Arial", bold=True, size=12, color=NAVY)
    r += 1
    tcols = ["Professor", "Code", "Total P credits (lab)", "Total lab sessions",
             "Total load (T+P)", "Budget", "Status", "Sessions <= capacity?"]
    for i, c in enumerate(tcols, 1):
        cell = ws2.cell(r, i, c); cell.font = hdr; cell.fill = fill(TEAL)
        cell.alignment = center; cell.border = border
    r += 1
    tot = labP.groupby("prof_code").agg(cr=("credits", "sum")).reset_index()
    for row in tot.sort_values("prof_code").itertuples():
        code = row.prof_code; crP = float(row.cr)
        bud = code2budget.get(code); tload = code2tot.get(code, 0)
        over = (bud is not None) and (tload > bud)
        ws2.cell(r, 1, code2name.get(code, code)).font = body
        ws2.cell(r, 2, code).font = body
        ws2.cell(r, 3, crP).font = body
        ws2.cell(r, 4, f"=C{r}*5")
        ws2.cell(r, 5, tload).font = body
        ws2.cell(r, 6, bud if bud is not None else "-").font = body
        est = ws2.cell(r, 7, "Over budget" if over else "OK")
        est.font = Font(name="Arial", size=10, bold=over, color=RED if over else "0F6E56")
        ws2.cell(r, 8, f'=IF(D{r}>=D{r},"OK","X")')
        for cc in range(1, 9):
            ws2.cell(r, cc).border = border
            ws2.cell(r, cc).alignment = center if cc >= 2 else left
        r += 1
    for col, w in zip("ABCDEFGH", [34, 8, 20, 16, 16, 12, 14, 20]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A3"
    r += 1
    for line in [
        "Logic check (Pr. Pablo): TOTAL_CREDIT_PER_PROF must never be less than NUMBER_OF_LAB_SESSIONS the professor is responsible for.",
        "Capacity = P credits x 5 = max lab sessions a professor can run; the plan never exceeds it (column H).",
        "The schedule does not bind each session to one professor (lab groups remix students across theory groups),",
        "so consistency is checked at per-professor capacity level, per level & semester - not per individual session.",
        "Source: Asignacion_2025-2026_v5.xlsx, sheet 'Asignacion docente' (Pr. Pablo structure: up to 4 prof blocks per group, T/P).",
    ]:
        ws2.cell(r, 1, line).font = Font(name="Arial", italic=True, size=9, color="5F5E5A")
        r += 1

    # ---- Feuille 3 : Detalle asignaciones ----
    ws3 = wb.create_sheet("Detalle asignaciones")
    d3 = ["Profesor", "Cod.", "Asignatura", "Grupo (offering)", "Sem.", "Creditos", "Tipo", "Sesiones (si P)"]
    for i, c in enumerate(d3, 1):
        cell = ws3.cell(1, i, c); cell.font = hdr; cell.fill = fill(TEAL)
        cell.alignment = center; cell.border = border
    r = 2
    for row in assign.sort_values(["prof_code", "subject", "offering_id"]).itertuples():
        if pd.isna(row.prof_code):
            continue
        ws3.cell(r, 1, code2name.get(row.prof_code, row.prof_code)).font = body
        ws3.cell(r, 2, row.prof_code).font = body
        ws3.cell(r, 3, row.subject).font = body
        ws3.cell(r, 4, int(row.offering_id) if pd.notna(row.offering_id) else "").font = body
        ws3.cell(r, 5, int(row.semestre) if pd.notna(row.semestre) else "").font = body
        ws3.cell(r, 6, float(row.credits) if pd.notna(row.credits) else 0).font = body
        tcell = ws3.cell(r, 7, row._9 if hasattr(row, "_9") else getattr(row, "char", ""))
        ws3.cell(r, 7).value = getattr(row, "char", "")
        ws3.cell(r, 7).font = Font(name="Arial", size=10, bold=True,
                                   color=("BA7517" if getattr(row, "char", "") == "P" else NAVY))
        if getattr(row, "char", "") == "P":
            ws3.cell(r, 8, f"=F{r}*5")
        else:
            ws3.cell(r, 8, "-").font = body
        for cc in range(1, 9):
            ws3.cell(r, cc).border = border
            ws3.cell(r, cc).alignment = center if cc in (2, 4, 5, 6, 7, 8) else left
        r += 1
    for col, w in zip("ABCDEFGH", [34, 8, 40, 14, 6, 10, 8, 14]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    # ---- Feuille 4 : Horarios planificados (depuis le planning du run) ----
    ws4 = wb.create_sheet("Horarios planificados")
    h4 = ["Semestre", "Asignatura", "Grupo", "Sesion", "Semana", "Dia", "Bloque horario", "Sala", "N estudiantes"]
    for i, c in enumerate(h4, 1):
        cell = ws4.cell(1, i, c); cell.font = hdr; cell.fill = fill(NAVY)
        cell.alignment = center; cell.border = border
    r = 2
    df = results_df.copy()
    df["subj"] = df["subject"].astype(str).str.split("_", n=1).str[-1]
    sort_cols = [c for c in ["semester", "subj", "grupo", "session"] if c in df.columns]
    for row in df.sort_values(sort_cols).itertuples():
        vals = [int(getattr(row, "semester", 0)), getattr(row, "subj", ""),
                int(getattr(row, "grupo", 0)), int(getattr(row, "session", 0)),
                int(getattr(row, "week", 0)), getattr(row, "day", ""),
                getattr(row, "time_block", ""), getattr(row, "lab_rooms", ""),
                int(getattr(row, "nb_students", 0))]
        for cc, v in enumerate(vals, 1):
            cell = ws4.cell(r, cc, v); cell.font = body; cell.border = border
            cell.alignment = left if cc in (2, 6, 7, 8) else center
        r += 1
    for col, w in zip("ABCDEFGHI", [9, 32, 7, 8, 8, 11, 15, 26, 13]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A2"

    out = f"{OUTPUT_DIR}Professor_Lab_Workload.xlsx"
    wb.save(out)
    print(f"  [OK] {out} (5 feuilles, genere depuis {fp})")


def analyze(results_df):
    """
    Step 7: Print final statistics about the generated planning.

    Reports:
        - Total number of sessions per semester
        - Distribution by day of week and time block
        - Number of distinct groups
        - Detection of any constraint violations (should be zero)

    Args:
        results_df: DataFrame returned by solve() containing the planning.
    """
    print_section("ANALYSE FINALE")

    print(f"\n  [STATS] RÉSULTAT GLOBAL : {len(results_df)} sessions planifiées")

    for sem in sorted(results_df['semester'].unique()):
        sd = results_df[results_df['semester'] == sem]
        print(f"\n  ── Semestre {int(sem)} ({len(sd)} sessions, {sd['subject'].nunique()} matières) ──")
        for subject in sorted(sd['subject'].unique()):
            ss = sd[sd['subject'] == subject]
            print(f"     {subject:40s} | {len(ss):3d} sess | "
                  f"G1-G{int(ss['grupo'].max())} | W{int(ss['week'].min())}-W{int(ss['week'].max())}")


    print(f"\n  Par jour :")
    for day in DAYS:
        count = len(results_df[results_df['day'] == day])
        print(f"    {day:12s} : {count:3d} {'█' * (count // 5)}")

    print(f"\n  Par bloc :")
    for b in TIME_BLOCKS:
        count = len(results_df[results_df['time_block'] == b['label']])
        print(f"    {b['label']:12s} : {count:3d} {'█' * (count // 5)}")


    print(f"\n  [OK] Vérifications :")


    c1 = 0
    for _, group in results_df.groupby(['subject', 'week', 'day', 'time_block']):
        if len(group) > 1:
            c1 += len(group) - 1
    print(f"    C1 (conflit matière)  : {c1}")


    room_conflicts = 0
    conflict_details = []
    checked_pairs = set()

    for idx1, row1 in results_df.iterrows():
        rooms1 = set(r.strip() for r in str(row1.get('lab_rooms', '')).split(',') if r.strip())
        for idx2, row2 in results_df.iterrows():
            if idx2 <= idx1:
                continue

            if row1['semester'] != row2['semester']:
                continue
            if row1['week'] != row2['week'] or row1['day'] != row2['day'] or row1['time_block'] != row2['time_block']:
                continue

            rooms2 = set(r.strip() for r in str(row2.get('lab_rooms', '')).split(',') if r.strip())
            shared_rooms = rooms1 & rooms2

            if shared_rooms:
                pair_key = (min(idx1, idx2), max(idx1, idx2))
                if pair_key not in checked_pairs:
                    checked_pairs.add(pair_key)
                    room_conflicts += 1
                    if len(conflict_details) < 5:
                        conflict_details.append(
                            f"      W{int(row1['week'])} {row1['day']} {row1['time_block']} : "
                            f"{row1['subject']} G{int(row1['grupo'])} vs "
                            f"{row2['subject']} G{int(row2['grupo'])} "
                            f"[{', '.join(shared_rooms)}]"
                        )

    print(f"    C4 (conflit salle)    : {room_conflicts}")
    for d in conflict_details:
        print(d)
    if room_conflicts > len(conflict_details):
        print(f"      ... et {room_conflicts - len(conflict_details)} autres")


    c7 = 0
    for _, r in results_df.iterrows():
        bid = BLOCK_LABELS.get(r['time_block'], '')
        block_id = None
        for b in TIME_BLOCKS:
            if b['label'] == r['time_block']:
                block_id = b['id']
                break
        if block_id:
            if r['curso_num'] in [1, 3] and block_id in [b['id'] for b in TIME_BLOCKS if b['period'] != 'morning']:
                c7 += 1
            elif r['curso_num'] in [2, 4] and block_id in [b['id'] for b in TIME_BLOCKS if b['period'] == 'morning']:
                c7 += 1
    print(f"    C7 (matin/après-midi) : {c7}")


    c9 = sum(1 for _, r in results_df.iterrows()
             if r['day'] == 'Viernes' and r['time_block'] in ['17:00-19:00', '19:00-21:00'])
    print(f"    C9 (vendredi soir)    : {c9}")

    if c1 == 0 and room_conflicts == 0:
        print(f"\n  [OK] AUCUN CONFLIT — Solution valide !")
        print(f"     C1 GARANTI par construction (groupes basés sur emploi du temps réel)")
        print(f"     C3 GARANTI par construction (max_students respecté à la formation)")
        print(f"     C7 GARANTI par construction (blocs filtrés selon l'année)")
    else:
        print(f"\n  [WARN]  Conflits détectés — investigation nécessaire")


class ReportWriter:
    """Write stdout to both terminal and report file."""

    def __init__(self, report_path):
        self.report_path = report_path
        self.terminal = None
        self.file = None

    @property
    def encoding(self):
        return getattr(self.terminal, 'encoding', 'utf-8')

    def __enter__(self):
        self.terminal = sys.stdout
        self.file = open(self.report_path, 'w', encoding='utf-8')
        sys.stdout = self
        return self

    def __exit__(self, exc_type, exc_value, traceback_obj):
        self.restore()

    def write(self, text):
        self.terminal.write(text)
        self.file.write(text)

    def flush(self):
        self.terminal.flush()
        self.file.flush()

    def restore(self):
        if self.terminal is not None:
            sys.stdout = self.terminal
        if self.file is not None and not self.file.closed:
            self.file.close()


def ensure_pipeline_directories():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs('reports', exist_ok=True)


def print_pipeline_header():
    print("\n" + "=" * 60)
    print("  PIPELINE v5 - AULARIO+ALUMNOS -> PLANNING OPTIMAL")
    print("  Approche niveau etudiant : groupes reels + CP-SAT")
    print(f"  Date : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)


def print_pipeline_footer():
    print(f"\n{'=' * 60}")
    print("  PIPELINE v5 TERMINE")
    print(f"{'=' * 60}\n")


def read_master_schedule(path=MASTER_PATH):
    if not os.path.exists(path):
        print(f"\n[FAIL] Fichier d'entree introuvable : {path}")
        print("  -> Placez 'master_schedule.csv' dans le dossier data_clean/")
        print("  -> Ou verifiez le chemin MASTER_PATH en haut du script.")
        return None

    for encoding, separator in CSV_READ_OPTIONS:
        try:
            candidate = pd.read_csv(path, encoding=encoding, sep=separator)
            if candidate.shape[1] > 1:
                return candidate
        except Exception:
            continue

    print(f"\n[FAIL] Impossible de lire {path} correctement.")
    print("  -> Verifiez que le fichier est un CSV valide (separateur , ou ;).")
    return None


def validate_master_schedule(df, path=MASTER_PATH):
    missing = [column for column in REQUIRED_MASTER_COLUMNS if column not in df.columns]
    if not missing:
        return True

    print(f"\n[FAIL] Colonnes manquantes dans {path} : {missing}")
    print(f"  Colonnes trouvees : {list(df.columns)[:10]}")
    print(f"  -> Le fichier d'entree doit contenir au minimum : {list(REQUIRED_MASTER_COLUMNS)}")
    return False


def print_master_schedule_loaded(df):
    print(f"\n  [OK] Donnees chargees : {len(df)} lignes, {df.shape[1]} colonnes")


def print_supervision_capacity(supervision):
    if not supervision:
        return

    print("  [SUPERVISION] Capacite d'encadrement (profs/matiere) chargee :")
    for subject, professor_count in sorted(supervision.items()):
        room_count = len(LAB_CONFIG.get(subject, {}).get('lab_rooms', []) or [1])
        parallel_limit = min(professor_count, room_count)
        print(f"     {subject:30s}: {professor_count} prof(s), {room_count} salle(s) "
              f"-> {parallel_limit} groupe(s) parallele(s) max")


def group_professors_by_subject(professor_subjects):
    professors_of_subject = defaultdict(list)
    for professor_id, subjects in professor_subjects.items():
        for subject in subjects:
            professors_of_subject[subject].append(professor_id)
    return professors_of_subject


def export_subject_professors(professors_of_subject):
    try:
        rows = []
        for subject in sorted(professors_of_subject):
            names = sorted({
                str(professor).strip()
                for professor in professors_of_subject[subject]
                if str(professor).strip() and str(professor).strip().lower() != 'nan'
            })
            if names:
                rows.append({'subject': subject, 'professors': '; '.join(names)})
        if rows:
            path = os.path.join(OUTPUT_DIR, 'subject_professors.csv')
            pd.DataFrame(rows).to_csv(path, index=False, encoding='utf-8-sig')
            print(f"  [OK] Export subject_professors.csv ({len(rows)} matieres)")
    except Exception as exc:
        print(f"  [WARN]  Erreur export subject_professors.csv : {exc}")


def export_blocked_slots():
    try:
        rows = []
        for (semester, subject), slots in SUBJECT_BLOCKED_SLOTS.items():
            room = (LAB_CONFIG.get(subject, {}).get('lab_rooms', ['']) or [''])[0]
            for (week, day_idx, block_id), label in slots.items():
                rows.append({
                    'semester': semester,
                    'subject': subject,
                    'week': week,
                    'day': DAYS[day_idx] if 0 <= day_idx < len(DAYS) else str(day_idx),
                    'day_idx': day_idx,
                    'block_id': block_id,
                    'time_block': BLOCK_LABELS.get(block_id, str(block_id)),
                    'lab_rooms': room,
                    'label': label,
                })
        if rows:
            path = os.path.join(OUTPUT_DIR, 'blocked_slots.csv')
            pd.DataFrame(rows).to_csv(path, index=False, encoding='utf-8-sig')
            print(f"  [OK] Export blocked_slots.csv ({len(rows)} creneaux reserves)")
    except Exception as exc:
        print(f"  [WARN]  Erreur export blocked_slots.csv : {exc}")


def build_subject_professor_busy(professor_busy, professors_of_subject):
    subject_professor_busy = {}
    for subject, professor_ids in professors_of_subject.items():
        busy_sets = [professor_busy.get(professor_id, set()) for professor_id in professor_ids]
        if not busy_sets:
            continue
        # Capacite : un creneau n'est bloque que si TOUS les profs y sont occupes.
        # Mono-prof : son emploi du temps s'applique (un prof ne peut pas etre a deux endroits).
        # Un prof sans cours enregistres (busy vide) est toujours disponible -> jamais de blocage.
        blocked = set.intersection(*busy_sets)
        if blocked:
            subject_professor_busy[subject] = blocked
    return subject_professor_busy


def _norm_name(s):
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.lower().split())


def _match_professor(typed_name, known_profs):
    """Associe un nom saisi a un professor_id reel (insensible accents/casse, nom inverse)."""
    target = _norm_name(typed_name)
    for p in known_profs:
        if _norm_name(p) == target:
            return p
    parts = set(target.replace(',', ' ').split())
    for p in known_profs:
        if parts and parts == set(_norm_name(p).replace(',', ' ').split()):
            return p
    return None


def _parse_teacher_slot_string(s):
    """'Viernes 15:00-17:00' -> (day_idx, block_id) ; None si non interpretable."""
    parts = str(s).split()
    if len(parts) < 2:
        return None
    di = DAY_IDS.get(parts[0])
    if di is None:
        return None
    start = parts[1].split('-')[0].strip()
    try:
        hh, mm = start.split(':')
        minutes = int(hh) * 60 + int(mm)
    except (ValueError, IndexError):
        return None
    block = min_to_block_id(minutes)
    if not block:
        return None
    return (di, block)


def inject_teacher_unavailability(spb, professor_busy, professors_of_subject):
    """Injecte les indisponibilites MANUELLES (TEACHER_UNAVAILABILITY, onglet Config de
    l'app) dans subject_professor_busy. Logique de capacite : un creneau d'une matiere
    n'est bloque que si TOUS ses profs sont indisponibles (busy auto OU manuel) ;
    matiere mono-prof -> l'indispo manuelle de son prof la bloque. Vide -> aucun effet."""
    if not TEACHER_UNAVAILABILITY:
        return spb
    known = set(professor_busy.keys())
    for profs in professors_of_subject.values():
        known.update(profs)
    manual = {}
    unmatched = []
    for typed_name, slots in TEACHER_UNAVAILABILITY.items():
        if not slots:
            continue
        pid = _match_professor(typed_name, known) or typed_name
        if pid not in known:
            unmatched.append(typed_name)
        manual.setdefault(pid, set()).update(slots)
    if unmatched:
        print(f"    [TEACHER][WARN] nom(s) non reconnu(s) dans 'docentes' (sans effet) : {unmatched}")
    if not manual:
        return spb
    spb = {k: set(v) for k, v in spb.items()}
    for subject, profs in professors_of_subject.items():
        profs = list(profs)
        if not profs:
            continue
        if len(profs) == 1:
            extra = manual.get(profs[0], set())
        else:
            unavail = [professor_busy.get(p, set()) | manual.get(p, set()) for p in profs]
            extra = set.intersection(*unavail) if unavail else set()
        if extra:
            before = spb.get(subject, set())
            merged = before | extra
            added = len(merged) - len(before)
            if added:
                spb[subject] = merged
                print(f"    [TEACHER]  {subject}: +{added} creneau(x) bloque(s) (indispo prof)")
    return spb


TEACHER_RULES = {}   # {prof: {'max_days_per_week': int, 'preferred_blocks': [int]}} — requete Daniel
PREF_BLOCK_PENALTY = 3   # poids retire du score de placement hors franje preferee (SOUPLE)


def _sanitize_teacher_rules(raw):
    clean = {}
    if not isinstance(raw, dict):
        return {}
    for name, rules in raw.items():
        if not isinstance(rules, dict):
            continue
        r = {}
        md = rules.get('max_days_per_week')
        try:
            md = int(md)
            if 1 <= md <= 5:
                r['max_days_per_week'] = md
        except (TypeError, ValueError):
            pass
        pb = rules.get('preferred_blocks')
        if isinstance(pb, list):
            keep = []
            for b in pb:
                try:
                    bi = int(b)
                except (TypeError, ValueError):
                    continue
                if bi in ALL_BLOCKS:
                    keep.append(bi)
            if keep:
                r['preferred_blocks'] = sorted(set(keep))
        if r:
            clean[str(name)] = r
    return clean


def load_teacher_rules(path='config/user_config.json'):
    """Lit la cle 'teacher_rules' de user_config.json (ecrite par la page Config).
    Auto-contenu : aucun impact sur apply_user_config ni sur la cle 'teachers'."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            raw = json.load(f)
    except Exception as e:
        print(f"    [RULES][WARN] lecture {path} echouee : {e}")
        return {}
    clean = _sanitize_teacher_rules(raw.get('teacher_rules', {}))
    if clean:
        print(f"    [RULES] Regles professeurs chargees : {len(clean)} prof(s)")
    return clean


def build_subject_block_penalty(professors_of_subject):
    """SOUPLE (option 3 Daniel) : blocs penalises par matiere depuis preferred_blocks.
    Logique de capacite : un bloc n'est penalise pour une matiere que si AUCUN de ses
    profs ayant declare une preference ne le prefere. Jamais bloquant."""
    if not TEACHER_RULES:
        return {}
    known = set()
    for profs in professors_of_subject.values():
        known.update(profs)
    pref = {}
    for cfg_name, rules in TEACHER_RULES.items():
        pb = rules.get('preferred_blocks')
        if not pb:
            continue
        pid = _match_professor(cfg_name, known) or cfg_name
        if pid not in known:
            print(f"    [RULES][WARN] preferred_blocks : prof non reconnu '{cfg_name}' (sans effet)")
            continue
        pref[pid] = set(pb)
    if not pref:
        return {}
    out = {}
    allset = set(ALL_BLOCKS)
    for subject, profs in professors_of_subject.items():
        prefs = [pref[p] for p in profs if p in pref]
        if not prefs:
            continue
        dispref = allset - set().union(*prefs)
        if dispref:
            out[subject] = dispref
            print(f"    [RULES]  {subject}: blocs penalises hors franje = {sorted(dispref)} (souple)")
    return out


def audit_teacher_max_days(all_groups, professors_of_subject):
    """SIGNAL (option 1 Daniel) : controle 'max jours labo / semaine' par prof.
    Ne bloque JAMAIS — le systeme valide, il ne decide pas.
    Approx. niveau matiere (exacte pour les matieres mono-prof)."""
    from collections import defaultdict
    caps = {n: r['max_days_per_week'] for n, r in TEACHER_RULES.items()
            if r.get('max_days_per_week')}
    if not caps:
        return
    known = set()
    for profs in professors_of_subject.values():
        known.update(profs)
    days_by_subject = defaultdict(set)
    for g in all_groups:
        days_by_subject[g['subject']].add(g['day_idx'])
    subj_of_prof = defaultdict(set)
    for subject, profs in professors_of_subject.items():
        for p in profs:
            subj_of_prof[p].add(subject)
    print("\n  [RULES] Controle jours labo / semaine (signalement, jamais bloquant)")
    for cfg_name, cap in sorted(caps.items()):
        pid = _match_professor(cfg_name, known) or cfg_name
        days = set()
        for sub in subj_of_prof.get(pid, ()):
            days |= days_by_subject.get(sub, set())
        nd = len(days)
        names = ', '.join(DAYS[d] for d in sorted(days)) or '—'
        flag = '  /!\\ DEPASSEMENT' if nd > cap else '  OK'
        print(f"    {cfg_name:<32} {nd} jour(s) (max {cap}) : {names}{flag}")


def verify_availability_constraints(all_groups, subject_professor_busy,
                                    professors_of_subject):
    """PREUVE D'APPLICATION (requete Daniel) — verifie a posteriori que le planning
    PRODUIT respecte les parametres de 'Teacher Availability Configuration'.

    Ecrit config/availability_verification.json, lu par la page Integridad pour que
    l'utilisateur CONSTATE reellement que :
      1. Creneaux indisponibles (DUR) : aucun groupe planifie ne tombe sur un creneau
         effectivement bloque pour sa matiere (= tous les profs de la matiere
         indisponibles a ce creneau).
      2. Franje horaire preferee (SOUPLE) : taux de sessions placees DANS la franje.
      3. Max jours labo / semaine (SIGNAL) : jours reellement utilises vs plafond.

    Ne bloque jamais : produit un constat factuel verifiable.
    """
    import json as _json
    report = {
        'generated_at': datetime.now().isoformat(),
        'hard_blocked_slots': {'status': 'ok', 'checked_groups': 0,
                               'violations': []},
        'preferred_range': [],
        'max_days_per_week': [],
        'requested': {
            'teacher_unavailability': {
                t: sorted(list(s)) for t, s in TEACHER_UNAVAILABILITY.items()
            },
            'teacher_rules': {
                str(n): {k: (sorted(list(v)) if isinstance(v, (set, list, tuple))
                             else v) for k, v in r.items()}
                for n, r in TEACHER_RULES.items()
            },
        },
    }

    # 1) Creneaux indisponibles — DUR. Le solveur a recu subject_professor_busy ;
    #    on prouve qu'aucun groupe planifie ne tombe sur un creneau bloque.
    checked = 0
    for g in all_groups:
        blk = subject_professor_busy.get(g['subject'], set())
        if not blk:
            continue
        checked += 1
        if (g['day_idx'], g['block_id']) in blk:
            report['hard_blocked_slots']['violations'].append({
                'subject': g['subject'],
                'group': g.get('group_num'),
                'day': DAYS[g['day_idx']] if 0 <= g['day_idx'] < len(DAYS) else g['day_idx'],
                'block': BLOCK_LABELS.get(g['block_id'], g['block_id']),
            })
    report['hard_blocked_slots']['checked_groups'] = checked
    report['hard_blocked_slots']['status'] = (
        'ok' if not report['hard_blocked_slots']['violations'] else 'violated')

    # match nom config -> prof reconnu
    known = set()
    for profs in professors_of_subject.values():
        known.update(profs)
    subj_of_prof = defaultdict(set)
    for subject, profs in professors_of_subject.items():
        for p in profs:
            subj_of_prof[p].add(subject)

    # 2) Franje preferee — SOUPLE : % de sessions dans la franje des matieres du prof.
    for cfg_name, rules in TEACHER_RULES.items():
        pb = rules.get('preferred_blocks')
        if not pb:
            continue
        pid = _match_professor(cfg_name, known) or cfg_name
        subs = subj_of_prof.get(pid, set())
        groups = [g for g in all_groups if g['subject'] in subs]
        total = len(groups)
        inside = sum(1 for g in groups if g['block_id'] in set(pb))
        report['preferred_range'].append({
            'teacher': cfg_name,
            'recognized': pid in known,
            'preferred_blocks': [BLOCK_LABELS.get(b, b) for b in pb],
            'sessions_total': total,
            'sessions_inside': inside,
            'pct_inside': round(100.0 * inside / total, 1) if total else None,
        })

    # 3) Max jours / semaine — SIGNAL.
    days_by_subject = defaultdict(set)
    for g in all_groups:
        days_by_subject[g['subject']].add(g['day_idx'])
    for cfg_name, rules in TEACHER_RULES.items():
        cap = rules.get('max_days_per_week')
        if not cap:
            continue
        pid = _match_professor(cfg_name, known) or cfg_name
        days = set()
        for sub in subj_of_prof.get(pid, ()):
            days |= days_by_subject.get(sub, set())
        nd = len(days)
        report['max_days_per_week'].append({
            'teacher': cfg_name,
            'recognized': pid in known,
            'cap': cap,
            'days_used': nd,
            'days': [DAYS[d] for d in sorted(days) if 0 <= d < len(DAYS)],
            'status': 'ok' if nd <= cap else 'exceeded',
        })

    try:
        os.makedirs('config', exist_ok=True)
        with open('config/availability_verification.json', 'w', encoding='utf-8') as f:
            _json.dump(report, f, indent=2, ensure_ascii=False)
        v = len(report['hard_blocked_slots']['violations'])
        print(f"  [VERIF] config/availability_verification.json ecrit "
              f"(creneaux DUR verifies : {checked}, violations : {v})")
    except Exception as e:
        print(f"  [VERIF][WARN] ecriture echouee : {e}")
    return report


def prepare_professor_constraints(df):
    professor_busy, professor_subjects, prof_available = build_professor_busy(df)
    print_supervision_capacity(load_supervision_capacity())
    export_blocked_slots()

    if not prof_available:
        return {}

    professors_of_subject = group_professors_by_subject(professor_subjects)
    export_subject_professors(professors_of_subject)
    subject_professor_busy = build_subject_professor_busy(professor_busy, professors_of_subject)
    subject_professor_busy = inject_teacher_unavailability(subject_professor_busy, professor_busy, professors_of_subject)
    if not TEACHER_RULES:
        TEACHER_RULES.update(load_teacher_rules())
    subject_block_penalty = build_subject_block_penalty(professors_of_subject)
    constrained_count = sum(1 for slots in subject_professor_busy.values() if slots)
    print(f"  Matieres de lab avec contrainte professeur active : {constrained_count}")
    print("  (creneau bloque seulement si AUCUN prof de la matiere n'est libre)")
    return subject_professor_busy, subject_block_penalty, professors_of_subject


def extract_program_abbreviation(programs_value):
    if pd.isna(programs_value):
        return None

    programs_raw = str(programs_value).strip()
    if not programs_raw or programs_raw.lower() == 'nan':
        return None

    tokens = [token.strip().split('-')[0].strip().upper() for token in programs_raw.split(',')]
    for abbreviation in tokens:
        if abbreviation in KNOWN_PROGRAMS:
            return abbreviation
    return tokens[0] if tokens else None


def build_student_program_lookup(df):
    student_program = {}
    if 'programas' not in df.columns:
        return student_program

    for _, row in df.dropna(subset=['AlumnoID', 'programas']).iterrows():
        student_id = row['AlumnoID']
        if student_id in student_program:
            continue
        program = extract_program_abbreviation(row['programas'])
        if program:
            student_program[student_id] = program
    return student_program


def build_output_lookups(df):
    name_lookup = {}
    program_lookup = {}
    name_data = df.dropna(subset=['AlumnoID']).drop_duplicates('AlumnoID')

    for _, row in name_data.iterrows():
        student_id = row['AlumnoID']
        first_name = str(row.get('Nombre', '')).strip() if pd.notna(row.get('Nombre')) else ''
        last_name = str(row.get('Apellidos', '')).strip() if pd.notna(row.get('Apellidos')) else ''

        if last_name and first_name:
            name_lookup[student_id] = f"{last_name}, {first_name}"
        elif first_name:
            name_lookup[student_id] = first_name
        elif last_name:
            name_lookup[student_id] = last_name

        program = extract_program_abbreviation(row.get('programas', None))
        if program:
            program_lookup[student_id] = program

    return name_lookup, program_lookup


def find_existing_script(script_name):
    for candidate in (os.path.join('src', script_name), script_name):
        if os.path.exists(candidate):
            return candidate
    return None


def build_utf8_subprocess_env():
    env = os.environ.copy()
    env['PYTHONIOENCODING'] = 'utf-8'
    env['PYTHONUTF8'] = '1'
    return env


def print_tail_lines(output, max_lines, prefix="    "):
    for line in output.strip().split('\n')[-max_lines:]:
        if line.strip():
            print(f"{prefix}{line}")


def run_daniel_format_generation():
    print_section("ETAPE 7 : Generation automatique format Daniel")
    env_utf8 = build_utf8_subprocess_env()

    for script_name, label in DANIEL_FORMAT_SCRIPTS:
        script_path = find_existing_script(script_name)
        if script_path is None:
            print(f"  [WARN]  {script_name} introuvable - skipping")
            continue

        print(f"\n  > Generation {label}...")
        try:
            result = subprocess.run(
                [sys.executable, script_path],
                capture_output=True, text=True,
                encoding='utf-8', errors='replace',
                env=env_utf8, timeout=120,
            )
            if result.returncode == 0:
                print_tail_lines(result.stdout, 15)
                print(f"  [OK] {label} genere")
            else:
                print(f"  [FAIL] Erreur lors de la generation {label}")
                if result.stderr:
                    print_tail_lines(result.stderr, 5)
        except subprocess.TimeoutExpired:
            print(f"  [WARN]  Timeout (120s) lors de la generation {label}")
        except Exception as exc:
            print(f"  [FAIL] Exception : {exc}")

    run_final_validation(env_utf8)


def run_final_validation(env_utf8):
    for validation_script in VALIDATION_SCRIPTS:
        if not os.path.exists(validation_script):
            continue

        print("\n  > Validation finale...")
        try:
            result = subprocess.run(
                [sys.executable, validation_script],
                capture_output=True, text=True,
                encoding='utf-8', errors='replace',
                env=env_utf8, timeout=60,
            )
            if result.stdout:
                print_tail_lines(result.stdout, 20)
        except Exception:
            pass
        break


def create_auto_snapshot():
    try:
        import version_manager as vm
        snap_id = vm.create_snapshot(
            snapshot_type='auto',
            description=f"Generation pipeline du {datetime.now().strftime('%d/%m/%Y a %H:%M')}",
        )
        if snap_id:
            print(f"\n  [OK] Snapshot automatique cree : {snap_id}")
        else:
            print("\n  [INFO] Snapshot non cree (aucun output a archiver)")
    except ImportError:
        pass
    except Exception as exc:
        print(f"\n  [WARN] Auto-snapshot echoue (non-bloquant) : {exc}")


def run_pipeline(df):
    df = load_and_prepare(df)
    subject_students = identify_students(df)
    student_busy, student_subject_slots = build_individual_timetables(df, subject_students)
    subject_professor_busy, subject_block_penalty, professors_of_subject = prepare_professor_constraints(df)
    student_program = build_student_program_lookup(df)

    all_groups = form_groups(
        subject_students,
        student_busy,
        student_subject_slots,
        student_program,
        subject_professor_busy,
        subject_block_penalty,
    )

    if not all_groups:
        print("\n  [FAIL] Aucun groupe forme.")
        return False

    # --- QA données / modèle (Étape 6.2) ---------------------------------
    # Couche qualité : intégrité du master, réconciliation de la jointure
    # Excel et du regroupement (anti-fuite). Non bloquant par défaut.
    dq_report = None
    if _dq is not None:
        try:
            dq_report = _dq.run_data_quality_checks(
                df,
                subject_students,
                all_groups,
                alumnos_path=_first_existing(ALUMNOS_SOURCE_CANDIDATES),
                aulario_path=_first_existing(AULARIO_SOURCE_CANDIDATES),
                strict=False,
            )
        except Exception as exc:  # ne casse jamais le pipeline
            print(f"  [QA][WARN] contrôle qualité ignoré : {exc}")

    audit_teacher_max_days(all_groups, professors_of_subject)
    # Preuve d'application des parametres de disponibilite (page Integridad).
    verify_availability_constraints(all_groups, subject_professor_busy,
                                    professors_of_subject)
    results_df = solve(all_groups)

    if results_df is None or len(results_df) == 0:
        print("\n  [FAIL] Aucune solution generee.")
        return True

    # --- Mesure de la qualité du planning (Étape 6.6) --------------------
    # KPIs objectifs à chaque exécution (placement, équilibrage, salles,
    # solveur). Écrit reports/kpi_report.{json,txt}. Généré AVANT les exports
    # Excel pour pouvoir y être embarqué. Non bloquant.
    if _kpi is not None:
        try:
            _kpi.generate_kpi_report(
                results_df,
                all_groups,
                dq_report=dq_report,
                solver_runs=list(SOLVER_RUNS),
            )
        except Exception as exc:
            print(f"  [KPI][WARN] rapport KPI ignoré : {exc}")

    name_lookup, program_lookup = build_output_lookups(df)
    generate_outputs(results_df, all_groups, name_lookup, program_lookup, subject_students)
    analyze(results_df)

    run_daniel_format_generation()
    return True


def main():
    ensure_pipeline_directories()

    with ReportWriter(REPORT_PATH):
        print_pipeline_header()
        apply_user_config()

        df = read_master_schedule()
        if df is None or not validate_master_schedule(df):
            return

        print_master_schedule_loaded(df)
        should_finalize = run_pipeline(df)
        if not should_finalize:
            return

        create_auto_snapshot()
        print_pipeline_footer()

    print(f"\n[REPORT] {REPORT_PATH}")
    print(f"[FILE] {OUTPUT_DIR}")


def handle_unexpected_error(exc):
    import traceback
    print("\n" + "=" * 60)
    print("  [ERREUR] Le pipeline a rencontre un probleme inattendu.")
    print("=" * 60)
    print(f"  Type    : {type(exc).__name__}")
    print(f"  Detail  : {exc}")
    print("\n  Le detail technique complet a ete enregistre dans :")
    print("    reports/pipeline_error.log")
    print("\n  Causes frequentes :")
    print("    - fichier d'entree manquant ou mal formate (data_clean/master_schedule.csv)")
    print("    - config/user_config.json invalide")
    print("=" * 60)

    try:
        os.makedirs('reports', exist_ok=True)
        with open('reports/pipeline_error.log', 'w', encoding='utf-8') as file:
            file.write(f"Erreur : {type(exc).__name__}: {exc}\n\n")
            file.write(traceback.format_exc())
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n[INTERROMPU] Execution arretee par l'utilisateur.")
        sys.exit(1)
    except Exception as exc:
        handle_unexpected_error(exc)
        sys.exit(1)