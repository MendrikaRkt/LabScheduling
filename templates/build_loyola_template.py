"""
build_loyola_template.py — regenerate the Loyola-branded Excel template.

The rendered artefact (``loyola_schedule_template.xlsx``) is a small workbook
with Loyola branding and ``{{TOKEN}}`` placeholders consumed by
:func:`export_manager.render_template`. Because .xlsx is a binary container we
keep this generator in the repo so the template can be rebuilt deterministically
(and reviewed as plain text) rather than committing an opaque binary blob only.

Run::

    python templates/build_loyola_template.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY = "FF003366"
GOLD = "FFFFCC00"
WHITE = "FFFFFFFF"
BLACK = "FF000000"
LIGHT = "FFF2F5FA"

_THIN = Side(style="thin", color="FFBFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def build() -> Workbook:
    """Build and return the Loyola template workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Title band.
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1,
                    value="Universidad Loyola - {{TITLE}}")
    title.font = Font(bold=True, size=16, color=WHITE)
    title.fill = _fill(NAVY)
    title.alignment = _CENTER
    ws.row_dimensions[1].height = 30

    # Meta band (gold accent labels + placeholder values).
    meta = [
        ("Semester", "{{SEMESTER}}"),
        ("Subject", "{{SUBJECT}}"),
        ("Group", "{{GROUP}}"),
        ("Professor", "{{PROFESSOR}}"),
        ("Room", "{{ROOM}}"),
        ("Day / Block", "{{DAY}} {{BLOCK}}"),
        ("Generated", "{{DATE}}"),
    ]
    r = 3
    for label, value in meta:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, color=BLACK)
        lc.fill = _fill(GOLD)
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=value)
        vc.border = _BORDER
        r += 1

    # Sessions table header (blank rows to be filled by the user / a later
    # data-merge step; the template only establishes branding + structure).
    header_row = r + 1
    headers = ["Session", "Week", "Day", "Time Block", "Room", "Students"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = _CENTER
        cell.border = _BORDER

    for i in range(1, 6):
        rr = header_row + i
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=rr, column=c, value="")
            cell.border = _BORDER
            if i % 2 == 1:
                cell.fill = _fill(LIGHT)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for col in "CDEF":
        ws.column_dimensions[col].width = 14

    ws.freeze_panes = f"A{header_row + 1}"
    return wb


def main() -> None:
    out = os.path.join(os.path.dirname(__file__), "loyola_schedule_template.xlsx")
    build().save(out)
    print(f"[SAVED] {out}")


if __name__ == "__main__":
    main()
