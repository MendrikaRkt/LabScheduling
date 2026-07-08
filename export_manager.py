"""
export_manager.py — Phase 3 export format & template manager.

WHY THIS EXISTS
---------------
Phase 3 introduces several *flavours* of Excel export plus a lightweight
template system. This module is the single entry point the UI (and any script)
calls to:

    * choose a format ('standard' | 'enhanced' | 'summary' | 'detailed') and
      translate it into the right builder + :class:`ExportOptions`;
    * load export preferences from ``config/export_preferences.yaml``;
    * discover, load, validate and render workbook *templates* from the
      ``templates/`` directory, substituting ``{{GROUP}}`` / ``{{PROFESSOR}}``
      / ``{{ROOM}}`` (and friends) placeholders.

It never modifies the validated exporters; 'standard' simply delegates to the
legacy :mod:`excel_export` so backward compatibility is preserved.

PUBLIC API
----------
    FORMAT_PRESETS: dict[str, ExportOptions]
    load_preferences() -> dict
    resolve_options(format_type, overrides=None) -> ExportOptions
    export_with_format(format_type, *, semester=None, out_path=None,
                       color_scheme=None) -> dict
    list_templates() -> list[str]
    load_template(name) -> openpyxl.Workbook
    extract_placeholders(wb) -> set[str]
    validate_template(name_or_wb) -> dict
    render_template(name, context, *, out_path=None) -> dict
"""

from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Union

import app_paths
import excel_export_enhanced as enhanced
from excel_export_enhanced import ExportOptions

# Placeholder syntax: {{TOKEN}} (upper snake, digits allowed).
_PLACEHOLDER_RE = re.compile(r"\{\{\s*([A-Z0-9_]+)\s*\}\}")

# Recognised template variables (documented set). Unknown placeholders are
# reported by validate_template but still substitutable via an explicit context.
KNOWN_PLACEHOLDERS: Set[str] = {
    "GROUP", "PROFESSOR", "ROOM", "SUBJECT", "SEMESTER",
    "DAY", "BLOCK", "DATE", "TITLE",
}

# ────────────────────────────────────────────────────────────
# Format presets — each maps to an ExportOptions profile.
# ────────────────────────────────────────────────────────────
FORMAT_PRESETS: Dict[str, ExportOptions] = {
    # 'standard' is handled specially (delegates to the legacy exporter); the
    # options here are only used if someone builds it via the enhanced engine.
    "standard": ExportOptions(
        color_coded_groups=True, legend=False,
        room_utilization=False, professor_workload=False,
        student_placement=False, time_slot_analysis=False,
        quality_metrics=False, conditional_formatting=False,
        data_validation=False, named_ranges=False, cell_comments=False,
    ),
    "summary": ExportOptions(
        color_coded_groups=True, legend=True,
        room_utilization=False, professor_workload=False,
        student_placement=True, time_slot_analysis=False,
        quality_metrics=True, conditional_formatting=True,
        data_validation=False, named_ranges=False, cell_comments=False,
    ),
    "enhanced": ExportOptions(),  # all features on (defaults)
    "detailed": ExportOptions(
        protect_sheets=True,  # everything on + protected sheets
    ),
}

VALID_FORMATS = tuple(FORMAT_PRESETS.keys())


# ════════════════════════════════════════════════════════════
# Preferences
# ════════════════════════════════════════════════════════════
_DEFAULT_PREFERENCES: Dict[str, Any] = {
    "default_format": "enhanced",
    "default_color_scheme": "loyola",
    "formats": {
        name: {k: getattr(opt, k) for k in ExportOptions.__dataclass_fields__}  # type: ignore[attr-defined]
        for name, opt in FORMAT_PRESETS.items()
    },
    "templates_dir": "templates",
}


def load_preferences() -> Dict[str, Any]:
    """Load ``config/export_preferences.yaml`` merged over sane defaults.

    Missing file or PyYAML errors fall back to the built-in defaults, so the
    manager always returns a usable configuration.
    """
    prefs = dict(_DEFAULT_PREFERENCES)
    path = app_paths.resolve_existing("config/export_preferences.yaml")
    if not path:
        return prefs
    try:
        import yaml

        with open(path, "r", encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}
        if isinstance(data, dict):
            prefs.update({k: v for k, v in data.items() if v is not None})
    except Exception:
        pass
    return prefs


# ════════════════════════════════════════════════════════════
# Format resolution
# ════════════════════════════════════════════════════════════
def resolve_options(
    format_type: str,
    overrides: Optional[Dict[str, Any]] = None,
) -> ExportOptions:
    """Return the :class:`ExportOptions` for a format, applying optional overrides.

    Preferences from ``export_preferences.yaml`` (if they define the format's
    flags) take precedence over the built-in preset; explicit ``overrides`` win
    over everything.
    """
    fmt = (format_type or "enhanced").lower()
    if fmt not in FORMAT_PRESETS:
        raise ValueError(
            f"Unknown format '{format_type}'. Valid: {', '.join(VALID_FORMATS)}"
        )

    base = FORMAT_PRESETS[fmt]
    merged = {k: getattr(base, k) for k in ExportOptions.__dataclass_fields__}  # type: ignore[attr-defined]

    prefs = load_preferences()
    pref_fmt = (prefs.get("formats") or {}).get(fmt)
    if isinstance(pref_fmt, dict):
        for k, v in pref_fmt.items():
            if k in merged:
                merged[k] = bool(v)

    if overrides:
        for k, v in overrides.items():
            if k in merged:
                merged[k] = bool(v)

    return ExportOptions(**merged)


def export_with_format(
    format_type: str,
    *,
    semester: Optional[int] = None,
    out_path: Optional[str] = None,
    color_scheme: Optional[str] = None,
    overrides: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Produce an export in the requested format.

    'standard' delegates to the validated :mod:`excel_export` (backward compat,
    Daniel-format workbooks). All other formats use the enhanced engine.

    Returns a result dict ``{ok, file(s), format, error}``.
    """
    fmt = (format_type or "enhanced").lower()
    if fmt not in FORMAT_PRESETS:
        return {"ok": False, "file": None, "format": fmt,
                "error": f"Unknown format '{format_type}'."}

    prefs = load_preferences()
    scheme = color_scheme or prefs.get("default_color_scheme", "loyola")

    if fmt == "standard":
        return _export_standard(semester)

    opts = resolve_options(fmt, overrides)
    result = enhanced.export_enhanced(
        semester=semester, out_path=out_path, options=opts, color_scheme=scheme,
    )
    result["format"] = fmt
    return result


def _export_standard(semester: Optional[int]) -> Dict[str, Any]:
    """Delegate to the legacy exporter, normalising the result shape."""
    import excel_export

    if semester is None:
        res = excel_export.generate_all()
    else:
        res = excel_export.generate_semester(semester)
    return {
        "ok": res.get("ok", False),
        "file": (res.get("files") or [None])[0],
        "files": res.get("files", []),
        "format": "standard",
        "error": res.get("error"),
        "log": res.get("log", ""),
    }


# ════════════════════════════════════════════════════════════
# Template engine
# ════════════════════════════════════════════════════════════
def _templates_dir() -> str:
    """Absolute path to the templates directory (workspace-first)."""
    existing = app_paths.resolve_existing("templates")
    return existing or app_paths.workspace_path("templates")


def list_templates() -> List[str]:
    """Return template file names (``*.xlsx``) available in ``templates/``."""
    root = _templates_dir()
    if not os.path.isdir(root):
        return []
    return sorted(
        f for f in os.listdir(root)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    )


def _resolve_template_path(name: str) -> Optional[str]:
    """Resolve a template name to an absolute path, if it exists."""
    if os.path.isabs(name) and os.path.isfile(name):
        return name
    candidate = os.path.join(_templates_dir(), name)
    if os.path.isfile(candidate):
        return candidate
    # allow passing without extension
    if not name.lower().endswith(".xlsx"):
        return _resolve_template_path(name + ".xlsx")
    return None


def load_template(name: str):
    """Load a template workbook by file name.

    Raises
    ------
    FileNotFoundError
        If no matching template exists.
    """
    from openpyxl import load_workbook

    path = _resolve_template_path(name)
    if not path:
        raise FileNotFoundError(f"Template not found: {name}")
    return load_workbook(path)


def extract_placeholders(wb) -> Set[str]:
    """Return the set of ``{{TOKEN}}`` placeholders present in a workbook."""
    found: Set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    found.update(_PLACEHOLDER_RE.findall(cell.value))
    return found


def validate_template(name_or_wb: Union[str, Any]) -> Dict[str, Any]:
    """Validate a template, returning a structured report.

    The report contains::

        {ok, placeholders, known, unknown, sheet_count, error}

    ``ok`` is False only for hard errors (unreadable file). Unknown
    placeholders are reported but do not, by themselves, fail validation.
    """
    try:
        wb = name_or_wb if not isinstance(name_or_wb, str) else load_template(name_or_wb)
    except FileNotFoundError as exc:
        return {"ok": False, "placeholders": [], "known": [], "unknown": [],
                "sheet_count": 0, "error": str(exc)}
    except Exception as exc:  # pragma: no cover - defensive
        return {"ok": False, "placeholders": [], "known": [], "unknown": [],
                "sheet_count": 0, "error": f"Cannot read template: {exc}"}

    placeholders = extract_placeholders(wb)
    unknown = sorted(placeholders - KNOWN_PLACEHOLDERS)
    known = sorted(placeholders & KNOWN_PLACEHOLDERS)
    return {
        "ok": True,
        "placeholders": sorted(placeholders),
        "known": known,
        "unknown": unknown,
        "sheet_count": len(wb.worksheets),
        "error": None,
    }


def render_template(
    name: str,
    context: Dict[str, Any],
    *,
    out_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Fill a template's ``{{TOKEN}}`` placeholders from ``context`` and save it.

    Any placeholder without a matching context key is left untouched (so a
    partially filled template stays diagnosable). ``DATE`` defaults to today
    when not supplied.

    Returns ``{ok, file, substituted, missing, error}``.
    """
    try:
        wb = load_template(name)
    except FileNotFoundError as exc:
        return {"ok": False, "file": None, "substituted": [], "missing": [],
                "error": str(exc)}

    ctx = {str(k).upper(): ("" if v is None else str(v)) for k, v in (context or {}).items()}
    ctx.setdefault("DATE", datetime.now().strftime("%Y-%m-%d"))

    substituted: Set[str] = set()
    missing: Set[str] = set()

    def _replace(match: "re.Match[str]") -> str:
        token = match.group(1)
        if token in ctx:
            substituted.add(token)
            return ctx[token]
        missing.add(token)
        return match.group(0)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = _PLACEHOLDER_RE.sub(_replace, cell.value)

    if out_path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(os.path.basename(name))[0]
        out_path = app_paths.workspace_path(
            "outputs", "optimization", f"{base}_rendered_{stamp}.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

    return {
        "ok": True,
        "file": out_path,
        "substituted": sorted(substituted),
        "missing": sorted(missing),
        "error": None,
    }
